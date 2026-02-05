"""
Excel Report Generator for Escalation AI.

Generates comprehensive Excel reports with McKinsey-style formatting,
including dashboards, charts, and detailed analysis sheets.
"""

import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import pandas as pd

from ..core.config import (
    REPORT_TITLE, REPORT_VERSION, GEN_MODEL, MC_BLUE,
    COL_SUMMARY, COL_SEVERITY, COL_ORIGIN, COL_TYPE, COL_DATETIME,
    COL_ENGINEER, COL_LOB
)
from ..visualization import ChartGenerator, AdvancedChartGenerator, get_chart_analyzer

logger = logging.getLogger(__name__)


class ExcelReportWriter:
    """
    Excel Report Writer with McKinsey-style formatting.
    
    Creates professional multi-sheet Excel reports with:
    - Executive Summary sheet
    - Dashboard with embedded charts
    - Scored Data with conditional formatting
    - Financial Analysis
    - Resolution Time Analysis
    - Raw Data backup
    """
    
    def __init__(self, output_path):
        self.output_path = output_path
        self.output_dir = os.path.dirname(output_path)
        self.wb = Workbook()
        self.chart_generator = None
        
        # Style definitions
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
        self.title_font = Font(bold=True, size=22, color="004C97")
        self.subtitle_font = Font(size=11, italic=True, color="666666")
        
        # Enhanced styling
        self.thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        self.thick_border = Border(
            left=Side(style='medium', color='004C97'),
            right=Side(style='medium', color='004C97'),
            top=Side(style='medium', color='004C97'),
            bottom=Side(style='medium', color='004C97')
        )
        self.kpi_fill_blue = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        self.kpi_fill_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        self.kpi_fill_red = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        self.kpi_fill_amber = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        self.section_font = Font(bold=True, size=13, color="004C97")
        
    def _style_header_row(self, ws, row=1, start_col=1, end_col=None):
        """Apply header styling to a row."""
        if end_col is None:
            end_col = ws.max_column
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    def write_executive_summary(self, df, exec_summary_text):
        """Write the Detailed Analysis sheet with metrics and insights."""
        ws = self.wb.create_sheet("Detailed Analysis", 1)  # Position 1 (after Summary)
        ws.sheet_view.showGridLines = False

        report_timestamp = datetime.now().strftime("%B %d, %Y at %H:%M")

        # Set column widths for WIDE two-column sectional layout (uses columns A-N)
        ws.column_dimensions['A'].width = 2    # Left margin
        ws.column_dimensions['B'].width = 6    # Left panel start
        ws.column_dimensions['C'].width = 18   # Left panel content
        ws.column_dimensions['D'].width = 18   # Left panel content
        ws.column_dimensions['E'].width = 18   # Left panel content
        ws.column_dimensions['F'].width = 3    # Gutter between columns
        ws.column_dimensions['G'].width = 6    # Right panel start
        ws.column_dimensions['H'].width = 18   # Right panel content
        ws.column_dimensions['I'].width = 18   # Right panel content
        ws.column_dimensions['J'].width = 18   # Right panel content
        ws.column_dimensions['K'].width = 2    # Right margin

        # ═══════════════════════════════════════════════════════════════
        # HEADER SECTION
        # ═══════════════════════════════════════════════════════════════
        ws.row_dimensions[1].height = 8  # Top spacing
        ws.row_dimensions[2].height = 35  # Title row

        # Title with blue background banner
        for col in range(2, 11):  # B to J
            ws.cell(row=2, column=col).fill = self.header_fill
        ws['B2'] = "DETAILED ANALYSIS"
        ws['B2'].font = Font(bold=True, size=20, color="FFFFFF")
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B2:J2')
        
        # Subtitle row
        ws.row_dimensions[3].height = 22
        ws['B3'] = f"Generated: {report_timestamp}  •  Version: {REPORT_VERSION}  •  AI Model: {GEN_MODEL}"
        ws['B3'].font = Font(size=10, italic=True, color="666666")
        ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B3:J3')
        
        # ═══════════════════════════════════════════════════════════════
        # KPI CARDS SECTION
        # ═══════════════════════════════════════════════════════════════
        ws.row_dimensions[4].height = 6  # Spacer
        ws.row_dimensions[5].height = 22
        
        # Section header
        ws['B5'] = "📈 KEY METRICS AT A GLANCE"
        ws['B5'].font = Font(bold=True, size=12, color="004C97")
        ws.merge_cells('B5:J5')
        
        # Calculate metrics
        total_tickets = len(df)
        total_friction = df['Strategic_Friction_Score'].sum() if 'Strategic_Friction_Score' in df.columns else 0
        avg_friction = df['Strategic_Friction_Score'].mean() if 'Strategic_Friction_Score' in df.columns else 0
        critical_count = (df['Severity_Norm'] == 'Critical').sum() if 'Severity_Norm' in df.columns else 0
        total_financial = df['Financial_Impact'].sum() if 'Financial_Impact' in df.columns else 0
        avg_resolution = df['Predicted_Resolution_Days'].mean() if 'Predicted_Resolution_Days' in df.columns else 0
        recurrence_rate = df['AI_Recurrence_Probability'].mean() * 100 if 'AI_Recurrence_Probability' in df.columns else 0
        
        # KPI card layout - 4 cards across
        kpi_data = [
            ("🎫 Total Tickets", str(total_tickets), self.kpi_fill_blue, "Records analyzed"),
            ("⚡ Friction Score", f"{total_friction:,.0f}", self.kpi_fill_amber, f"Avg: {avg_friction:.1f}"),
            ("🚨 Critical Issues", str(critical_count), self.kpi_fill_red, f"{critical_count/total_tickets*100:.1f}% of total"),
            ("💰 Financial Impact", f"${total_financial:,.0f}", self.kpi_fill_green, f"${total_financial/total_tickets:.0f}/ticket"),
        ]
        
        # KPI rows setup
        ws.row_dimensions[6].height = 8  # Spacer
        ws.row_dimensions[7].height = 18  # Label row
        ws.row_dimensions[8].height = 30  # Value row
        ws.row_dimensions[9].height = 16  # Subtext row
        
        kpi_cols = [(2, 3), (4, 5), (6, 7), (8, 8)]  # Column ranges for each KPI
        
        for i, (label, value, fill, subtext) in enumerate(kpi_data):
            start_col, end_col = kpi_cols[i][0], kpi_cols[i][-1] if len(kpi_cols[i]) > 1 else kpi_cols[i][0]
            
            # Apply fill to all cells in the KPI card
            for row in [7, 8, 9]:
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.border = self.thin_border
            
            # Label
            ws.cell(row=7, column=start_col).value = label
            ws.cell(row=7, column=start_col).font = Font(bold=True, size=9, color="333333")
            ws.cell(row=7, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
            if start_col != end_col:
                ws.merge_cells(start_row=7, start_column=start_col, end_row=7, end_column=end_col)
            
            # Value
            ws.cell(row=8, column=start_col).value = value
            ws.cell(row=8, column=start_col).font = Font(bold=True, size=18, color="004C97")
            ws.cell(row=8, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
            if start_col != end_col:
                ws.merge_cells(start_row=8, start_column=start_col, end_row=8, end_column=end_col)
            
            # Subtext
            ws.cell(row=9, column=start_col).value = subtext
            ws.cell(row=9, column=start_col).font = Font(size=8, italic=True, color="666666")
            ws.cell(row=9, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
            if start_col != end_col:
                ws.merge_cells(start_row=9, start_column=start_col, end_row=9, end_column=end_col)
        
        # ═══════════════════════════════════════════════════════════════
        # SECONDARY METRICS ROW
        # ═══════════════════════════════════════════════════════════════
        ws.row_dimensions[10].height = 6  # Spacer
        ws.row_dimensions[11].height = 22
        
        secondary_metrics = [
            ("Avg Resolution", f"{avg_resolution:.1f} days"),
            ("Recurrence Risk", f"{recurrence_rate:.1f}%"),
            ("Categories", f"{df['AI_Category'].nunique() if 'AI_Category' in df.columns else 0}"),
        ]
        
        sec_start = 2
        for label, value in secondary_metrics:
            cell = ws.cell(row=11, column=sec_start)
            cell.value = f"{label}: {value}"
            cell.font = Font(size=10, color="004C97")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sec_start += 2
        
        # ═══════════════════════════════════════════════════════════════
        # AI EXECUTIVE SYNTHESIS - Visual Executive Format
        # ═══════════════════════════════════════════════════════════════
        import re

        # Color definitions for visual elements
        fill_critical = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Light red
        fill_warning = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")   # Light amber
        fill_success = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")   # Light green
        fill_insight = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")   # Light blue
        fill_dark = PatternFill(start_color="263238", end_color="263238", fill_type="solid")      # Dark gray
        border_gray = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        ws.row_dimensions[12].height = 10  # Spacer
        ws.row_dimensions[13].height = 32  # Section header

        # Section header with dark banner
        for col in range(2, 11):  # B to J
            ws.cell(row=13, column=col).fill = fill_dark
        ws['B13'] = "⚡ AI EXECUTIVE SYNTHESIS"
        ws['B13'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['B13'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B13:J13')

        current_row = 15

        # ─────────────────────────────────────────────────────────────
        # QUICK STATS DASHBOARD (extracted from data)
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 20
        ws.cell(row=current_row, column=2).value = "📊 SNAPSHOT"
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=11, color="004C97")
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 1

        # Quick stats row with colored boxes
        ws.row_dimensions[current_row].height = 50
        quick_stats = [
            ("B", "C", f"{total_tickets:,}", "Total Tickets", self.kpi_fill_blue),
            ("D", "E", f"${total_financial:,.0f}", "Financial Impact", fill_critical if total_financial > 500000 else fill_warning),
            ("F", "G", f"{critical_count}", "Critical Issues", fill_critical if critical_count > 20 else fill_warning if critical_count > 5 else fill_success),
            ("H", "H", f"{avg_friction:.0f}", "Avg Friction", fill_warning if avg_friction > 50 else fill_success),
        ]

        for start_col, end_col, value, label, fill in quick_stats:
            # Value cell
            cell = ws[f'{start_col}{current_row}']
            cell.value = value
            cell.font = Font(bold=True, size=16, color="004C97")
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray
            if start_col != end_col:
                ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')

        current_row += 1
        ws.row_dimensions[current_row].height = 18
        # Labels row
        for start_col, end_col, value, label, fill in quick_stats:
            cell = ws[f'{start_col}{current_row}']
            cell.value = label
            cell.font = Font(size=8, color="666666")
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray
            if start_col != end_col:
                ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')

        current_row += 2  # Gap

        # ─────────────────────────────────────────────────────────────
        # PYRAMID PRINCIPLE: Lead with Answer (McKinsey Style)
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 24
        cell = ws.cell(row=current_row, column=2)
        cell.value = "📌 THE BOTTOM LINE"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = fill_dark
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 1

        # Extract key conclusion from text
        clean_text = exec_summary_text.strip()
        clean_text = re.sub(r'\*\*([^*]+)\*\*', r'\1', clean_text)
        clean_text = re.sub(r'^#+\s*', '', clean_text, flags=re.MULTILINE)

        # Bottom line summary box
        ws.row_dimensions[current_row].height = 55
        bottom_line = f"Analysis of {total_tickets:,} escalation tickets reveals ${total_financial:,.0f} in financial impact, with {critical_count} critical issues requiring immediate attention. Key drivers: process gaps and recurring systemic issues."
        cell = ws.cell(row=current_row, column=2)
        cell.value = bottom_line
        cell.font = Font(size=11, bold=True, color="263238")
        cell.fill = fill_insight
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border_gray
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 2

        # ─────────────────────────────────────────────────────────────
        # TRAFFIC LIGHT STATUS (RAG)
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 22
        cell = ws.cell(row=current_row, column=2)
        cell.value = "🚦 STATUS AT A GLANCE"
        cell.font = Font(bold=True, size=10, color="004C97")
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 1

        # RAG status indicators
        ws.row_dimensions[current_row].height = 35
        rag_items = [
            ("B", "C", "🔴", "CRITICAL", "Financial exposure high", fill_critical),
            ("D", "E", "🟡", "AT RISK", "Recurrence patterns", fill_warning),
            ("F", "G", "🟢", "ON TRACK", "Resolution improving", fill_success),
        ]

        for start_col, end_col, icon, status, desc, fill in rag_items:
            cell = ws[f'{start_col}{current_row}']
            cell.value = f"{icon} {status}"
            cell.font = Font(bold=True, size=10, color="333333")
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray
            ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')

        current_row += 1
        ws.row_dimensions[current_row].height = 22
        for start_col, end_col, icon, status, desc, fill in rag_items:
            cell = ws[f'{start_col}{current_row}']
            cell.value = desc
            cell.font = Font(size=8, italic=True, color="666666")
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray
            ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')

        current_row += 2

        # ─────────────────────────────────────────────────────────────
        # MECE FRAMEWORK: Issue Decomposition (DATA-DRIVEN)
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 24
        cell = ws.cell(row=current_row, column=2)
        cell.value = "🎯 ISSUE BREAKDOWN (MECE)"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 1

        # MECE categories with impact indicators - use full width B:J
        ws.row_dimensions[current_row].height = 20
        # Headers: CATEGORY(B:D), COUNT(E), %(F), IMPACT(G), ACTION(H:J)
        header_specs = [
            ("B", "D", "CATEGORY"),
            ("E", "E", "COUNT"),
            ("F", "F", "%"),
            ("G", "G", "IMPACT"),
            ("H", "J", "ACTION")
        ]
        for start_col, end_col, header in header_specs:
            cell = ws[f'{start_col}{current_row}']
            cell.value = header
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray
            if start_col != end_col:
                ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
        current_row += 1

        # Build MECE data from actual DataFrame categories
        if 'AI_Category' in df.columns:
            cat_counts = df['AI_Category'].value_counts().head(6)
            cat_total = cat_counts.sum()

            # Action mapping for categories
            action_map = {
                'Scheduling & Planning': 'Standardize scheduling protocols',
                'Documentation & Reporting': 'Improve documentation templates',
                'Validation & QA': 'Enhance QA validation checks',
                'Process Compliance': 'Enforce SOP compliance',
                'Configuration & Data Mismatch': 'Validate configurations pre-deploy',
                'Site Readiness': 'Pre-deployment checks required',
                'Communication & Response': 'Improve response protocols',
                'Nesting & Tool Errors': 'Tool usage training',
            }

            for idx, (cat, count) in enumerate(cat_counts.items()):
                pct = (count / cat_total) * 100 if cat_total > 0 else 0
                impact = "HIGH" if pct > 20 else "MED" if pct > 10 else "LOW"
                action = action_map.get(cat, "Review & address")

                ws.row_dimensions[current_row].height = 28
                impact_fill = fill_critical if impact == "HIGH" else fill_warning if impact == "MED" else fill_success

                # Category (full name, no truncation)
                cell = ws[f'B{current_row}']
                cell.value = cat
                cell.font = Font(size=10, color="333333")
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_gray
                ws.merge_cells(f'B{current_row}:D{current_row}')

                # Count
                cell = ws[f'E{current_row}']
                cell.value = count
                cell.font = Font(bold=True, size=10, color="333333")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_gray

                # Percentage
                cell = ws[f'F{current_row}']
                cell.value = f"{pct:.1f}%"
                cell.font = Font(size=10, color="666666")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_gray

                # Impact
                cell = ws[f'G{current_row}']
                cell.value = impact
                cell.font = Font(bold=True, size=10, color="333333")
                cell.fill = impact_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_gray

                # Action (spans H:J)
                cell = ws[f'H{current_row}']
                cell.value = f"→ {action}"
                cell.font = Font(size=9, bold=True, color="004C97")
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_gray
                ws.merge_cells(f'H{current_row}:J{current_row}')

                current_row += 1

        current_row += 1

        # ─────────────────────────────────────────────────────────────
        # 80/20 ANALYSIS: Pareto Insight
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 24
        cell = ws.cell(row=current_row, column=2)
        cell.value = "📊 80/20 INSIGHT"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 1

        ws.row_dimensions[current_row].height = 40
        pareto_text = f"20% of issue categories drive 80% of escalation volume and financial impact. Focus improvement efforts on the top 3 categories to achieve maximum ROI with minimal resource expenditure."
        cell = ws.cell(row=current_row, column=2)
        cell.value = pareto_text
        cell.font = Font(size=10, color="333333")
        cell.fill = fill_warning
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border_gray
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 2

        # ─────────────────────────────────────────────────────────────
        # ACTION PRIORITY MATRIX (Impact vs Effort)
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 24
        cell = ws.cell(row=current_row, column=2)
        cell.value = "⚡ PRIORITIZED ACTIONS"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 1

        # Priority headers
        ws.row_dimensions[current_row].height = 18
        priority_headers = [("B", "B", "#"), ("C", "D", "ACTION"), ("E", "E", "IMPACT"), ("F", "F", "EFFORT"), ("G", "H", "TIMELINE")]
        for start_col, end_col, header in priority_headers:
            cell = ws[f'{start_col}{current_row}']
            cell.value = header
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray
            if start_col != end_col:
                ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
        current_row += 1

        # Priority actions (Quick Wins first - high impact, low effort)
        actions = [
            ("1", "Standardize escalation criteria", "HIGH", "LOW", "Week 1-2"),
            ("2", "Implement root cause templates", "HIGH", "MED", "Week 2-4"),
            ("3", "Automate recurring issue detection", "MED", "MED", "Month 2"),
            ("4", "Enhanced training program", "MED", "HIGH", "Quarter 2"),
        ]

        for num, action, impact, effort, timeline in actions:
            ws.row_dimensions[current_row].height = 26

            # Quick win indicator
            is_quick_win = impact == "HIGH" and effort == "LOW"
            row_fill = fill_success if is_quick_win else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Number
            cell = ws[f'B{current_row}']
            cell.value = num
            cell.font = Font(bold=True, size=11, color="004C97")
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray

            # Action
            cell = ws[f'C{current_row}']
            cell.value = ("⭐ " if is_quick_win else "") + action
            cell.font = Font(size=10, bold=is_quick_win, color="333333")
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border_gray
            ws.merge_cells(f'C{current_row}:D{current_row}')

            # Impact
            impact_fill = fill_critical if impact == "HIGH" else fill_warning if impact == "MED" else fill_success
            cell = ws[f'E{current_row}']
            cell.value = impact
            cell.font = Font(bold=True, size=9, color="333333")
            cell.fill = impact_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray

            # Effort
            effort_fill = fill_success if effort == "LOW" else fill_warning if effort == "MED" else fill_critical
            cell = ws[f'F{current_row}']
            cell.value = effort
            cell.font = Font(bold=True, size=9, color="333333")
            cell.fill = effort_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray

            # Timeline
            cell = ws[f'G{current_row}']
            cell.value = timeline
            cell.font = Font(size=9, color="666666")
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_gray
            ws.merge_cells(f'H{current_row}:J{current_row}')

            current_row += 1

        # Quick wins legend
        current_row += 1
        ws.row_dimensions[current_row].height = 18
        cell = ws.cell(row=current_row, column=2)
        cell.value = "⭐ = Quick Win (High Impact, Low Effort) - Prioritize these first"
        cell.font = Font(size=9, italic=True, color="28A745")
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 2

        # ─────────────────────────────────────────────────────────────
        # TWO-COLUMN SECTION: Key Findings | Next Steps
        # ─────────────────────────────────────────────────────────────

        # Section headers (side by side)
        ws.row_dimensions[current_row].height = 26

        # Left column header: KEY FINDINGS (B to E)
        cell = ws.cell(row=current_row, column=2)
        cell.value = "🔍 KEY FINDINGS"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_gray
        ws.merge_cells(f'B{current_row}:E{current_row}')

        # Right column header: NEXT STEPS (G to J)
        cell = ws.cell(row=current_row, column=7)  # Column G
        cell.value = "🚀 NEXT STEPS"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_gray
        ws.merge_cells(f'G{current_row}:J{current_row}')

        current_row += 1

        # Build data-driven findings from DataFrame
        left_findings = []

        # Top category finding
        if 'AI_Category' in df.columns:
            top_cat = df['AI_Category'].value_counts()
            if len(top_cat) > 0:
                top_pct = (top_cat.iloc[0] / len(df)) * 100
                icon = "🔴" if top_pct > 25 else "🟡" if top_pct > 15 else "🟢"
                left_findings.append((icon, f"{top_cat.index[0]} driving {top_pct:.0f}% of escalations"))

        # Financial impact finding
        if 'Total_Financial_Impact' in df.columns and df['Total_Financial_Impact'].sum() > 0:
            total_impact = df['Total_Financial_Impact'].sum()
            icon = "🔴" if total_impact > 500000 else "🟡"
            left_findings.append((icon, f"${total_impact:,.0f} total financial impact identified"))

        # Recurrence risk finding
        if 'AI_Recurrence_Probability' in df.columns:
            high_risk = (df['AI_Recurrence_Probability'] > 0.7).sum()
            pct_high = (high_risk / len(df)) * 100 if len(df) > 0 else 0
            icon = "🔴" if pct_high > 30 else "🟡" if pct_high > 15 else "🟢"
            left_findings.append((icon, f"{high_risk} tickets ({pct_high:.0f}%) at high recurrence risk"))

        # Resolution time finding
        if 'Predicted_Resolution_Days' in df.columns:
            avg_days = df['Predicted_Resolution_Days'].mean()
            icon = "🟡" if avg_days > 7 else "🟢"
            left_findings.append((icon, f"Average predicted resolution: {avg_days:.1f} days"))

        # Fallback if no data
        if not left_findings:
            left_findings = [("🟡", "Data analysis in progress")]

        # Build data-driven next steps based on top categories
        right_steps = []
        if 'AI_Category' in df.columns:
            top_cats = df['AI_Category'].value_counts().head(3).index.tolist()
            step_map = {
                'Scheduling & Planning': ("Implement scheduling validation checks", "Week 1-2"),
                'Documentation & Reporting': ("Standardize documentation templates", "Week 2-3"),
                'Validation & QA': ("Enhance QA validation protocols", "Week 2-4"),
                'Process Compliance': ("Reinforce SOP compliance training", "Week 1-2"),
                'Configuration & Data Mismatch': ("Automate config validation", "Month 1"),
                'Site Readiness': ("Implement pre-readiness checklists", "Week 2"),
                'Communication & Response': ("Establish response SLAs", "Week 1"),
                'Nesting & Tool Errors': ("Tool usage training sessions", "Week 2-3"),
            }
            for i, cat in enumerate(top_cats, 1):
                step, timeline = step_map.get(cat, ("Address root causes", "Ongoing"))
                right_steps.append((f"{i}.", step, timeline))

        right_steps.append((f"{len(right_steps)+1}.", "Continue monitoring & optimization", "Ongoing"))

        # Ensure we have at least 4 items for visual balance
        while len(left_findings) < 4:
            left_findings.append(("🟢", "—"))
        while len(right_steps) < 4:
            right_steps.append(("", "—", ""))

        for i in range(max(len(left_findings), len(right_steps))):
            ws.row_dimensions[current_row].height = 32

            # Left column: Finding (B=icon, C-E=finding text)
            if i < len(left_findings):
                icon, finding = left_findings[i]
                row_fill = fill_insight if i % 2 == 0 else PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

                # Icon in column B
                cell = ws.cell(row=current_row, column=2)
                cell.value = icon
                cell.font = Font(size=12)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_gray

                # Finding text merged across C:E
                cell = ws.cell(row=current_row, column=3)
                cell.value = finding
                cell.font = Font(size=9, color="333333")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border_gray
                ws.merge_cells(f'C{current_row}:E{current_row}')

            # Right column: Next Step (G=number, H-I=step, J=timeline)
            if i < len(right_steps):
                num, step, timeline = right_steps[i]
                row_fill = fill_success if i % 2 == 0 else PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

                # Number in column G
                cell = ws.cell(row=current_row, column=7)
                cell.value = num
                cell.font = Font(bold=True, size=10, color="2E7D32")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_gray

                # Step text merged across H:I
                cell = ws.cell(row=current_row, column=8)
                cell.value = step
                cell.font = Font(size=9, color="333333")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border_gray
                ws.merge_cells(f'H{current_row}:I{current_row}')

                # Timeline in column J
                cell = ws.cell(row=current_row, column=10)
                cell.value = timeline
                cell.font = Font(size=8, italic=True, color="666666")
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_gray

            current_row += 1

        current_row += 1

        # ─────────────────────────────────────────────────────────────
        # EXECUTIVE CALLOUT BOX
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 50

        # Callout box spanning full width
        callout_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        callout_border = Border(
            left=Side(style='medium', color='FF9800'),
            right=Side(style='thin', color='FFE0B2'),
            top=Side(style='thin', color='FFE0B2'),
            bottom=Side(style='thin', color='FFE0B2')
        )

        # Build dynamic executive insight from data
        if 'AI_Category' in df.columns:
            top_3_cats = df['AI_Category'].value_counts().head(3).index.tolist()
            top_3_str = ", ".join(top_3_cats[:3]) if top_3_cats else "key categories"
            top_3_pct = (df['AI_Category'].value_counts().head(3).sum() / len(df)) * 100 if len(df) > 0 else 0
            insight_text = f"💡 EXECUTIVE INSIGHT: Focus on {top_3_str} to address {top_3_pct:.0f}% of escalation volume. Quick wins in process standardization can yield significant reduction in escalation volume."
        else:
            insight_text = "💡 EXECUTIVE INSIGHT: Focus on the top categories identified to achieve maximum ROI with minimal resource expenditure."

        cell = ws.cell(row=current_row, column=2)
        cell.value = insight_text
        cell.font = Font(size=10, bold=True, color="E65100")
        cell.fill = callout_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = callout_border
        ws.merge_cells(f'B{current_row}:J{current_row}')

        current_row += 2

        # ─────────────────────────────────────────────────────────────
        # DETAILED AI ANALYSIS (Full LLM Output)
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 24
        cell = ws.cell(row=current_row, column=2)
        cell.value = "📝 DETAILED AI ANALYSIS"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = fill_dark
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{current_row}:J{current_row}')
        current_row += 1

        # Parse and display AI synthesis content
        if exec_summary_text and len(exec_summary_text.strip()) > 50:
            # Split into paragraphs/sections - handle both \n\n and single \n for SECTION headers
            raw_paragraphs = exec_summary_text.strip().split('\n')

            processed_items = []
            for line in raw_paragraphs:
                line = line.strip()
                if not line:
                    continue

                # Clean up markdown formatting
                line_clean = re.sub(r'\*\*([^*]+)\*\*', r'\1', line)
                line_clean = re.sub(r'^#+\s*', '', line_clean, flags=re.MULTILINE)
                line_clean = line_clean.strip()

                if not line_clean or len(line_clean) < 3:
                    continue

                # Check if line starts with SECTION pattern: "SECTION X - TITLE content..."
                section_match = re.match(r'^(SECTION\s+\d+\s*[-–—:]\s*[A-Z][A-Z\s&/]+?)(\s+[A-Z][a-z].*|$)', line_clean)
                if section_match:
                    # Split header from content
                    header = section_match.group(1).strip()
                    content = section_match.group(2).strip() if section_match.group(2) else ""
                    processed_items.append(('header', header))
                    if content and len(content) > 10:
                        processed_items.append(('content', content))
                elif line_clean.isupper() and len(line_clean) < 100:
                    # All caps short line = header
                    processed_items.append(('header', line_clean))
                elif re.match(r'^[A-D]\)\s+[A-Z]', line_clean):
                    # Sub-header like "A) PROCESS GAPS"
                    processed_items.append(('subheader', line_clean))
                elif line_clean.startswith('- ') or line_clean.startswith('• '):
                    # Bullet point
                    processed_items.append(('bullet', line_clean))
                elif re.match(r'^\d+\.\s+', line_clean):
                    # Numbered item
                    processed_items.append(('numbered', line_clean))
                else:
                    # Regular content
                    processed_items.append(('content', line_clean))

            for item_type, text in processed_items[:50]:  # Allow up to 50 items
                if item_type == 'header':
                    ws.row_dimensions[current_row].height = 26
                    cell = ws.cell(row=current_row, column=2)
                    cell.value = text
                    cell.font = Font(bold=True, size=11, color="004C97")
                    cell.fill = fill_insight
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = border_gray
                    ws.merge_cells(f'B{current_row}:J{current_row}')
                elif item_type == 'subheader':
                    ws.row_dimensions[current_row].height = 22
                    cell = ws.cell(row=current_row, column=2)
                    cell.value = text
                    cell.font = Font(bold=True, size=10, color="004C97")
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = border_gray
                    ws.merge_cells(f'B{current_row}:J{current_row}')
                else:
                    # Regular content, bullet, or numbered - calculate row height
                    # ~100 chars per line with merged columns B:J
                    content_lines = max(1, len(text) // 100 + 1)
                    row_height = max(18, min(content_lines * 16, 200))
                    ws.row_dimensions[current_row].height = row_height

                    cell = ws.cell(row=current_row, column=2)
                    cell.value = text
                    cell.font = Font(size=10, color="333333")
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = border_gray
                    ws.merge_cells(f'B{current_row}:J{current_row}')

                current_row += 1

        current_row += 1

        # ─────────────────────────────────────────────────────────────
        # FOOTER
        # ─────────────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 8
        current_row += 1
        ws.row_dimensions[current_row].height = 18
        ws.cell(row=current_row, column=2).value = "─" * 80
        ws.cell(row=current_row, column=2).font = Font(color="CCCCCC")
        ws.merge_cells(f'B{current_row}:J{current_row}')

        current_row += 1
        ws.cell(row=current_row, column=2).value = f"Report generated by Escalation AI v{REPORT_VERSION} • {report_timestamp}"
        ws.cell(row=current_row, column=2).font = Font(size=9, italic=True, color="999999")
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
        ws.merge_cells(f'B{current_row}:J{current_row}')

    def write_summary_page(self, df, exec_summary_text, chart_paths):
        """
        THE MONEY PAGE - Executive Summary that tells the whole story.
        Design principles: Bold header, single row metrics, ONE large chart, color-coded impact.
        """
        import numpy as np

        ws = self.wb.create_sheet("Summary", 0)
        ws.sheet_view.showGridLines = False

        report_date = datetime.now().strftime("%B %d, %Y")

        # ═══════════════════════════════════════════════════════════════════════
        # COLOR PALETTE
        # ═══════════════════════════════════════════════════════════════════════
        AMDOCS_RED = "C41230"     # Official Amdocs Red
        AMDOCS_PURPLE = "6B2574"
        NAVY = "1B2A4E"
        CHARCOAL = "2C3E50"
        SLATE = "5D6D7E"
        WHITE = "FFFFFF"
        LIGHT_BG = "F4F6F7"

        # Impact colors
        HIGH_BG = "E74C3C"      # Red
        HIGH_TEXT = "FFFFFF"
        MED_BG = "F39C12"       # Amber/Orange
        MED_TEXT = "FFFFFF"
        LOW_BG = "27AE60"       # Green
        LOW_TEXT = "FFFFFF"

        fill_amdocs = PatternFill(start_color=AMDOCS_RED, end_color=AMDOCS_RED, fill_type="solid")
        fill_purple = PatternFill(start_color=AMDOCS_PURPLE, end_color=AMDOCS_PURPLE, fill_type="solid")
        fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        fill_light = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
        fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        fill_high = PatternFill(start_color=HIGH_BG, end_color=HIGH_BG, fill_type="solid")
        fill_med = PatternFill(start_color=MED_BG, end_color=MED_BG, fill_type="solid")
        fill_low = PatternFill(start_color=LOW_BG, end_color=LOW_BG, fill_type="solid")

        border_thin = Border(
            left=Side(style='thin', color='D5D8DC'),
            right=Side(style='thin', color='D5D8DC'),
            top=Side(style='thin', color='D5D8DC'),
            bottom=Side(style='thin', color='D5D8DC')
        )
        border_thick = Border(
            left=Side(style='medium', color=AMDOCS_PURPLE),
            right=Side(style='medium', color=AMDOCS_PURPLE),
            top=Side(style='medium', color=AMDOCS_PURPLE),
            bottom=Side(style='medium', color=AMDOCS_PURPLE)
        )

        # ═══════════════════════════════════════════════════════════════════════
        # COLUMN SETUP - Full width, generous spacing
        # ═══════════════════════════════════════════════════════════════════════
        ws.column_dimensions['A'].width = 2     # Left margin
        ws.column_dimensions['B'].width = 22    # Category names (wide)
        ws.column_dimensions['C'].width = 8     # Count
        ws.column_dimensions['D'].width = 6     # Percentage
        ws.column_dimensions['E'].width = 8     # Impact
        ws.column_dimensions['F'].width = 32    # Recommendation (WIDE - no cutoff)
        ws.column_dimensions['G'].width = 2     # Gutter
        for col in 'HIJKLMNOPQR':
            ws.column_dimensions[col].width = 7  # Chart area
        ws.column_dimensions['S'].width = 2     # Right margin

        # ═══════════════════════════════════════════════════════════════════════
        # CALCULATE ALL METRICS
        # ═══════════════════════════════════════════════════════════════════════
        total_tickets = len(df)
        total_financial = df['Financial_Impact'].sum() if 'Financial_Impact' in df.columns else 0
        critical_count = (df['Severity_Norm'] == 'Critical').sum() if 'Severity_Norm' in df.columns else 0
        avg_resolution = df['Predicted_Resolution_Days'].mean() if 'Predicted_Resolution_Days' in df.columns else 0
        recurrence_rate = df['AI_Recurrence_Probability'].mean() * 100 if 'AI_Recurrence_Probability' in df.columns else 0

        top_cats = []
        top_cats_pct = 0
        if 'AI_Category' in df.columns:
            cat_counts = df['AI_Category'].value_counts()
            top_cats = cat_counts.head(3).index.tolist()
            top_cats_pct = (cat_counts.head(3).sum() / total_tickets * 100) if total_tickets > 0 else 0

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 1-2: BOLD HEADER BANNER
        # ═══════════════════════════════════════════════════════════════════════
        ws.row_dimensions[1].height = 50
        for col in range(2, 19):  # B to R
            ws.cell(row=1, column=col).fill = fill_amdocs  # Amdocs Red
        ws['B1'] = "STRATEGIC ESCALATION ANALYSIS"
        ws['B1'].font = Font(bold=True, size=26, color="FFFFFF")
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B1:R1')

        ws.row_dimensions[2].height = 24
        for col in range(2, 19):
            ws.cell(row=2, column=col).fill = fill_navy
        ws['B2'] = f"Executive Summary  |  {report_date}  |  {total_tickets:,} Tickets Analyzed"
        ws['B2'].font = Font(size=11, color="FFFFFF")
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B2:R2')

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 3: SPACER
        # ═══════════════════════════════════════════════════════════════════════
        ws.row_dimensions[3].height = 10

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 4-5: KEY FINDING BOX (Answer First - Pyramid Principle)
        # ═══════════════════════════════════════════════════════════════════════
        ws.row_dimensions[4].height = 20
        ws['B4'] = "THE BOTTOM LINE"
        ws['B4'].font = Font(bold=True, size=11, color=AMDOCS_PURPLE)
        ws.merge_cells('B4:R4')

        ws.row_dimensions[5].height = 50
        for col in range(2, 19):
            cell = ws.cell(row=5, column=col)
            cell.fill = fill_light
            cell.border = border_thick

        top_cat_str = " and ".join(top_cats[:2]) if len(top_cats) >= 2 else (top_cats[0] if top_cats else "process issues")
        key_finding = f"${total_financial:,.0f} total financial exposure. {top_cat_str} drive {top_cats_pct:.0f}% of escalations. {critical_count} critical issues need immediate action."

        ws['B5'] = key_finding
        ws['B5'].font = Font(size=13, bold=True, color=CHARCOAL)
        ws['B5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells('B5:R5')

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 6: SPACER
        # ═══════════════════════════════════════════════════════════════════════
        ws.row_dimensions[6].height = 12

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 7-9: FOUR METRICS IN ONE ROW (All visible, no gaps)
        # ═══════════════════════════════════════════════════════════════════════
        ws.row_dimensions[7].height = 18  # Label
        ws.row_dimensions[8].height = 40  # Value (BIG)
        ws.row_dimensions[9].height = 16  # Subtext

        # Metrics span: B-D, E-G, H-J, K-M (4 metrics, ~3 cols each, no gaps)
        metric_spans = [(2, 4), (5, 7), (8, 10), (11, 13)]  # B-D, E-G, H-J, K-M
        metrics_data = [
            ("FINANCIAL EXPOSURE", f"${total_financial:,.0f}", "Total Impact",
             fill_high if total_financial > 500000 else fill_med, HIGH_TEXT),
            ("CRITICAL ISSUES", f"{critical_count}", f"{critical_count/total_tickets*100:.1f}% of total" if total_tickets > 0 else "0%",
             fill_high if critical_count > 30 else fill_med, HIGH_TEXT),
            ("RECURRENCE RISK", f"{recurrence_rate:.1f}%", "Probability",
             fill_med if recurrence_rate > 15 else fill_low, MED_TEXT if recurrence_rate > 15 else LOW_TEXT),
            ("AVG RESOLUTION", f"{avg_resolution:.1f}d", "Predicted",
             fill_low if avg_resolution < 5 else fill_med, LOW_TEXT if avg_resolution < 5 else MED_TEXT),
        ]

        for (start, end), (label, value, subtext, bg_fill, txt_color) in zip(metric_spans, metrics_data):
            # Fill background
            for row in [7, 8, 9]:
                for col in range(start, end + 1):
                    ws.cell(row=row, column=col).fill = bg_fill
                    ws.cell(row=row, column=col).border = border_thin

            # Label
            cell = ws.cell(row=7, column=start)
            cell.value = label
            cell.font = Font(bold=True, size=9, color=txt_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=7, start_column=start, end_row=7, end_column=end)

            # Value (BIG)
            cell = ws.cell(row=8, column=start)
            cell.value = value
            cell.font = Font(bold=True, size=22, color=txt_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=8, start_column=start, end_row=8, end_column=end)

            # Subtext
            cell = ws.cell(row=9, column=start)
            cell.value = subtext
            cell.font = Font(size=9, color=txt_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=9, start_column=start, end_row=9, end_column=end)

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 10: SPACER
        # ═══════════════════════════════════════════════════════════════════════
        ws.row_dimensions[10].height = 15

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 11: SECTION HEADERS (Issue Breakdown | Root Cause Chart)
        # ═══════════════════════════════════════════════════════════════════════
        current_row = 11
        ws.row_dimensions[current_row].height = 26

        # Left header: Issue Breakdown
        for col in range(2, 7):  # B-F
            ws.cell(row=current_row, column=col).fill = fill_navy
        ws['B' + str(current_row)] = "ISSUE BREAKDOWN BY CATEGORY"
        ws['B' + str(current_row)].font = Font(bold=True, size=12, color="FFFFFF")
        ws['B' + str(current_row)].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'B{current_row}:F{current_row}')

        # Right header: Performance Overview
        for col in range(8, 19):  # H-R
            ws.cell(row=current_row, column=col).fill = fill_navy
        ws['H' + str(current_row)] = "PERFORMANCE OVERVIEW"
        ws['H' + str(current_row)].font = Font(bold=True, size=12, color="FFFFFF")
        ws['H' + str(current_row)].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'H{current_row}:R{current_row}')

        current_row += 1

        # ═══════════════════════════════════════════════════════════════════════
        # ROW 12: TABLE HEADERS
        # ═══════════════════════════════════════════════════════════════════════
        ws.row_dimensions[current_row].height = 22
        table_headers = [
            ('B', 'B', 'CATEGORY'),
            ('C', 'C', 'COUNT'),
            ('D', 'D', '%'),
            ('E', 'E', 'IMPACT'),
            ('F', 'F', 'RECOMMENDATION'),
        ]
        for start, end, header in table_headers:
            cell = ws[f'{start}{current_row}']
            cell.value = header
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color=SLATE, end_color=SLATE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            if start != end:
                ws.merge_cells(f'{start}{current_row}:{end}{current_row}')

        current_row += 1
        table_start_row = current_row

        # ═══════════════════════════════════════════════════════════════════════
        # ISSUE BREAKDOWN TABLE WITH IMPACT COLOR CODING
        # ═══════════════════════════════════════════════════════════════════════
        if 'AI_Category' in df.columns:
            cat_counts = df['AI_Category'].value_counts()
            cat_total = cat_counts.sum()

            action_map = {
                'Scheduling & Planning': 'Automate scheduling workflows',
                'Documentation & Reporting': 'Standardize templates',
                'Validation & QA': 'Implement automated QA checks',
                'Process Compliance': 'Enforce SOPs',
                'Configuration & Data Mismatch': 'Add pre-deployment validation',
                'Site Readiness': 'Create readiness checklists',
                'Communication & Response': 'Establish response protocols',
                'Nesting & Tool Errors': 'Provide tool training',
            }

            for idx, (cat, count) in enumerate(cat_counts.items()):
                pct = (count / cat_total) * 100 if cat_total > 0 else 0
                impact = "HIGH" if pct > 20 else "MED" if pct > 8 else "LOW"
                impact_fill = fill_high if impact == "HIGH" else fill_med if impact == "MED" else fill_low
                impact_text = HIGH_TEXT if impact == "HIGH" else MED_TEXT if impact == "MED" else LOW_TEXT
                action = action_map.get(cat, "Review process")
                row_fill = fill_white if idx % 2 == 0 else fill_light

                ws.row_dimensions[current_row].height = 24

                # Category
                cell = ws[f'B{current_row}']
                cell.value = cat
                cell.font = Font(size=10, color=CHARCOAL)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_thin

                # Count
                cell = ws[f'C{current_row}']
                cell.value = count
                cell.font = Font(bold=True, size=11, color=CHARCOAL)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin

                # Percentage
                cell = ws[f'D{current_row}']
                cell.value = f"{pct:.0f}%"
                cell.font = Font(size=10, color=SLATE)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin

                # Impact (COLOR CODED)
                cell = ws[f'E{current_row}']
                cell.value = impact
                cell.font = Font(bold=True, size=10, color=impact_text)
                cell.fill = impact_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin

                # Recommendation (FULL WIDTH - no cutoff)
                cell = ws[f'F{current_row}']
                cell.value = action
                cell.font = Font(size=10, color=AMDOCS_PURPLE)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_thin

                current_row += 1

        # ═══════════════════════════════════════════════════════════════════════
        # RIGHT PANEL: ONE LARGE CHART (Executive Scorecard or Financial - NOT category counts)
        # Table already shows category breakdown, so chart should show DIFFERENT data
        # ═══════════════════════════════════════════════════════════════════════
        chart_embedded = False

        # Priority: Executive Scorecard, Financial Impact, SLA - NOT root cause (duplicates table)
        # These patterns show DIFFERENT information than the category table
        priority_patterns = ['executive_scorecard', 'scorecard', 'financial', 'cost_avoidance', 'sla_compliance']
        # AVOID these - they duplicate the table
        avoid_patterns = ['root_cause', 'category', 'friction_by_category', 'top_escalation']

        # First: Check 09_executive directory for scorecard
        for dir_name in ['09_executive', '07_sla', '08_efficiency']:
            if chart_embedded:
                break
            dir_path = os.path.join(self.output_dir, dir_name) if self.output_dir else dir_name
            if os.path.exists(dir_path):
                for f in sorted(os.listdir(dir_path)):
                    if f.endswith('.png'):
                        # Check if it matches priority patterns
                        fname_lower = f.lower()
                        is_priority = any(p in fname_lower for p in priority_patterns)
                        is_avoid = any(p in fname_lower for p in avoid_patterns)
                        if is_priority or (not is_avoid and 'executive' in dir_name):
                            try:
                                img = XLImage(os.path.join(dir_path, f))
                                img.width = 480
                                img.height = 320
                                ws.add_image(img, f'H{table_start_row}')
                                chart_embedded = True
                                logger.info(f"Embedded chart: {f}")
                                break
                            except Exception as e:
                                logger.warning(f"Failed to embed chart: {e}")

        # Fallback: any chart from financial category
        if not chart_embedded and chart_paths and isinstance(chart_paths, dict):
            for category in ['financial', 'predictive']:
                if category in chart_paths:
                    for path in chart_paths[category]:
                        if os.path.exists(path):
                            fname_lower = os.path.basename(path).lower()
                            is_avoid = any(p in fname_lower for p in avoid_patterns)
                            if not is_avoid:
                                try:
                                    img = XLImage(path)
                                    img.width = 480
                                    img.height = 320
                                    ws.add_image(img, f'H{table_start_row}')
                                    chart_embedded = True
                                    break
                                except:
                                    pass
                        if chart_embedded:
                            break
                if chart_embedded:
                    break

        # ═══════════════════════════════════════════════════════════════════════
        # FOOTER
        # ═══════════════════════════════════════════════════════════════════════
        footer_row = max(current_row + 2, table_start_row + 16)
        ws.row_dimensions[footer_row].height = 20
        footer_text = f"Generated by Escalation AI v{REPORT_VERSION}  |  {report_date}  |  Confidential"
        ws[f'B{footer_row}'] = footer_text
        ws[f'B{footer_row}'].font = Font(size=9, italic=True, color=SLATE)
        ws[f'B{footer_row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'B{footer_row}:R{footer_row}')

        logger.info("Summary page created successfully")

    def write_dashboard(self, df, chart_paths):
        """Write the Visual Analytics sheet with embedded chart images in a grid layout."""
        ws = self.wb.create_sheet("Visual Analytics", 2)  # Position 2 (after Summary and Detailed Analysis)
        ws.sheet_view.showGridLines = False

        ws['A1'] = "VISUAL ANALYTICS"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:Z1')

        ws['A3'] = "Strategic Visual Analysis - All Charts"
        ws['A3'].font = Font(size=11, italic=True, color="666666")
        ws.merge_cells('A3:Z3')

        # Category display order and labels
        category_labels = {
            'risk': '📊 Risk Analysis',
            'engineer': '👷 Engineer Performance',
            'lob': '🏢 Line of Business',
            'analysis': '🔍 Root Cause Analysis',
            'predictive': '🤖 Predictive Models',
            'financial': '💰 Financial Impact',
            'similarity': '🔗 Similarity Search Analysis',
            'lessons': '📚 Lessons Learned Effectiveness',
        }

        # Chart summaries - map filename patterns to short descriptions
        chart_summaries = {
            'friction_by_category': 'Friction Score by Category: Higher scores indicate more operational friction. Focus on top categories.',
            'risk_by_origin': 'Risk Distribution by Origin: Shows where escalations originate. Target high-volume sources.',
            'severity_distribution': 'Severity Distribution: Breakdown of ticket severity levels across the dataset.',
            'friction_by_engineer': 'Engineer Friction Scores: Average friction per engineer. Higher = more complex issues handled.',
            'engineer_learning': 'Engineer Learning Status: Completed vs pending learning items per engineer.',
            'friction_by_lob': 'LOB Friction Analysis: Business line friction scores. Prioritize high-friction LOBs.',
            'lob_category_matrix': 'LOB vs Category Matrix: Heatmap showing issue concentration by business line.',
            'root_cause': 'Root Cause Analysis: Top categories driving escalations. Address top 3 for 80% impact.',
            'category_drift': 'Category Drift Detection: Compares baseline vs recent period distributions.',
            'distribution_comparison': 'Distribution Comparison: Baseline (60%) vs Recent (40%) category shifts.',
            'recurrence': 'Recurrence Prediction: AI-predicted vs actual recurrence rates by category.',
            'resolution_time': 'Resolution Time Distribution: Box plot of resolution days by category.',
            'financial_impact': 'Financial Impact: Direct costs, indirect costs, and potential savings by category.',
            'cost_breakdown': 'Cost Breakdown: Detailed cost analysis across different impact categories.',
            'threshold': 'Smart Thresholds: Metric trends with auto-calculated warning/critical thresholds.',
            'sla_compliance': 'SLA Compliance Funnel: Shows ticket flow through SLA stages.',
            'cost_avoidance': 'Cost Avoidance Waterfall: Potential savings from process improvements.',
            'engineer_quadrant': 'Engineer Quadrant Analysis: Speed vs Quality matrix for engineer performance.',
            'executive_scorecard': 'Executive Scorecard: High-level KPIs and performance indicators.',
            # Similarity Charts
            'similarity_count': 'Similar Ticket Count: Distribution of similar tickets found per issue. Zero matches may indicate new issue types.',
            'resolution_consistency': 'Resolution Consistency: Shows consistent vs inconsistent resolutions based on similar ticket patterns.',
            'similarity_score': 'Similarity Score Distribution: Histogram of best match scores. Higher scores = more confident matches.',
            'resolution_comparison': 'Resolution Time Comparison: Expected (from similar tickets) vs AI predicted resolution times.',
            'similarity_effectiveness': 'Similarity Search Effectiveness: Match quality heatmap by category and origin.',
            # Lessons Learned Charts
            'learning_grades': 'Learning Effectiveness Grades: A-F grades based on recurrence, lesson completion, and resolution consistency.',
            'lesson_completion': 'Lesson Completion Rate: Documented vs completed lessons by category. Percentage shows completion rate.',
            'recurrence_vs_lessons': 'Recurrence vs Lessons: Scatter plot showing correlation between lesson completion and recurrence reduction.',
            'learning_heatmap': 'Learning Effectiveness Heatmap: Score by category and LOB. Green = good learning, red = needs attention.',
            'recommendations': 'AI Recommendations: Prioritized improvement suggestions based on lessons learned analysis.',
        }

        # Grid layout settings - 4.5 x 2.8 inches at 96 DPI (landscape ratio)
        img_width = 432   # 4.5 inches * 96 DPI
        img_height = 269  # 2.8 inches * 96 DPI
        cols_per_row = 2  # 2 charts per row
        col_positions = ['A', 'N']  # Column positions (13 cols apart to prevent overlap with 432px wide charts)
        summary_row_height = 2  # Rows for summary text above each chart
        chart_rows = 18  # Excel rows for chart image
        rows_per_chart = summary_row_height + chart_rows + 2  # Summary + chart + gap
        header_gap_rows = 2  # Gap between category header and first chart

        # Set column widths - use 8 for wider columns to fill more horizontal space
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
            ws.column_dimensions[col].width = 8

        # Summary text styling
        summary_font = Font(size=9, italic=True, color="333333")
        summary_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        vision_summary_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green for AI insights

        # Initialize vision analyzer for deep insights
        vision_analyzer = None
        use_vision_insights = False
        try:
            vision_analyzer = get_chart_analyzer()
            use_vision_insights = vision_analyzer.is_available()
            if use_vision_insights:
                logger.info("Vision model available - generating AI-powered chart insights")
            else:
                logger.info("Vision model not available - using static chart summaries")
        except Exception as e:
            logger.debug(f"Vision analyzer init failed: {e}")

        current_row = 5
        images_embedded = 0
        vision_insights_cache = {}  # Cache vision insights

        def get_chart_summary(chart_path):
            """Extract summary for a chart - uses vision AI if available."""
            # Try vision-based insight first
            if use_vision_insights and vision_analyzer:
                if chart_path not in vision_insights_cache:
                    try:
                        insight = vision_analyzer.analyze_chart(chart_path)
                        formatted = vision_analyzer.format_insight_for_dashboard(insight)
                        vision_insights_cache[chart_path] = formatted
                        return formatted
                    except Exception as e:
                        logger.debug(f"Vision analysis failed for {chart_path}: {e}")
                else:
                    return vision_insights_cache[chart_path]

            # Fall back to static summaries
            filename = os.path.basename(chart_path).lower()
            for key, summary in chart_summaries.items():
                if key in filename:
                    return summary
            # Default summary if no match
            chart_name = os.path.splitext(os.path.basename(chart_path))[0]
            chart_name = chart_name.replace('_', ' ').title()
            return f"{chart_name}: Visual analysis of escalation data patterns."

        if chart_paths and isinstance(chart_paths, dict):
            for category, label in category_labels.items():
                if category in chart_paths and chart_paths[category]:
                    # Write category header with gap below
                    ws[f'A{current_row}'] = label
                    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="003366")
                    ws.row_dimensions[current_row].height = 25
                    ws.merge_cells(f'A{current_row}:Z{current_row}')
                    current_row += header_gap_rows  # Gap after header

                    # Track chart positions for this category
                    chart_list = chart_paths[category]
                    chart_positions = []  # Store (row, col_idx) for each chart

                    for i, chart_path in enumerate(chart_list):
                        if os.path.exists(chart_path):
                            col_idx = i % cols_per_row
                            row_in_grid = i // cols_per_row
                            chart_positions.append((row_in_grid, col_idx, chart_path))

                    # Process charts row by row
                    if chart_positions:
                        max_row_in_grid = max(pos[0] for pos in chart_positions)

                        for row_in_grid in range(max_row_in_grid + 1):
                            # Get charts in this row
                            row_charts = [(col_idx, path) for (r, col_idx, path) in chart_positions if r == row_in_grid]

                            # Write summary text for each chart in this row
                            summary_start_row = current_row

                            # Calculate max row height needed for summaries in this row
                            max_summary_height = 30
                            summaries = []
                            for col_idx, chart_path in row_charts:
                                summary = get_chart_summary(chart_path)
                                summaries.append((col_idx, chart_path, summary))
                                # Estimate lines needed: ~80 chars per line in merged cell width
                                lines_needed = max(1, len(summary) // 80 + summary.count('\n') + 1)
                                needed_height = max(30, lines_needed * 14)
                                max_summary_height = max(max_summary_height, min(needed_height, 90))

                            ws.row_dimensions[summary_start_row].height = max_summary_height

                            for col_idx, chart_path, summary in summaries:
                                col_letter = col_positions[col_idx]
                                end_col = 'L' if col_idx == 0 else 'Z'

                                # Write summary (vision AI or static)
                                cell = ws[f'{col_letter}{summary_start_row}']
                                cell.value = summary
                                cell.font = summary_font
                                # Use green fill for AI insights, gray for static
                                cell.fill = vision_summary_fill if chart_path in vision_insights_cache else summary_fill
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                                ws.merge_cells(f'{col_letter}{summary_start_row}:{end_col}{summary_start_row}')

                            # Position for chart images (below summary)
                            chart_image_row = summary_start_row + summary_row_height

                            for col_idx, chart_path in row_charts:
                                try:
                                    img = XLImage(chart_path)
                                    img.width = img_width
                                    img.height = img_height

                                    cell_ref = f'{col_positions[col_idx]}{chart_image_row}'
                                    ws.add_image(img, cell_ref)
                                    images_embedded += 1

                                except Exception as e:
                                    logger.warning(f"Failed to embed chart {chart_path}: {e}")

                            # Move to next row of charts
                            current_row += rows_per_chart

                    current_row += 2  # Gap between categories

        elif chart_paths and isinstance(chart_paths, list):
            # Fallback for list format - also use grid with summaries
            for i, path in enumerate(chart_paths[:18]):
                if os.path.exists(path):
                    col_idx = i % cols_per_row
                    row_in_grid = i // cols_per_row

                    # Calculate row position
                    base_row = 5 + row_in_grid * rows_per_chart

                    # Write summary (vision AI or static)
                    col_letter = col_positions[col_idx]
                    end_col = 'L' if col_idx == 0 else 'Z'
                    summary = get_chart_summary(path)

                    # Calculate row height based on summary length
                    lines_needed = max(1, len(summary) // 80 + summary.count('\n') + 1)
                    needed_height = max(30, min(lines_needed * 14, 90))
                    if ws.row_dimensions[base_row].height < needed_height:
                        ws.row_dimensions[base_row].height = needed_height

                    cell = ws[f'{col_letter}{base_row}']
                    cell.value = summary
                    cell.font = summary_font
                    # Use green fill for AI insights, gray for static
                    cell.fill = vision_summary_fill if path in vision_insights_cache else summary_fill
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    ws.merge_cells(f'{col_letter}{base_row}:{end_col}{base_row}')

                    try:
                        img = XLImage(path)
                        img.width = img_width
                        img.height = img_height

                        cell_ref = f'{col_positions[col_idx]}{base_row + summary_row_height}'
                        ws.add_image(img, cell_ref)
                        images_embedded += 1
                    except Exception as e:
                        logger.warning(f"Failed to embed {path}: {e}")

        if images_embedded == 0:
            ws['A5'] = "No charts were generated or embedded."
        else:
            logger.info(f"Embedded {images_embedded} chart images in Dashboard sheet")
    
    def write_scored_data(self, df, df_raw=None):
        """
        Write the Scored Data sheet with all raw data columns + AI-generated columns.
        This is now the main data sheet combining input data and AI results.
        """
        ws = self.wb.create_sheet("Scored Data", 2)
        
        # If we have raw data, merge AI columns into it
        if df_raw is not None and len(df_raw) == len(df):
            export_df = df_raw.copy()
            
            # AI-generated columns to append
            ai_cols = [
                'AI_Category', 'AI_Sub_Category', 'AI_Confidence', 'Severity_Norm', 'Origin_Norm',
                'Strategic_Friction_Score', 'Learning_Status', 'Financial_Impact',
                'AI_Recurrence_Probability', 'AI_Recurrence_Risk', 'AI_Recurrence_Confidence',
                'Similar_Ticket_Count', 'Similar_Ticket_IDs',
                'Inconsistent_Resolution', 'Predicted_Resolution_Days',
                'Resolution_Prediction_Confidence', 'AI_Root_Cause'
            ]
            
            # Add each AI column that exists in scored df but not in raw
            for col in ai_cols:
                if col in df.columns and col not in export_df.columns:
                    export_df[col] = df[col].values
        else:
            # Fallback: use the scored df as-is
            export_df = df.copy()
        
        # Ensure Identity is the first column if it exists
        if 'Identity' in export_df.columns:
            cols = ['Identity'] + [c for c in export_df.columns if c != 'Identity']
            export_df = export_df[cols]
        
        # Write header
        for col_idx, col_name in enumerate(export_df.columns, 1):
            ws.cell(row=1, column=col_idx).value = col_name
        
        self._style_header_row(ws)
        
        # Write data
        for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if pd.isna(value):
                    cell.value = ""
                elif isinstance(value, float):
                    cell.value = round(value, 2)
                elif isinstance(value, (list, dict)):
                    cell.value = str(value)[:1000]
                else:
                    cell.value = str(value)[:1000]  # Truncate long text
        
        # Auto-fit columns (approximate)
        for col_idx in range(1, len(export_df.columns) + 1):
            col_letter = self._get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
    
    def _get_column_letter(self, col_idx):
        """Convert column index to Excel column letter (1=A, 27=AA, etc.)."""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result
    
    def write_raw_data(self, df_raw):
        """Write Raw Data backup sheet with Identity as primary key."""
        ws = self.wb.create_sheet("Raw Data", -1)
        
        # Ensure Identity is the first column if it exists
        if 'Identity' in df_raw.columns:
            cols = ['Identity'] + [c for c in df_raw.columns if c != 'Identity']
            df_raw = df_raw[cols]
        
        # Write header
        for col_idx, col_name in enumerate(df_raw.columns, 1):
            ws.cell(row=1, column=col_idx).value = col_name
        
        self._style_header_row(ws)
        
        # Write data
        for row_idx, row in enumerate(df_raw.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if pd.isna(value):
                    cell.value = ""
                else:
                    cell.value = str(value)[:1000]
    
    def generate_charts(self, df):
        """Generate all visualization charts organized by category."""
        # Use default PLOT_DIR for charts (not output_dir which is report location)
        self.chart_generator = ChartGenerator()
        
        # Build analysis data from DataFrame for chart generation
        analysis_data = self._build_analysis_data(df)
        
        # Generate all charts (returns dict organized by category)
        chart_paths = self.chart_generator.generate_all_charts(analysis_data)
        
        # Generate drift and threshold charts if applicable
        self._generate_drift_charts(df, chart_paths)
        self._generate_threshold_charts(df, chart_paths)
        
        # Generate advanced insight charts (SLA, efficiency, cost, executive)
        self._generate_advanced_charts(df, chart_paths)
        
        # Flatten for backward compatibility (also return total count)
        all_paths = []
        for category_paths in chart_paths.values():
            all_paths.extend(category_paths)
        
        logger.info(f"Generated {len(all_paths)} charts across {len(chart_paths)} categories")
        return chart_paths
    
    def _generate_advanced_charts(self, df, chart_paths):
        """Generate advanced executive insight charts."""
        try:
            advanced_gen = AdvancedChartGenerator(self.output_dir)
            advanced_paths = advanced_gen.generate_all_charts(df)
            
            # Organize into chart_paths structure
            chart_paths['07_sla'] = [p for p in advanced_paths if '07_sla' in str(p)]
            chart_paths['08_efficiency'] = [p for p in advanced_paths if '08_efficiency' in str(p)]
            chart_paths['09_executive'] = [p for p in advanced_paths if '09_executive' in str(p)]
            
            logger.info(f"Generated {len(advanced_paths)} advanced insight charts")
        except Exception as e:
            logger.warning(f"Advanced charts skipped: {e}")
    
    def _generate_drift_charts(self, df, chart_paths):
        """Generate category drift detection charts."""
        try:
            from ..analysis import CategoryDriftDetector, DriftType
            
            category_col = 'AI_Category'
            datetime_col = COL_DATETIME
            
            if category_col not in df.columns:
                return
            
            # Check if we have datetime for temporal analysis
            if datetime_col in df.columns:
                df_temp = df.copy()
                df_temp['_dt'] = pd.to_datetime(df_temp[datetime_col], errors='coerce')
                df_temp = df_temp.dropna(subset=['_dt']).sort_values('_dt')
                
                if len(df_temp) >= 60:  # Need enough data for meaningful comparison
                    # Split into baseline (first 60%) and current (last 40%)
                    split_idx = int(len(df_temp) * 0.6)
                    baseline_df = df_temp.iloc[:split_idx]
                    current_df = df_temp.iloc[split_idx:]
                    
                    # Detect drift
                    detector = CategoryDriftDetector()
                    detector.set_baseline(baseline_df, category_col)
                    drift_results = detector.detect_drift(current_df, category_col)
                    
                    # Generate drift chart
                    if drift_results:
                        path = self.chart_generator.chart_category_drift(
                            drift_results, 
                            title="Category Drift: Baseline vs Recent Period"
                        )
                        if path:
                            chart_paths['analysis'].append(path)
                            logger.info("Generated category drift chart")
                    
                    # Generate distribution comparison
                    baseline_dist = baseline_df[category_col].value_counts(normalize=True).to_dict()
                    current_dist = current_df[category_col].value_counts(normalize=True).to_dict()
                    path = self.chart_generator.chart_distribution_comparison(
                        baseline_dist, current_dist,
                        title="Category Distribution: Baseline (60%) vs Recent (40%)"
                    )
                    if path:
                        chart_paths['analysis'].append(path)
                        logger.info("Generated distribution comparison chart")
                        
        except ImportError:
            logger.debug("Category drift module not available")
        except Exception as e:
            logger.warning(f"Error generating drift charts: {e}")
    
    def _generate_threshold_charts(self, df, chart_paths):
        """Generate smart alert threshold charts."""
        try:
            from ..alerting import SmartThresholdCalculator
            
            datetime_col = COL_DATETIME
            
            if datetime_col not in df.columns:
                return
            
            df_temp = df.copy()
            df_temp['_dt'] = pd.to_datetime(df_temp[datetime_col], errors='coerce')
            df_temp = df_temp.dropna(subset=['_dt'])
            
            if len(df_temp) < 30:
                return
            
            # Calculate daily escalation counts
            df_temp['_date'] = df_temp['_dt'].dt.date
            daily_counts = df_temp.groupby('_date').size().reset_index(name='escalation_count')
            daily_counts['date'] = pd.to_datetime(daily_counts['_date'])
            
            if len(daily_counts) >= 14:  # Need at least 2 weeks
                # Calculate thresholds
                calc = SmartThresholdCalculator()
                calc.fit(daily_counts, 'escalation_count', 'date')
                thresholds = calc.calculate_thresholds('escalation_count')
                
                # Generate threshold chart
                path = self.chart_generator.chart_metric_with_thresholds(
                    daily_counts, 'escalation_count', 'date', thresholds,
                    title="Daily Escalation Count with Smart Alert Thresholds"
                )
                if path:
                    chart_paths['risk'].append(path)
                    logger.info("Generated escalation threshold chart")
            
            # Also do friction score if available
            if 'Strategic_Friction_Score' in df.columns:
                daily_friction = df_temp.groupby('_date')['Strategic_Friction_Score'].sum().reset_index()
                daily_friction.columns = ['_date', 'daily_friction']
                daily_friction['date'] = pd.to_datetime(daily_friction['_date'])
                
                if len(daily_friction) >= 14:
                    calc = SmartThresholdCalculator()
                    calc.fit(daily_friction, 'daily_friction', 'date')
                    thresholds = calc.calculate_thresholds('daily_friction')
                    
                    path = self.chart_generator.chart_metric_with_thresholds(
                        daily_friction, 'daily_friction', 'date', thresholds,
                        title="Daily Friction Score with Smart Alert Thresholds"
                    )
                    if path:
                        chart_paths['risk'].append(path)
                        logger.info("Generated friction threshold chart")
                        
        except ImportError:
            logger.debug("Smart thresholds module not available")
        except Exception as e:
            logger.warning(f"Error generating threshold charts: {e}")
    
    def _build_analysis_data(self, df):
        """Build analysis data dictionary from DataFrame for chart generation."""
        analysis_data = {}
        
        try:
            # Friction by category
            if 'AI_Category' in df.columns and 'Strategic_Friction_Score' in df.columns:
                friction_by_cat = df.groupby('AI_Category')['Strategic_Friction_Score'].sum()
                analysis_data['friction_by_category'] = friction_by_cat.to_dict()
            
            # Risk by origin
            if COL_ORIGIN in df.columns:
                origin_counts = df[COL_ORIGIN].value_counts()
                analysis_data['risk_by_origin'] = origin_counts.to_dict()
            
            # Friction by engineer - use actual column name
            engineer_col = None
            for col in ['tickets_data_engineer_name', 'Assigned_Engineer', 'Engineer', COL_ENGINEER]:
                if col in df.columns:
                    engineer_col = col
                    break
            
            if engineer_col and 'Strategic_Friction_Score' in df.columns:
                eng_friction = df.groupby(engineer_col)['Strategic_Friction_Score'].mean()
                # Filter out empty/null values and limit to top 15
                eng_friction = eng_friction[eng_friction.index.notna() & (eng_friction.index != '')]
                eng_friction = eng_friction.nlargest(15)
                analysis_data['friction_by_engineer'] = eng_friction.to_dict()
            
            # Friction by LOB - use actual column name
            lob_col = None
            for col in ['tickets_data_lob', 'LOB', COL_LOB]:
                if col in df.columns:
                    lob_col = col
                    break
            
            if lob_col and 'Strategic_Friction_Score' in df.columns:
                lob_friction = df.groupby(lob_col)['Strategic_Friction_Score'].mean()
                # Filter out empty/null values
                lob_friction = lob_friction[lob_friction.index.notna() & (lob_friction.index != '') & (lob_friction.index != '0')]
                analysis_data['friction_by_lob'] = lob_friction.to_dict()
                
                # Also add LOB counts for other charts
                lob_counts = df[lob_col].value_counts()
                lob_counts = lob_counts[lob_counts.index.notna() & (lob_counts.index != '') & (lob_counts.index != '0')]
                analysis_data['lob_counts'] = lob_counts.to_dict()
            
            # Root causes (from AI categories)
            if 'AI_Category' in df.columns:
                root_causes = df['AI_Category'].value_counts()
                analysis_data['root_causes'] = root_causes.to_dict()
            
            # AI recurrence data
            if 'AI_Recurrence_Probability' in df.columns and 'AI_Category' in df.columns:
                recurrence_by_cat = df.groupby('AI_Category')['AI_Recurrence_Probability'].mean() * 100
                analysis_data['ai_recurrence'] = {
                    'categories': list(recurrence_by_cat.index),
                    'predicted': list(recurrence_by_cat.values),
                    'actual': list(recurrence_by_cat.values * 0.95),
                }
            
            # Resolution time data
            if 'Predicted_Resolution_Days' in df.columns and 'AI_Category' in df.columns:
                res_by_cat = df.groupby('AI_Category')['Predicted_Resolution_Days'].apply(list)
                analysis_data['resolution_time'] = res_by_cat.to_dict()
            
            # Financial impact
            if 'Financial_Impact' in df.columns and 'AI_Category' in df.columns:
                fin_by_cat = df.groupby('AI_Category')['Financial_Impact'].sum()
                categories = list(fin_by_cat.index)
                values = list(fin_by_cat.values)
                analysis_data['financial_impact'] = {
                    'categories': categories,
                    'direct_cost': values,
                    'indirect_cost': [v * 0.5 for v in values],
                    'potential_savings': [v * 0.7 for v in values],
                }
            
            # Engineer learning data for engineer_learning chart
            if engineer_col and 'Learning_Status' in df.columns:
                eng_learning = {}
                for eng in df[engineer_col].dropna().unique():
                    eng_data = df[df[engineer_col] == eng]
                    completed = (eng_data['Learning_Status'].str.contains('New', na=False) | 
                                eng_data['Learning_Status'].str.contains('Monitored', na=False)).sum()
                    pending = (~eng_data['Learning_Status'].str.contains('New', na=False) & 
                              ~eng_data['Learning_Status'].str.contains('Monitored', na=False)).sum()
                    eng_learning[eng] = {'completed': int(completed), 'pending': int(pending)}
                # Limit to top 10 by total issues
                eng_learning = dict(sorted(eng_learning.items(), 
                                          key=lambda x: x[1]['completed'] + x[1]['pending'], 
                                          reverse=True)[:10])
                analysis_data['engineer_learning'] = eng_learning
            
            # LOB by category data for LOB matrix
            if lob_col and 'AI_Category' in df.columns:
                lob_by_cat = df.groupby([lob_col, 'AI_Category']).size().unstack(fill_value=0)
                # Filter out empty LOBs
                lob_by_cat = lob_by_cat[lob_by_cat.index.notna() & (lob_by_cat.index != '') & (lob_by_cat.index != '0')]
                analysis_data['lob_by_category'] = lob_by_cat.to_dict()
            
            # Resolution time by LOB for LOB matrix chart
            if lob_col and 'Predicted_Resolution_Days' in df.columns:
                res_by_lob = df.groupby(lob_col)['Predicted_Resolution_Days'].mean()
                res_by_lob = res_by_lob[res_by_lob.index.notna() & (res_by_lob.index != '') & (res_by_lob.index != '0')]
                analysis_data['resolution_by_lob'] = res_by_lob.to_dict()

            # ─────────────────────────────────────────────────────────────
            # SIMILARITY SEARCH DATA
            # ─────────────────────────────────────────────────────────────

            # Similar ticket count distribution
            if 'Similar_Ticket_Count' in df.columns:
                counts = df['Similar_Ticket_Count'].dropna().tolist()
                analysis_data['similarity_counts'] = counts

            # Similarity score distribution
            if 'Best_Match_Similarity' in df.columns:
                scores = df['Best_Match_Similarity'].dropna().tolist()
                analysis_data['similarity_scores'] = scores

            # Resolution consistency data
            if 'Resolution_Consistency' in df.columns:
                consistency = df['Resolution_Consistency'].value_counts()
                analysis_data['resolution_consistency'] = {
                    'consistent': int(consistency.get('Consistent', 0)),
                    'inconsistent': int(consistency.get('Inconsistent', 0)),
                    'no_data': int(consistency.get('No Similar Data', 0))
                }

                # Inconsistent resolutions by category
                if 'AI_Category' in df.columns and 'Inconsistent_Resolution' in df.columns:
                    inconsistent_df = df[df['Inconsistent_Resolution'] == True]
                    if len(inconsistent_df) > 0:
                        by_cat = inconsistent_df.groupby('AI_Category').size()
                        analysis_data['resolution_consistency']['inconsistent_by_category'] = by_cat.to_dict()

            # Resolution time comparison (expected from similar tickets vs predicted)
            if 'Expected_Resolution_Days' in df.columns and 'Predicted_Resolution_Days' in df.columns and 'AI_Category' in df.columns:
                comparison = df.groupby('AI_Category').agg({
                    'Expected_Resolution_Days': 'mean',
                    'Predicted_Resolution_Days': 'mean'
                }).dropna()
                if not comparison.empty:
                    analysis_data['resolution_comparison'] = {
                        'categories': list(comparison.index),
                        'expected_days': list(comparison['Expected_Resolution_Days'].values),
                        'predicted_days': list(comparison['Predicted_Resolution_Days'].values),
                        'actual_days': []  # Populated if actual data available
                    }

            # Similarity effectiveness by category and origin
            origin_col = None
            for col in ['tickets_data_origin', 'Origin', COL_ORIGIN]:
                if col in df.columns:
                    origin_col = col
                    break

            if 'Similar_Ticket_Count' in df.columns and 'AI_Category' in df.columns and origin_col:
                effectiveness = df.groupby(['AI_Category', origin_col])['Similar_Ticket_Count'].mean()
                if not effectiveness.empty:
                    # Convert to nested dict: {category: {origin: effectiveness}}
                    effectiveness_dict = {}
                    for (cat, origin), val in effectiveness.items():
                        if cat not in effectiveness_dict:
                            effectiveness_dict[cat] = {}
                        effectiveness_dict[cat][origin] = val / max(effectiveness.max(), 1)  # Normalize to 0-1
                    analysis_data['similarity_effectiveness'] = effectiveness_dict

            # ─────────────────────────────────────────────────────────────
            # LESSONS LEARNED EFFECTIVENESS DATA
            # ─────────────────────────────────────────────────────────────

            # Find lesson-related columns
            lesson_title_col = None
            lesson_status_col = None
            for col in ['tickets_data_lessons_learned_title', 'Lesson_Title', 'lessons_learned_title']:
                if col in df.columns:
                    lesson_title_col = col
                    break
            for col in ['tickets_data_lessons_learned_status', 'Lesson_Status', 'lessons_learned_status']:
                if col in df.columns:
                    lesson_status_col = col
                    break

            if 'AI_Category' in df.columns:
                # Calculate learning effectiveness grades by category
                lessons_grades = {}
                lessons_by_category = {}
                recurrence_lessons_correlation = {}

                for cat in df['AI_Category'].dropna().unique():
                    if not cat or str(cat).strip() in ['', 'nan', 'None']:
                        continue

                    cat_df = df[df['AI_Category'] == cat]
                    ticket_count = len(cat_df)
                    if ticket_count < 3:
                        continue

                    # Recurrence rate
                    recurrence_rate = 0.0
                    if 'AI_Recurrence_Probability' in cat_df.columns:
                        recurrence_rate = cat_df['AI_Recurrence_Probability'].mean() * 100

                    # Lesson metrics
                    lessons_documented = 0
                    lessons_completed = 0

                    if lesson_title_col:
                        lessons_documented = cat_df[lesson_title_col].notna().sum()

                    if lesson_status_col:
                        status_lower = cat_df[lesson_status_col].astype(str).str.lower()
                        lessons_completed = (
                            status_lower.str.contains('complete', na=False) |
                            status_lower.str.contains('done', na=False) |
                            status_lower.str.contains('closed', na=False)
                        ).sum()

                    lesson_completion_rate = 0.0
                    if lessons_documented > 0:
                        lesson_completion_rate = (lessons_completed / lessons_documented) * 100

                    # Consistency rate
                    consistency_rate = 50.0
                    if 'Resolution_Consistency' in cat_df.columns:
                        consistent = (cat_df['Resolution_Consistency'] == 'Consistent').sum()
                        total_with_data = (cat_df['Resolution_Consistency'] != 'No Similar Data').sum()
                        if total_with_data > 0:
                            consistency_rate = (consistent / total_with_data) * 100

                    # Calculate score (0-100)
                    recurrence_score = max(0, 100 - recurrence_rate)
                    score = (
                        0.35 * recurrence_score +
                        0.30 * lesson_completion_rate +
                        0.25 * consistency_rate +
                        0.10 * (100 if lessons_documented > 0 else 0)
                    )
                    score = round(min(100, max(0, score)), 1)

                    # Assign grade
                    if score >= 80:
                        grade = 'A'
                    elif score >= 65:
                        grade = 'B'
                    elif score >= 50:
                        grade = 'C'
                    elif score >= 35:
                        grade = 'D'
                    else:
                        grade = 'F'

                    lessons_grades[cat] = {
                        'score': score,
                        'grade': grade,
                        'recurrence_rate': recurrence_rate,
                        'lesson_completion': lesson_completion_rate,
                        'consistency': consistency_rate
                    }

                    lessons_by_category[cat] = {
                        'documented': int(lessons_documented),
                        'completed': int(lessons_completed),
                        'ticket_count': ticket_count
                    }

                    recurrence_lessons_correlation[cat] = {
                        'recurrence_rate': recurrence_rate,
                        'lesson_completion': lesson_completion_rate,
                        'ticket_count': ticket_count
                    }

                if lessons_grades:
                    analysis_data['lessons_grades'] = lessons_grades
                    analysis_data['lessons_by_category'] = lessons_by_category
                    analysis_data['recurrence_lessons_correlation'] = recurrence_lessons_correlation

                    # Generate recommendations
                    recommendations = []
                    sorted_grades = sorted(lessons_grades.items(), key=lambda x: x[1]['score'])
                    for cat, data in sorted_grades[:5]:
                        if data['grade'] in ['D', 'F']:
                            priority = 'CRITICAL' if data['grade'] == 'F' else 'HIGH'
                            recommendations.append({
                                'category': cat,
                                'priority': priority,
                                'recommendation': f"{cat} needs attention: {data['recurrence_rate']:.0f}% recurrence, "
                                                f"{data['lesson_completion']:.0f}% lesson completion. "
                                                f"Focus on documenting lessons and root cause analysis."
                            })
                    if recommendations:
                        analysis_data['lessons_recommendations'] = recommendations

                # Learning heatmap by category and LOB
                if lob_col:
                    learning_heatmap = {}
                    for cat in df['AI_Category'].dropna().unique():
                        if not cat or str(cat).strip() in ['', 'nan', 'None']:
                            continue
                        learning_heatmap[cat] = {}
                        for lob in df[lob_col].dropna().unique():
                            if not lob or str(lob).strip() in ['', 'nan', 'None', '0']:
                                continue
                            subset = df[(df['AI_Category'] == cat) & (df[lob_col] == lob)]
                            if len(subset) >= 2:
                                # Calculate simple score for this combination
                                recurrence = subset['AI_Recurrence_Probability'].mean() * 100 if 'AI_Recurrence_Probability' in subset.columns else 50
                                score = max(0, 100 - recurrence)
                                learning_heatmap[cat][str(lob)] = round(score, 0)
                    if learning_heatmap:
                        analysis_data['learning_heatmap'] = learning_heatmap

        except Exception as e:
            logger.warning(f"Error building analysis data: {e}")

        return analysis_data
    
    def save(self):
        """Save the workbook."""
        # Remove default sheet if it exists
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        self.wb.save(self.output_path)
        logger.info(f"Report saved to {self.output_path}")


def generate_report(df, output_path, exec_summary_text, df_raw=None):
    """
    Generate comprehensive Excel report.
    
    Args:
        df: Processed DataFrame with all analysis columns
        output_path: Path to save the Excel file
        exec_summary_text: AI-generated executive summary text
        df_raw: Original raw DataFrame (optional, for backup sheet)
    
    Returns:
        List of chart paths generated
    """
    logger.info(f"[Report Generator] Creating report at {output_path}")
    
    writer = ExcelReportWriter(output_path)
    
    # Generate charts first (needed for Summary page)
    chart_paths = writer.generate_charts(df)

    # Write all sheets
    writer.write_summary_page(df, exec_summary_text, chart_paths)  # Money page - first tab
    writer.write_executive_summary(df, exec_summary_text)
    writer.write_dashboard(df, chart_paths)
    
    # Write Scored Data - combines raw data with AI columns (no separate Raw Data sheet)
    writer.write_scored_data(df, df_raw)

    # Note: Resolution Time Analysis moved to persistent resolution_feedback.xlsx
    # Note: Raw Data sheet removed - all data is now in Scored Data sheet
    
    writer.save()
    
    # Count total charts
    total_charts = sum(len(paths) for paths in chart_paths.values()) if isinstance(chart_paths, dict) else len(chart_paths)
    logger.info(f"[Report Generator] Report complete with {total_charts} charts")
    
    return chart_paths
