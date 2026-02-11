"""
Price Catalog System for Financial Impact Calculation.

=== PURPOSE ===
This module is the single source of truth for translating escalation
categories into dollar-denominated financial impact.  It reads all pricing
data from an Excel workbook (``price_catalog.xlsx``) so that finance teams
and managers can tune costs, multipliers, and premiums without modifying
any Python code.

=== ARCHITECTURE ===
The PriceCatalog class is a singleton (accessed via ``get_price_catalog()``)
that loads and caches data from the following sheets in price_catalog.xlsx:

  1. **Category Costs** -- base cost per 8-category classification bucket.
     Each row defines material cost, labour hours, and hourly rate.
  2. **Sub-Category Costs** -- granular per-sub-category overrides (~60 rows).
     Provides more precise costing when the AI classifier assigns a sub-cat.
  3. **Keyword Patterns** -- regex-based cost overrides (highest priority).
     If the ticket description matches a pattern, its costs override all else.
  4. **Severity Multipliers** -- severity-level cost scaling factors.
     Critical=2.5x, High=1.75x, Medium=1.25x, Low=1.0x.
  5. **Origin Premiums** -- percentage adders based on who caused the issue.
     External=20%, Vendor=15%, Customer=10%, Process=5%, Technical=0%.
  6. **Business Multipliers** -- named multipliers for financial metrics:
     revenue_at_risk, opportunity_cost, customer_impact, sla_penalty,
     prevention_rate, cost_avoidance_rate.

=== COSTING FORMULA ===
  Total_Impact = (Material + Labor_Hours x Hourly_Rate)
                 x Severity_Multiplier
                 x (1 + Origin_Premium)

=== LOOKUP PRIORITY ===
When ``calculate_financial_impact()`` is called for a ticket:
  1. Check keyword patterns (regex match against description).  Highest
     priority because these are hand-crafted for known cost outliers.
  2. Check sub-category costs (category + sub_category pair).
  3. Check category costs (category only).
  4. Fall back to the "Unclassified" category row.
  5. If nothing matches, return $0 with a "no_match" source audit trail.

=== DATA FLOW ===
  Input:  price_catalog.xlsx on disk (created from template if missing).
  Load:   PriceCatalog.load_catalog() reads all sheets into Python dicts.
  Use:    Scoring engine calls calculate_financial_impact() per ticket.
          Financial metrics module calls get_business_multiplier() for
          broader financial calculations (ROI, SLA penalties, etc.).
  Output: Per-ticket cost dict with breakdown + audit trail (source field).

Excel-based pricing catalog for calculating financial impact of escalations.
"""

import os
import re
import logging
from typing import Dict, List, Optional, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from ..core.config import PRICE_CATALOG_FILE, DEFAULT_HOURLY_RATE, SUB_CATEGORIES

logger = logging.getLogger(__name__)


class PriceCatalog:
    """
    Excel-based pricing catalog for calculating financial impact of escalations.

    This class is the single source of truth for all financial calculations
    in the Escalation AI pipeline.  It supports:
    - Keyword pattern-based cost overrides (regex)
    - Category and sub-category pricing lookup
    - Severity multipliers (Critical/High/Medium/Low)
    - Origin premiums (External/Vendor/Customer/Process/Technical)
    - Business multipliers for financial metrics calculations

    All data is loaded from price_catalog.xlsx.  If the file does not exist,
    a template is auto-generated with default values.

    Typical usage::

        catalog = get_price_catalog()
        result = catalog.calculate_financial_impact(
            category='Site Readiness',
            sub_category='BH Not Ready',
            severity='Critical',
            origin='External',
            description='BH not actualized in MB'
        )
        print(result['total_impact'])  # Dollar amount
        print(result['source'])        # Audit trail
    """

    def __init__(self, catalog_path: str = PRICE_CATALOG_FILE):
        """Initialise the price catalog.

        Args:
            catalog_path: Path to the price_catalog.xlsx file.  Defaults to
                          the PRICE_CATALOG_FILE constant from config.py.
        """
        self.catalog_path = catalog_path

        # Data stores populated by load_catalog()
        self.category_costs: Dict[str, Dict] = {}           # {category_name: {material_cost, labor_hours, hourly_rate}}
        self.sub_category_costs: Dict[str, Dict[str, Dict]] = {}  # {category: {sub_cat: {material_cost, labor_hours, hourly_rate}}}
        self.keyword_costs: List[Dict] = []                  # [{pattern, category_override, material_cost, labor_hours, priority}]
        self.severity_multipliers: Dict[str, float] = {}     # {severity_lower: multiplier}
        self.origin_premiums: Dict[str, float] = {}          # {origin_lower: premium_fraction}
        self.business_multipliers: Dict[str, float] = {}     # {name_lower: value}
        self.is_loaded = False

    def create_template(self) -> str:
        """Create a blank price catalog Excel template with sample data.

        This is called automatically by ``load_catalog()`` if the file does
        not exist.  It generates a fully-populated template with sensible
        defaults for all 8 categories, ~60 sub-categories, keyword patterns,
        severity multipliers, origin premiums, and business multipliers.

        The template uses $20/hr labour rate and $0 material cost (rework
        labour cost model) across all categories.  Labour hours vary by
        category complexity -- Site Readiness has the most (8-10 hrs) while
        Communication & Response has the least (0.5 hrs).

        Returns:
            str: Path to the created template file.

        Raises:
            Exception: If the file cannot be written (permissions, disk space).
        """
        try:
            wb = Workbook()

            # ==========================================
            # Sheet 1: Instructions
            # ==========================================
            # Human-readable guide explaining the workbook structure and formula.
            ws_instructions = wb.active
            ws_instructions.title = "Instructions"

            instructions = [
                ["PRICE CATALOG - INSTRUCTIONS"],
                [""],
                ["This workbook defines the financial impact calculations for escalation analysis."],
                [""],
                ["SHEETS:"],
                ["1. Category Costs - Base costs per escalation category"],
                ["2. Sub-Category Costs - Granular costs per sub-category"],
                ["3. Keyword Patterns - Regex patterns for specific cost overrides"],
                ["4. Severity Multipliers - Cost multipliers based on severity level"],
                ["5. Origin Premiums - Additional percentage costs based on origin type"],
                ["6. Business Multipliers - Configurable multipliers for business metrics"],
                [""],
                ["FORMULA:"],
                ["Total_Impact = (Material + Labor_Hours × Hourly_Rate) × Severity_Mult × (1 + Origin_Premium)"],
            ]

            for row in instructions:
                ws_instructions.append(row)

            ws_instructions['A1'].font = Font(bold=True, size=14)
            ws_instructions.column_dimensions['A'].width = 80

            # ==========================================
            # Sheet 2: Category Costs
            # ==========================================
            # One row per 8-category classification bucket + "Unclassified" fallback.
            # Material_Cost is $0 for all (rework labour cost model -- no physical materials).
            # Labor_Hours reflects the average rework effort for that category.
            # Hourly_Rate is a blended loaded cost rate ($20/hr default).
            ws_category = wb.create_sheet("Category Costs")
            category_headers = ["Category", "Material_Cost", "Labor_Hours", "Hourly_Rate", "Notes"]
            ws_category.append(category_headers)

            # 8-category system - values match actual price_catalog.xlsx
            # $20/hr rate, $0 material (rework labor cost only)
            category_data = [
                ["Scheduling & Planning", 0, 2.0, 20, "TI scheduling, FE coordination"],
                ["Documentation & Reporting", 0, 1.5, 20, "Snapshots, E911, CBN reports"],
                ["Validation & QA", 0, 3.0, 20, "Precheck, postcheck, VSWR validation"],
                ["Process Compliance", 0, 2.5, 20, "SOP adherence, escalation process"],
                ["Configuration & Data Mismatch", 0, 4.0, 20, "Port matrix, RET, TAC mismatch"],
                ["Site Readiness", 0, 8.0, 20, "BH actualization, MW readiness"],
                ["Communication & Response", 0, 0.5, 20, "Delayed replies, follow-ups"],
                ["Nesting & Tool Errors", 0, 3.0, 20, "NSA/NSI nesting, RIOT, FCI tools"],
                ["Unclassified", 0, 0.2, 20, "Default for unknown"],
            ]

            for row in category_data:
                ws_category.append(row)

            self._style_header(ws_category, len(category_headers))
            ws_category.column_dimensions['A'].width = 30
            ws_category.column_dimensions['F'].width = 35

            # ==========================================
            # Sheet 3: Sub-Category Costs
            # ==========================================
            # Granular per-sub-category pricing.  When both category and sub-category
            # match, these values override the category-level costs.
            # Labour hours are calibrated per sub-type complexity:
            #   - Simple doc issues: 1.0-1.5 hrs
            #   - Moderate rework: 2.0-4.0 hrs
            #   - Complex site/config issues: 5.0-10.0 hrs
            ws_subcat = wb.create_sheet("Sub-Category Costs")
            subcat_headers = ["Category", "Sub_Category", "Material_Cost", "Labor_Hours", "Hourly_Rate", "Notes"]
            ws_subcat.append(subcat_headers)

            # Sub-category costs - granular pricing per sub-type
            # Values use $20/hr base rate; labor hours vary by complexity
            subcat_data = [
                # --- Scheduling & Planning ---
                ["Scheduling & Planning", "No TI Entry", 0, 3.0, 20, "Site not scheduled in TI"],
                ["Scheduling & Planning", "Schedule Not Followed", 0, 2.0, 20, "FE logged on wrong day"],
                ["Scheduling & Planning", "Weekend Schedule Issue", 0, 3.0, 20, "Weekend scheduling error"],
                ["Scheduling & Planning", "Ticket Status Issue", 0, 1.5, 20, "Ticket in wrong bucket/closeout"],
                ["Scheduling & Planning", "Premature Scheduling", 0, 4.0, 20, "Scheduled without BH/MW ready"],
                # --- Documentation & Reporting ---
                ["Documentation & Reporting", "Missing Snapshot", 0, 1.5, 20, "CBN/validation snapshot missing"],
                ["Documentation & Reporting", "Missing Attachment", 0, 1.0, 20, "Pre-check logs not attached"],
                ["Documentation & Reporting", "Incorrect Reporting", 0, 2.0, 20, "RTT/data incorrectly reported"],
                ["Documentation & Reporting", "Wrong Site ID", 0, 2.0, 20, "Different site ID in mail"],
                ["Documentation & Reporting", "Incomplete Snapshot", 0, 1.5, 20, "Lemming snap missing sectors"],
                ["Documentation & Reporting", "Missing Information", 0, 1.5, 20, "PSAP/Live CT details missing"],
                ["Documentation & Reporting", "Wrong Attachment", 0, 1.0, 20, "Wrong file attached"],
                ["Documentation & Reporting", "Incorrect Status", 0, 2.5, 20, "Pass/Fail status wrong"],
                # --- Validation & QA ---
                ["Validation & QA", "Incomplete Validation", 0, 3.5, 20, "BH fields not checked"],
                ["Validation & QA", "Missed Issue", 0, 4.0, 20, "SFP/fiber issue not identified"],
                ["Validation & QA", "Missed Check", 0, 3.0, 20, "Cell status not verified"],
                ["Validation & QA", "No Escalation", 0, 2.5, 20, "Issue captured but not escalated"],
                ["Validation & QA", "Missed Degradation", 0, 5.0, 20, "KPI degradation not detected"],
                ["Validation & QA", "Wrong Tool Usage", 0, 2.5, 20, "AEHC swap report misused"],
                ["Validation & QA", "Incomplete Testing", 0, 3.5, 20, "E911/VoNR testing incomplete"],
                # --- Process Compliance ---
                ["Process Compliance", "Process Violation", 0, 4.0, 20, "IX without BH actualized"],
                ["Process Compliance", "Wrong Escalation", 0, 3.0, 20, "Escalated to wrong vendor/NTAC"],
                ["Process Compliance", "Wrong Bucket", 0, 2.5, 20, "Ticket in preliminary design"],
                ["Process Compliance", "Missed Step", 0, 3.0, 20, "Forgot to unlock/share precheck"],
                ["Process Compliance", "Missing Ticket", 0, 2.5, 20, "PAG ticket not created"],
                ["Process Compliance", "Process Non-Compliance", 0, 3.5, 20, "Guidelines not followed"],
                # --- Configuration & Data Mismatch ---
                ["Configuration & Data Mismatch", "Port Matrix Mismatch", 0, 5.0, 20, "RET count mismatch with PMX"],
                ["Configuration & Data Mismatch", "RET Naming", 0, 4.0, 20, "Extra/missing letter in naming"],
                ["Configuration & Data Mismatch", "RET Swap", 0, 5.0, 20, "Alpha/Beta RET swapped"],
                ["Configuration & Data Mismatch", "TAC Mismatch", 0, 4.0, 20, "TAC causing RIOT red"],
                ["Configuration & Data Mismatch", "CIQ/SCF Mismatch", 0, 5.0, 20, "CIQ and SCF not matching"],
                ["Configuration & Data Mismatch", "RFDS Mismatch", 0, 5.0, 20, "RFDS and SCF mismatch"],
                ["Configuration & Data Mismatch", "Missing Documents", 0, 3.0, 20, "RFDS/Port Matrix missing in TI"],
                ["Configuration & Data Mismatch", "Config Error", 0, 4.0, 20, "NRPLMNSET/BWP not defined"],
                # --- Site Readiness ---
                # These have the highest labour hours because site-level issues
                # often require truck rolls, FE re-dispatches, and multi-team coordination.
                ["Site Readiness", "BH Not Ready", 0, 10.0, 20, "Backhaul not actualized in MB"],
                ["Site Readiness", "MW Not Ready", 0, 8.0, 20, "Microwave link not ready"],
                ["Site Readiness", "Material Missing", 0, 6.0, 20, "SFP/AMID not available"],
                ["Site Readiness", "Site Down", 0, 8.0, 20, "Site shut down during IX"],
                ["Site Readiness", "BH Status Issue", 0, 5.0, 20, "BH status incorrectly filled"],
                ["Site Readiness", "Site Complexity", 0, 8.0, 20, "MW chain/generator issues"],
                # --- Communication & Response ---
                # Lowest labour hours: communication issues are quick to resolve
                # once identified, but they cause cascading delays upstream.
                ["Communication & Response", "Delayed Response", 0, 0.5, 20, "Late reply to GC/FE query"],
                ["Communication & Response", "Delayed Deliverable", 0, 1.0, 20, "FE waited 3-4hrs for EOD"],
                ["Communication & Response", "No Proactive Update", 0, 0.5, 20, "Not answering delay queries"],
                ["Communication & Response", "No Communication", 0, 0.5, 20, "Schedule not communicated to FE"],
                ["Communication & Response", "Training Issue", 0, 2.0, 20, "FE competency/knowledge gap"],
                # --- Nesting & Tool Errors ---
                ["Nesting & Tool Errors", "Wrong Nest Type", 0, 4.0, 20, "Nested as NSA when SA required"],
                ["Nesting & Tool Errors", "Improper Extension", 0, 3.0, 20, "Nest extended during follow-up"],
                ["Nesting & Tool Errors", "Missing Nesting", 0, 3.0, 20, "Site not nested before activity"],
                ["Nesting & Tool Errors", "HW Issue", 0, 5.0, 20, "GPS SFP/RET antenna failure"],
                ["Nesting & Tool Errors", "Rework", 0, 5.0, 20, "SCF prep and IX rework needed"],
                ["Nesting & Tool Errors", "Post-OA Degradation", 0, 6.0, 20, "Congestion/DCR after on-air"],
                ["Nesting & Tool Errors", "Delayed Audit", 0, 2.0, 20, "Audit done 5 days late"],
            ]

            for row in subcat_data:
                ws_subcat.append(row)

            self._style_header(ws_subcat, len(subcat_headers))
            ws_subcat.column_dimensions['A'].width = 30
            ws_subcat.column_dimensions['B'].width = 25
            ws_subcat.column_dimensions['F'].width = 30

            # ==========================================
            # Sheet 4: Keyword Patterns
            # ==========================================
            # Regex-based cost overrides.  These fire BEFORE category/sub-category
            # lookups.  Sorted by Priority (lower number = higher priority).
            # Use case: known expensive issues (e.g. "BH not actualized") that
            # should always get a specific cost regardless of how the AI classifies them.
            ws_keywords = wb.create_sheet("Keyword Patterns")
            keyword_headers = ["Keyword_Pattern", "Category_Override", "Material_Cost", "Labor_Hours", "Priority", "Notes"]
            ws_keywords.append(keyword_headers)

            # Keyword patterns for cost overrides - $20/hr, $0 material
            keyword_data = [
                [".*BH.*not.*actual.*", "Site Readiness", 0, 10, 1, "Backhaul not actualized"],
                [".*port.*matrix.*mismatch.*", "Configuration & Data Mismatch", 0, 6, 1, "Port matrix issues"],
                [".*RET.*naming.*", "Configuration & Data Mismatch", 0, 4, 2, "RET naming errors"],
                [".*not.*schedul.*", "Scheduling & Planning", 0, 3, 2, "Scheduling issues"],
                [".*snapshot.*missing.*", "Documentation & Reporting", 0, 1.5, 3, "Missing documentation"],
                [".*RIOT.*red.*", "Nesting & Tool Errors", 0, 4, 2, "RIOT validation failure"],
            ]

            for row in keyword_data:
                ws_keywords.append(row)

            self._style_header(ws_keywords, len(keyword_headers))

            # ==========================================
            # Sheet 5: Severity Multipliers
            # ==========================================
            # Cost scaling based on ticket severity.  Higher severity = higher
            # total cost because the business impact and urgency of rework increases.
            ws_severity = wb.create_sheet("Severity Multipliers")
            severity_headers = ["Severity_Level", "Cost_Multiplier", "Description"]
            ws_severity.append(severity_headers)

            severity_data = [
                ["Critical", 2.5, "Network down, major outage"],
                ["High", 1.75, "Significant degradation"],
                ["Medium", 1.25, "Moderate impact"],
                ["Low", 1.0, "Minor issue, no multiplier"],
            ]

            for row in severity_data:
                ws_severity.append(row)

            self._style_header(ws_severity, len(severity_headers))

            # ==========================================
            # Sheet 6: Origin Premiums
            # ==========================================
            # Percentage-based cost adders based on the origin/cause of the issue.
            # External and vendor issues carry premiums because they typically
            # involve more coordination, SLA implications, and customer impact.
            ws_origin = wb.create_sheet("Origin Premiums")
            origin_headers = ["Origin_Type", "Premium_Percentage", "Description"]
            ws_origin.append(origin_headers)

            origin_data = [
                ["Vendor", 0.15, "15% premium for vendor-caused issues"],
                ["Process", 0.05, "5% premium for internal process failures"],
                ["External", 0.20, "20% premium for external/uncontrollable factors"],
                ["Customer", 0.10, "10% premium for customer-initiated issues"],
                ["Technical", 0.0, "No premium for standard technical issues"],
            ]

            for row in origin_data:
                ws_origin.append(row)

            self._style_header(ws_origin, len(origin_headers))

            # ==========================================
            # Sheet 7: Business Multipliers
            # ==========================================
            # Named multipliers consumed by the financial metrics module
            # (escalation_ai.financial.metrics) for broader financial analysis.
            # These translate raw escalation costs into business-level exposure.
            ws_biz = wb.create_sheet("Business Multipliers")
            biz_headers = ["Multiplier_Name", "Value", "Description"]
            ws_biz.append(biz_headers)

            biz_data = [
                # revenue_at_risk: $1 of escalation cost implies $2.50 of revenue exposure
                # (customer churn, SLA credits, downstream rework)
                ["revenue_at_risk", 2.5, "Multiplier for downstream revenue impact (total_cost × this)"],

                # opportunity_cost: 35% of escalation cost represents productive work
                # that could have been done instead of fighting fires
                ["opportunity_cost", 0.35, "Opportunity cost as fraction of total cost"],

                # customer_impact: external/customer-facing issues carry 1.5x premium
                # due to reputation risk and potential contract penalties
                ["customer_impact", 1.5, "Multiplier for customer-facing issue costs"],

                # sla_penalty: 20% of critical-ticket costs are at risk of SLA penalties
                ["sla_penalty", 0.2, "SLA penalty rate for critical issues"],

                # prevention_rate: 80% of preventable costs can realistically be
                # eliminated through process improvements (conservative estimate)
                ["prevention_rate", 0.8, "Expected prevention rate for ROI calculation (0-1)"],

                # cost_avoidance_rate: 70% of repeat-pattern costs can be avoided
                # by investing in root-cause fixes
                ["cost_avoidance_rate", 0.7, "Expected cost avoidance rate for repeat issues (0-1)"],
            ]

            for row in biz_data:
                ws_biz.append(row)

            self._style_header(ws_biz, len(biz_headers))
            ws_biz.column_dimensions['A'].width = 25
            ws_biz.column_dimensions['C'].width = 55

            wb.save(self.catalog_path)
            logger.info(f"✓ Price catalog template created: {self.catalog_path}")
            return self.catalog_path

        except Exception as e:
            logger.error(f"Failed to create price catalog template: {e}")
            raise

    def _style_header(self, ws, col_count):
        """Apply header styling to a worksheet.

        Fills the first row with a dark blue background and white bold text
        to visually distinguish headers from data rows.

        Args:
            ws:        The openpyxl worksheet to style.
            col_count: Number of columns in the header row.
        """
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

    def load_catalog(self) -> bool:
        """Load pricing data from the Excel catalog into memory.

        Reads all six data sheets from price_catalog.xlsx and populates:
        - self.category_costs       (Category Costs sheet)
        - self.keyword_costs        (Keyword Patterns sheet, sorted by priority)
        - self.severity_multipliers (Severity Multipliers sheet)
        - self.sub_category_costs   (Sub-Category Costs sheet)
        - self.origin_premiums      (Origin Premiums sheet)
        - self.business_multipliers (Business Multipliers sheet)

        If the file does not exist, ``create_template()`` is called first to
        generate a default template so the pipeline can still run.

        Returns:
            True on success, False on failure.  On failure, the catalog
            remains in its previous state (or empty if never loaded).
        """
        if not os.path.exists(self.catalog_path):
            logger.warning(f"Price catalog not found: {self.catalog_path}")
            logger.info("Creating template price catalog...")
            self.create_template()

        try:
            # data_only=True reads cached formula results rather than formulas
            wb = load_workbook(self.catalog_path, data_only=True)

            # ----------------------------------------------------------
            # Load Category Costs (base cost per 8-category bucket)
            # ----------------------------------------------------------
            if "Category Costs" in wb.sheetnames:
                ws = wb["Category Costs"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Skip empty rows
                        category = str(row[0]).strip()
                        self.category_costs[category] = {
                            'material_cost': float(row[1] or 0),
                            'labor_hours': float(row[2] or 0),
                            'hourly_rate': float(row[3] or DEFAULT_HOURLY_RATE),
                        }
                logger.info(f"  → Loaded {len(self.category_costs)} category cost entries")

            # ----------------------------------------------------------
            # Load Keyword Patterns (regex-based cost overrides)
            # ----------------------------------------------------------
            # Sorted by priority (ascending) so that higher-priority rules
            # are checked first in _match_keyword_pattern().
            if "Keyword Patterns" in wb.sheetnames:
                ws = wb["Keyword Patterns"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Skip empty rows
                        self.keyword_costs.append({
                            'pattern': str(row[0]).strip(),
                            'category_override': str(row[1] or '').strip(),
                            'material_cost': float(row[2] or 0),
                            'labor_hours': float(row[3] or 0),
                            'priority': int(row[4] or 99),
                        })
                # Sort: lower priority number = checked first
                self.keyword_costs.sort(key=lambda x: x['priority'])
                logger.info(f"  → Loaded {len(self.keyword_costs)} keyword pattern rules")

            # ----------------------------------------------------------
            # Load Severity Multipliers
            # ----------------------------------------------------------
            # Keys are stored lowercase for case-insensitive lookup.
            if "Severity Multipliers" in wb.sheetnames:
                ws = wb["Severity Multipliers"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        severity = str(row[0]).strip().lower()
                        self.severity_multipliers[severity] = float(row[1] or 1.0)
                logger.info(f"  → Loaded {len(self.severity_multipliers)} severity multipliers")

            # ----------------------------------------------------------
            # Load Sub-Category Costs
            # ----------------------------------------------------------
            # Nested dict: {category: {sub_category: {costs...}}}
            if "Sub-Category Costs" in wb.sheetnames:
                ws = wb["Sub-Category Costs"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:  # Both category and sub-category required
                        category = str(row[0]).strip()
                        sub_cat = str(row[1]).strip()
                        if category not in self.sub_category_costs:
                            self.sub_category_costs[category] = {}
                        self.sub_category_costs[category][sub_cat] = {
                            'material_cost': float(row[2] or 0),
                            'labor_hours': float(row[3] or 0),
                            'hourly_rate': float(row[4] or DEFAULT_HOURLY_RATE),
                        }
                total_sub = sum(len(v) for v in self.sub_category_costs.values())
                logger.info(f"  → Loaded {total_sub} sub-category cost entries across {len(self.sub_category_costs)} categories")

            # ----------------------------------------------------------
            # Load Origin Premiums
            # ----------------------------------------------------------
            # Keys stored lowercase for case-insensitive lookup.
            if "Origin Premiums" in wb.sheetnames:
                ws = wb["Origin Premiums"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        origin = str(row[0]).strip().lower()
                        self.origin_premiums[origin] = float(row[1] or 0)
                logger.info(f"  → Loaded {len(self.origin_premiums)} origin premium rules")

            # ----------------------------------------------------------
            # Load Business Multipliers
            # ----------------------------------------------------------
            # Named multipliers used by the financial metrics module.
            # Keys stored lowercase for case-insensitive lookup.
            if "Business Multipliers" in wb.sheetnames:
                ws = wb["Business Multipliers"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        name = str(row[0]).strip().lower()
                        self.business_multipliers[name] = float(row[1] or 1.0)
                logger.info(f"  → Loaded {len(self.business_multipliers)} business multipliers")

            wb.close()
            self.is_loaded = True
            logger.info(f"✓ Price catalog loaded successfully")
            return True

        except Exception as e:
            logger.error(f"Failed to load price catalog: {e}")
            return False

    def _match_keyword_pattern(self, text: str) -> Optional[Dict]:
        """Check if text matches any keyword pattern (regex).

        Iterates through keyword_costs (sorted by priority, lowest first)
        and returns the first matching rule.  Uses case-insensitive regex
        matching via ``re.search()``.

        Args:
            text: The ticket description to match against patterns.

        Returns:
            The matching keyword rule dict, or None if no pattern matches.
        """
        if not text:
            return None

        text_lower = text.lower()
        for rule in self.keyword_costs:
            try:
                if re.search(rule['pattern'], text_lower, re.IGNORECASE):
                    return rule
            except re.error:
                # Invalid regex in the spreadsheet; log and skip
                logger.warning(f"Invalid regex pattern: {rule['pattern']}")
        return None

    def calculate_financial_impact(
        self,
        category: str,
        severity: str = "Medium",
        origin: str = "Technical",
        description: str = "",
        sub_category: str = "",
        **kwargs
    ) -> Dict[str, Any]:
        """Calculate total financial impact (rework cost) for an escalation.

        This is the main costing method called by the scoring engine for
        every ticket.  It applies the following formula:

            Total = (Material + Labor_Hours x Hourly_Rate)
                    x Severity_Multiplier
                    x (1 + Origin_Premium)

        The base costs (material, labour) are resolved through a priority-
        based lookup chain:

        1. **Keyword pattern match** (highest priority):
           If the ticket description matches a regex in the Keyword Patterns
           sheet, use that rule's costs.  This catches known expensive issues.
        2. **Sub-category cost**:
           If both category and sub_category are provided and found in the
           Sub-Category Costs sheet, use those granular costs.
        3. **Category cost**:
           If the category exists in the Category Costs sheet, use those costs.
        4. **Unclassified fallback**:
           Use the "Unclassified" row from Category Costs as a last resort.
        5. **No match**:
           Return $0 with a "no_match" source for audit trail transparency.

        No delay cost is included -- this is a rework-only cost model because
        the dataset does not contain reliable delay duration data.

        Args:
            category:     AI-assigned category (one of the 8 categories).
            severity:     Ticket severity level (Critical/High/Medium/Low).
            origin:       Origin/cause type (External/Vendor/Customer/Process/Technical).
            description:  Free-text ticket description (for keyword matching).
            sub_category: AI-assigned sub-category (optional, for granular lookup).
            **kwargs:     Reserved for future expansion.

        Returns:
            Dict with complete cost breakdown and audit trail::

                {
                    'material_cost': float,      # Material component
                    'labor_cost': float,          # Labor_Hours x Hourly_Rate
                    'labor_hours': float,         # Raw hours
                    'hourly_rate': float,         # Rate used
                    'base_cost': float,           # material + labor (pre-multipliers)
                    'severity_multiplier': float, # Multiplier applied
                    'origin_premium': float,      # Premium fraction applied
                    'total_impact': float,        # Final dollar amount
                    'source': str,                # Audit trail (which lookup was used)
                    'keyword_match': str|None,    # Matched pattern if keyword hit
                }
        """
        # Ensure catalog is loaded before computing
        if not self.is_loaded:
            self.load_catalog()

        source = 'unknown'

        # ----------------------------------------------------------
        # Priority 1: Keyword pattern match (regex against description)
        # ----------------------------------------------------------
        keyword_match = self._match_keyword_pattern(description)

        if keyword_match:
            # Keyword match overrides all other lookups
            material_cost = keyword_match['material_cost']
            labor_hours = keyword_match['labor_hours']
            hourly_rate = DEFAULT_HOURLY_RATE
            source = f"keyword:{keyword_match['pattern']}"

        # ----------------------------------------------------------
        # Priority 2: Sub-category cost (category + sub-category pair)
        # ----------------------------------------------------------
        elif (sub_category and category in self.sub_category_costs
              and sub_category in self.sub_category_costs[category]):
            sub_costs = self.sub_category_costs[category][sub_category]
            material_cost = sub_costs.get('material_cost', 0)
            labor_hours = sub_costs.get('labor_hours', 2)
            hourly_rate = sub_costs.get('hourly_rate', DEFAULT_HOURLY_RATE)
            source = f"sub_category:{category}/{sub_category}"

        # ----------------------------------------------------------
        # Priority 3: Category cost (category-level fallback)
        # ----------------------------------------------------------
        elif category in self.category_costs:
            cat_costs = self.category_costs[category]
            material_cost = cat_costs.get('material_cost', 0)
            labor_hours = cat_costs.get('labor_hours', 2)
            hourly_rate = cat_costs.get('hourly_rate', DEFAULT_HOURLY_RATE)
            source = f"category:{category}"

        # ----------------------------------------------------------
        # Priority 4: Unclassified fallback
        # ----------------------------------------------------------
        elif 'Unclassified' in self.category_costs:
            cat_costs = self.category_costs['Unclassified']
            material_cost = cat_costs.get('material_cost', 0)
            labor_hours = cat_costs.get('labor_hours', 2)
            hourly_rate = cat_costs.get('hourly_rate', DEFAULT_HOURLY_RATE)
            source = f"fallback:Unclassified (requested:{category})"

        # ----------------------------------------------------------
        # Priority 5: No match at all
        # ----------------------------------------------------------
        else:
            # Catalog loaded but category not found and no Unclassified - fail loud
            logger.warning(f"No pricing data for category '{category}' and no Unclassified fallback")
            material_cost = 0
            labor_hours = 0
            hourly_rate = DEFAULT_HOURLY_RATE
            source = f"no_match:{category}"

        # ----------------------------------------------------------
        # Compute rework cost: material + labour (no delay component)
        # ----------------------------------------------------------
        labor_cost = labor_hours * hourly_rate
        base_cost = material_cost + labor_cost

        # ----------------------------------------------------------
        # Apply severity multiplier
        # ----------------------------------------------------------
        # Lookup is case-insensitive (keys stored lowercase)
        severity_key = severity.lower() if severity else 'medium'
        severity_mult = self.severity_multipliers.get(severity_key, 1.0)
        if severity_key not in self.severity_multipliers:
            logger.warning(f"Unknown severity '{severity}', using multiplier 1.0")

        # ----------------------------------------------------------
        # Apply origin premium
        # ----------------------------------------------------------
        # Origin premium is additive: (1 + premium).  E.g. External = 1.20x
        origin_key = origin.lower() if origin else 'technical'
        origin_premium = self.origin_premiums.get(origin_key, 0.0)

        # ----------------------------------------------------------
        # Final calculation: base x severity x (1 + origin_premium)
        # ----------------------------------------------------------
        total_impact = base_cost * severity_mult * (1 + origin_premium)

        # Sanity check: cost should never be negative
        if total_impact < 0:
            logger.error(f"Negative financial impact calculated: {total_impact} for {category}/{sub_category}")
            total_impact = 0

        return {
            'material_cost': round(material_cost, 2),
            'labor_cost': round(labor_cost, 2),
            'labor_hours': labor_hours,
            'hourly_rate': hourly_rate,
            'base_cost': round(base_cost, 2),
            'severity_multiplier': severity_mult,
            'origin_premium': origin_premium,
            'total_impact': round(total_impact, 2),
            'source': source,
            'keyword_match': keyword_match['pattern'] if keyword_match else None,
        }

    def get_business_multiplier(self, name: str, default: float = 1.0) -> float:
        """Get a business multiplier from price_catalog.xlsx.

        Business multipliers are used by the financial metrics module to
        translate raw escalation costs into broader financial exposure
        (revenue at risk, opportunity cost, SLA penalties, etc.).

        Args:
            name:    Multiplier name (case-insensitive).
            default: Fallback value if the multiplier is not found.

        Returns:
            The multiplier value, or ``default`` if not configured.
        """
        if not self.is_loaded:
            self.load_catalog()
        return self.business_multipliers.get(name.lower(), default)

    def get_benchmark_costs(self) -> Dict[str, float]:
        """Calculate benchmark cost statistics from the loaded catalog.

        Computes industry-comparison metrics by averaging across all
        category costs and severity multipliers in the loaded catalog.

        Benchmark definitions:
        - avg_per_ticket:  average base cost x average severity multiplier
        - best_in_class:   lowest base cost across all categories
        - industry_avg:    same as avg_per_ticket (catalog represents industry norms)
        - laggard:         highest base cost x highest severity multiplier
        - hourly_rate:     the hourly rate from the first category entry

        Returns:
            Dict with avg_per_ticket, best_in_class, industry_avg, laggard,
            and hourly_rate.  All derived from price_catalog.xlsx data.
        """
        if not self.is_loaded:
            self.load_catalog()

        if not self.category_costs:
            return {'avg_per_ticket': 0, 'best_in_class': 0, 'industry_avg': 0, 'laggard': 0, 'hourly_rate': DEFAULT_HOURLY_RATE}

        # Compute base cost (material + labour) for each category
        base_costs = []
        for cat_data in self.category_costs.values():
            material = cat_data.get('material_cost', 0)
            labor = cat_data.get('labor_hours', 0) * cat_data.get('hourly_rate', DEFAULT_HOURLY_RATE)
            base_costs.append(material + labor)

        # Average base cost across all categories
        avg_base = sum(base_costs) / len(base_costs) if base_costs else 0

        # Average severity multiplier (blended across all severity levels)
        avg_sev = sum(self.severity_multipliers.values()) / len(self.severity_multipliers) if self.severity_multipliers else 1.0

        return {
            'avg_per_ticket': round(avg_base * avg_sev, 2),
            'best_in_class': round(min(base_costs), 2),
            'industry_avg': round(avg_base * avg_sev, 2),
            'laggard': round(max(base_costs) * max(self.severity_multipliers.values(), default=1.0), 2),
            'hourly_rate': next(iter(self.category_costs.values()), {}).get('hourly_rate', DEFAULT_HOURLY_RATE),
        }

    def get_catalog_summary(self) -> str:
        """Get summary of loaded pricing data for AI context.

        Produces a concise text summary of what is loaded in the catalog,
        suitable for injection into AI prompts or log messages.

        Returns:
            Multi-line string summarising the loaded catalog state.
            Returns "Price catalog not loaded." if load_catalog() has not
            been called or failed.
        """
        if not self.is_loaded:
            return "Price catalog not loaded."

        total_sub = sum(len(v) for v in self.sub_category_costs.values())
        summary_lines = [
            f"PRICE CATALOG SUMMARY:",
            f"- {len(self.category_costs)} category cost entries",
            f"- {total_sub} sub-category cost entries",
            f"- {len(self.keyword_costs)} keyword pattern rules",
            f"- {len(self.severity_multipliers)} severity multipliers",
            f"- {len(self.origin_premiums)} origin premium rules",
            f"- {len(self.business_multipliers)} business multipliers",
        ]

        return "\n".join(summary_lines)


# ==========================================
# SINGLETON ACCESS
# ==========================================
# The price catalog is a module-level singleton.  All modules that need
# pricing data call get_price_catalog() which returns the same instance.
# This ensures the Excel file is read at most once per load_catalog() call
# and all modules see consistent pricing data.

# Global price catalog instance (lazily initialised by get_price_catalog())
_price_catalog: Optional[PriceCatalog] = None

def get_price_catalog() -> PriceCatalog:
    """Get or create the price catalog singleton.

    On first call, creates a new PriceCatalog instance (but does NOT load
    it -- callers must call ``load_catalog()`` explicitly or rely on
    ``calculate_financial_impact()`` to auto-load on first use).

    Returns:
        The global PriceCatalog singleton instance.
    """
    global _price_catalog
    if _price_catalog is None:
        _price_catalog = PriceCatalog()
    return _price_catalog
