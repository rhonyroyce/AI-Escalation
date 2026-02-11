"""
Human-in-the-loop feedback and learning system.
Enables iterative improvement of AI classification through user corrections.

Architecture Overview
=====================
This module implements two independent feedback loops that close the gap between
AI predictions and domain-expert knowledge:

1. **FeedbackLearning** -- Classification Feedback Loop
   - *Output artefact*: ``classification_feedback.xlsx`` (multi-sheet workbook)
   - *Editable column*: ``Corrected_Category`` (dropdown referencing the
     "Category Reference" sheet so the user can also **add** categories)
   - *Learning mechanism*: **centroid adjustment** -- user corrections are
     embedded and blended with the original anchor centroids using a weighted
     average controlled by ``FEEDBACK_WEIGHT`` (from config):

       adjusted = (1 - w) * anchor_centroid + w * feedback_centroid

     where ``w = FEEDBACK_WEIGHT`` (typically 0.3) and ``feedback_centroid``
     is the mean embedding of all user-corrected examples for that category.
   - Entirely new categories added to the "Category Reference" tab are
     bootstrapped from their keywords -- each keyword is embedded and the
     mean embedding becomes the new category centroid.

2. **ResolutionFeedbackLearning** -- Resolution Time Feedback Loop
   - *Output artefact*: ``resolution_feedback.xlsx`` (multi-sheet workbook)
   - *Editable columns*: ``Human_Expected_Days`` and ``Actual_Resolution_Days``
   - *Learning mechanism*: the resolution-time predictor can query this
     feedback at inference time to incorporate human estimates (see
     ``resolution_time.py``).

Both systems follow a **non-destructive merge** pattern:
- On every pipeline run the Excel files are regenerated with the latest
  predictions, but any values the user has already entered in the editable
  columns are preserved by matching on a stable row key (ticket ID + text
  prefix for classifications, ``Identity`` for resolution feedback).
- This allows continuous, incremental learning across runs without the
  user losing previous work.

Data Flow
---------
::

  Pipeline run N
    ├── save_for_review(df, output_dir)   ← exports predictions + preserves
    │       │                                   existing user edits
    │       └──→  classification_feedback.xlsx / resolution_feedback.xlsx
    │
  [User reviews and edits the Excel file]
    │
  Pipeline run N+1
    ├── load_feedback(ai)                 ← reads corrections, embeds them
    ├── adjust_centroids(original)        ← blends feedback into anchors
    │       │
    │       └──→  improved centroids used by Classifier
    └── save_for_review(...)              ← exports updated predictions,
                                              still preserving user edits

Module Dependencies
-------------------
- ``openpyxl`` for Excel I/O (reading/writing), formatting (fonts, fills,
  data-validation dropdowns), and freeze-panes.
- ``core.config`` for path constants, embedding weight, anchor definitions,
  sub-category taxonomy, report versioning, and confidence thresholds.
"""

import os
import logging
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from ..core.config import (
    FEEDBACK_FILE, RESOLUTION_FEEDBACK_FILE, FEEDBACK_WEIGHT, ANCHORS, SUB_CATEGORIES,
    REPORT_VERSION, MIN_CLASSIFICATION_CONFIDENCE
)

logger = logging.getLogger(__name__)


# =============================================================================
# Classification Feedback Loop
# =============================================================================

class FeedbackLearning:
    """
    Human-in-the-loop learning system for classification improvement.

    This class manages the full lifecycle of classification feedback:

    1. **Export** -- ``save_for_review()`` writes the current AI classifications
       to a formatted Excel workbook with dropdown-based correction columns.
    2. **Ingest** -- ``load_feedback()`` reads user corrections and custom
       categories from the workbook, computes their embeddings via the AI
       provider, and stores them in ``self.corrections``.
    3. **Adjust** -- ``adjust_centroids()`` blends the user-correction
       embeddings into the original anchor centroids so the classifier can
       incorporate domain-expert knowledge on the next run.

    Workflow (from the user's perspective):
        1. Run the analysis pipeline -- generates ``classification_feedback.xlsx``.
        2. Open the workbook and review AI categories in the "Classifications"
           sheet; select the correct category from the ``Corrected_Category``
           dropdown where the AI was wrong.
        3. Optionally add **new** categories in the "Category Reference" tab
           (name + comma-separated example keywords).
        4. Save the workbook and re-run the pipeline -- the AI will load the
           corrections, re-embed the texts, and shift the category centroids
           toward the user's examples.

    Attributes:
        feedback_path (str):
            Filesystem path to the ``classification_feedback.xlsx`` workbook.
            Defaults to the value of ``FEEDBACK_FILE`` from config.
        corrections (Dict[str, List[Dict]]):
            Mapping from category name to a list of correction records, each
            containing ``text`` (str), ``embedding`` (np.ndarray), and
            ``original`` (str -- the AI's original category or 'custom_category'
            for keyword-bootstrapped entries).
        custom_categories (Dict[str, List[str]]):
            Categories that the user has added to the "Category Reference" tab
            which are **not** already in the built-in ``ANCHORS`` dict.
        stats (dict):
            Tracking counters: ``loaded`` (total corrections ingested),
            ``categories_enhanced`` (distinct categories with at least one
            correction), ``custom_categories`` (user-created categories).
    """

    def __init__(self, feedback_path: Optional[str] = None):
        self.feedback_path = feedback_path or FEEDBACK_FILE
        # corrections: category_name -> [{text, embedding, original}, ...]
        self.corrections: Dict[str, List[Dict]] = {}  # category -> list of {text, embedding}
        # custom_categories: category_name -> [keyword_strings] (from Category Reference tab)
        self.custom_categories: Dict[str, List[str]] = {}  # category -> list of keywords (from Category Reference)
        self.stats = {'loaded': 0, 'categories_enhanced': 0, 'custom_categories': 0}

    # -------------------------------------------------------------------------
    # Category Reference Loading
    # -------------------------------------------------------------------------

    def load_category_reference(self) -> Dict[str, List[str]]:
        """
        Load categories from the "Category Reference" tab of the feedback workbook.

        This method merges the built-in ``ANCHORS`` (hardcoded in config) with
        any categories the user has manually added to the Excel file.  For each
        row in the "Category Reference" sheet:

        - If the category name is **not** in ``ANCHORS``, it is treated as a
          brand-new user-defined category.  Its keywords (comma-separated in
          the ``Example_Keywords`` column) are stored; if no keywords are
          provided the category name itself (lowercased) is used as a fallback.
        - If the category name **is** already in ``ANCHORS``, only genuinely
          *new* keywords (those not already in the anchor list) are appended.

        Returns:
            Dict[str, List[str]]: Merged dictionary mapping every known
            category name to its keyword list (anchors + user additions).
        """
        # Start with a mutable copy of the built-in anchor keywords
        all_categories = {cat: list(keywords) for cat, keywords in ANCHORS.items()}

        if not os.path.exists(self.feedback_path):
            return all_categories

        try:
            df_ref = pd.read_excel(self.feedback_path, sheet_name='Category Reference')

            for _, row in df_ref.iterrows():
                cat_name = str(row.get('Available_Categories', '')).strip()
                keywords_str = str(row.get('Example_Keywords', '')).strip()

                if cat_name and cat_name.lower() not in ['', 'nan', 'none']:
                    # Parse the comma-separated keyword string into a list
                    keywords = [k.strip() for k in keywords_str.split(',') if k.strip()]

                    if cat_name not in all_categories:
                        # ---- Brand-new user-added category ----
                        # If the user didn't supply keywords, fall back to the
                        # lowercased category name as a minimal keyword anchor.
                        all_categories[cat_name] = keywords if keywords else [cat_name.lower()]
                        self.custom_categories[cat_name] = keywords if keywords else [cat_name.lower()]
                        logger.info(f"  → Loaded custom category '{cat_name}' with {len(keywords)} keywords")
                    elif keywords and cat_name in ANCHORS:
                        # ---- Existing category with potential new keywords ----
                        # Deduplicate: only append keywords that aren't already
                        # part of the original anchor set.
                        existing_keywords = set(ANCHORS[cat_name])
                        new_keywords = [k for k in keywords if k not in existing_keywords]
                        if new_keywords:
                            all_categories[cat_name] = list(ANCHORS[cat_name]) + new_keywords

            if self.custom_categories:
                self.stats['custom_categories'] = len(self.custom_categories)
                logger.info(f"✓ Loaded {len(self.custom_categories)} custom categories from Category Reference")

        except Exception as e:
            logger.warning(f"Could not load Category Reference: {e}")

        return all_categories

    # -------------------------------------------------------------------------
    # Feedback Ingestion
    # -------------------------------------------------------------------------

    def load_feedback(self, ai) -> bool:
        """
        Load user feedback from the Excel workbook and compute embeddings.

        This is the primary ingestion entry-point called early in the pipeline.
        It performs two distinct loading phases:

        **Phase 1 -- Custom Categories (from "Category Reference" tab)**

        For each user-added category (i.e. not already in ``ANCHORS``), the
        method calls ``ai.get_embeddings_batch(keywords)`` to convert the
        keyword strings into embedding vectors.  These embeddings are stored
        in ``self.corrections`` exactly like user corrections, but with
        ``original='custom_category'`` so they can be distinguished if needed.

        **Phase 2 -- Classification Corrections (from "Classifications" tab)**

        Each row where ``Corrected_Category`` differs from ``AI_Category`` is
        treated as a user correction.  The corrected text is embedded and
        grouped by corrected category in ``self.corrections``.

        Args:
            ai: The AI provider instance (must expose
                ``get_embeddings_batch(texts: List[str]) -> List[np.ndarray]``).

        Returns:
            bool: ``True`` if at least one correction or custom category was
            loaded; ``False`` otherwise.
        """
        if not os.path.exists(self.feedback_path):
            logger.info(f"No feedback file found at {self.feedback_path} - using default anchors")
            return False

        has_feedback = False

        # ---- Phase 1: Custom categories from Category Reference tab --------
        try:
            all_categories = self.load_category_reference()

            # Create embeddings for custom categories (not in ANCHORS)
            if self.custom_categories:
                logger.info(f"Creating embeddings for {len(self.custom_categories)} custom categories...")
                for cat_name, keywords in self.custom_categories.items():
                    if cat_name not in self.corrections:
                        self.corrections[cat_name] = []

                    # Embed each keyword string -- the resulting vectors will
                    # be averaged later in adjust_centroids() to form the
                    # category's centroid in embedding space.
                    keyword_embeddings = ai.get_embeddings_batch(keywords)
                    for kw, emb in zip(keywords, keyword_embeddings):
                        if emb is not None:
                            self.corrections[cat_name].append({
                                'text': kw,
                                'embedding': emb,
                                'original': 'custom_category'  # flag: came from Category Reference, not a correction
                            })

                    logger.info(f"  → Created {len(keyword_embeddings)} embeddings for '{cat_name}'")
                has_feedback = True

        except Exception as e:
            logger.warning(f"Could not load custom categories: {e}")

        # ---- Phase 2: User corrections from Classifications sheet -----------
        try:
            df_feedback = pd.read_excel(self.feedback_path, sheet_name='Classifications')

            # Collect rows where the user has entered a corrected category that
            # differs from what the AI originally assigned.
            corrections_list = []
            for _, row in df_feedback.iterrows():
                original = str(row.get('AI_Category', '')).strip()
                corrected = str(row.get('Corrected_Category', '')).strip()
                text = str(row.get('Text', '')).strip()

                # Only count as a correction if: (a) a corrected value exists,
                # (b) it's not a null/blank artefact, (c) it differs from the
                # AI category, and (d) there is associated text to embed.
                if corrected and corrected.lower() not in ['', 'nan', 'none'] and corrected != original and text:
                    corrections_list.append({
                        'text': text,
                        'category': corrected,
                        'original': original
                    })

            if corrections_list:
                logger.info(f"Found {len(corrections_list)} user corrections in feedback file")

                # Batch-embed all corrected texts in one call for efficiency
                texts = [c['text'] for c in corrections_list]
                embeddings = ai.get_embeddings_batch(texts)

                # Group correction embeddings by the user's corrected category
                for corr, emb in zip(corrections_list, embeddings):
                    cat = corr['category']
                    if cat not in self.corrections:
                        self.corrections[cat] = []
                    self.corrections[cat].append({
                        'text': corr['text'],
                        'embedding': emb,
                        'original': corr['original']
                    })

                self.stats['loaded'] = len(corrections_list)
                has_feedback = True
            else:
                logger.info("No corrections found in Classifications sheet")

        except Exception as e:
            logger.warning(f"Failed to load feedback corrections: {e}")

        # ---- Update aggregate stats -----------------------------------------
        self.stats['categories_enhanced'] = len(self.corrections)

        if has_feedback:
            logger.info(f"✓ Loaded feedback: {self.stats['loaded']} corrections, "
                       f"{self.stats['custom_categories']} custom categories, "
                       f"{self.stats['categories_enhanced']} total categories enhanced")
            for cat, items in self.corrections.items():
                logger.info(f"  - {cat}: {len(items)} examples")

        return has_feedback

    # -------------------------------------------------------------------------
    # Centroid Adjustment (core learning mechanism)
    # -------------------------------------------------------------------------

    def adjust_centroids(self, original_centroids: Dict[str, np.ndarray]) -> Dict[str, np.ndarray]:
        """
        Blend user feedback embeddings with the original anchor centroids.

        This is the core learning step that translates user corrections into
        improved classification behaviour.  For every category that has
        accumulated feedback:

        1. Compute the **feedback centroid** as the element-wise mean of all
           correction/keyword embedding vectors for that category.
        2. Produce the **adjusted centroid** via weighted linear interpolation::

               adjusted = (1 - FEEDBACK_WEIGHT) * original_centroid
                        +      FEEDBACK_WEIGHT  * feedback_centroid

           where ``FEEDBACK_WEIGHT`` (default 0.3 from config) controls how
           aggressively user feedback shifts the centroid.  A value of 0.3
           means the original anchor still dominates (70%) while user examples
           contribute 30%.

        For **new categories** (present in ``self.corrections`` but absent from
        ``original_centroids``), the centroid is simply the mean of the
        feedback embeddings -- there is no original anchor to blend with.

        Args:
            original_centroids (Dict[str, np.ndarray]): The current category
                centroids as produced by the classifier's anchor embedding step.

        Returns:
            Dict[str, np.ndarray]: Updated centroid dictionary with all original
            categories (adjusted or unchanged) plus any new categories.
        """
        if not self.corrections:
            return original_centroids

        adjusted = {}

        # ---- Pass 1: Adjust existing categories ----------------------------
        for category, centroid in original_centroids.items():
            if category in self.corrections:
                # Collect all feedback embedding vectors for this category
                feedback_vecs = [c['embedding'] for c in self.corrections[category]]
                # Mean of the user-provided examples forms the feedback centroid
                feedback_centroid = np.mean(feedback_vecs, axis=0)

                # Weighted blend: original anchor dominates, feedback nudges
                # toward user-preferred region of embedding space.
                adjusted[category] = (
                    (1 - FEEDBACK_WEIGHT) * centroid +
                    FEEDBACK_WEIGHT * feedback_centroid
                )
                logger.info(f"  → Adjusted '{category}' centroid with {len(feedback_vecs)} user examples")
            else:
                # No feedback for this category -- keep the original centroid
                adjusted[category] = centroid

        # ---- Pass 2: Create centroids for brand-new categories ---------------
        # These are categories that only exist in user feedback (either from
        # corrections or from the Category Reference tab) and have no
        # pre-existing anchor centroid.
        for category in self.corrections:
            if category not in adjusted:
                feedback_vecs = [c['embedding'] for c in self.corrections[category]]
                # With no prior anchor, the centroid is purely the mean of
                # user-supplied embeddings.
                adjusted[category] = np.mean(feedback_vecs, axis=0)
                logger.info(f"  → Created NEW category '{category}' from {len(feedback_vecs)} user examples")

        return adjusted

    # -------------------------------------------------------------------------
    # Export for User Review
    # -------------------------------------------------------------------------

    def save_for_review(self, df: pd.DataFrame, output_dir: str) -> str:
        """
        Export classifications to a formatted Excel workbook for user review.

        The workbook contains four sheets:

        1. **Instructions** -- step-by-step guide for the reviewer, auto-
           generated list of available categories, and version/timestamp.
        2. **Classifications** -- one row per ticket with AI predictions and
           a ``Corrected_Category`` dropdown column for user corrections.
        3. **Category Reference** -- editable list of category names and
           example keywords.  Users can add rows here to create new categories.
        4. **Sub-Category Reference** -- read-only reference of the sub-
           category taxonomy (from ``SUB_CATEGORIES`` config).

        **Non-destructive merge**: If a previous feedback file exists, any
        ``Corrected_Category`` and ``Notes`` values the user has already
        entered are matched by a composite key (ticket ID + first 200 chars
        of text, lowercased) and carried forward into the new export.  This
        ensures users never lose their previous corrections.

        The workbook is written to **two** locations:
        - ``output_dir/classification_feedback.xlsx`` (the report output)
        - ``self.feedback_path`` (the working-directory copy read on next run)

        Args:
            df (pd.DataFrame): The fully-processed ticket DataFrame with at
                minimum ``Combined_Text``, ``AI_Category``, ``AI_Sub_Category``,
                ``AI_Confidence``, etc.
            output_dir (str): Directory to write the output workbook into.

        Returns:
            str: The full path to the exported workbook in ``output_dir``.
        """

        # ---- Preserve existing user corrections -----------------------------
        # Build a lookup dict from the previous feedback file keyed on
        # (ID + truncated-text) so we can restore any corrections/notes
        # the user entered in a prior run.
        existing_corrections = {}
        existing_notes = {}
        if os.path.exists(self.feedback_path):
            try:
                df_existing = pd.read_excel(self.feedback_path, sheet_name='Classifications')
                for _, row in df_existing.iterrows():
                    # Composite key: row ID + first 200 chars of text (lowercased)
                    # This handles cases where row indices may shift between runs.
                    text_key = str(row.get('Text', ''))[:200].strip().lower()
                    row_id = str(row.get('ID', ''))
                    key = f"{row_id}_{text_key}"

                    corrected = str(row.get('Corrected_Category', '')).strip()
                    notes = str(row.get('Notes', '')).strip()

                    if corrected and corrected.lower() not in ['', 'nan', 'none']:
                        existing_corrections[key] = corrected
                    if notes and notes.lower() not in ['', 'nan', 'none']:
                        existing_notes[key] = notes

                logger.info(f"Loaded {len(existing_corrections)} existing corrections from feedback file")
            except Exception as e:
                logger.warning(f"Could not load existing feedback for preservation: {e}")

        # ---- Build the feedback rows ----------------------------------------
        # Each row includes the AI's predictions plus editable columns for
        # the reviewer.
        feedback_rows = []
        for idx, row in df.iterrows():
            text_truncated = str(row.get('Combined_Text', ''))[:500]
            # Match key uses the same 200-char prefix as above for consistency
            text_key = text_truncated[:200].strip().lower()
            row_id = str(idx)
            key = f"{row_id}_{text_key}"

            # Carry forward previously entered corrections/notes (if any)
            preserved_correction = existing_corrections.get(key, '')
            preserved_notes = existing_notes.get(key, '')

            feedback_rows.append({
                'ID': row_id,
                'Text': text_truncated,
                'AI_Category': row.get('AI_Category', 'Unclassified'),
                'AI_Sub_Category': row.get('AI_Sub_Category', 'General'),
                'AI_Confidence': round(float(row.get('AI_Confidence', 0)), 3),
                'Root_Cause_Category': row.get('Root_Cause_Category', 'Unclassified'),
                'PM_Recurrence_Risk': row.get('PM_Recurrence_Risk_Norm', 'Unknown'),
                'AI_Recurrence_Risk': row.get('AI_Recurrence_Risk', 'Unknown'),
                'AI_Recurrence_Prob': f"{row.get('AI_Recurrence_Probability', 0)*100:.0f}%",
                'PM_Prediction_Accuracy': row.get('PM_Prediction_Accuracy', 'Pending'),
                'LOB': row.get('LOB', 'Unknown'),
                'Corrected_Category': preserved_correction,  # editable by the user
                'Notes': preserved_notes                      # editable by the user
            })

        df_feedback = pd.DataFrame(feedback_rows)
        export_path = os.path.join(output_dir, 'classification_feedback.xlsx')

        # ---- Build the Category Reference sheet -----------------------------
        # Merge built-in anchors with any user-added categories so the
        # dropdown always reflects the full taxonomy.
        all_categories = self._build_category_reference()
        categories_df = pd.DataFrame({
            'Available_Categories': list(all_categories.keys()),
            # Show the first 5 keywords as a preview for each category
            'Example_Keywords': [', '.join(kw[:5]) for kw in all_categories.values()]
        })

        # ---- Build the Sub-Category Reference sheet -------------------------
        sub_cat_rows = []
        for cat, sub_cats in SUB_CATEGORIES.items():
            for sub_cat, keywords in sub_cats.items():
                sub_cat_rows.append({
                    'Category': cat,
                    'Sub_Category': sub_cat,
                    'Keywords': ', '.join(keywords[:5])
                })
        sub_categories_df = pd.DataFrame(sub_cat_rows)

        # ---- Write the primary export workbook (to output_dir) --------------
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            # Instructions sheet (first tab for discoverability)
            instructions_df = self._create_instructions_df(all_categories)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

            # Classifications sheet (the main review/correction sheet)
            df_feedback.to_excel(writer, sheet_name='Classifications', index=False)

            # Category reference sheet (editable -- users can add new rows)
            categories_df.to_excel(writer, sheet_name='Category Reference', index=False)

            # Sub-Category reference sheet (read-only taxonomy reference)
            sub_categories_df.to_excel(writer, sheet_name='Sub-Category Reference', index=False)

            # Apply visual formatting (header colours, column widths,
            # dropdown validation, low-confidence highlighting)
            self._format_feedback_workbook(writer, df_feedback, all_categories)

        # ---- Write the working-directory copy (for next pipeline run) -------
        # This copy is what load_feedback() reads on the next run.  It must
        # contain all four sheets with the same structure.
        try:
            with pd.ExcelWriter(self.feedback_path, engine='openpyxl') as wd_writer:
                # Instructions sheet
                instructions_df = self._create_instructions_df(all_categories)
                instructions_df.to_excel(wd_writer, sheet_name='Instructions', index=False)

                # Classifications sheet
                df_feedback.to_excel(wd_writer, sheet_name='Classifications', index=False)

                # Category reference sheet
                categories_df.to_excel(wd_writer, sheet_name='Category Reference', index=False)

                # Sub-Category reference sheet
                sub_categories_df.to_excel(wd_writer, sheet_name='Sub-Category Reference', index=False)

                # Apply formatting
                self._format_feedback_workbook(wd_writer, df_feedback, all_categories)

            logger.info(f"✓ Feedback file updated at: {self.feedback_path}")
        except Exception as e:
            logger.warning(f"Could not save feedback to working directory: {e}")

        logger.info(f"✓ Feedback Excel saved to: {export_path}")
        return export_path

    # -------------------------------------------------------------------------
    # Internal Helpers
    # -------------------------------------------------------------------------

    def _build_category_reference(self) -> Dict[str, List[str]]:
        """
        Build the merged category reference from built-in anchors + user additions.

        This is used when constructing the "Category Reference" sheet for
        export.  It reads the existing feedback file (if present) and merges
        any user-added categories or keywords back into the reference so they
        are not lost when the workbook is regenerated.

        The merge logic mirrors ``load_category_reference()`` but is used at
        *export* time rather than *ingest* time.

        Returns:
            Dict[str, List[str]]: Complete category -> keywords mapping.
        """
        # Start with a mutable copy of the built-in anchor keywords
        all_categories = {cat: list(keywords) for cat, keywords in ANCHORS.items()}

        # Load existing user-added categories from feedback file
        if os.path.exists(self.feedback_path):
            try:
                df_ref = pd.read_excel(self.feedback_path, sheet_name='Category Reference')

                for _, row in df_ref.iterrows():
                    cat_name = str(row.get('Available_Categories', '')).strip()
                    keywords_str = str(row.get('Example_Keywords', '')).strip()

                    if cat_name and cat_name.lower() not in ['', 'nan', 'none']:
                        # Parse comma-separated keywords
                        keywords = [k.strip() for k in keywords_str.split(',') if k.strip()]

                        if cat_name not in all_categories:
                            # User-added category -- preserve it so it appears
                            # in the regenerated workbook and dropdown.
                            all_categories[cat_name] = keywords if keywords else [cat_name.lower()]
                            logger.info(f"  → Preserved user category: '{cat_name}'")
                        elif keywords:
                            # Existing category -- check for newly added keywords
                            existing = set(ANCHORS.get(cat_name, []))
                            new_kw = [k for k in keywords if k not in existing]
                            if new_kw:
                                all_categories[cat_name] = list(ANCHORS[cat_name]) + new_kw

            except Exception as e:
                logger.debug(f"Could not load existing Category Reference: {e}")

        return all_categories

    def _create_instructions_df(self, all_categories: Dict[str, List[str]] = None) -> pd.DataFrame:
        """
        Create a single-column DataFrame that renders as an instructions sheet.

        The instructions explain the human-in-the-loop workflow, how to
        correct categories, how to add new categories, and list all available
        categories.  The sheet also includes a generation timestamp and the
        report version string.

        Args:
            all_categories (dict, optional): Category -> keywords mapping used
                to render the "AVAILABLE CATEGORIES" list.  Falls back to the
                built-in ``ANCHORS`` if not provided.

        Returns:
            pd.DataFrame: Single-column DataFrame ("Instructions") suitable
            for writing to an Excel sheet.
        """
        if all_categories is None:
            all_categories = ANCHORS

        return pd.DataFrame({
            'Instructions': [
                'HOW TO IMPROVE AI CLASSIFICATION',
                '',
                '1. Go to the "Classifications" sheet',
                '2. Review each row - check if AI_Category is correct',
                '3. If WRONG: Select the correct category from "Corrected_Category" dropdown',
                '4. If CORRECT: Leave "Corrected_Category" blank',
                '5. Save this file',
                '6. Run the analysis again - AI will learn from your corrections!',
                '',
                'ADDING NEW CATEGORIES:',
                '1. Go to the "Category Reference" sheet',
                '2. Add a new row with category name in "Available_Categories"',
                '3. Add example keywords in "Example_Keywords" (comma-separated)',
                '4. Save and re-run - the new category will be available!',
                '',
                'AVAILABLE CATEGORIES:',
                *[f'  • {cat}' for cat in all_categories.keys()],
                '',
                'TIP: Yellow rows = Low confidence (<50%) - priority for review!',
                '',
                f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
                f'Version: {REPORT_VERSION}'
            ]
        })

    def _format_feedback_workbook(self, writer, df_feedback: pd.DataFrame,
                                    all_categories: Dict[str, List[str]] = None):
        """
        Apply visual formatting and data validation to the feedback workbook.

        Formatting applied:
        - **Column widths** sized to the expected content of each column.
        - **Data validation dropdown** on the ``Corrected_Category`` column
          (column K) that references the "Category Reference" sheet range
          ``$A$2:$A$100``.  Using a sheet reference (rather than an inline
          list) means the dropdown automatically picks up new categories the
          user adds to that tab.
        - **Header row styling** with white-on-navy font and fill.
        - **Low-confidence highlighting** (yellow fill) on rows where
          ``AI_Confidence < 0.5`` to draw the reviewer's attention to the
          tickets most likely to need correction.
        - **Freeze panes** on row 2 so the header stays visible during
          scrolling.
        - Formatting of the "Category Reference" sheet with column widths,
          header styling, and freeze panes.

        Args:
            writer: An open ``pd.ExcelWriter`` with 'openpyxl' engine.
            df_feedback (pd.DataFrame): The feedback DataFrame (used to
                determine row count for validation ranges and highlighting).
            all_categories (dict, optional): Category -> keywords mapping;
                not directly used for formatting (the dropdown references the
                sheet), but retained for potential future use.

        Note:
            The openpyxl ``DataValidation`` parameter ``showDropDown=False``
            is counter-intuitive: setting it to ``False`` actually **shows**
            the dropdown arrow in Excel.  This is a known openpyxl quirk.
        """
        ws = writer.sheets['Classifications']

        # ---- Column widths --------------------------------------------------
        widths = {'A': 8, 'B': 80, 'C': 25, 'D': 12, 'E': 18, 'F': 16,
                  'G': 22, 'H': 14, 'I': 20, 'J': 12, 'K': 25, 'L': 30}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # ---- Category dropdown validation -----------------------------------
        # Reference the "Category Reference" sheet's Available_Categories
        # column (A2:A100).  A large range (100 rows) is used so users can
        # add many categories without needing to edit the formula.  Empty
        # cells in the range are silently ignored by Excel's dropdown.
        if all_categories is None:
            all_categories = ANCHORS

        # Use a reference to the Category Reference sheet's Available_Categories column
        # Use a large range (A2:A100) so users can add more categories without editing the formula
        # Empty cells in the range are ignored by Excel's dropdown
        category_validation = DataValidation(
            type='list',
            formula1="'Category Reference'!$A$2:$A$100",
            allow_blank=True,
            showDropDown=False  # In openpyxl, False actually shows the dropdown
        )
        # Apply validation to the Corrected_Category column (K) for all data rows
        category_validation.add(f'K2:K{len(df_feedback) + 1}')
        ws.add_data_validation(category_validation)

        # ---- Header row styling (navy background, white text) ---------------
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # ---- Low-confidence row highlighting --------------------------------
        # Rows where AI_Confidence (column D, index 4) is below 0.5 are
        # filled with bright yellow to draw the reviewer's attention.
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row_idx in range(2, len(df_feedback) + 2):
            try:
                confidence = float(ws.cell(row=row_idx, column=4).value or 0)
                if confidence < 0.5:
                    # Highlight the entire row (columns A through L)
                    for col_idx in range(1, 13):
                        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
            except (ValueError, TypeError):
                pass

        # Freeze the header row so it stays visible while scrolling
        ws.freeze_panes = 'A2'

        # ---- Format the Category Reference sheet ----------------------------
        if 'Category Reference' in writer.sheets:
            ws_ref = writer.sheets['Category Reference']
            ws_ref.column_dimensions['A'].width = 30   # Available_Categories
            ws_ref.column_dimensions['B'].width = 60   # Example_Keywords

            # Apply the same header styling for visual consistency
            for cell in ws_ref[1]:
                cell.font = header_font
                cell.fill = header_fill

            ws_ref.freeze_panes = 'A2'


# =============================================================================
# Singleton Access
# =============================================================================

# Global feedback learner instance (lazily initialised)
_feedback_learner: Optional[FeedbackLearning] = None

def get_feedback_learner() -> FeedbackLearning:
    """
    Get or create the global FeedbackLearning singleton.

    Returns:
        FeedbackLearning: The shared instance used across the pipeline.
    """
    global _feedback_learner
    if _feedback_learner is None:
        _feedback_learner = FeedbackLearning()
    return _feedback_learner


# =============================================================================
# Resolution Time Feedback Loop
# =============================================================================

class ResolutionFeedbackLearning:
    """
    Human-in-the-loop learning system for resolution time prediction improvement.

    This is the resolution-time counterpart to ``FeedbackLearning``.  While the
    classification loop adjusts *category centroids*, this loop collects
    human estimates of how long a ticket *should* take to resolve, giving the
    resolution-time predictor an additional signal to learn from.

    Workflow (from the user's perspective):
        1. Run the analysis pipeline -- generates/updates
           ``resolution_feedback.xlsx``.
        2. Open the workbook and review the AI-predicted resolution days.
        3. Enter a human estimate in ``Human_Expected_Days`` for tickets where
           the AI prediction looks off.
        4. When tickets are closed, fill in ``Actual_Resolution_Days``.
        5. Save the workbook and re-run -- the pipeline will incorporate
           the human estimates into its predictions.

    Attributes:
        feedback_path (str):
            Filesystem path to ``resolution_feedback.xlsx``.  Defaults to
            ``RESOLUTION_FEEDBACK_FILE`` from config.
        feedback_data (Dict[str, Dict]):
            Mapping from ticket ``Identity`` to a record dict with keys:
            ``AI_Category``, ``AI_Sub_Category``, ``Predicted_Resolution_Days``,
            ``Human_Expected_Days`` (Optional[float]),
            ``Actual_Resolution_Days`` (Optional[float]), ``Notes``.
        stats (dict):
            Tracking counters: ``loaded`` (total records ingested),
            ``with_human_feedback`` (records with a non-null
            ``Human_Expected_Days``).
    """

    def __init__(self, feedback_path: Optional[str] = None):
        self.feedback_path = feedback_path or RESOLUTION_FEEDBACK_FILE
        # feedback_data: Identity -> {AI_Category, Predicted, Human, Actual, Notes}
        self.feedback_data: Dict[str, Dict] = {}  # identity -> feedback record
        self.stats = {'loaded': 0, 'with_human_feedback': 0}

    # -------------------------------------------------------------------------
    # Feedback Ingestion
    # -------------------------------------------------------------------------

    def load_feedback(self) -> Dict[str, Dict]:
        """
        Load existing resolution time feedback from the Excel workbook.

        Reads the "Resolution Feedback" sheet and populates ``self.feedback_data``
        keyed by the ticket ``Identity`` field.  For each record, the method
        distinguishes between ``Human_Expected_Days`` (user estimate) and
        ``Actual_Resolution_Days`` (ground truth from ticket closure).

        Returns:
            Dict[str, Dict]: The populated ``self.feedback_data`` mapping, also
            stored as an instance attribute for later queries via
            ``get_human_expected_days()``.
        """
        if not os.path.exists(self.feedback_path):
            logger.info(f"No resolution feedback file found at {self.feedback_path}")
            return {}

        try:
            df_feedback = pd.read_excel(self.feedback_path, sheet_name='Resolution Feedback')

            for _, row in df_feedback.iterrows():
                identity = str(row.get('Identity', '')).strip()
                if not identity or identity.lower() in ['', 'nan', 'none']:
                    continue

                human_days = row.get('Human_Expected_Days')
                actual_days = row.get('Actual_Resolution_Days')

                self.feedback_data[identity] = {
                    'AI_Category': row.get('AI_Category', ''),
                    'AI_Sub_Category': row.get('AI_Sub_Category', ''),
                    'Predicted_Resolution_Days': row.get('Predicted_Resolution_Days'),
                    # Store None if the cell was empty (pd.isna) rather than
                    # carrying forward NaN values.
                    'Human_Expected_Days': human_days if pd.notna(human_days) else None,
                    'Actual_Resolution_Days': actual_days if pd.notna(actual_days) else None,
                    'Notes': row.get('Notes', '')
                }

                if pd.notna(human_days):
                    self.stats['with_human_feedback'] += 1

            self.stats['loaded'] = len(self.feedback_data)

            if self.feedback_data:
                logger.info(f"✓ Loaded resolution feedback: {self.stats['loaded']} records, "
                           f"{self.stats['with_human_feedback']} with human feedback")

        except Exception as e:
            logger.warning(f"Failed to load resolution feedback: {e}")

        return self.feedback_data

    def get_human_expected_days(self, identity: str) -> Optional[float]:
        """
        Look up the human-provided expected resolution days for a specific ticket.

        This is called by ``resolution_time.py``'s prediction pipeline to
        optionally blend human estimates into its output.

        Args:
            identity (str): The unique ticket identifier (e.g. "INC12345").

        Returns:
            Optional[float]: The human's estimate in days, or ``None`` if no
            feedback exists for this ticket.
        """
        if identity in self.feedback_data:
            return self.feedback_data[identity].get('Human_Expected_Days')
        return None

    # -------------------------------------------------------------------------
    # Export for User Review
    # -------------------------------------------------------------------------

    def save_for_review(self, df: pd.DataFrame, output_dir: str) -> str:
        """
        Export resolution predictions to a formatted Excel workbook for review.

        The workbook contains two sheets:

        1. **Instructions** -- step-by-step guide explaining the columns and
           how to provide feedback.
        2. **Resolution Feedback** -- one row per ticket with columns for
           AI predictions, human estimates, actual resolution times, and notes.

        **Non-destructive merge**: Existing ``Human_Expected_Days``,
        ``Actual_Resolution_Days``, and ``Notes`` values from a prior run are
        preserved by matching on the ticket ``Identity`` string.

        The workbook is written to two locations (just like the classification
        feedback): the report ``output_dir`` and the ``self.feedback_path``
        working directory copy.

        Args:
            df (pd.DataFrame): The fully-processed ticket DataFrame.
            output_dir (str): Directory to write the output workbook into.

        Returns:
            str: The full path to the exported workbook in ``output_dir``.
        """
        # ---- Preserve existing user feedback --------------------------------
        existing_feedback = {}
        if os.path.exists(self.feedback_path):
            try:
                df_existing = pd.read_excel(self.feedback_path, sheet_name='Resolution Feedback')
                for _, row in df_existing.iterrows():
                    identity = str(row.get('Identity', '')).strip()
                    if identity and identity.lower() not in ['', 'nan', 'none']:
                        human_days = row.get('Human_Expected_Days')
                        actual_days = row.get('Actual_Resolution_Days')
                        notes = row.get('Notes', '')

                        existing_feedback[identity] = {
                            'Human_Expected_Days': human_days if pd.notna(human_days) else None,
                            'Actual_Resolution_Days': actual_days if pd.notna(actual_days) else None,
                            'Notes': notes if pd.notna(notes) and str(notes).lower() not in ['', 'nan'] else ''
                        }

                logger.info(f"Loaded {len(existing_feedback)} existing resolution feedback entries")
            except Exception as e:
                logger.warning(f"Could not load existing resolution feedback: {e}")

        # ---- Build feedback rows --------------------------------------------
        feedback_rows = []
        for idx, row in df.iterrows():
            identity = str(row.get('Identity', idx))

            # Carry forward previously entered human feedback (if any)
            existing = existing_feedback.get(identity, {})

            feedback_rows.append({
                'Identity': identity,
                'AI_Category': row.get('AI_Category', 'Unclassified'),
                'AI_Sub_Category': row.get('AI_Sub_Category', 'General'),
                # Round predictions to 2 decimal places for readability;
                # show blank if the prediction is missing.
                'Predicted_Resolution_Days': round(float(row.get('Predicted_Resolution_Days', 0)), 2)
                    if pd.notna(row.get('Predicted_Resolution_Days')) else '',
                'Human_Expected_Days': existing.get('Human_Expected_Days', ''),
                # Actual days: prefer the current run's data, fall back to
                # the previously stored value.
                'Actual_Resolution_Days': row.get('Actual_Resolution_Days', existing.get('Actual_Resolution_Days', '')),
                'Resolution_Prediction_Confidence': round(float(row.get('Resolution_Prediction_Confidence', 0)), 2)
                    if pd.notna(row.get('Resolution_Prediction_Confidence')) else '',
                'Notes': existing.get('Notes', '')
            })

        df_feedback = pd.DataFrame(feedback_rows)
        export_path = os.path.join(output_dir, 'resolution_feedback.xlsx')

        # ---- Create Instructions sheet --------------------------------------
        instructions_df = pd.DataFrame({
            'Instructions': [
                'RESOLUTION TIME FEEDBACK',
                '',
                'This file helps improve AI resolution time predictions through your feedback.',
                '',
                'HOW TO PROVIDE FEEDBACK:',
                '1. Go to the "Resolution Feedback" sheet',
                '2. Review each row - check if Predicted_Resolution_Days is reasonable',
                '3. Enter your expected resolution time in "Human_Expected_Days"',
                '4. When tickets are closed, enter "Actual_Resolution_Days"',
                '5. Save this file',
                '6. Next analysis run will learn from your feedback!',
                '',
                'COLUMNS:',
                '• Identity: Unique ticket identifier',
                '• AI_Category: AI-assigned category',
                '• AI_Sub_Category: AI-assigned sub-category',
                '• Predicted_Resolution_Days: AI prediction',
                '• Human_Expected_Days: YOUR estimate (edit this!)',
                '• Actual_Resolution_Days: Real resolution time',
                '• Resolution_Prediction_Confidence: AI confidence (0-1)',
                '• Notes: Any additional notes',
                '',
                f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
                f'Version: {REPORT_VERSION}'
            ]
        })

        # ---- Write the primary export workbook ------------------------------
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            df_feedback.to_excel(writer, sheet_name='Resolution Feedback', index=False)
            self._format_feedback_workbook(writer, df_feedback)

        # ---- Write the working-directory copy -------------------------------
        try:
            with pd.ExcelWriter(self.feedback_path, engine='openpyxl') as wd_writer:
                instructions_df.to_excel(wd_writer, sheet_name='Instructions', index=False)
                df_feedback.to_excel(wd_writer, sheet_name='Resolution Feedback', index=False)
                self._format_feedback_workbook(wd_writer, df_feedback)
            logger.info(f"✓ Resolution feedback file updated at: {self.feedback_path}")
        except Exception as e:
            logger.warning(f"Could not save resolution feedback to working directory: {e}")

        logger.info(f"✓ Resolution feedback Excel saved to: {export_path}")
        return export_path

    # -------------------------------------------------------------------------
    # Internal Formatting
    # -------------------------------------------------------------------------

    def _format_feedback_workbook(self, writer, df_feedback: pd.DataFrame):
        """
        Apply visual formatting to the resolution feedback workbook.

        Formatting applied:
        - **Column widths** sized to the expected content of each column
          (Identity, categories, prediction values, notes).
        - **Header row styling** with white-on-navy font and fill (matching
          the classification feedback workbook for visual consistency).
        - **Editable-column highlighting** -- columns ``Human_Expected_Days``
          (E) and ``Notes`` (H) are given a light-yellow background to signal
          to the user that these are the columns they should edit.
        - **Freeze panes** on row 2.

        Args:
            writer: An open ``pd.ExcelWriter`` with 'openpyxl' engine.
            df_feedback (pd.DataFrame): Used to determine the row count for
                applying cell-level formatting.
        """
        ws = writer.sheets['Resolution Feedback']

        # ---- Column widths --------------------------------------------------
        widths = {'A': 15, 'B': 25, 'C': 25, 'D': 22, 'E': 20, 'F': 22, 'G': 28, 'H': 40}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # ---- Header row styling (navy background, white text) ---------------
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # ---- Editable column highlighting -----------------------------------
        # Light yellow fill on the two columns the user is expected to edit,
        # making it visually obvious which cells need their attention.
        editable_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        for row_idx in range(2, len(df_feedback) + 2):
            ws.cell(row=row_idx, column=5).fill = editable_fill  # Human_Expected_Days
            ws.cell(row=row_idx, column=8).fill = editable_fill  # Notes

        # Freeze the header row
        ws.freeze_panes = 'A2'


# =============================================================================
# Singleton Access (Resolution)
# =============================================================================

# Global resolution feedback learner instance (lazily initialised)
_resolution_feedback_learner: Optional[ResolutionFeedbackLearning] = None

def get_resolution_feedback_learner() -> ResolutionFeedbackLearning:
    """
    Get or create the global ResolutionFeedbackLearning singleton.

    Returns:
        ResolutionFeedbackLearning: The shared instance used across the pipeline.
    """
    global _resolution_feedback_learner
    if _resolution_feedback_learner is None:
        _resolution_feedback_learner = ResolutionFeedbackLearning()
    return _resolution_feedback_learner
