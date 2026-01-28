"""
Similar Ticket Finder - AI-powered semantic similarity analysis.

Uses embeddings to find semantically similar historical tickets,
then compares resolutions to identify patterns and inconsistencies.

GPU-accelerated with RAPIDS cuML when available.
"""

import os
import json
import logging
import numpy as np
import pandas as pd
from datetime import datetime
from collections import Counter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ..core.config import (
    COL_SUMMARY, COL_SEVERITY, COL_DATETIME, COL_CLOSE_DATE,
    COL_ROOT_CAUSE, COL_LESSON_TITLE, COL_ENGINEER, COL_LOB,
    COL_TYPE, COL_ORIGIN, COL_RESOLUTION_DATE,
    SIMILARITY_FEEDBACK_PATH, USE_GPU
)
from ..core.gpu_utils import (
    GPUSimilaritySearch,
    cosine_similarity_gpu,
    batch_cosine_similarity_gpu,
    is_gpu_available
)

logger = logging.getLogger(__name__)


class SimilarTicketFinder:
    """
    AI-powered Similar Ticket Finder with Resolution Comparison.
    
    Uses embeddings to find semantically similar historical tickets,
    then compares resolutions to identify patterns and inconsistencies.
    
    Key Features:
    - Embedding-based semantic similarity (GPU-accelerated)
    - Resolution pattern analysis across similar tickets
    - Days to resolution comparison
    - Human-in-the-loop feedback for similarity validation
    """
    
    def __init__(self, ai_engine=None, top_k=5, similarity_threshold=0.75):
        """
        Initialize the Similar Ticket Finder.
        
        Args:
            ai_engine: AI engine instance for embeddings
            top_k: Number of similar tickets to find per query
            similarity_threshold: Minimum cosine similarity to consider (0-1)
        """
        self.ai = ai_engine
        self.top_k = top_k
        self.similarity_threshold = similarity_threshold
        self.embedding_cache = {}
        self.resolution_patterns = {}
        self.feedback_data = {}
        
        # GPU-accelerated similarity search index
        self.use_gpu = USE_GPU and is_gpu_available()
        self.gpu_search_index = None
        self.index_embeddings = None
        self.index_ids = None
        
        # Load existing feedback if available
        self._load_feedback()
        
        # Resolution-related column names
        self.resolution_columns = [
            COL_LESSON_TITLE,
            'Action_Required',
            'Root_Cause_Category',
            COL_ROOT_CAUSE,
        ]
        
        logger.info(f"[Similar Ticket Finder] Initialized (GPU: {self.use_gpu})")
    
    def _load_feedback(self):
        """Load human feedback on similarity matches from file."""
        try:
            if os.path.exists(SIMILARITY_FEEDBACK_PATH):
                with open(SIMILARITY_FEEDBACK_PATH, 'r') as f:
                    self.feedback_data = json.load(f)
                logger.info(f"[Similar Ticket Finder] Loaded {len(self.feedback_data)} feedback entries")
        except Exception as e:
            logger.warning(f"[Similar Ticket Finder] Could not load feedback: {e}")
            self.feedback_data = {}
    
    def _save_feedback(self):
        """Save human feedback to file for future runs."""
        try:
            with open(SIMILARITY_FEEDBACK_PATH, 'w') as f:
                json.dump(self.feedback_data, f, indent=2)
            logger.info(f"[Similar Ticket Finder] Saved {len(self.feedback_data)} feedback entries")
        except Exception as e:
            logger.warning(f"[Similar Ticket Finder] Could not save feedback: {e}")
    
    def record_feedback(self, ticket_id, similar_ticket_id, is_similar, notes=""):
        """Record human feedback on whether two tickets are actually similar."""
        feedback_key = f"{ticket_id}|{similar_ticket_id}"
        self.feedback_data[feedback_key] = {
            'ticket_id': str(ticket_id),
            'similar_ticket_id': str(similar_ticket_id),
            'is_similar': is_similar,
            'notes': notes,
            'timestamp': datetime.now().isoformat()
        }
        self._save_feedback()
    
    def get_feedback_adjusted_similarity(self, ticket_id, similar_ticket_id, base_similarity):
        """Adjust similarity score based on human feedback."""
        feedback_key = f"{ticket_id}|{similar_ticket_id}"
        reverse_key = f"{similar_ticket_id}|{ticket_id}"
        
        feedback = self.feedback_data.get(feedback_key) or self.feedback_data.get(reverse_key)
        
        if feedback:
            if feedback['is_similar']:
                return min(1.0, base_similarity + 0.2)
            else:
                return max(0.0, base_similarity - 0.5)
        
        return base_similarity
    
    def _calculate_days_to_resolution(self, row):
        """Calculate days from issue open to resolution."""
        try:
            open_date = None
            close_date = None
            
            if 'Issue_Date' in row.index and pd.notna(row.get('Issue_Date')):
                open_date = pd.to_datetime(row['Issue_Date'], errors='coerce')
            elif COL_DATETIME in row.index and pd.notna(row.get(COL_DATETIME)):
                open_date = pd.to_datetime(row[COL_DATETIME], errors='coerce')
            
            if COL_CLOSE_DATE in row.index and pd.notna(row.get(COL_CLOSE_DATE)):
                close_date = pd.to_datetime(row[COL_CLOSE_DATE], errors='coerce')
            elif 'Close_Date' in row.index and pd.notna(row.get('Close_Date')):
                close_date = pd.to_datetime(row['Close_Date'], errors='coerce')
            elif 'Resolution_Date' in row.index and pd.notna(row.get('Resolution_Date')):
                close_date = pd.to_datetime(row['Resolution_Date'], errors='coerce')
            
            if open_date and close_date and pd.notna(open_date) and pd.notna(close_date):
                delta = (close_date - open_date).days
                return max(0, delta)
            
        except Exception:
            pass
        
        return None
    
    def build_gpu_index(self, historical_df):
        """
        Build GPU-accelerated similarity search index from historical data.
        
        This pre-computes embeddings and builds a fast nearest neighbor index
        for efficient similarity search.
        """
        from tqdm import tqdm
        
        # Always build embeddings for caching, even if not using GPU for similarity search
        logger.info("[Similar Ticket Finder] Building similarity index...")
        print("  🧠 Pre-computing embeddings for similarity search...")
        
        embeddings = []
        ids = []
        
        for idx, row in tqdm(historical_df.iterrows(), total=len(historical_df), 
                             desc="  Embedding tickets", ncols=80):
            text = str(row.get(COL_SUMMARY, row.get('Text', ''))).strip()
            if not text:
                continue
            
            embedding = self._get_embedding(text, f"hist_{idx}")
            if embedding is not None:
                embeddings.append(embedding)
                ids.append(idx)
        
        if embeddings:
            self.index_embeddings = np.array(embeddings, dtype=np.float32)
            self.index_ids = ids
            
            if self.use_gpu:
                # Build GPU search index
                self.gpu_search_index = GPUSimilaritySearch(
                    use_gpu=self.use_gpu,
                    metric='cosine',
                    n_neighbors=min(self.top_k * 2, len(embeddings))
                )
                self.gpu_search_index.fit(self.index_embeddings)
                print(f"  ✅ GPU index built with {len(embeddings)} embeddings")
            else:
                print(f"  ✅ CPU index built with {len(embeddings)} embeddings")
            
            logger.info(f"[Similar Ticket Finder] Index built with {len(embeddings)} embeddings")
    
    def _cosine_similarity(self, vec1, vec2):
        """Calculate cosine similarity between two vectors (GPU-accelerated)."""
        if vec1 is None or vec2 is None:
            return 0.0
        use_gpu = USE_GPU and is_gpu_available()
        return cosine_similarity_gpu(vec1, vec2, use_gpu=use_gpu)
    
    def _get_embedding(self, text, ticket_id=None):
        """Get embedding for text, with caching."""
        if not text or pd.isna(text):
            return None
        
        text = str(text).strip()
        if not text:
            return None
        
        cache_key = ticket_id if ticket_id else hash(text)
        if cache_key in self.embedding_cache:
            return self.embedding_cache[cache_key]
        
        if self.ai:
            embedding = self.ai.get_embedding(text)
            if embedding is not None and len(embedding) > 0:
                self.embedding_cache[cache_key] = embedding
                return embedding
        
        return None
    
    def _get_resolution_text(self, row):
        """Extract resolution/action text from a ticket row."""
        resolution_parts = []
        
        if COL_LESSON_TITLE in row.index:
            lesson = row.get(COL_LESSON_TITLE)
            if lesson and not pd.isna(lesson) and str(lesson).strip():
                resolution_parts.append(f"Lesson: {lesson}")
        
        if 'Action_Required' in row.index:
            action = row.get('Action_Required')
            if action and not pd.isna(action) and str(action).strip() and action != 'Monitor':
                resolution_parts.append(f"Action: {action}")
        
        root_cause = None
        if 'Root_Cause_Category' in row.index:
            root_cause = row.get('Root_Cause_Category')
        elif COL_ROOT_CAUSE in row.index:
            root_cause = row.get(COL_ROOT_CAUSE)
        
        if root_cause and not pd.isna(root_cause) and str(root_cause).strip():
            resolution_parts.append(f"Root Cause: {root_cause}")
        
        return " | ".join(resolution_parts) if resolution_parts else "No resolution documented"
    
    def _analyze_resolution_consistency(self, similar_tickets):
        """Analyze if similar tickets had consistent resolutions."""
        if len(similar_tickets) < 2:
            return {
                'consistency_score': 1.0,
                'status': 'N/A - Not enough similar tickets',
                'unique_approaches': 1,
                'dominant_approach': None,
                'insight': None,
                'resolution_time': None
            }
        
        root_causes = []
        actions = []
        lessons = []
        resolution_times = []
        
        for ticket in similar_tickets:
            rc = ticket.get('root_cause', '')
            if rc and rc != 'Unknown':
                root_causes.append(rc)
            
            action = ticket.get('action', '')
            if action and action != 'Monitor' and action != 'No resolution documented':
                actions.append(action)
            
            lesson = ticket.get('lesson', '')
            if lesson:
                lessons.append(lesson)
            
            days = ticket.get('days_to_resolution')
            if days is not None:
                resolution_times.append(days)
        
        unique_root_causes = len(set(root_causes)) if root_causes else 0
        unique_actions = len(set(actions)) if actions else 0
        
        total_unique = unique_root_causes + unique_actions
        total_items = len(root_causes) + len(actions)
        
        if total_items == 0:
            consistency_score = 0.0
            status = '⚠️ No resolution data available'
        elif total_unique <= 1:
            consistency_score = 1.0
            status = '✅ Highly consistent resolutions'
        elif total_unique <= 2:
            consistency_score = 0.7
            status = '🟡 Mostly consistent with minor variations'
        else:
            consistency_score = max(0.3, 1.0 - (total_unique / total_items))
            status = '🔴 Inconsistent resolutions - review needed'
        
        all_approaches = root_causes + actions
        dominant = None
        if all_approaches:
            counts = Counter(all_approaches)
            dominant = counts.most_common(1)[0][0] if counts else None
        
        # Analyze resolution times
        resolution_time_analysis = None
        if resolution_times:
            avg_time = np.mean(resolution_times)
            min_time = min(resolution_times)
            max_time = max(resolution_times)
            std_time = np.std(resolution_times) if len(resolution_times) > 1 else 0
            
            resolution_time_analysis = {
                'avg_days': round(avg_time, 1),
                'min_days': round(min_time, 1),
                'max_days': round(max_time, 1),
                'std_days': round(std_time, 1),
                'sample_size': len(resolution_times),
                'time_variance': 'High' if std_time > avg_time * 0.5 else 'Low'
            }
        
        insight = None
        if consistency_score < 0.5 and len(similar_tickets) >= 3:
            insight = f"⚠️ ALERT: {len(similar_tickets)} similar tickets resolved {unique_root_causes + unique_actions} different ways. Consider standardizing approach."
        elif consistency_score >= 0.8 and lessons:
            insight = f"✅ Consistent approach: {dominant if dominant else lessons[0][:50]}"
        
        if resolution_time_analysis and resolution_time_analysis['time_variance'] == 'High':
            time_insight = f"⏱️ Resolution time varies significantly ({resolution_time_analysis['min_days']}-{resolution_time_analysis['max_days']} days)"
            insight = f"{insight}\n{time_insight}" if insight else time_insight
        
        return {
            'consistency_score': consistency_score,
            'status': status,
            'unique_approaches': total_unique,
            'dominant_approach': dominant,
            'insight': insight,
            'has_lessons': len(lessons) > 0,
            'resolution_time': resolution_time_analysis
        }
    
    def _build_similar_ticket_result(self, idx, hist_row, query_id, adjusted_similarity, base_similarity):
        """Build a similar ticket result dictionary."""
        hist_text = str(hist_row.get(COL_SUMMARY, hist_row.get('Text', ''))).strip()
        hist_ticket_id = hist_row.get('ID', str(idx))
        resolution_text = self._get_resolution_text(hist_row)
        days_to_resolution = self._calculate_days_to_resolution(hist_row)
        
        feedback_key = f"{query_id}|{hist_ticket_id}"
        reverse_key = f"{hist_ticket_id}|{query_id}"
        has_feedback = feedback_key in self.feedback_data or reverse_key in self.feedback_data
        feedback_status = ""
        if has_feedback:
            fb = self.feedback_data.get(feedback_key) or self.feedback_data.get(reverse_key)
            if fb:
                feedback_status = "✅ Confirmed" if fb.get('is_similar') else "❌ Rejected"
        
        return {
            'index': idx,
            'id': hist_ticket_id,
            'text': hist_text[:200] + '...' if len(hist_text) > 200 else hist_text,
            'similarity': adjusted_similarity,
            'base_similarity': base_similarity,
            'category': hist_row.get('AI_Category', 'Unknown'),
            'severity': hist_row.get('Severity_Norm', hist_row.get(COL_SEVERITY, 'Unknown')),
            'root_cause': hist_row.get('Root_Cause_Category', hist_row.get(COL_ROOT_CAUSE, 'Unknown')),
            'resolution': resolution_text,
            'lesson': hist_row.get(COL_LESSON_TITLE, ''),
            'action': hist_row.get('Action_Required', ''),
            'date': hist_row.get('Issue_Date', hist_row.get(COL_DATETIME, '')),
            'engineer': hist_row.get('Engineer', hist_row.get(COL_ENGINEER, '')),
            'lob': hist_row.get('LOB', hist_row.get(COL_LOB, '')),
            'friction_score': hist_row.get('Strategic_Friction_Score', 0),
            'recurred': hist_row.get('Recurrence_Actual', hist_row.get('Is_Repeat', 'Unknown')),
            'days_to_resolution': days_to_resolution,
            'feedback_status': feedback_status
        }
    
    def find_similar(self, query_row, historical_df, exclude_self=True):
        """Find tickets similar to the query ticket (GPU-accelerated when available)."""
        query_id = query_row.get(COL_SUMMARY, query_row.get('ID', ''))
        query_text = str(query_row.get(COL_SUMMARY, query_row.get('Text', ''))).strip()
        
        if not query_text:
            return []
        
        query_embedding = self._get_embedding(query_text, f"query_{hash(query_text)}")
        if query_embedding is None:
            return []
        
        similar_tickets = []
        
        # Use GPU index if available for fast search
        if self.gpu_search_index is not None and self.index_ids is not None:
            # Fast GPU-based nearest neighbor search
            similarities, indices = self.gpu_search_index.search(query_embedding, k=self.top_k * 2)
            
            for sim, local_idx in zip(similarities, indices):
                if local_idx >= len(self.index_ids):
                    continue
                    
                idx = self.index_ids[local_idx]
                if idx not in historical_df.index:
                    continue
                    
                hist_row = historical_df.loc[idx]
                
                if exclude_self:
                    hist_id = hist_row.get(COL_SUMMARY, hist_row.get('ID', ''))
                    if hist_id == query_id:
                        continue
                
                base_similarity = float(sim)
                hist_ticket_id = hist_row.get('ID', str(idx))
                adjusted_similarity = self.get_feedback_adjusted_similarity(
                    str(query_id), str(hist_ticket_id), base_similarity
                )
                
                if adjusted_similarity >= self.similarity_threshold:
                    similar_tickets.append(
                        self._build_similar_ticket_result(idx, hist_row, query_id, 
                                                          adjusted_similarity, base_similarity)
                    )
            
            similar_tickets.sort(key=lambda x: x['similarity'], reverse=True)
            return similar_tickets[:self.top_k]
        
        # Fallback: CPU-based row-by-row search
        for idx, hist_row in historical_df.iterrows():
            if exclude_self:
                hist_id = hist_row.get(COL_SUMMARY, hist_row.get('ID', ''))
                if hist_id == query_id:
                    continue
            
            hist_text = str(hist_row.get(COL_SUMMARY, hist_row.get('Text', ''))).strip()
            if not hist_text:
                continue
            
            hist_embedding = self._get_embedding(hist_text, f"hist_{idx}")
            if hist_embedding is None:
                continue
            
            base_similarity = self._cosine_similarity(query_embedding, hist_embedding)
            
            hist_ticket_id = hist_row.get('ID', str(idx))
            adjusted_similarity = self.get_feedback_adjusted_similarity(
                str(query_id), str(hist_ticket_id), base_similarity
            )
            
            if adjusted_similarity >= self.similarity_threshold:
                resolution_text = self._get_resolution_text(hist_row)
                days_to_resolution = self._calculate_days_to_resolution(hist_row)
                
                feedback_key = f"{query_id}|{hist_ticket_id}"
                reverse_key = f"{hist_ticket_id}|{query_id}"
                has_feedback = feedback_key in self.feedback_data or reverse_key in self.feedback_data
                feedback_status = ""
                if has_feedback:
                    fb = self.feedback_data.get(feedback_key) or self.feedback_data.get(reverse_key)
                    if fb:
                        feedback_status = "✅ Confirmed" if fb.get('is_similar') else "❌ Rejected"
                
                similar_tickets.append({
                    'index': idx,
                    'id': hist_ticket_id,
                    'text': hist_text[:200] + '...' if len(hist_text) > 200 else hist_text,
                    'similarity': adjusted_similarity,
                    'base_similarity': base_similarity,
                    'category': hist_row.get('AI_Category', 'Unknown'),
                    'severity': hist_row.get('Severity_Norm', hist_row.get(COL_SEVERITY, 'Unknown')),
                    'root_cause': hist_row.get('Root_Cause_Category', hist_row.get(COL_ROOT_CAUSE, 'Unknown')),
                    'resolution': resolution_text,
                    'lesson': hist_row.get(COL_LESSON_TITLE, ''),
                    'action': hist_row.get('Action_Required', ''),
                    'date': hist_row.get('Issue_Date', hist_row.get(COL_DATETIME, '')),
                    'engineer': hist_row.get('Engineer', hist_row.get(COL_ENGINEER, '')),
                    'lob': hist_row.get('LOB', hist_row.get(COL_LOB, '')),
                    'friction_score': hist_row.get('Strategic_Friction_Score', 0),
                    'recurred': hist_row.get('Recurrence_Actual', hist_row.get('Is_Repeat', 'Unknown')),
                    'days_to_resolution': days_to_resolution,
                    'feedback_status': feedback_status
                })
        
        similar_tickets.sort(key=lambda x: x['similarity'], reverse=True)
        return similar_tickets[:self.top_k]
    
    def analyze_ticket(self, query_row, historical_df):
        """Full analysis of a ticket: find similar and compare resolutions."""
        similar = self.find_similar(query_row, historical_df)
        
        if not similar:
            return {
                'similar_tickets': [],
                'match_count': 0,
                'resolution_analysis': None,
                'recommendation': 'No sufficiently similar historical tickets found.',
                'confidence': 'Low',
                'expected_resolution_days': None
            }
        
        resolution_analysis = self._analyze_resolution_consistency(similar)
        
        if resolution_analysis['consistency_score'] >= 0.8 and resolution_analysis['dominant_approach']:
            recommendation = f"✅ RECOMMENDED: Follow established approach - {resolution_analysis['dominant_approach']}"
            confidence = 'High'
        elif resolution_analysis['consistency_score'] >= 0.5:
            recommendation = f"🟡 SUGGESTED: Review past resolutions. Most common: {resolution_analysis['dominant_approach']}"
            confidence = 'Medium'
        else:
            recommendation = "🔴 CAUTION: Inconsistent past resolutions. Consult with team lead."
            confidence = 'Low'
        
        expected_days = None
        if resolution_analysis.get('resolution_time'):
            time_info = resolution_analysis['resolution_time']
            expected_days = time_info['avg_days']
            recommendation += f"\n\n⏱️ EXPECTED: {time_info['avg_days']} days (range: {time_info['min_days']}-{time_info['max_days']} days)"
        
        lessons = [t['lesson'] for t in similar if t.get('lesson')]
        if lessons:
            recommendation += f"\n\n📚 LESSONS LEARNED:\n" + "\n".join([f"  • {l[:100]}" for l in lessons[:3]])
        
        recurrence_info = [t for t in similar if t.get('recurred') in ['Yes', True, 1]]
        if recurrence_info:
            recommendation += f"\n\n⚠️ WARNING: {len(recurrence_info)} of {len(similar)} similar tickets recurred!"
        
        return {
            'similar_tickets': similar,
            'match_count': len(similar),
            'avg_similarity': np.mean([t['similarity'] for t in similar]),
            'resolution_analysis': resolution_analysis,
            'recommendation': recommendation,
            'confidence': confidence,
            'expected_resolution_days': expected_days
        }
    
    def process_all_tickets(self, df, progress_callback=None):
        """Find similar tickets for all tickets in the dataset."""
        from tqdm import tqdm
        
        logger.info(f"[Similar Ticket Finder] Processing {len(df)} tickets...")
        
        df = df.copy()
        
        # Build GPU index first for fast similarity search
        # This pre-computes all embeddings once
        self.build_gpu_index(df)
        
        # Initialize new columns
        df['Similar_Ticket_Count'] = 0
        df['Best_Match_Similarity'] = 0.0
        df['Resolution_Consistency'] = 'N/A'
        df['Similar_Ticket_IDs'] = ''
        df['Expected_Resolution_Days'] = np.nan
        df['Avg_Similar_Resolution_Days'] = np.nan
        df['Resolution_Recommendation'] = ''
        df['Similarity_Feedback'] = ''
        
        processed = 0
        total = len(df)
        
        # Use tqdm progress bar
        for idx, row in tqdm(df.iterrows(), total=total, desc="  Analyzing tickets", ncols=80):
            analysis = self.analyze_ticket(row, df)
            
            df.at[idx, 'Similar_Ticket_Count'] = analysis['match_count']
            df.at[idx, 'Best_Match_Similarity'] = analysis.get('avg_similarity', 0)
            
            if analysis['resolution_analysis']:
                df.at[idx, 'Resolution_Consistency'] = analysis['resolution_analysis']['status']
                
                if analysis['resolution_analysis'].get('resolution_time'):
                    time_info = analysis['resolution_analysis']['resolution_time']
                    df.at[idx, 'Avg_Similar_Resolution_Days'] = time_info['avg_days']
            
            if analysis.get('expected_resolution_days') is not None:
                df.at[idx, 'Expected_Resolution_Days'] = analysis['expected_resolution_days']
            
            if analysis['similar_tickets']:
                similar_info = []
                for t in analysis['similar_tickets'][:3]:
                    ticket_str = str(t['id'])
                    if t.get('days_to_resolution') is not None:
                        ticket_str += f" ({t['days_to_resolution']}d)"
                    similar_info.append(ticket_str)
                df.at[idx, 'Similar_Ticket_IDs'] = ', '.join(similar_info)
            
            df.at[idx, 'Resolution_Recommendation'] = analysis['recommendation'].split('\n')[0]
            
            processed += 1
            if progress_callback and processed % 10 == 0:
                progress_callback(processed, total)
        
        has_matches = (df['Similar_Ticket_Count'] > 0).sum()
        inconsistent = df['Resolution_Consistency'].str.contains('Inconsistent', na=False).sum()
        has_time_estimate = df['Expected_Resolution_Days'].notna().sum()
        
        logger.info(f"[Similar Ticket Finder] Complete:")
        logger.info(f"  → {has_matches}/{total} tickets have similar historical matches")
        logger.info(f"  → {inconsistent} tickets have inconsistent resolution patterns")
        logger.info(f"  → {has_time_estimate} tickets have resolution time estimates")
        
        return df
    
    def export_for_feedback(self, df, output_path):
        """Export similar ticket matches for human review with dropdowns."""
        feedback_rows = []
        
        for idx, row in df.iterrows():
            ticket_id = row.get('ID', str(idx))
            similar_ids_str = row.get('Similar_Ticket_IDs', '')
            similarity = row.get('Best_Match_Similarity', 0)
            category = row.get('AI_Category', 'Unknown')
            
            actual_days = self._calculate_days_to_resolution(row)
            
            if similar_ids_str:
                for similar_info in similar_ids_str.split(', '):
                    similar_id = similar_info.split('(')[0].strip()
                    
                    feedback_key = f"{ticket_id}|{similar_id}"
                    existing = self.feedback_data.get(feedback_key, {})
                    existing_feedback = ""
                    if existing:
                        existing_feedback = "Similar" if existing.get('is_similar') else "Not Similar"
                    
                    feedback_rows.append({
                        'Ticket_ID': ticket_id,
                        'Ticket_Summary': str(row.get(COL_SUMMARY, ''))[:100],
                        'Category': category,
                        'Similar_Ticket_ID': similar_id,
                        'AI_Similarity': f"{similarity:.0%}",
                        'Is_Similar': existing_feedback,
                        'Actual_Resolution_Days': actual_days if actual_days else '',
                        'Expected_Resolution_Days': '',
                        'Notes': existing.get('notes', '')
                    })
        
        feedback_df = pd.DataFrame(feedback_rows)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            instructions = pd.DataFrame({
                'Instructions': [
                    'SIMILARITY FEEDBACK & RESOLUTION TIME INPUT',
                    '',
                    '📋 SIMILARITY FEEDBACK:',
                    '1. Review each ticket pair in the "Similarity Feedback" sheet',
                    '2. In "Is_Similar" column, use DROPDOWN to select:',
                    '   - "Similar" if the tickets ARE actually similar',
                    '   - "Not Similar" if the tickets are NOT related',
                    '3. AI learns from your feedback on next run!',
                    '',
                    f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
                ]
            })
            instructions.to_excel(writer, sheet_name='Instructions', index=False)
            feedback_df.to_excel(writer, sheet_name='Similarity Feedback', index=False)
            
            ws = writer.sheets['Similarity Feedback']
            
            similarity_validation = DataValidation(
                type='list',
                formula1='"Similar,Not Similar"',
                allow_blank=True,
                showDropDown=False
            )
            similarity_validation.error = 'Please select Similar or Not Similar'
            similarity_validation.errorTitle = 'Invalid Selection'
            
            if len(feedback_df) > 0:
                similarity_validation.add(f'F2:F{len(feedback_df) + 1}')
                ws.add_data_validation(similarity_validation)
            
            for col_idx in range(1, 10):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
            
            for row_idx in range(2, len(feedback_df) + 2):
                cell = ws.cell(row=row_idx, column=6)
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        logger.info(f"[Similar Ticket Finder] Feedback file exported to {output_path}")
        return output_path
    
    def load_feedback_from_excel(self, excel_path):
        """Load human feedback from the Similar Tickets feedback Excel file."""
        resolution_expectations = {}
        
        try:
            if not os.path.exists(excel_path):
                return 0, resolution_expectations
            
            feedback_df = pd.read_excel(excel_path, sheet_name='Similarity Feedback')
            
            loaded = 0
            for _, row in feedback_df.iterrows():
                ticket_id = str(row.get('Ticket_ID', ''))
                similar_id = str(row.get('Similar_Ticket_ID', ''))
                category = str(row.get('Category', 'Unknown'))
                
                feedback = str(row.get('Is_Similar', row.get('Human_Feedback', ''))).strip().lower()
                notes = str(row.get('Notes', ''))
                
                is_similar = None
                if feedback in ['similar', 'correct']:
                    is_similar = True
                elif feedback in ['not similar', 'wrong']:
                    is_similar = False
                
                if ticket_id and similar_id and is_similar is not None:
                    self.record_feedback(ticket_id, similar_id, is_similar, notes)
                    loaded += 1
                
                expected_days = row.get('Expected_Resolution_Days', None)
                if pd.notna(expected_days) and expected_days != '':
                    try:
                        exp_days = float(expected_days)
                        if category not in resolution_expectations:
                            resolution_expectations[category] = []
                        resolution_expectations[category].append(exp_days)
                    except (ValueError, TypeError):
                        pass
            
            for cat in resolution_expectations:
                resolution_expectations[cat] = np.mean(resolution_expectations[cat])
            
            logger.info(f"[Similar Ticket Finder] Loaded {loaded} feedback entries")
            return loaded, resolution_expectations
            
        except Exception as e:
            logger.warning(f"[Similar Ticket Finder] Could not load Excel feedback: {e}")
            return 0, resolution_expectations


# Global similar ticket finder instance
similar_ticket_finder = None


def apply_similar_ticket_analysis(df, ai_engine=None):
    """Apply similar ticket analysis to the dataframe."""
    global similar_ticket_finder
    
    logger.info("[Similar Ticket Finder] Initializing similar ticket analysis...")
    
    similar_ticket_finder = SimilarTicketFinder(
        ai_engine=ai_engine,
        top_k=5,
        similarity_threshold=0.70
    )
    
    similarity_feedback_excel = "similarity_feedback.xlsx"
    if os.path.exists(similarity_feedback_excel):
        similar_ticket_finder.load_feedback_from_excel(similarity_feedback_excel)
    
    df = similar_ticket_finder.process_all_tickets(df)
    
    try:
        similar_ticket_finder.export_for_feedback(df, similarity_feedback_excel)
    except Exception as e:
        logger.warning(f"Could not export similarity feedback file: {e}")
    
    return df
