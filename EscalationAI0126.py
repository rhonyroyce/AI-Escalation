import pandas as pd
import numpy as np
import re
import os
import requests
import warnings
import logging
import tkinter as tk
from tkinter import filedialog, messagebox
from tqdm import tqdm
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, roc_auc_score
import pickle
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Optional, List, Dict
from datetime import datetime
import json

# Suppress warnings
warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='   > [%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

# ==========================================
# CONFIGURATION
# ==========================================
OLLAMA_BASE_URL = "http://localhost:11434"
EMBED_MODEL = "qwen3-embedding:8b"  # Best semantic understanding for classification
GEN_MODEL = "gemma3:27b"              # Clean professional output, no thinking tags

# STRATEGIC WEIGHTS (McKinsey Framework)
WEIGHTS = {
    'BASE_SEVERITY': {'Critical': 100, 'Major': 50, 'Minor': 10, 'Default': 5},
    'TYPE_MULTIPLIER': {'Escalations': 1.5, 'Concerns': 1.0, 'Lessons Learned': 0.0},
    'ORIGIN_MULTIPLIER': {'External': 2.5, 'Internal': 1.0},
    'IMPACT_MULTIPLIER': {'High': 2.0, 'Low': 1.1, 'None': 1.0}
}

# SIMILARITY THRESHOLD FOR RECIDIVISM DETECTION
# Lowered thresholds based on observed max similarity scores (~0.62)
SIMILARITY_THRESHOLD_HIGH = 0.60   # High confidence match (was 0.80)
SIMILARITY_THRESHOLD_MEDIUM = 0.50  # Medium confidence - worth flagging (was 0.70)
KEYWORD_OVERLAP_THRESHOLD = 0.35    # 35% keyword overlap = likely related (was 0.5)

# CLASSIFICATION CONFIDENCE THRESHOLD
MIN_CLASSIFICATION_CONFIDENCE = 0.25  # Below this = "Unclassified"

# RECIDIVISM PENALTY MULTIPLIERS
RECIDIVISM_PENALTY_HIGH = 1.5      # 50% score increase for confirmed repeats
RECIDIVISM_PENALTY_MEDIUM = 1.25   # 25% score increase for possible repeats

# FEEDBACK/REINFORCEMENT LEARNING SETTINGS
FEEDBACK_FILE = "classification_feedback.xlsx"
FEEDBACK_WEIGHT = 0.4  # How much weight to give user feedback vs original anchors (0-1)

# PRICE CATALOG SETTINGS
PRICE_CATALOG_FILE = "price_catalog.xlsx"
DEFAULT_HOURLY_RATE = 20.0  # $/hour for labor if not specified
DEFAULT_DELAY_COST = 500.0  # $/hour of delay (SLA penalty, revenue loss)

# REPORT METADATA
REPORT_VERSION = "2.2"
REPORT_TITLE = "STRATEGIC FRICTION ANALYSIS"

# CATEGORIZATION ANCHORS (The "Bucket" Definitions for AI Classification)
# Comprehensive 11-category system for telecom escalations
ANCHORS = {
    # Equipment & Infrastructure
    "RF & Antenna Issues": [
        "antenna misalignment", "vswr alarm", "rru fault", "radio failure", 
        "sector down", "rf interference", "antenna swap", "feeder cable",
        "bbu fault", "baseband", "carrier down", "cell outage"
    ],
    "Transmission & Backhaul": [
        "fiber cut", "microwave link down", "transmission failure", "mw fade",
        "ethernet fault", "capacity exhaust", "latency", "packet loss",
        "backhaul", "transport", "ipsec", "vpn down", "circuit down"
    ],
    "Power & Environment": [
        "power outage", "battery failure", "rectifier fault", "generator issue",
        "ac failure", "high temperature", "equipment smoke", "cooling failure",
        "ups fault", "breaker trip", "fuel empty", "solar panel", "cabinet alarm"
    ],
    
    # Access & Field Operations
    "Site Access & Logistics": [
        "keys missing", "gate locked", "access denied", "landlord issue",
        "permit expired", "security clearance", "escort required", "site unsafe",
        "site inaccessible", "road blocked", "no access", "tower climb"
    ],
    "Contractor & Vendor Issues": [
        "crew no show", "wrong crew", "incomplete work", "material shortage",
        "vendor delay", "subcontractor issue", "quality defect", "rework required",
        "parts missing", "tool shortage", "training gap", "crew late"
    ],
    
    # Technical & Software
    "Configuration & Integration": [
        "parameter mismatch", "wrong ip", "integration error", "script failed",
        "alarm suppressed", "neighbor list", "handover failure", "config rollback",
        "software bug", "feature activation", "license issue", "template error"
    ],
    "OSS/NMS & Systems": [
        "oss fault", "nms unreachable", "provisioning error", "database sync",
        "element manager", "snmp trap", "discovery failed", "inventory mismatch",
        "mediation", "ticketing system", "monitoring gap", "correlation failure",
        "nesting", "nest extension", "nsi", "si nesting", "cell planning", 
        "network planning", "pci conflict", "antenna tilt", "coverage optimization"
    ],
    
    # Process & Communication
    "Process & Documentation": [
        "paperwork missing", "approval delay", "incorrect data", "process gap",
        "sow mismatch", "change window", "notification failure", "handoff issue",
        "documentation error", "method statement", "safety violation", "audit finding"
    ],
    "Communication & Coordination": [
        "miscommunication", "escalation delay", "wrong contact", "no response",
        "scheduling conflict", "timezone issue", "language barrier", "email missed",
        "handover gap", "shift change", "notification delay", "stakeholder"
    ],
    
    # External Factors
    "Weather & Natural Events": [
        "flood", "hurricane", "storm", "lightning", "extreme heat", "ice", 
        "wind damage", "earthquake", "wildfire", "snow", "fog", "monsoon"
    ],
    "Third-Party & External": [
        "theft", "vandalism", "fiber cut by third party", "construction damage",
        "utility outage", "road closure", "civil unrest", "regulatory hold",
        "permit rejection", "zoning issue", "public complaint", "legal dispute"
    ]
}

# COLUMN NAME CONSTANTS
COL_SEVERITY = 'tickets_data_severity'
COL_TYPE = 'tickets_data_type_1'
COL_ORIGIN = 'tickets_data_escalation_origin'
COL_IMPACT = 'tickets_data_business_impact_pm'
COL_SUMMARY = 'tickets_data_issue_summary'
COL_CATEGORY = 'tickets_data_issue_category_1'
COL_DATETIME = 'tickets_data_issue_datetime'
COL_CLOSE_DATE = 'tickets_data_close_datetime'  # When ticket was resolved/closed
COL_RESOLUTION_DATE = 'tickets_data_close_datetime'  # Alias for resolution date
COL_RESOLUTION_NOTES = 'tickets_data_resolution_notes'  # Resolution notes/comments
COL_ENGINEER = 'tickets_data_engineer_name'  # Engineer responsible for issue
COL_LOB = 'tickets_data_lob'  # Line of Business
COL_LESSON_TITLE = 'tickets_data_lessons_learned_title'  # Lesson title if logged
COL_LESSON_STATUS = 'tickets_data_lessons_learned_status'  # Lesson status
COL_ROOT_CAUSE = 'tickets_data_root_cause'  # Root cause category (Human Error, Non Amdocs, Process Gap, etc.)
COL_RECURRENCE_RISK = 'tickets_data_risk_for_recurrence_pm'  # PM's assessment of recurrence risk (Low/High)

# Similarity feedback file path
SIMILARITY_FEEDBACK_PATH = "similarity_feedback.json"

# Root Cause Categories for analysis
ROOT_CAUSE_CATEGORIES = {
    'Human Error': ['human error', 'operator error', 'manual error', 'user error', 'mistake'],
    'External Party': ['non amdocs', 'external', 'vendor', 'third party', '3rd party', 'customer caused'],
    'Process Gap': ['process gap', 'process issue', 'sop missing', 'procedure gap', 'workflow issue'],
    'System/Technical': ['system error', 'technical issue', 'software bug', 'hardware failure', 'system failure'],
    'Training Gap': ['training', 'knowledge gap', 'skill gap', 'lack of training'],
    'Communication': ['communication', 'miscommunication', 'information gap', 'handoff issue'],
    'Resource': ['resource', 'understaffed', 'capacity', 'bandwidth'],
}

# Engineer Recidivism Thresholds
ENGINEER_REPEAT_THRESHOLD = 3  # Flag engineers with 3+ issues
LOB_RISK_THRESHOLD = 5  # Flag LOBs with 5+ issues

REQUIRED_COLUMNS = [COL_SEVERITY, COL_TYPE, COL_ORIGIN]

MC_BLUE = '#004C97'

# ==========================================
# 1. AI WRAPPERS (HYBRID ARCHITECTURE)
# ==========================================
class OllamaBrain:
    """Handles both Embedding (Left Brain) and Generation (Right Brain)"""
   
    def __init__(self):
        self.embed_model = EMBED_MODEL
        self.gen_model = GEN_MODEL
        self._embed_dim = None

    def get_embedding(self, text):
        """Get vector for a single string"""
        if pd.isna(text) or text == "": return np.zeros(self.get_dim())
        try:
            res = requests.post(f"{OLLAMA_BASE_URL}/api/embed",
                              json={"model": self.embed_model, "input": str(text)},
                              timeout=30)
            if res.status_code == 200:
                vec = res.json().get('embedding') or res.json().get('embeddings', [[]])[0]
                return np.array(vec)
        except Exception as e:
            logger.warning(f"Embedding failed: {e}")
        return np.zeros(self.get_dim())

    def get_embeddings_batch(self, texts: List[str]) -> List[np.ndarray]:
        """Get vectors for multiple strings in one API call"""
        if not texts:
            return []
        
        # Filter out empty texts, keep track of indices
        valid_indices = []
        valid_texts = []
        for i, text in enumerate(texts):
            if pd.isna(text) or text == "":
                continue
            valid_indices.append(i)
            valid_texts.append(str(text))
        
        # Initialize result with zero vectors
        result = [np.zeros(self.get_dim()) for _ in texts]
        
        if not valid_texts:
            return result
        
        try:
            res = requests.post(f"{OLLAMA_BASE_URL}/api/embed",
                              json={"model": self.embed_model, "input": valid_texts},
                              timeout=120)
            if res.status_code == 200:
                embeddings = res.json().get('embeddings', [])
                for idx, vec in zip(valid_indices, embeddings):
                    result[idx] = np.array(vec)
        except Exception as e:
            logger.warning(f"Batch embedding failed: {e}")
        
        return result

    def get_dim(self):
        if self._embed_dim: return self._embed_dim
        # Test call to get dimension
        v = self.get_embedding("test")
        self._embed_dim = len(v)
        return self._embed_dim

    def _strip_thinking_tags(self, text):
        """Remove <think>...</think> blocks from LLM output (qwen3 thinking mode)"""
        if not text:
            return text
        # Remove everything between <think> and </think> tags (including the tags)
        cleaned = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL | re.IGNORECASE)
        return cleaned.strip()

    def generate_synthesis(self, context_text):
        """Use the LLM to write a comprehensive executive summary"""
        prompt = f"""You are a Principal Consultant at McKinsey & Company, specializing in Telecom Operations Risk Management. You have been engaged to analyze escalation data for a major telecommunications network deployment project.

ROLE & EXPERTISE:
- You are an expert in identifying systemic operational failures
- You understand telecom infrastructure (RAN, transmission, fiber, site access)
- You focus on strategic risks that threaten project timelines, costs, and reputation
- You communicate with C-suite executives who need actionable insights

ANALYSIS FRAMEWORK:
1. PATTERN RECOGNITION: Identify recurring failure modes and root causes
2. RISK STRATIFICATION: Distinguish between isolated incidents vs systemic issues
3. ORGANIZATIONAL LEARNING: Assess whether past lessons are being applied
4. BUSINESS IMPACT: Translate technical issues to business consequences

DATA CONTEXT (Analyzed Escalation Report):
{context_text}

YOUR TASK:
Produce a comprehensive Executive Risk Assessment with the following sections:

1. CRITICAL ALERT (1-2 sentences): The single most urgent finding that requires immediate executive attention.

2. KEY FINDINGS (3-4 bullet points): Major patterns, systemic issues, or concerning trends discovered in the data.

3. ROOT CAUSE HYPOTHESIS: Based on the data patterns, what underlying organizational or process issues might be driving these escalations?

4. RECOMMENDED ACTIONS (2-3 specific actions): Concrete steps leadership should take this week to address the risks.

5. RISK OUTLOOK: One sentence on what will happen if these issues are not addressed.

FORMATTING RULES:
- Write in clear, professional business English
- Be specific - reference actual categories, percentages, and issue types from the data
- Be direct and urgent where warranted
- Do NOT use markdown formatting (no #, *, or bullet symbols)
- Use plain text with clear section headers
- Total length: 250-400 words
"""
        try:
            res = requests.post(f"{OLLAMA_BASE_URL}/api/generate",
                              json={
                                  "model": self.gen_model, 
                                  "prompt": prompt, 
                                  "stream": False,
                                  "options": {
                                      "num_predict": 800  # Allow longer output for detailed analysis
                                  }
                              },
                              timeout=120)
            if res.status_code == 200:
                raw_response = res.json()['response'].strip()
                # Clean up any thinking tags that may have leaked through
                cleaned = self._strip_thinking_tags(raw_response)
                return cleaned if cleaned else self._generate_fallback_summary(context_text)
        except Exception as e:
            logger.error(f"AI Synthesis Failed: {e}")
            return self._generate_fallback_summary(context_text)
        return self._generate_fallback_summary(context_text)

    def _generate_fallback_summary(self, context_text):
        """Generate a basic summary when AI is unavailable"""
        logger.warning("Using fallback summary generation (AI unavailable)")
        
        # Extract key numbers from context
        import re
        total_match = re.search(r'Total Weighted Friction Score: ([\d,]+)', context_text)
        external_match = re.search(r'External.*?: ([\d.]+)%', context_text)
        repeat_match = re.search(r'Confirmed Repeat Offenses: (\d+)', context_text)
        
        total = total_match.group(1) if total_match else "N/A"
        external = external_match.group(1) if external_match else "N/A"
        repeats = repeat_match.group(1) if repeat_match else "0"
        
        return (
            f"[AUTO-GENERATED SUMMARY - AI Unavailable]\n\n"
            f"This report analyzed escalation data with a total weighted friction score of {total}. "
            f"External-facing issues account for {external}% of total risk. "
            f"There are {repeats} confirmed repeat offenses indicating potential organizational learning gaps. "
            f"Please review the detailed data and charts below for full analysis."
        )

    def unload(self):
        # Unload both models to free VRAM
        for m in [self.embed_model, self.gen_model]:
            try:
                requests.post(f"{OLLAMA_BASE_URL}/api/generate", json={"model": m, "keep_alive": 0})
            except Exception as e:
                logger.warning(f"Failed to unload model {m}: {e}")

def check_models(ai):
    """Quick probe to ensure required AI models exist and are working"""
    logger.info("Checking AI model availability...")
    
    # Check embedding model
    try:
        test_vec = ai.get_embedding("test connection")
        if len(test_vec) == 0 or np.all(test_vec == 0):
            raise ValueError("Embedding returned zero vector")
        logger.info(f"✓ Embedding model '{EMBED_MODEL}' is active (dim={len(test_vec)})")
    except Exception as e:
        logger.error(f"Embedding model check failed: {e}")
        messagebox.showerror(
            "Model Not Found", 
            f"Embedding model '{EMBED_MODEL}' is not available.\n\n"
            f"Please run:\n  ollama pull {EMBED_MODEL}\n\n"
            f"Error: {e}"
        )
        return False
    
    # Check generation model (optional - just warn, don't block)
    try:
        res = requests.post(
            f"{OLLAMA_BASE_URL}/api/generate",
            json={"model": GEN_MODEL, "prompt": "test", "stream": False},
            timeout=10
        )
        if res.status_code == 200:
            logger.info(f"✓ Generation model '{GEN_MODEL}' is active")
        else:
            logger.warning(f"Generation model '{GEN_MODEL}' may not be available (status: {res.status_code})")
    except Exception as e:
        logger.warning(f"Generation model '{GEN_MODEL}' check failed: {e}. Executive summary may not work.")
    
    return True

def check_ollama_server():
    """Check if Ollama server is running"""
    try:
        res = requests.get(f"{OLLAMA_BASE_URL}/", timeout=3)
        return True
    except Exception as e:
        logger.error(f"Ollama server not reachable: {e}")
        messagebox.showerror(
            "Ollama Not Running",
            f"Cannot connect to Ollama at {OLLAMA_BASE_URL}\n\n"
            "Please ensure Ollama is running:\n  ollama serve"
        )
        return False

def clean_text(text):
    if pd.isna(text): return ""
    text = str(text).replace("_x000D_", " ").replace("\n", " ")
    return re.sub(r'\s+', ' ', text).strip()

def validate_columns(df, required_cols):
    """Validate that required columns exist in dataframe"""
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        logger.warning(f"Missing columns: {missing}. Some features may be limited.")
        return False
    return True

# ==========================================
# 2. FEEDBACK/REINFORCEMENT LEARNING SYSTEM
# ==========================================

class FeedbackLearning:
    """
    Human-in-the-loop learning system for classification improvement.
    
    Workflow:
    1. Run analysis → generates classification_feedback.xlsx
    2. User reviews and corrects categories in the Excel file (easy editing!)
    3. Next run loads corrections and adjusts centroids
    """
    
    def __init__(self, feedback_path: Optional[str] = None):
        self.feedback_path = feedback_path or os.path.join(os.path.dirname(__file__), FEEDBACK_FILE)
        self.corrections: Dict[str, List[Dict]] = {}  # category -> list of {text, embedding}
        self.stats = {'loaded': 0, 'categories_enhanced': 0}
    
    def load_feedback(self, ai) -> bool:
        """Load user feedback from Excel file and compute embeddings for corrections"""
        if not os.path.exists(self.feedback_path):
            logger.info(f"No feedback file found at {self.feedback_path} - using default anchors")
            return False
        
        try:
            # Read the Classifications sheet
            df_feedback = pd.read_excel(self.feedback_path, sheet_name='Classifications')
            
            # Extract corrections (where user filled in Corrected_Category)
            corrections_list = []
            for _, row in df_feedback.iterrows():
                original = str(row.get('AI_Category', '')).strip()
                corrected = str(row.get('Corrected_Category', '')).strip()
                text = str(row.get('Text', '')).strip()
                
                # Only use if user provided a correction (not empty, not 'nan', different from original)
                if corrected and corrected.lower() not in ['', 'nan', 'none'] and corrected != original and text:
                    corrections_list.append({
                        'text': text,
                        'category': corrected,
                        'original': original
                    })
            
            if not corrections_list:
                logger.info("Feedback file found but no corrections to apply")
                return False
            
            logger.info(f"Found {len(corrections_list)} user corrections in feedback file")
            
            # Compute embeddings for corrected texts
            texts = [c['text'] for c in corrections_list]
            embeddings = ai.get_embeddings_batch(texts)
            
            # Group by corrected category
            for corr, emb in zip(corrections_list, embeddings):
                cat = corr['category']
                if cat not in self.corrections:
                    self.corrections[cat] = []
                self.corrections[cat].append({
                    'text': corr['text'],
                    'embedding': emb,
                    'original': corr['original']
                })
            
            self.stats['loaded'] = len(corrections_list)
            self.stats['categories_enhanced'] = len(self.corrections)
            
            logger.info(f"✓ Loaded {self.stats['loaded']} corrections across {self.stats['categories_enhanced']} categories")
            for cat, items in self.corrections.items():
                logger.info(f"  - {cat}: {len(items)} examples")
            
            return True
            
        except Exception as e:
            logger.warning(f"Failed to load feedback file: {e}")
            return False
    
    def adjust_centroids(self, original_centroids: Dict[str, np.ndarray]) -> Dict[str, np.ndarray]:
        """Blend user feedback with original anchor centroids"""
        if not self.corrections:
            return original_centroids
        
        adjusted = {}
        
        for category, centroid in original_centroids.items():
            if category in self.corrections:
                # Compute centroid from user feedback
                feedback_vecs = [c['embedding'] for c in self.corrections[category]]
                feedback_centroid = np.mean(feedback_vecs, axis=0)
                
                # Blend: weighted average of original and feedback
                adjusted[category] = (
                    (1 - FEEDBACK_WEIGHT) * centroid + 
                    FEEDBACK_WEIGHT * feedback_centroid
                )
                logger.info(f"  → Adjusted '{category}' centroid with {len(feedback_vecs)} user examples")
            else:
                adjusted[category] = centroid
        
        # Handle new categories from feedback that don't exist in anchors
        for category in self.corrections:
            if category not in adjusted:
                feedback_vecs = [c['embedding'] for c in self.corrections[category]]
                adjusted[category] = np.mean(feedback_vecs, axis=0)
                logger.info(f"  → Created NEW category '{category}' from {len(feedback_vecs)} user examples")
        
        return adjusted
    
    def save_for_review(self, df: pd.DataFrame, output_dir: str):
        """Export classifications to Excel for easy user review and correction"""
        
        # Create feedback dataframe
        feedback_rows = []
        for idx, row in df.iterrows():
            feedback_rows.append({
                'ID': str(idx),
                'Text': str(row.get('Combined_Text', ''))[:500],  # Truncate for readability
                'AI_Category': row.get('AI_Category', 'Unclassified'),
                'AI_Confidence': round(float(row.get('AI_Confidence', 0)), 3),
                'Root_Cause_Category': row.get('Root_Cause_Category', 'Unclassified'),
                'PM_Recurrence_Risk': row.get('PM_Recurrence_Risk_Norm', 'Unknown'),
                'AI_Recurrence_Risk': row.get('AI_Recurrence_Risk', 'Unknown'),
                'AI_Recurrence_Prob': f"{row.get('AI_Recurrence_Probability', 0)*100:.0f}%",
                'PM_Prediction_Accuracy': row.get('PM_Prediction_Accuracy', 'Pending'),
                'LOB': row.get('LOB', 'Unknown'),
                'Corrected_Category': '',  # User fills this in - LEFT BLANK (dropdown)
                'Notes': ''  # Optional user notes
            })
        
        df_feedback = pd.DataFrame(feedback_rows)
        
        # Create Excel workbook with formatting
        export_path = os.path.join(output_dir, 'classification_feedback.xlsx')
        
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            # Sheet 1: Instructions
            instructions_df = pd.DataFrame({
                'Instructions': [
                    'HOW TO IMPROVE AI CLASSIFICATION',
                    '',
                    '1. Go to the "Classifications" sheet',
                    '2. Review each row - check if AI_Category is correct',
                    '3. If WRONG: Select the correct category from "Corrected_Category" dropdown (Column K)',
                    '4. If CORRECT: Leave "Corrected_Category" blank',
                    '5. Save this file',
                    '6. Run the analysis again - AI will learn from your corrections!',
                    '',
                    'COLUMN DESCRIPTIONS:',
                    '  A) ID: Unique ticket/issue identifier',
                    '  B) Text: Issue description',
                    '  C) AI_Category: AI-assigned category',
                    '  D) AI_Confidence: Confidence score (0-1)',
                    '  E) Root_Cause_Category: PM-reported root cause classification',
                    '  F) PM_Recurrence_Risk: PM prediction of recurrence risk',
                    '  G) PM_Prediction_Accuracy: Whether PM prediction matched actual outcome',
                    '  H) LOB: Line of Business',
                    '  I) AI_Recurrence_Risk: ML-predicted recurrence risk (High/Medium/Low)',
                    '  J) AI_Recurrence_Prob: Probability of recurrence (0-1)',
                    '  K) Corrected_Category: YOUR CORRECTION (dropdown or type new)',
                    '  L) Notes: Add any additional notes',
                    '',
                    'SIMILAR TICKET COLUMNS (in main report):',
                    '  • Similar_Ticket_Count: Number of similar historical tickets found',
                    '  • Best_Match_Similarity: Highest similarity score (0-100%)',
                    '  • Resolution_Consistency: Whether similar tickets were resolved consistently',
                    '  • Expected_Resolution_Days: Predicted resolution time from similar tickets',
                    '  • Avg_Similar_Resolution_Days: Average days to resolve similar tickets',
                    '',
                    'HUMAN FEEDBACK FOR SIMILARITY MATCHING:',
                    '  1. Open "similarity_feedback.xlsx" (auto-generated)',
                    '  2. Review ticket pairs in the "Similarity Feedback" sheet',
                    '  3. Mark as "Correct" or "Wrong" in Human_Feedback column',
                    '  4. Save and re-run - AI adjusts similarity scoring!',
                    '',
                    'AVAILABLE CATEGORIES (use dropdown in Corrected_Category):',
                    *[f'  • {cat}' for cat in ANCHORS.keys()],
                    '',
                    'TIP: Yellow rows = Low confidence (<50%) - priority for review!',
                    'TIP: You can type a NEW category name if none fit!',
                    'TIP: AI_Recurrence_Risk is ML-predicted; PM_Recurrence_Risk is PM-reported',
                    '',
                    f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
                    f'Version: {REPORT_VERSION}'
                ]
            })
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Sheet 2: Classifications (main editing sheet)
            df_feedback.to_excel(writer, sheet_name='Classifications', index=False)
            
            # Sheet 3: Category Reference
            categories_df = pd.DataFrame({
                'Available_Categories': list(ANCHORS.keys()),
                'Example_Keywords': [', '.join(phrases[:3]) for phrases in ANCHORS.values()]
            })
            categories_df.to_excel(writer, sheet_name='Category Reference', index=False)
            
            # Format the Classifications sheet
            workbook = writer.book
            ws = writer.sheets['Classifications']
            
            # Set column widths for new expanded structure (12 columns now)
            ws.column_dimensions['A'].width = 8   # ID
            ws.column_dimensions['B'].width = 80  # Text
            ws.column_dimensions['C'].width = 25  # AI_Category
            ws.column_dimensions['D'].width = 12  # AI_Confidence
            ws.column_dimensions['E'].width = 18  # Root_Cause_Category
            ws.column_dimensions['F'].width = 16  # PM_Recurrence_Risk
            ws.column_dimensions['G'].width = 22  # AI_Recurrence_Risk
            ws.column_dimensions['H'].width = 14  # AI_Recurrence_Prob
            ws.column_dimensions['I'].width = 20  # PM_Prediction_Accuracy
            ws.column_dimensions['J'].width = 12  # LOB
            ws.column_dimensions['K'].width = 25  # Corrected_Category (dropdown)
            ws.column_dimensions['L'].width = 30  # Notes
            
            # Create dropdown list for Corrected_Category column (Column K now)
            category_list = ','.join(list(ANCHORS.keys()))
            category_validation = DataValidation(
                type='list',
                formula1=f'"{category_list}"',
                allow_blank=True,
                showDropDown=False,  # False = show dropdown arrow
                showErrorMessage=True,
                errorTitle='Invalid Category',
                error='Please select from the dropdown or type a new category.',
                promptTitle='Select Category',
                prompt='Choose a category or type a new one'
            )
            category_validation.add(f'K2:K{len(df_feedback) + 1}')
            ws.add_data_validation(category_validation)
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Highlight low confidence rows (< 0.5) in yellow
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row_idx in range(2, len(df_feedback) + 2):
                try:
                    confidence = float(ws.cell(row=row_idx, column=4).value or 0)
                    if confidence < 0.5:
                        for col_idx in range(1, 13):  # Updated to 12 columns (A-L)
                            ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
                except (ValueError, TypeError):
                    pass
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        # Also save to working directory for next run
        df_feedback.to_excel(self.feedback_path, sheet_name='Classifications', index=False)
        
        # Re-open and add the full formatting to the working directory copy
        try:
            with pd.ExcelWriter(self.feedback_path, engine='openpyxl') as writer:
                instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
                df_feedback.to_excel(writer, sheet_name='Classifications', index=False)
                categories_df.to_excel(writer, sheet_name='Category Reference', index=False)
        except Exception as e:
            logger.warning(f"Could not save formatted feedback to working directory: {e}")
        
        logger.info(f"✓ Feedback Excel saved to: {export_path}")
        logger.info(f"  → Edit 'Corrected_Category' column and re-run to improve AI")
        
        return export_path


# Global feedback learner instance
_feedback_learner: Optional[FeedbackLearning] = None

def get_feedback_learner() -> FeedbackLearning:
    """Get or create the feedback learner singleton"""
    global _feedback_learner
    if _feedback_learner is None:
        _feedback_learner = FeedbackLearning()
    return _feedback_learner


# ==========================================
# 2B. PRICE CATALOG SYSTEM
# ==========================================

class PriceCatalog:
    """
    Excel-based pricing catalog for calculating financial impact of escalations.
    Supports keyword patterns, category-based pricing, and severity multipliers.
    """
    
    def __init__(self, catalog_path: str = PRICE_CATALOG_FILE):
        self.catalog_path = catalog_path
        self.category_costs: Dict[str, Dict] = {}       # Category -> {material, labor_hours, delay_cost}
        self.keyword_costs: List[Dict] = []              # Pattern-based pricing rules
        self.severity_multipliers: Dict[str, float] = {} # Severity -> multiplier
        self.origin_premiums: Dict[str, float] = {}      # Origin -> premium percentage
        self.is_loaded = False
    
    def create_template(self) -> str:
        """Create a blank price catalog Excel template with sample data"""
        try:
            wb = Workbook()
            
            # ===== Sheet 1: Instructions =====
            ws_instructions = wb.active
            ws_instructions.title = "Instructions"
            
            instructions = [
                ["PRICE CATALOG - INSTRUCTIONS"],
                [""],
                ["This workbook defines the financial impact calculations for escalation analysis."],
                [""],
                ["SHEETS:"],
                ["1. Category Costs - Base costs per escalation category"],
                ["2. Keyword Patterns - Regex patterns for specific cost overrides"],
                ["3. Severity Multipliers - Cost multipliers based on severity level"],
                ["4. Origin Premiums - Additional percentage costs based on origin type"],
                [""],
                ["USAGE:"],
                ["- Material_Cost: Direct material/equipment costs in dollars"],
                ["- Labor_Hours: Estimated labor hours to resolve"],
                ["- Hourly_Rate: Labor cost per hour (default used if not specified)"],
                ["- Delay_Cost_Per_Hour: Business impact cost for each hour of delay"],
                [""],
                ["KEYWORD PATTERNS:"],
                ["- Use regex patterns to match specific issues"],
                ["- Example: '.*antenna.*damage.*' matches any antenna damage issue"],
                ["- Patterns are case-insensitive"],
                ["- Keyword costs OVERRIDE category costs when matched"],
                [""],
                ["FORMULA:"],
                ["Total_Impact = (Material_Cost + Labor_Hours × Hourly_Rate + Delay_Cost) × Severity_Mult × (1 + Origin_Premium)"],
            ]
            
            for row in instructions:
                ws_instructions.append(row)
            
            # Format header
            ws_instructions['A1'].font = Font(bold=True, size=14)
            ws_instructions.column_dimensions['A'].width = 80
            
            # ===== Sheet 2: Category Costs =====
            ws_category = wb.create_sheet("Category Costs")
            category_headers = ["Category", "Material_Cost", "Labor_Hours", "Hourly_Rate", "Delay_Cost_Per_Hour", "Notes"]
            ws_category.append(category_headers)
            
            # Sample category data based on ANCHORS
            category_data = [
                ["RF & Antenna Systems", 2500, 8, 150, 500, "Antenna repairs, RF optimization"],
                ["Transmission & Backhaul", 3000, 12, 175, 750, "Fiber, microwave, transport"],
                ["Power & Environment", 1500, 6, 125, 400, "Battery, rectifier, HVAC"],
                ["Site Access & Logistics", 500, 4, 100, 200, "Keys, permits, scheduling"],
                ["Contractor & Vendor Issues", 1000, 8, 125, 350, "Third-party coordination"],
                ["Configuration & Integration", 800, 10, 150, 450, "Software, parameters"],
                ["OSS/NMS & Systems", 600, 6, 175, 400, "Monitoring, alarms"],
                ["Process & Documentation", 200, 4, 100, 150, "MOPs, procedures"],
                ["Communication & Coordination", 100, 2, 100, 100, "Updates, notifications"],
                ["Weather & Natural Events", 5000, 16, 150, 1000, "Storm damage, disasters"],
                ["Third-Party & External", 2000, 8, 125, 500, "Utility, landlord issues"],
                ["Unclassified", 1000, 6, 125, 300, "Default for unknown"],
            ]
            
            for row in category_data:
                ws_category.append(row)
            
            # Format header row
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col in range(1, len(category_headers) + 1):
                cell = ws_category.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            # Adjust column widths
            ws_category.column_dimensions['A'].width = 30
            ws_category.column_dimensions['F'].width = 35
            
            # ===== Sheet 3: Keyword Patterns =====
            ws_keywords = wb.create_sheet("Keyword Patterns")
            keyword_headers = ["Keyword_Pattern", "Category_Override", "Material_Cost", "Labor_Hours", "Priority", "Notes"]
            ws_keywords.append(keyword_headers)
            
            # Sample keyword patterns
            keyword_data = [
                [".*antenna.*replace.*", "RF & Antenna Systems", 8000, 16, 1, "Full antenna replacement"],
                [".*fiber.*cut.*", "Transmission & Backhaul", 5000, 24, 1, "Emergency fiber repair"],
                [".*battery.*fail.*", "Power & Environment", 3000, 8, 2, "Battery bank replacement"],
                [".*generator.*", "Power & Environment", 2000, 12, 2, "Generator issues"],
                [".*microwave.*", "Transmission & Backhaul", 4000, 16, 2, "Microwave link problems"],
                [".*schedul.*delay.*", "Site Access & Logistics", 300, 2, 3, "Scheduling conflicts"],
                [".*permit.*", "Site Access & Logistics", 500, 4, 3, "Permit-related delays"],
                [".*software.*upgrade.*", "Configuration & Integration", 200, 8, 3, "Software updates"],
            ]
            
            for row in keyword_data:
                ws_keywords.append(row)
            
            # Format header row
            for col in range(1, len(keyword_headers) + 1):
                cell = ws_keywords.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            ws_keywords.column_dimensions['A'].width = 25
            ws_keywords.column_dimensions['B'].width = 25
            ws_keywords.column_dimensions['F'].width = 30
            
            # ===== Sheet 4: Severity Multipliers =====
            ws_severity = wb.create_sheet("Severity Multipliers")
            severity_headers = ["Severity_Level", "Cost_Multiplier", "Description"]
            ws_severity.append(severity_headers)
            
            severity_data = [
                ["Critical", 2.5, "Network down, major outage"],
                ["High", 1.75, "Significant degradation"],
                ["Medium", 1.25, "Moderate impact"],
                ["Low", 1.0, "Minor issue, no multiplier"],
            ]
            
            for row in severity_data:
                ws_severity.append(row)
            
            for col in range(1, len(severity_headers) + 1):
                cell = ws_severity.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            ws_severity.column_dimensions['A'].width = 20
            ws_severity.column_dimensions['C'].width = 35
            
            # ===== Sheet 5: Origin Premiums =====
            ws_origin = wb.create_sheet("Origin Premiums")
            origin_headers = ["Origin_Type", "Premium_Percentage", "Description"]
            ws_origin.append(origin_headers)
            
            origin_data = [
                ["Vendor", 0.15, "15% premium for vendor-caused issues"],
                ["Process", 0.05, "5% premium for internal process failures"],
                ["External", 0.20, "20% premium for external/uncontrollable factors"],
                ["Customer", 0.10, "10% premium for customer-initiated issues"],
                ["Technical", 0.0, "No premium for standard technical issues"],
            ]
            
            for row in origin_data:
                ws_origin.append(row)
            
            for col in range(1, len(origin_headers) + 1):
                cell = ws_origin.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            ws_origin.column_dimensions['A'].width = 20
            ws_origin.column_dimensions['C'].width = 45
            
            # Save workbook
            wb.save(self.catalog_path)
            logger.info(f"✓ Price catalog template created: {self.catalog_path}")
            return self.catalog_path
            
        except Exception as e:
            logger.error(f"Failed to create price catalog template: {e}")
            raise
    
    def load_catalog(self) -> bool:
        """Load pricing data from Excel catalog"""
        if not os.path.exists(self.catalog_path):
            logger.warning(f"Price catalog not found: {self.catalog_path}")
            logger.info("Creating template price catalog...")
            self.create_template()
        
        try:
            wb = load_workbook(self.catalog_path, data_only=True)
            
            # Load Category Costs
            if "Category Costs" in wb.sheetnames:
                ws = wb["Category Costs"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Category name exists
                        category = str(row[0]).strip()
                        self.category_costs[category] = {
                            'material_cost': float(row[1] or 0),
                            'labor_hours': float(row[2] or 0),
                            'hourly_rate': float(row[3] or DEFAULT_HOURLY_RATE),
                            'delay_cost_per_hour': float(row[4] or 0),
                        }
                logger.info(f"  → Loaded {len(self.category_costs)} category cost entries")
            
            # Load Keyword Patterns
            if "Keyword Patterns" in wb.sheetnames:
                ws = wb["Keyword Patterns"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Pattern exists
                        self.keyword_costs.append({
                            'pattern': str(row[0]).strip(),
                            'category_override': str(row[1] or '').strip(),
                            'material_cost': float(row[2] or 0),
                            'labor_hours': float(row[3] or 0),
                            'priority': int(row[4] or 99),
                        })
                # Sort by priority (lower = higher priority)
                self.keyword_costs.sort(key=lambda x: x['priority'])
                logger.info(f"  → Loaded {len(self.keyword_costs)} keyword pattern rules")
            
            # Load Severity Multipliers
            if "Severity Multipliers" in wb.sheetnames:
                ws = wb["Severity Multipliers"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        severity = str(row[0]).strip().lower()
                        self.severity_multipliers[severity] = float(row[1] or 1.0)
                logger.info(f"  → Loaded {len(self.severity_multipliers)} severity multipliers")
            
            # Load Origin Premiums
            if "Origin Premiums" in wb.sheetnames:
                ws = wb["Origin Premiums"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        origin = str(row[0]).strip().lower()
                        self.origin_premiums[origin] = float(row[1] or 0)
                logger.info(f"  → Loaded {len(self.origin_premiums)} origin premium rules")
            
            wb.close()
            self.is_loaded = True
            logger.info(f"✓ Price catalog loaded successfully")
            return True
            
        except Exception as e:
            logger.error(f"Failed to load price catalog: {e}")
            return False
    
    def _match_keyword_pattern(self, text: str) -> Optional[Dict]:
        """Check if text matches any keyword pattern, return first match by priority"""
        if not text:
            return None
        
        text_lower = text.lower()
        for rule in self.keyword_costs:
            try:
                if re.search(rule['pattern'], text_lower, re.IGNORECASE):
                    return rule
            except re.error:
                logger.warning(f"Invalid regex pattern: {rule['pattern']}")
        return None
    
    def calculate_financial_impact(
        self,
        category: str,
        severity: str = "Medium",
        origin: str = "Technical",
        description: str = "",
        delay_hours: float = 4.0
    ) -> Dict[str, float]:
        """
        Calculate total financial impact for an escalation.
        
        Returns dict with breakdown:
        - material_cost
        - labor_cost  
        - delay_cost
        - severity_multiplier
        - origin_premium
        - total_impact
        """
        if not self.is_loaded:
            self.load_catalog()
        
        # Check for keyword pattern match first (overrides category)
        keyword_match = self._match_keyword_pattern(description)
        
        if keyword_match:
            material_cost = keyword_match['material_cost']
            labor_hours = keyword_match['labor_hours']
            hourly_rate = DEFAULT_HOURLY_RATE
            delay_cost_per_hour = self.category_costs.get(
                keyword_match.get('category_override', category), {}
            ).get('delay_cost_per_hour', 300)
        else:
            # Use category-based costs
            cat_costs = self.category_costs.get(category, self.category_costs.get('Unclassified', {}))
            material_cost = cat_costs.get('material_cost', 1000)
            labor_hours = cat_costs.get('labor_hours', 6)
            hourly_rate = cat_costs.get('hourly_rate', DEFAULT_HOURLY_RATE)
            delay_cost_per_hour = cat_costs.get('delay_cost_per_hour', 300)
        
        # Calculate base costs
        labor_cost = labor_hours * hourly_rate
        delay_cost = delay_hours * delay_cost_per_hour
        base_cost = material_cost + labor_cost + delay_cost
        
        # Apply severity multiplier
        severity_mult = self.severity_multipliers.get(severity.lower(), 1.0)
        
        # Apply origin premium
        origin_premium = self.origin_premiums.get(origin.lower(), 0.0)
        
        # Calculate total
        total_impact = base_cost * severity_mult * (1 + origin_premium)
        
        return {
            'material_cost': round(material_cost, 2),
            'labor_cost': round(labor_cost, 2),
            'delay_cost': round(delay_cost, 2),
            'base_cost': round(base_cost, 2),
            'severity_multiplier': severity_mult,
            'origin_premium': origin_premium,
            'total_impact': round(total_impact, 2),
            'keyword_match': keyword_match['pattern'] if keyword_match else None,
        }
    
    def get_catalog_summary(self) -> str:
        """Get summary of loaded pricing data for AI context"""
        if not self.is_loaded:
            return "Price catalog not loaded."
        
        summary_lines = [
            f"PRICE CATALOG SUMMARY:",
            f"- {len(self.category_costs)} category cost entries",
            f"- {len(self.keyword_costs)} keyword pattern rules",
            f"- {len(self.severity_multipliers)} severity multipliers",
            f"- {len(self.origin_premiums)} origin premium rules",
        ]
        
        # Add top 3 most expensive categories
        if self.category_costs:
            sorted_cats = sorted(
                self.category_costs.items(),
                key=lambda x: x[1].get('material_cost', 0) + x[1].get('labor_hours', 0) * x[1].get('hourly_rate', 125),
                reverse=True
            )[:3]
            summary_lines.append("\nTop 3 highest-cost categories:")
            for cat, costs in sorted_cats:
                total = costs.get('material_cost', 0) + costs.get('labor_hours', 0) * costs.get('hourly_rate', 125)
                summary_lines.append(f"  - {cat}: ${total:,.0f} base cost")
        
        return "\n".join(summary_lines)


# Global price catalog instance
_price_catalog: Optional[PriceCatalog] = None

def get_price_catalog() -> PriceCatalog:
    """Get or create the price catalog singleton"""
    global _price_catalog
    if _price_catalog is None:
        _price_catalog = PriceCatalog()
    return _price_catalog


# ==========================================
# 3. AI CLASSIFICATION ENGINE
# ==========================================

# Cache for anchor centroids (computed once per session)
_anchor_centroids_cache: Dict[str, np.ndarray] = {}

def get_anchor_centroids(ai) -> Dict[str, np.ndarray]:
    """Get or compute anchor centroids with caching and feedback adjustment"""
    global _anchor_centroids_cache
    
    if _anchor_centroids_cache:
        logger.info("Using cached anchor centroids...")
        return _anchor_centroids_cache
    
    logger.info("Computing anchor centroids for categories (will be cached)...")
    
    # Batch all anchor phrases for efficiency
    all_phrases = []
    phrase_to_category = []
    for category, phrases in ANCHORS.items():
        for phrase in phrases:
            all_phrases.append(phrase)
            phrase_to_category.append(category)
    
    # Single batch call for all anchor phrases
    all_vecs = ai.get_embeddings_batch(all_phrases)
    
    # Group by category and compute centroids
    category_vecs: Dict[str, List[np.ndarray]] = {cat: [] for cat in ANCHORS.keys()}
    for phrase, category, vec in zip(all_phrases, phrase_to_category, all_vecs):
        category_vecs[category].append(vec)
    
    for category, vecs in category_vecs.items():
        _anchor_centroids_cache[category] = np.mean(vecs, axis=0)
    
    logger.info(f"Computed centroids for {len(_anchor_centroids_cache)} categories")
    
    # Apply feedback adjustments if available
    feedback_learner = get_feedback_learner()
    if feedback_learner.stats['loaded'] > 0:
        logger.info("Applying user feedback adjustments to centroids...")
        adjusted = feedback_learner.adjust_centroids(_anchor_centroids_cache)
        _anchor_centroids_cache.clear()
        _anchor_centroids_cache.update(adjusted)
        logger.info(f"Final category count: {len(_anchor_centroids_cache)}")
    
    return _anchor_centroids_cache

def classify_rows(df, ai):
    """
    Categorize rows using AI embeddings and anchor-based similarity.
    
    Includes consistency improvements:
    - Keyword-based pre-classification for known patterns
    - Tie-breaking for close similarity scores
    - Stable sorting of categories to ensure deterministic results
    """
    logger.info("[AI Engine] Categorizing rows using Embeddings...")
    df = df.copy()
    
    # Get anchor centroids (cached after first call)
    anchor_centroids = get_anchor_centroids(ai)
    
    # Define keyword patterns for deterministic pre-classification
    # These override embedding-based classification for known patterns
    KEYWORD_OVERRIDES = {
        "OSS/NMS & Systems": [
            r'\bnest(?:ing|ed)\b', r'\bnsi\b', r'\bsi\s+nest', r'\bcell\s*planning\b',
            r'\bpci\s*conflict\b', r'\bprovision', r'\binventory\b', r'\boss\b', r'\bnms\b'
        ],
        "RF & Antenna Issues": [
            r'\bantenna\b', r'\bvswr\b', r'\brru\b', r'\bradio\b', r'\bsector\b',
            r'\bcell\s*down\b', r'\boutage\b', r'\brf\b', r'\bbaseband\b'
        ],
        "Transmission & Backhaul": [
            r'\bfiber\b', r'\bmicrowave\b', r'\btransmission\b', r'\bbackhaul\b',
            r'\bcircuit\b', r'\blatency\b', r'\bpacket\s*loss\b'
        ],
        "Power & Environment": [
            r'\bpower\b', r'\bbattery\b', r'\bgenerator\b', r'\brectifier\b',
            r'\bac\s*fail', r'\bcooling\b', r'\btemperature\b'
        ],
        "Site Access & Logistics": [
            r'\baccess\b.*\bden', r'\bkey\b', r'\bgate\b', r'\blandlord\b',
            r'\bpermit\b', r'\bescort\b', r'\binaccessible\b'
        ],
    }
    
    def keyword_classify(text):
        """Check if text matches any keyword patterns for pre-classification"""
        if not text:
            return None
        text_lower = str(text).lower()
        
        for category, patterns in KEYWORD_OVERRIDES.items():
            for pattern in patterns:
                if re.search(pattern, text_lower):
                    return category
        return None
    
    # Batch compute embeddings for all texts
    logger.info(f"Computing embeddings for {len(df)} rows...")
    all_texts = df['Combined_Text'].tolist()
    all_vecs = ai.get_embeddings_batch(all_texts)
    
    cats = []
    scores = []
    
    # Sort categories alphabetically for deterministic tie-breaking
    sorted_categories = sorted(anchor_centroids.keys())
    
    # Classify each row
    for idx, vec in enumerate(tqdm(all_vecs, desc="   > Classifying")):
        text = all_texts[idx]
        
        # Step 1: Try keyword-based classification first
        keyword_cat = keyword_classify(text)
        if keyword_cat:
            # Verify with embedding (must be at least 0.3 similar)
            sim = cosine_similarity(vec.reshape(1, -1), 
                                   anchor_centroids[keyword_cat].reshape(1, -1))[0][0]
            if sim >= 0.3:
                cats.append(keyword_cat)
                scores.append(sim)
                continue
        
        # Step 2: Embedding-based classification
        category_scores = []
        for cat in sorted_categories:
            anchor_vec = anchor_centroids[cat]
            sim = cosine_similarity(vec.reshape(1, -1), anchor_vec.reshape(1, -1))[0][0]
            category_scores.append((cat, sim))
        
        # Sort by score descending, then by category name for tie-breaking
        category_scores.sort(key=lambda x: (-x[1], x[0]))
        
        best_cat, best_score = category_scores[0]
        second_cat, second_score = category_scores[1] if len(category_scores) > 1 else ("", 0.0)
        
        # Check for close scores (within 0.03 = essentially a tie)
        if second_cat and second_score > 0 and (best_score - second_score) < 0.03:
            # Use keyword overlap as tie-breaker
            text_lower = str(text).lower()
            best_keywords = sum(1 for kw in ANCHORS.get(best_cat, []) if kw in text_lower)
            second_keywords = sum(1 for kw in ANCHORS.get(second_cat, []) if kw in text_lower)
            
            if second_keywords > best_keywords:
                best_cat, best_score = second_cat, second_score
        
        # Apply minimum confidence threshold
        if best_score < MIN_CLASSIFICATION_CONFIDENCE:
            best_cat = "Unclassified"
        
        cats.append(best_cat)
        scores.append(best_score)
    
    df['AI_Category'] = cats
    df['AI_Confidence'] = scores
    
    # Log category distribution
    logger.info(f"Classification complete. Distribution: {df['AI_Category'].value_counts().to_dict()}")
    
    return df

# ==========================================
# 3. SCORING ENGINE
# ==========================================
def calculate_strategic_friction(df):
    logger.info("[Strategic Engine] Applying Multi-Variable Risk Scoring...")
    df = df.copy()
    
    # Validate columns
    validate_columns(df, REQUIRED_COLUMNS)
   
    # Normalize with safe column access
    df['Severity_Norm'] = df[COL_SEVERITY].astype(str).str.title().str.strip() if COL_SEVERITY in df.columns else 'Default'
    df['Type_Norm'] = df[COL_TYPE].astype(str).str.title().str.strip() if COL_TYPE in df.columns else ''
    df['Origin_Norm'] = df[COL_ORIGIN].astype(str).str.title().str.strip() if COL_ORIGIN in df.columns else ''
    df['Impact_Norm'] = df[COL_IMPACT].fillna('None').astype(str).str.title().str.strip() if COL_IMPACT in df.columns else 'None'

    # Load price catalog for financial impact
    price_catalog = get_price_catalog()
    if not price_catalog.is_loaded:
        price_catalog.load_catalog()

    def get_score(row):
        # 1. Base Score (Severity)
        base = WEIGHTS['BASE_SEVERITY'].get(row['Severity_Norm'], 5)
       
        # 2. Type Multiplier (Escalation > Concern)
        m_type = 1.0
        if 'Escalation' in row['Type_Norm']: m_type = WEIGHTS['TYPE_MULTIPLIER']['Escalations']
        elif 'Lesson' in row['Type_Norm']: m_type = 0.0 # Lessons are not risks
       
        # 3. Origin Multiplier (External > Internal)
        m_origin = WEIGHTS['ORIGIN_MULTIPLIER'].get(row['Origin_Norm'], 1.0)
       
        # 4. Impact Multiplier
        m_impact = 1.0
        if 'High' in row['Impact_Norm']: m_impact = WEIGHTS['IMPACT_MULTIPLIER']['High']
       
        return base * m_type * m_origin * m_impact

    def get_financial_impact(row):
        """Calculate financial impact using price catalog"""
        category = row.get('AI_Category', 'Unclassified')
        severity = row['Severity_Norm']
        origin = row['Origin_Norm']
        description = str(row.get(COL_SUMMARY, ''))
        
        impact = price_catalog.calculate_financial_impact(
            category=category,
            severity=severity,
            origin=origin,
            description=description,
            delay_hours=4.0  # Default assumption
        )
        return impact['total_impact']

    df['Strategic_Friction_Score'] = df.apply(get_score, axis=1)
    
    # Add financial impact column
    df['Financial_Impact'] = df.apply(get_financial_impact, axis=1)
    logger.info(f"  → Total estimated financial impact: ${df['Financial_Impact'].sum():,.2f}")
    
    # =========================================
    # ADDITIONAL METRICS FOR SCORED DATA
    # =========================================
    
    # 1. Risk Tier (High/Medium/Low based on friction score)
    def get_risk_tier(score):
        if score >= 150:
            return "Critical"
        elif score >= 75:
            return "High"
        elif score >= 25:
            return "Medium"
        else:
            return "Low"
    
    df['Risk_Tier'] = df['Strategic_Friction_Score'].apply(get_risk_tier)
    
    # 2. Engineer Accountability (for escalations/concerns)
    if COL_ENGINEER in df.columns:
        df['Engineer'] = df[COL_ENGINEER].fillna('Unknown').astype(str).str.strip()
        
        # Count issues per engineer for "repeat offender" detection
        engineer_counts = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])].groupby('Engineer').size()
        df['Engineer_Issue_Count'] = df['Engineer'].map(engineer_counts).fillna(0).astype(int)
        
        # Calculate engineer's total friction contribution
        engineer_friction = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])].groupby('Engineer')['Strategic_Friction_Score'].sum()
        df['Engineer_Total_Friction'] = df['Engineer'].map(engineer_friction).fillna(0)
        
        # Flag repeat offenders (3+ issues)
        df['Engineer_Flag'] = df['Engineer_Issue_Count'].apply(
            lambda x: '🔴 Repeat Offender' if x >= 5 else ('🟡 Multiple Issues' if x >= 3 else '')
        )
        
        logger.info(f"  → Engineer accountability tracked for {df['Engineer'].nunique()} unique engineers")
    
    # 3. Days Since Issue (if datetime available)
    if COL_DATETIME in df.columns:
        df['Issue_Date'] = pd.to_datetime(df[COL_DATETIME], errors='coerce')
        df['Days_Since_Issue'] = (pd.Timestamp.now() - df['Issue_Date']).dt.days
        df['Days_Since_Issue'] = df['Days_Since_Issue'].fillna(-1).astype(int)
        
        # Aging flag
        df['Aging_Status'] = df['Days_Since_Issue'].apply(
            lambda x: '🔴 >30 days' if x > 30 else ('🟡 >14 days' if x > 14 else ('🟢 Recent' if x >= 0 else 'Unknown'))
        )
    
    # 4. Is Human Error flag (based on origin being internal + escalation/concern type)
    df['Is_Human_Error'] = (
        (df['Origin_Norm'] == 'Internal') & 
        (df['Type_Norm'].isin(['Escalations', 'Concerns']))
    ).map({True: 'Yes', False: 'No'})
    
    # 4a. Root Cause Classification (from tickets_data_root_cause)
    df['Root_Cause_Category'] = 'Unclassified'
    df['Root_Cause_Original'] = ''
    
    if COL_ROOT_CAUSE in df.columns:
        df['Root_Cause_Original'] = df[COL_ROOT_CAUSE].fillna('').astype(str).str.strip()
        
        # Classify root cause based on keywords
        def classify_root_cause(root_cause_text):
            if pd.isna(root_cause_text) or not root_cause_text:
                return 'Unclassified'
            text_lower = str(root_cause_text).lower()
            
            for category, keywords in ROOT_CAUSE_CATEGORIES.items():
                for keyword in keywords:
                    if keyword in text_lower:
                        return category
            return 'Other'
        
        df['Root_Cause_Category'] = df['Root_Cause_Original'].apply(classify_root_cause)
        
        # Update Is_Human_Error if root cause explicitly says Human Error
        human_error_mask = df['Root_Cause_Category'] == 'Human Error'
        df.loc[human_error_mask, 'Is_Human_Error'] = 'Yes'
        
        # Update for External Party (Non Amdocs = External origin)
        external_mask = df['Root_Cause_Category'] == 'External Party'
        df.loc[external_mask, 'Is_Human_Error'] = 'External'
        
        logger.info(f"  → Root cause classified: {df['Root_Cause_Category'].value_counts().to_dict()}")
    
    # 4b. PM Recurrence Risk Assessment
    df['PM_Recurrence_Risk'] = 'Unknown'
    df['PM_Recurrence_Risk_Norm'] = 'Unknown'
    
    if COL_RECURRENCE_RISK in df.columns:
        df['PM_Recurrence_Risk'] = df[COL_RECURRENCE_RISK].fillna('Unknown').astype(str).str.strip()
        
        # Normalize recurrence risk
        def normalize_recurrence_risk(risk_text):
            if pd.isna(risk_text) or not risk_text:
                return 'Unknown'
            text_lower = str(risk_text).lower().strip()
            
            if text_lower in ['high', 'yes', 'likely', 'probable', 'very high']:
                return 'High'
            elif text_lower in ['medium', 'moderate', 'possible', 'maybe']:
                return 'Medium'
            elif text_lower in ['low', 'no', 'unlikely', 'none', 'very low', 'minimal']:
                return 'Low'
            else:
                return 'Unknown'
        
        df['PM_Recurrence_Risk_Norm'] = df['PM_Recurrence_Risk'].apply(normalize_recurrence_risk)
        logger.info(f"  → PM Recurrence risk: {df['PM_Recurrence_Risk_Norm'].value_counts().to_dict()}")
    
    # 5. Priority Score (composite for sorting - higher = more urgent)
    # Combines friction, recency, and human error factor
    df['Priority_Score'] = df['Strategic_Friction_Score']
    if 'Days_Since_Issue' in df.columns:
        # Recent issues get boosted priority
        df['Priority_Score'] = df['Priority_Score'] * (1 + (30 - df['Days_Since_Issue'].clip(0, 30)) / 100)
    
    # 6. Action Required flag
    def get_action_required(row):
        actions = []
        if row['Risk_Tier'] in ['Critical', 'High']:
            actions.append('Immediate Review')
        if row.get('Is_Human_Error') == 'Yes':
            actions.append('Training Review')
        if row.get('Engineer_Flag') and 'Repeat' in str(row.get('Engineer_Flag', '')):
            actions.append('Performance Discussion')
        if row.get('Learning_Status') == 'Confirmed Repeat':
            actions.append('Process Fix Required')
        return ' | '.join(actions) if actions else 'Monitor'
    
    df['Action_Required'] = df.apply(get_action_required, axis=1)
    
    logger.info(f"  → Added {len([c for c in df.columns if c in ['Risk_Tier', 'Engineer', 'Is_Human_Error', 'Priority_Score', 'Action_Required']])} new metric columns")
    
    return df

# ==========================================
# 4. RECIDIVISM AUDIT (Enhanced Semantic Matching + Engineer Tracking)
# ==========================================

def analyze_engineer_patterns(df):
    """
    Analyze engineer-based recidivism patterns:
    1. Track engineers with multiple issues (across any category)
    2. Cross-reference with lessons learned
    3. Calculate learning gap (issues caused vs lessons logged)
    
    Returns: DataFrame with engineer pattern columns added
    """
    logger.info("[Engineer Analysis] Analyzing engineer patterns...")
    df = df.copy()
    
    # Initialize new columns
    df['Engineer_Issue_History'] = ''
    df['Engineer_Lessons_Logged'] = 0
    df['Engineer_Learning_Gap'] = 0
    df['Engineer_Is_Repeat_Offender'] = False
    df['Engineer_Learning_Score'] = 1.0  # 1.0 = learning from mistakes, 0 = not learning
    
    # Check if engineer column exists and has data
    if COL_ENGINEER not in df.columns:
        logger.warning(f"Column '{COL_ENGINEER}' not found. Engineer pattern analysis skipped.")
        return df
    
    # Clean engineer names
    df['_engineer_clean'] = df[COL_ENGINEER].fillna('Unknown').astype(str).str.strip().str.lower()
    
    # Identify lessons vs incidents
    type_col = df[COL_TYPE].astype(str) if COL_TYPE in df.columns else pd.Series([''] * len(df))
    is_lesson = type_col.str.contains('Lesson', case=False, na=False)
    
    # Also check for lesson title/status columns
    has_lesson_title = False
    if COL_LESSON_TITLE in df.columns:
        has_lesson_title = df[COL_LESSON_TITLE].notna() & (df[COL_LESSON_TITLE].astype(str).str.strip() != '')
    
    has_lesson_status = False
    if COL_LESSON_STATUS in df.columns:
        has_lesson_status = df[COL_LESSON_STATUS].notna() & (df[COL_LESSON_STATUS].astype(str).str.strip() != '')
    
    # Mark rows that are or have lessons
    df['_has_lesson'] = is_lesson | has_lesson_title | has_lesson_status
    
    # Build engineer profiles
    engineer_profiles = {}
    
    for idx, row in df.iterrows():
        engineer = row['_engineer_clean']
        if engineer == 'unknown' or engineer == '':
            continue
        
        if engineer not in engineer_profiles:
            engineer_profiles[engineer] = {
                'issues': [],
                'lessons': [],
                'categories': set(),
                'total_friction': 0.0
            }
        
        category = str(row.get('AI_Category', row.get(COL_CATEGORY, 'Unknown')))
        friction = float(row.get('Strategic_Friction_Score', 0))
        
        if row['_has_lesson']:
            engineer_profiles[engineer]['lessons'].append({
                'idx': idx,
                'category': category,
                'title': row.get(COL_LESSON_TITLE, '')
            })
        else:
            # This is an issue/incident
            engineer_profiles[engineer]['issues'].append({
                'idx': idx,
                'category': category,
                'friction': friction
            })
            engineer_profiles[engineer]['categories'].add(category)
            engineer_profiles[engineer]['total_friction'] += friction
    
    # Log engineer summary
    engineers_with_issues = {k: v for k, v in engineer_profiles.items() if len(v['issues']) > 0}
    logger.info(f"  Found {len(engineers_with_issues)} engineers with issues")
    
    # Identify repeat offenders and calculate metrics
    repeat_offenders = []
    
    for engineer, profile in engineer_profiles.items():
        issue_count = len(profile['issues'])
        lesson_count = len(profile['lessons'])
        
        if issue_count == 0:
            continue
        
        # Calculate learning gap and score
        learning_gap = max(0, issue_count - lesson_count)
        learning_score = lesson_count / issue_count if issue_count > 0 else 1.0
        learning_score = min(1.0, learning_score)  # Cap at 1.0
        
        # Determine if repeat offender
        is_repeat_offender = issue_count >= ENGINEER_REPEAT_THRESHOLD
        
        # Build issue history string
        issue_history = ', '.join(sorted(profile['categories']))[:200]  # Truncate for cell size
        
        if is_repeat_offender:
            repeat_offenders.append({
                'engineer': engineer,
                'issue_count': issue_count,
                'lesson_count': lesson_count,
                'learning_gap': learning_gap,
                'categories': list(profile['categories'])
            })
        
        # Update all rows for this engineer
        for issue in profile['issues']:
            idx = issue['idx']
            df.at[idx, 'Engineer_Issue_History'] = issue_history
            df.at[idx, 'Engineer_Lessons_Logged'] = lesson_count
            df.at[idx, 'Engineer_Learning_Gap'] = learning_gap
            df.at[idx, 'Engineer_Is_Repeat_Offender'] = is_repeat_offender
            df.at[idx, 'Engineer_Learning_Score'] = round(learning_score, 2)
        
        # Also update lesson rows for this engineer
        for lesson in profile['lessons']:
            idx = lesson['idx']
            df.at[idx, 'Engineer_Issue_History'] = issue_history
            df.at[idx, 'Engineer_Lessons_Logged'] = lesson_count
            df.at[idx, 'Engineer_Learning_Gap'] = learning_gap
            df.at[idx, 'Engineer_Is_Repeat_Offender'] = is_repeat_offender
            df.at[idx, 'Engineer_Learning_Score'] = round(learning_score, 2)
    
    # Log repeat offenders
    if repeat_offenders:
        logger.warning(f"  ⚠ Found {len(repeat_offenders)} repeat offender engineer(s):")
        for offender in sorted(repeat_offenders, key=lambda x: x['issue_count'], reverse=True)[:5]:
            logger.warning(f"    - {offender['engineer'].title()}: {offender['issue_count']} issues, "
                          f"{offender['lesson_count']} lessons, gap: {offender['learning_gap']}")
    else:
        logger.info("  ✓ No repeat offender engineers identified")
    
    # Clean up temporary columns
    df.drop(columns=['_engineer_clean', '_has_lesson'], inplace=True, errors='ignore')
    
    return df


def analyze_lob_patterns(df):
    """
    McKinsey-Style Line of Business (LOB) Strategic Analysis:
    
    Analyzes organizational performance at the LOB level with key metrics:
    1. LOB Risk Profile - Friction concentration by business unit
    2. LOB Operational Efficiency - Issue density and resolution patterns
    3. LOB Learning Culture - Lessons logged vs issues ratio
    4. LOB Talent Quality - Engineer performance within each LOB
    5. Cross-LOB Comparisons - Benchmarking and best practices identification
    
    Returns: DataFrame with LOB pattern columns added
    """
    logger.info("[LOB Analysis] McKinsey-style organizational pattern analysis...")
    df = df.copy()
    
    # Initialize new LOB columns
    df['LOB'] = 'Unknown'
    df['LOB_Issue_Count'] = 0
    df['LOB_Total_Friction'] = 0.0
    df['LOB_Avg_Friction'] = 0.0
    df['LOB_Engineer_Count'] = 0
    df['LOB_Lessons_Logged'] = 0
    df['LOB_Learning_Rate'] = 0.0  # Lessons per issue
    df['LOB_Risk_Tier'] = 'Low'
    df['LOB_Efficiency_Score'] = 0.0  # Inverse of friction per engineer
    df['LOB_Repeat_Offender_Rate'] = 0.0  # % of engineers who are repeat offenders
    
    # Check if LOB column exists
    if COL_LOB not in df.columns:
        logger.warning(f"Column '{COL_LOB}' not found. LOB pattern analysis skipped.")
        return df
    
    # Clean LOB names
    df['LOB'] = df[COL_LOB].fillna('Unknown').astype(str).str.strip()
    df.loc[df['LOB'] == '', 'LOB'] = 'Unknown'
    
    # Identify lessons vs incidents
    type_col = df[COL_TYPE].astype(str) if COL_TYPE in df.columns else pd.Series([''] * len(df))
    is_lesson = type_col.str.contains('Lesson', case=False, na=False)
    
    # Also check for lesson columns
    has_lesson_title = False
    if COL_LESSON_TITLE in df.columns:
        has_lesson_title = df[COL_LESSON_TITLE].notna() & (df[COL_LESSON_TITLE].astype(str).str.strip() != '')
    
    df['_is_lesson'] = is_lesson | has_lesson_title
    
    # Build LOB profiles
    lob_profiles = {}
    
    for idx, row in df.iterrows():
        lob = row['LOB']
        if lob == 'Unknown':
            continue
        
        if lob not in lob_profiles:
            lob_profiles[lob] = {
                'issues': [],
                'lessons': [],
                'engineers': set(),
                'repeat_offender_engineers': set(),
                'categories': {},
                'total_friction': 0.0,
                'financial_impact': 0.0
            }
        
        category = str(row.get('AI_Category', row.get(COL_CATEGORY, 'Unknown')))
        friction = float(row.get('Strategic_Friction_Score', 0))
        engineer = str(row.get('Engineer', row.get(COL_ENGINEER, ''))).strip().lower()
        is_repeat = row.get('Engineer_Is_Repeat_Offender', False)
        financial = float(row.get('Financial_Impact', 0))
        
        if row['_is_lesson']:
            lob_profiles[lob]['lessons'].append(idx)
        else:
            lob_profiles[lob]['issues'].append(idx)
            lob_profiles[lob]['total_friction'] += friction
            lob_profiles[lob]['financial_impact'] += financial
            
            # Track categories within LOB
            if category not in lob_profiles[lob]['categories']:
                lob_profiles[lob]['categories'][category] = 0
            lob_profiles[lob]['categories'][category] += 1
        
        # Track engineers
        if engineer:
            lob_profiles[lob]['engineers'].add(engineer)
            if is_repeat:
                lob_profiles[lob]['repeat_offender_engineers'].add(engineer)
    
    # Calculate LOB metrics and rankings
    lob_rankings = []
    
    for lob, profile in lob_profiles.items():
        issue_count = len(profile['issues'])
        lesson_count = len(profile['lessons'])
        engineer_count = len(profile['engineers'])
        repeat_count = len(profile['repeat_offender_engineers'])
        
        if issue_count == 0:
            continue
        
        # Calculate metrics
        avg_friction = profile['total_friction'] / issue_count if issue_count > 0 else 0
        learning_rate = lesson_count / issue_count if issue_count > 0 else 0
        repeat_offender_rate = repeat_count / engineer_count if engineer_count > 0 else 0
        
        # Efficiency Score: Lower friction per engineer = higher efficiency
        # Scale: 100 = best, 0 = worst
        friction_per_engineer = profile['total_friction'] / engineer_count if engineer_count > 0 else profile['total_friction']
        # Normalize (assuming max reasonable friction per engineer is 5000)
        efficiency_score = max(0, min(100, 100 - (friction_per_engineer / 50)))
        
        # Risk Tier based on multiple factors
        risk_score = (
            (issue_count / 10) * 0.3 +  # Issue volume (max 10)
            (avg_friction / 100) * 0.3 +  # Average friction (max 100)
            (1 - learning_rate) * 0.2 +  # Learning gap
            repeat_offender_rate * 0.2  # Repeat offender prevalence
        )
        
        if risk_score >= 0.7:
            risk_tier = 'Critical'
        elif risk_score >= 0.5:
            risk_tier = 'High'
        elif risk_score >= 0.3:
            risk_tier = 'Medium'
        else:
            risk_tier = 'Low'
        
        # Top issue categories for this LOB
        top_categories = sorted(profile['categories'].items(), key=lambda x: x[1], reverse=True)[:3]
        top_cats_str = ', '.join([f"{cat}({cnt})" for cat, cnt in top_categories])
        
        lob_rankings.append({
            'lob': lob,
            'issue_count': issue_count,
            'lesson_count': lesson_count,
            'engineer_count': engineer_count,
            'total_friction': profile['total_friction'],
            'avg_friction': avg_friction,
            'financial_impact': profile['financial_impact'],
            'learning_rate': learning_rate,
            'efficiency_score': efficiency_score,
            'repeat_offender_rate': repeat_offender_rate,
            'risk_tier': risk_tier,
            'top_categories': top_cats_str
        })
        
        # Update dataframe for all rows in this LOB
        for idx in profile['issues'] + profile['lessons']:
            df.at[idx, 'LOB_Issue_Count'] = issue_count
            df.at[idx, 'LOB_Total_Friction'] = round(profile['total_friction'], 2)
            df.at[idx, 'LOB_Avg_Friction'] = round(avg_friction, 2)
            df.at[idx, 'LOB_Engineer_Count'] = engineer_count
            df.at[idx, 'LOB_Lessons_Logged'] = lesson_count
            df.at[idx, 'LOB_Learning_Rate'] = round(learning_rate, 2)
            df.at[idx, 'LOB_Risk_Tier'] = risk_tier
            df.at[idx, 'LOB_Efficiency_Score'] = round(efficiency_score, 1)
            df.at[idx, 'LOB_Repeat_Offender_Rate'] = round(repeat_offender_rate * 100, 1)
    
    # Log LOB summary (McKinsey executive style)
    logger.info(f"  Found {len(lob_profiles)} Lines of Business")
    
    # Sort by friction for executive summary
    lob_rankings.sort(key=lambda x: x['total_friction'], reverse=True)
    
    critical_lobs = [l for l in lob_rankings if l['risk_tier'] == 'Critical']
    high_risk_lobs = [l for l in lob_rankings if l['risk_tier'] == 'High']
    
    if critical_lobs:
        logger.warning(f"  🔴 CRITICAL RISK LOBs ({len(critical_lobs)}):")
        for lob in critical_lobs[:3]:
            logger.warning(f"     - {lob['lob']}: {lob['issue_count']} issues, "
                          f"${lob['financial_impact']:,.0f} impact, "
                          f"{lob['repeat_offender_rate']*100:.0f}% repeat offenders")
    
    if high_risk_lobs:
        logger.info(f"  🟠 HIGH RISK LOBs ({len(high_risk_lobs)}): {', '.join([l['lob'] for l in high_risk_lobs[:5]])}")
    
    # Identify best performing LOB (for benchmarking)
    best_lobs = sorted([l for l in lob_rankings if l['issue_count'] >= 3], 
                       key=lambda x: x['efficiency_score'], reverse=True)
    if best_lobs:
        best = best_lobs[0]
        logger.info(f"  ✅ BEST PRACTICE LOB: {best['lob']} "
                   f"(Efficiency: {best['efficiency_score']:.0f}/100, "
                   f"Learning Rate: {best['learning_rate']:.0%})")
    
    # Clean up
    df.drop(columns=['_is_lesson'], inplace=True, errors='ignore')
    
    # Store LOB rankings for use in reporting
    df.attrs['lob_rankings'] = lob_rankings
    
    return df

def extract_keywords(text):
    """Extract meaningful keywords from text for overlap comparison"""
    if pd.isna(text) or not text:
        return set()
    # Remove common stop words and extract key terms
    stop_words = {'the', 'a', 'an', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
                  'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could', 
                  'should', 'may', 'might', 'must', 'shall', 'can', 'need', 'dare',
                  'to', 'of', 'in', 'for', 'on', 'with', 'at', 'by', 'from', 'as',
                  'into', 'through', 'during', 'before', 'after', 'above', 'below',
                  'between', 'under', 'again', 'further', 'then', 'once', 'here',
                  'there', 'when', 'where', 'why', 'how', 'all', 'each', 'few', 'more',
                  'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own',
                  'same', 'so', 'than', 'too', 'very', 'just', 'and', 'but', 'if', 'or',
                  'because', 'until', 'while', 'this', 'that', 'these', 'those', 'it'}
    
    # Tokenize and filter
    words = re.findall(r'\b[a-zA-Z]{3,}\b', str(text).lower())
    keywords = {w for w in words if w not in stop_words}
    return keywords

def calculate_keyword_overlap(text1, text2):
    """Calculate Jaccard similarity of keywords between two texts"""
    kw1 = extract_keywords(text1)
    kw2 = extract_keywords(text2)
    
    if not kw1 or not kw2:
        return 0.0
    
    intersection = len(kw1 & kw2)
    union = len(kw1 | kw2)
    
    return intersection / union if union > 0 else 0.0

def enrich_text_for_embedding(text, category=None):
    """Expand text with contextual information for better embedding"""
    if pd.isna(text) or not text:
        return ""
    
    enriched = str(text)
    
    # Add category context if available
    if category and not pd.isna(category):
        enriched = f"{category}: {enriched}"
    
    # Expand common telecom abbreviations
    expansions = {
        'rru': 'remote radio unit equipment',
        'bbu': 'baseband unit equipment', 
        'oss': 'operations support system',
        'nms': 'network management system',
        'ran': 'radio access network',
        'lte': 'long term evolution cellular',
        'nr': 'new radio 5g',
        'mw': 'microwave transmission',
        'tx': 'transmission',
        'rx': 'receive reception',
        'rf': 'radio frequency',
        'vswr': 'voltage standing wave ratio antenna',
        'gps': 'global positioning system synchronization',
        'ptp': 'precision time protocol synchronization'
    }
    
    text_lower = enriched.lower()
    for abbrev, expansion in expansions.items():
        if abbrev in text_lower:
            enriched = f"{enriched} ({expansion})"
    
    return enriched

def audit_learning(df, ai):
    """
    Enhanced recidivism detection using multi-strategy semantic matching:
    1. Keyword overlap detection (fast, catches obvious duplicates)
    2. Embedding similarity (semantic understanding)
    3. Combined scoring with tiered confidence levels
    """
    logger.info("[Audit Engine] Enhanced recidivism detection starting...")
    df = df.copy()
    df['Learning_Status'] = "New Issue"
    df['Recidivism_Score'] = 0.0
    df['Matched_Lesson_ID'] = None
   
    # Separate Lessons from Incidents
    type_col = df[COL_TYPE].astype(str) if COL_TYPE in df.columns else pd.Series([''] * len(df))
    lessons = df[type_col.str.contains('Lesson', case=False, na=False)]
    incidents = df[~df.index.isin(lessons.index)]
   
    if lessons.empty: 
        logger.info("No lessons found in dataset, skipping recidivism check.")
        return df
    
    logger.info(f"Found {len(lessons)} lessons and {len(incidents)} incidents to analyze.")

    # Enrich lesson texts for better embedding
    logger.info("Enriching lesson texts for semantic analysis...")
    lesson_enriched_texts = []
    for _, l_row in lessons.iterrows():
        enriched = enrich_text_for_embedding(
            l_row.get('Combined_Text', ''),
            l_row.get(COL_CATEGORY, None)
        )
        lesson_enriched_texts.append(enriched)
    
    # Pre-calculate lesson vectors using batch embedding
    logger.info(f"Computing embeddings for {len(lessons)} lessons...")
    lesson_vecs = ai.get_embeddings_batch(lesson_enriched_texts)
    
    # Build lesson data structures
    lesson_data_list = []
    for i, (lesson_idx, l_row) in enumerate(lessons.iterrows()):
        lesson_data_list.append({
            'id': l_row.get('id', f'L{lesson_idx}'),
            'idx': lesson_idx,
            'vec': lesson_vecs[i],
            'text': l_row.get('Combined_Text', ''),
            'enriched_text': lesson_enriched_texts[i],
            'date': pd.to_datetime(l_row.get(COL_DATETIME), errors='coerce'),
            'keywords': extract_keywords(l_row.get('Combined_Text', ''))
        })
    
    # Enrich incident texts
    logger.info("Enriching incident texts for semantic analysis...")
    incident_enriched_texts = []
    for _, row in incidents.iterrows():
        enriched = enrich_text_for_embedding(
            row.get('Combined_Text', ''),
            row.get(COL_CATEGORY, None)
        )
        incident_enriched_texts.append(enriched)

    # Pre-calculate incident vectors using batch embedding
    logger.info(f"Computing embeddings for {len(incidents)} incidents...")
    incident_vecs = ai.get_embeddings_batch(incident_enriched_texts)
    
    # Build incident data with vectors
    incident_data_list = []
    for i, (idx, row) in enumerate(incidents.iterrows()):
        incident_data_list.append({
            'idx': idx,
            'vec': incident_vecs[i],
            'text': row.get('Combined_Text', ''),
            'enriched_text': incident_enriched_texts[i],
            'date': pd.to_datetime(row.get(COL_DATETIME), errors='coerce'),
            'keywords': extract_keywords(row.get('Combined_Text', ''))
        })

    # Statistics tracking
    matches_found = 0
    high_confidence = 0
    medium_confidence = 0
    
    # Check Incidents against Lessons with multi-strategy matching
    for inc in tqdm(incident_data_list, desc="   > Cross-referencing incidents"):
        best_match = None
        best_combined_score = 0
        
        for l_data in lesson_data_list:
            # Skip if incident is before or same time as lesson (can't repeat what hasn't been learned)
            if pd.notna(inc['date']) and pd.notna(l_data['date']):
                if inc['date'] <= l_data['date']:
                    continue
            
            # Strategy 1: Keyword overlap (fast check)
            keyword_score = calculate_keyword_overlap(inc['text'], l_data['text'])
            
            # Strategy 2: Embedding similarity
            embedding_score = 0.0
            if np.any(inc['vec'] != 0) and np.any(l_data['vec'] != 0):
                embedding_score = cosine_similarity(
                    inc['vec'].reshape(1, -1), 
                    l_data['vec'].reshape(1, -1)
                )[0][0]
            
            # Combined score: weighted average (embeddings weighted higher)
            combined_score = (embedding_score * 0.7) + (keyword_score * 0.3)
            
            if combined_score > best_combined_score:
                best_combined_score = combined_score
                best_match = {
                    'lesson_id': l_data['id'],
                    'embedding_score': embedding_score,
                    'keyword_score': keyword_score,
                    'combined_score': combined_score
                }
        
        # Apply tiered matching
        if best_match:
            df.at[inc['idx'], 'Recidivism_Score'] = best_match['combined_score']
            
            if best_match['combined_score'] >= SIMILARITY_THRESHOLD_HIGH:
                # High confidence match
                df.at[inc['idx'], 'Learning_Status'] = f"REPEAT OFFENSE (Lesson #{best_match['lesson_id']})"
                df.at[inc['idx'], 'Matched_Lesson_ID'] = best_match['lesson_id']
                df.at[inc['idx'], 'Strategic_Friction_Score'] *= RECIDIVISM_PENALTY_HIGH
                matches_found += 1
                high_confidence += 1
                
            elif best_match['combined_score'] >= SIMILARITY_THRESHOLD_MEDIUM:
                # Medium confidence - flag for review
                df.at[inc['idx'], 'Learning_Status'] = f"POSSIBLE REPEAT (Lesson #{best_match['lesson_id']}, {best_match['combined_score']:.0%} match)"
                df.at[inc['idx'], 'Matched_Lesson_ID'] = best_match['lesson_id']
                df.at[inc['idx'], 'Strategic_Friction_Score'] *= RECIDIVISM_PENALTY_MEDIUM
                matches_found += 1
                medium_confidence += 1
    
    # Log results
    logger.info(f"Recidivism analysis complete:")
    logger.info(f"  - High confidence matches: {high_confidence}")
    logger.info(f"  - Medium confidence (review): {medium_confidence}")
    logger.info(f"  - Total potential repeats: {matches_found}")
    
    # Debug: Show score distribution if no matches found
    if matches_found == 0:
        max_score = df['Recidivism_Score'].max()
        avg_score = df['Recidivism_Score'].mean()
        logger.warning(f"  - No matches found. Max similarity score: {max_score:.3f}, Avg: {avg_score:.3f}")
        logger.warning(f"  - Consider lowering thresholds if scores are close to {SIMILARITY_THRESHOLD_MEDIUM}")
    
    # ========================================
    # PHASE 2: Engineer-Based Recidivism
    # ========================================
    logger.info("[Audit Engine] Phase 2: Engineer pattern analysis...")
    df = analyze_engineer_patterns(df)
    
    # Apply penalty for repeat offender engineers (even if issue type is new)
    engineer_penalty_applied = 0
    for idx, row in df.iterrows():
        if row.get('Engineer_Is_Repeat_Offender', False):
            # Apply a smaller penalty for engineer repeat (1.2x) vs issue repeat (1.5x)
            current_status = str(row.get('Learning_Status', ''))
            if 'REPEAT' not in current_status:
                # Only add engineer flag if not already marked as issue repeat
                engineer_name = str(row.get('Engineer', row.get(COL_ENGINEER, 'Unknown')))[:20]
                learning_gap = row.get('Engineer_Learning_Gap', 0)
                df.at[idx, 'Learning_Status'] = f"ENGINEER REPEAT ({engineer_name}, gap:{learning_gap})"
                df.at[idx, 'Strategic_Friction_Score'] *= 1.2  # 20% penalty for repeat offender
                engineer_penalty_applied += 1
    
    if engineer_penalty_applied > 0:
        logger.info(f"  - Applied engineer repeat penalty to {engineer_penalty_applied} issues")
    
    # Summary statistics
    total_repeat_offenders = df['Engineer_Is_Repeat_Offender'].sum()
    avg_learning_gap = df['Engineer_Learning_Gap'].mean() if 'Engineer_Learning_Gap' in df.columns else 0
    logger.info(f"  - Total rows from repeat offender engineers: {total_repeat_offenders}")
    logger.info(f"  - Average learning gap (issues - lessons): {avg_learning_gap:.1f}")
    
    # ========================================
    # PHASE 3: LOB (Line of Business) Analysis
    # ========================================
    logger.info("[Audit Engine] Phase 3: LOB strategic analysis...")
    df = analyze_lob_patterns(df)
    
    # ========================================
    # PHASE 4: PM Recurrence Risk vs Actual Recidivism Validation
    # ========================================
    logger.info("[Audit Engine] Phase 4: PM Prediction Accuracy Analysis...")
    df = analyze_recurrence_accuracy(df)
    
    # ========================================
    # PHASE 5: AI Recurrence Prediction (ML Model)
    # ========================================
    logger.info("[Audit Engine] Phase 5: AI Recurrence Prediction...")
    df = apply_recurrence_predictions(df)
    
    # ========================================
    # PHASE 6: Similar Ticket Analysis & Resolution Comparison
    # ========================================
    logger.info("[Audit Engine] Phase 6: Similar Ticket Analysis...")
    
    # Load human expectations from similarity feedback file
    human_expectations = {}
    similarity_feedback_excel = "similarity_feedback.xlsx"
    if os.path.exists(similarity_feedback_excel):
        try:
            feedback_result = SimilarTicketFinder(ai_engine=ai).load_feedback_from_excel(similarity_feedback_excel)
            if isinstance(feedback_result, tuple) and len(feedback_result) == 2:
                _, human_expectations = feedback_result
                if human_expectations:
                    logger.info(f"[Audit Engine] Loaded {len(human_expectations)} human resolution expectations")
        except Exception as e:
            logger.debug(f"Could not load expectations: {e}")
    
    df = apply_similar_ticket_analysis(df, ai_engine=ai)
    
    # ========================================
    # PHASE 7: Resolution Time Prediction
    # ========================================
    logger.info("[Audit Engine] Phase 7: Resolution Time Prediction...")
    df = apply_resolution_time_prediction(df, human_expectations=human_expectations)
               
    return df


def analyze_recurrence_accuracy(df):
    """
    Compare PM's recurrence risk prediction with actual recidivism findings.
    
    This identifies:
    1. PM Accuracy Rate - How often PM's prediction matches reality
    2. False Positives - PM said High risk but no recurrence found
    3. False Negatives - PM said Low risk but recurrence happened (CRITICAL)
    4. Risk Calibration Score - Overall prediction quality
    
    Returns: DataFrame with recurrence analysis columns added
    """
    df = df.copy()
    
    # Initialize columns
    df['Recurrence_Actual'] = 'No'  # Based on recidivism analysis
    df['PM_Prediction_Accuracy'] = 'N/A'
    df['Risk_Calibration'] = 'N/A'
    
    # Determine actual recurrence from recidivism analysis
    # If Learning_Status contains REPEAT or recidivism score is high, it's a recurrence
    if 'Learning_Status' in df.columns:
        repeat_mask = df['Learning_Status'].astype(str).str.contains('REPEAT', case=False, na=False)
        df.loc[repeat_mask, 'Recurrence_Actual'] = 'Yes'
    
    if 'Recidivism_Score' in df.columns:
        high_score_mask = df['Recidivism_Score'] >= SIMILARITY_THRESHOLD_MEDIUM
        df.loc[high_score_mask, 'Recurrence_Actual'] = 'Yes'
    
    # Check if PM recurrence risk column exists
    if 'PM_Recurrence_Risk_Norm' not in df.columns:
        logger.info("  PM Recurrence Risk column not found, skipping accuracy analysis")
        return df
    
    # Compare PM prediction vs actual
    def evaluate_prediction(row):
        pm_risk = row.get('PM_Recurrence_Risk_Norm', 'Unknown')
        actual = row.get('Recurrence_Actual', 'No')
        
        if pm_risk == 'Unknown':
            return 'No Prediction'
        
        if pm_risk == 'High' and actual == 'Yes':
            return '✅ Correct (High→Recurred)'
        elif pm_risk == 'Low' and actual == 'No':
            return '✅ Correct (Low→No Recurrence)'
        elif pm_risk == 'Medium' and actual in ['Yes', 'No']:
            return '⚪ Moderate (Medium Risk)'
        elif pm_risk == 'High' and actual == 'No':
            return '🟡 Overestimate (High→No Recurrence)'
        elif pm_risk == 'Low' and actual == 'Yes':
            return '🔴 MISSED (Low→Recurred!)'
        else:
            return 'Inconclusive'
    
    df['PM_Prediction_Accuracy'] = df.apply(evaluate_prediction, axis=1)
    
    # Risk Calibration Score
    def calculate_calibration(row):
        pm_risk = row.get('PM_Recurrence_Risk_Norm', 'Unknown')
        recid_score = row.get('Recidivism_Score', 0)
        
        if pm_risk == 'Unknown':
            return 'N/A'
        
        # Compare PM assessment with AI recidivism score
        if pm_risk == 'High':
            if recid_score >= 0.5:
                return 'Well Calibrated'
            elif recid_score >= 0.3:
                return 'Slightly Over'
            else:
                return 'Significantly Over'
        elif pm_risk == 'Low':
            if recid_score < 0.3:
                return 'Well Calibrated'
            elif recid_score < 0.5:
                return 'Slightly Under'
            else:
                return 'UNDER-ESTIMATED!'
        else:  # Medium
            if 0.3 <= recid_score <= 0.6:
                return 'Well Calibrated'
            else:
                return 'Review Needed'
    
    df['Risk_Calibration'] = df.apply(calculate_calibration, axis=1)
    
    # Log summary statistics
    if 'PM_Recurrence_Risk_Norm' in df.columns:
        accuracy_counts = df['PM_Prediction_Accuracy'].value_counts()
        calibration_counts = df['Risk_Calibration'].value_counts()
        
        # Calculate accuracy metrics
        total_predictions = len(df[df['PM_Recurrence_Risk_Norm'] != 'Unknown'])
        correct = len(df[df['PM_Prediction_Accuracy'].str.contains('Correct', na=False)])
        missed = len(df[df['PM_Prediction_Accuracy'].str.contains('MISSED', na=False)])
        
        if total_predictions > 0:
            accuracy_rate = (correct / total_predictions) * 100
            miss_rate = (missed / total_predictions) * 100
            
            logger.info(f"  PM Prediction Accuracy: {accuracy_rate:.1f}% ({correct}/{total_predictions})")
            if missed > 0:
                logger.warning(f"  ⚠ CRITICAL: {missed} issues where PM said Low risk but actually recurred!")
            
            # Store accuracy metrics for reporting
            df.attrs['pm_accuracy'] = {
                'total_predictions': total_predictions,
                'correct': correct,
                'missed': missed,
                'accuracy_rate': accuracy_rate,
                'miss_rate': miss_rate
            }
    
    return df


# ==========================================
# 4B. PREDICTIVE RECURRENCE MODEL (ML-Based)
# ==========================================

# Model persistence path
RECURRENCE_MODEL_PATH = "recurrence_predictor.pkl"
RECURRENCE_ENCODERS_PATH = "recurrence_encoders.pkl"

class RecurrencePredictor:
    """
    Machine Learning model to predict ticket recurrence probability.
    
    Uses historical data to learn patterns and predict which new tickets
    are likely to recur within 30 days, allowing proactive intervention.
    
    Features used:
    - Severity level
    - Issue type (Escalation/Concern/etc.)
    - Origin (Internal/External)
    - Root cause category
    - Engineer history (repeat offender status)
    - LOB risk tier
    - AI category
    - Text embedding similarity to past recurring issues
    """
    
    def __init__(self):
        self.model = None
        self.encoders = {}
        self.feature_columns = []
        self.is_trained = False
        self.model_metrics = {}
        
    def _prepare_features(self, df, fit_encoders=False):
        """
        Prepare feature matrix for training/prediction.
        
        Args:
            df: DataFrame with ticket data
            fit_encoders: If True, fit new encoders (training mode)
            
        Returns:
            X: Feature matrix (numpy array)
            feature_names: List of feature column names
        """
        feature_df = pd.DataFrame()
        
        # Numeric features (direct use)
        numeric_features = [
            'Strategic_Friction_Score',
            'AI_Confidence',
            'Engineer_Issue_Count',
            'Days_Since_Issue',
            'Recidivism_Score'
        ]
        
        for feat in numeric_features:
            if feat in df.columns:
                feature_df[feat] = pd.to_numeric(df[feat], errors='coerce').fillna(0)
        
        # Categorical features (need encoding)
        categorical_features = {
            'Severity_Norm': ['Critical', 'High', 'Medium', 'Low', 'Unknown'],
            'Type_Norm': ['Escalations', 'Concerns', 'Lessons Learned', 'Unknown'],
            'Origin_Norm': ['Internal', 'External', 'Unknown'],
            'Root_Cause_Category': ['Human Error', 'External Party', 'Process Gap', 
                                     'System/Technical', 'Training Gap', 'Communication', 
                                     'Resource', 'Other', 'Unclassified'],
            'AI_Category': list(ANCHORS.keys()) + ['Unclassified'],
            'LOB_Risk_Tier': ['Critical', 'High', 'Medium', 'Low', 'Unknown'],
            'Is_Human_Error': ['Yes', 'No', 'External']
        }
        
        for feat, categories in categorical_features.items():
            if feat in df.columns:
                if fit_encoders:
                    # Create and fit new encoder
                    le = LabelEncoder()
                    # Fit on all possible categories to handle unseen values
                    le.fit(categories + ['Unknown'])
                    self.encoders[feat] = le
                
                if feat in self.encoders:
                    # Transform, handling unseen values
                    values = df[feat].fillna('Unknown').astype(str)
                    values = values.apply(lambda x: x if x in self.encoders[feat].classes_ else 'Unknown')
                    feature_df[f'{feat}_encoded'] = self.encoders[feat].transform(values)
        
        # Binary flags
        if 'Engineer_Flag' in df.columns:
            feature_df['Is_Repeat_Offender'] = df['Engineer_Flag'].apply(
                lambda x: 1 if 'Repeat' in str(x) else 0
            )
        
        if 'Aging_Status' in df.columns:
            feature_df['Is_Aged'] = df['Aging_Status'].apply(
                lambda x: 1 if '>30' in str(x) else (0.5 if '>14' in str(x) else 0)
            )
        
        # Embedding-based features (if available)
        if 'embedding' in df.columns:
            # Calculate mean embedding distance to known recurring issues
            # (Simplified: use first 10 dimensions of embedding as features)
            try:
                embeddings = np.vstack(df['embedding'].values)
                for i in range(min(10, embeddings.shape[1])):
                    feature_df[f'emb_dim_{i}'] = embeddings[:, i]
            except Exception:
                pass  # Skip if embeddings not available
        
        self.feature_columns = list(feature_df.columns)
        return feature_df.values, self.feature_columns
    
    def train(self, df, min_samples=50):
        """
        Train the recurrence prediction model on historical data.
        
        Args:
            df: DataFrame with historical ticket data including recurrence outcomes
            min_samples: Minimum samples required to train
            
        Returns:
            dict: Training metrics
        """
        logger.info("[Recurrence Predictor] Training predictive model...")
        
        # Determine target variable (actual recurrence)
        if 'Recurrence_Actual' not in df.columns:
            logger.warning("No Recurrence_Actual column found. Cannot train.")
            return {'error': 'No recurrence data available'}
        
        # Filter to rows with known outcomes
        train_df = df[df['Recurrence_Actual'].isin(['Yes', 'No'])].copy()
        
        if len(train_df) < min_samples:
            logger.warning(f"Insufficient training data ({len(train_df)} samples). Need {min_samples}+")
            return {'error': f'Need at least {min_samples} samples, have {len(train_df)}'}
        
        # Prepare features
        X, feature_names = self._prepare_features(train_df, fit_encoders=True)
        y = (train_df['Recurrence_Actual'] == 'Yes').astype(int)
        
        # Handle class imbalance
        recurrence_rate = y.mean()
        logger.info(f"  Historical recurrence rate: {recurrence_rate:.1%}")
        
        # Split data
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42, stratify=y if len(y.unique()) > 1 else None
        )
        
        # Train model (Gradient Boosting for better probability calibration)
        self.model = GradientBoostingClassifier(
            n_estimators=100,
            max_depth=4,
            learning_rate=0.1,
            min_samples_split=10,
            random_state=42
        )
        
        self.model.fit(X_train, y_train)
        
        # Evaluate
        y_pred = self.model.predict(X_test)
        y_proba = self.model.predict_proba(X_test)[:, 1]
        
        # Calculate metrics
        try:
            auc_score = roc_auc_score(y_test, y_proba) if len(y_test.unique()) > 1 else 0.5
        except Exception:
            auc_score = 0.5
        
        accuracy = (y_pred == y_test).mean()
        
        # Feature importance
        feature_importance = dict(zip(feature_names, self.model.feature_importances_))
        top_features = sorted(feature_importance.items(), key=lambda x: x[1], reverse=True)[:5]
        
        self.model_metrics = {
            'accuracy': accuracy,
            'auc_roc': auc_score,
            'train_samples': len(X_train),
            'test_samples': len(X_test),
            'recurrence_rate': recurrence_rate,
            'top_features': top_features
        }
        
        self.is_trained = True
        
        # Log results
        logger.info(f"  ✓ Model trained on {len(X_train)} samples")
        logger.info(f"  → Accuracy: {accuracy:.1%}")
        logger.info(f"  → AUC-ROC: {auc_score:.3f}")
        logger.info(f"  → Top predictive features:")
        for feat, importance in top_features:
            logger.info(f"      • {feat}: {importance:.3f}")
        
        return self.model_metrics
    
    def predict(self, df):
        """
        Predict recurrence probability for tickets.
        
        Args:
            df: DataFrame with ticket data
            
        Returns:
            DataFrame with prediction columns added
        """
        df = df.copy()
        
        # Initialize prediction columns
        df['AI_Recurrence_Probability'] = 0.0
        df['AI_Recurrence_Risk'] = 'Unknown'
        df['AI_Recurrence_Confidence'] = 'Low'
        
        if not self.is_trained:
            logger.info("[Recurrence Predictor] Model not trained, skipping predictions")
            return df
        
        try:
            # Prepare features
            X, _ = self._prepare_features(df, fit_encoders=False)
            
            # Predict probabilities
            probabilities = self.model.predict_proba(X)[:, 1]
            
            df['AI_Recurrence_Probability'] = probabilities
            
            # Categorize risk level
            def categorize_risk(prob):
                if prob >= 0.7:
                    return '🔴 High Risk (>70%)'
                elif prob >= 0.5:
                    return '🟠 Elevated (50-70%)'
                elif prob >= 0.3:
                    return '🟡 Moderate (30-50%)'
                else:
                    return '🟢 Low (<30%)'
            
            df['AI_Recurrence_Risk'] = df['AI_Recurrence_Probability'].apply(categorize_risk)
            
            # Confidence based on model certainty (distance from 0.5)
            def get_confidence(prob):
                certainty = abs(prob - 0.5) * 2  # 0 to 1 scale
                if certainty >= 0.6:
                    return 'High'
                elif certainty >= 0.3:
                    return 'Medium'
                else:
                    return 'Low'
            
            df['AI_Recurrence_Confidence'] = df['AI_Recurrence_Probability'].apply(get_confidence)
            
            # Log summary
            high_risk = (df['AI_Recurrence_Probability'] >= 0.5).sum()
            logger.info(f"[Recurrence Predictor] Predictions complete:")
            logger.info(f"  → {high_risk} tickets flagged as high recurrence risk (≥50%)")
            logger.info(f"  → Average recurrence probability: {probabilities.mean():.1%}")
            
        except Exception as e:
            logger.warning(f"[Recurrence Predictor] Prediction failed: {e}")
        
        return df
    
    def save(self, model_path=None, encoders_path=None):
        """Save trained model and encoders to disk."""
        model_path = model_path or RECURRENCE_MODEL_PATH
        encoders_path = encoders_path or RECURRENCE_ENCODERS_PATH
        
        if self.is_trained:
            with open(model_path, 'wb') as f:
                pickle.dump({
                    'model': self.model,
                    'feature_columns': self.feature_columns,
                    'metrics': self.model_metrics
                }, f)
            
            with open(encoders_path, 'wb') as f:
                pickle.dump(self.encoders, f)
            
            logger.info(f"[Recurrence Predictor] Model saved to {model_path}")
    
    def load(self, model_path=None, encoders_path=None):
        """Load trained model and encoders from disk."""
        model_path = model_path or RECURRENCE_MODEL_PATH
        encoders_path = encoders_path or RECURRENCE_ENCODERS_PATH
        
        try:
            if os.path.exists(model_path) and os.path.exists(encoders_path):
                with open(model_path, 'rb') as f:
                    data = pickle.load(f)
                    self.model = data['model']
                    self.feature_columns = data['feature_columns']
                    self.model_metrics = data.get('metrics', {})
                
                with open(encoders_path, 'rb') as f:
                    self.encoders = pickle.load(f)
                
                self.is_trained = True
                logger.info(f"[Recurrence Predictor] Model loaded from {model_path}")
                return True
        except Exception as e:
            logger.warning(f"[Recurrence Predictor] Could not load model: {e}")
        
        return False
    
    def get_risk_factors(self, row):
        """
        Explain why a specific ticket is flagged as high risk.
        
        Returns: List of contributing risk factors
        """
        factors = []
        
        if row.get('AI_Recurrence_Probability', 0) >= 0.5:
            # Check individual risk indicators
            if row.get('Is_Human_Error') == 'Yes':
                factors.append("Human error root cause (historically high recurrence)")
            
            if row.get('Engineer_Issue_Count', 0) >= 3:
                factors.append(f"Engineer has {row.get('Engineer_Issue_Count')} prior issues")
            
            if row.get('Root_Cause_Category') == 'Process Gap':
                factors.append("Process gap issues tend to recur until fixed")
            
            if row.get('Severity_Norm') in ['Critical', 'High']:
                factors.append("High severity issues often indicate systemic problems")
            
            if row.get('Recidivism_Score', 0) >= 0.5:
                factors.append(f"Similar issue occurred before (similarity: {row.get('Recidivism_Score', 0):.0%})")
            
            if row.get('LOB_Risk_Tier') in ['Critical', 'High']:
                factors.append(f"LOB has elevated risk profile")
        
        return factors if factors else ["No specific factors identified"]


# Global predictor instance
recurrence_predictor = RecurrencePredictor()


def apply_recurrence_predictions(df, train_if_possible=True):
    """
    Apply recurrence predictions to the dataframe.
    
    This function:
    1. Tries to load a pre-trained model
    2. If not available and data permits, trains a new model
    3. Applies predictions to all tickets
    
    Args:
        df: DataFrame with ticket data
        train_if_possible: Whether to train if no model exists
        
    Returns:
        DataFrame with prediction columns added
    """
    global recurrence_predictor
    
    logger.info("[Recurrence Predictor] Initializing AI-based recurrence prediction...")
    
    # Try to load existing model
    if recurrence_predictor.load():
        logger.info("  Using pre-trained model")
    elif train_if_possible:
        # Train new model if we have recurrence data
        if 'Recurrence_Actual' in df.columns:
            metrics = recurrence_predictor.train(df)
            if 'error' not in metrics:
                recurrence_predictor.save()
        else:
            logger.info("  No historical recurrence data - will train on next run with outcomes")
    
    # Apply predictions
    df = recurrence_predictor.predict(df)
    
    return df


# ==========================================
# 4B. SIMILAR TICKET FINDER
# ==========================================
class SimilarTicketFinder:
    """
    AI-powered Similar Ticket Finder with Resolution Comparison.
    
    Uses embeddings to find semantically similar historical tickets,
    then compares resolutions to identify patterns and inconsistencies.
    
    Key Features:
    - Embedding-based semantic similarity (not just keyword matching)
    - Resolution pattern analysis across similar tickets
    - Days to resolution comparison
    - Identifies resolution inconsistencies (same problem, different solutions)
    - Human-in-the-loop feedback for similarity validation
    - Calculates resolution effectiveness scores
    - Provides actionable recommendations
    """
    
    def __init__(self, ai_engine=None, top_k=5, similarity_threshold=0.75):
        """
        Initialize the Similar Ticket Finder.
        
        Args:
            ai_engine: AI engine instance for embeddings (or uses global)
            top_k: Number of similar tickets to find per query
            similarity_threshold: Minimum cosine similarity to consider (0-1)
        """
        self.ai = ai_engine
        self.top_k = top_k
        self.similarity_threshold = similarity_threshold
        self.embedding_cache = {}
        self.resolution_patterns = {}
        self.feedback_data = {}  # Human feedback on similarity matches
        
        # Load existing feedback if available
        self._load_feedback()
        
        # Resolution-related column names (may vary by dataset)
        self.resolution_columns = [
            COL_LESSON_TITLE,      # Lesson learned = resolution insight
            'Action_Required',     # Recommended action
            'Root_Cause_Category', # Root cause classification
            COL_ROOT_CAUSE,        # Original root cause field
        ]
        
        logger.info("[Similar Ticket Finder] Initialized")
    
    def _load_feedback(self):
        """Load human feedback on similarity matches from file."""
        try:
            if os.path.exists(SIMILARITY_FEEDBACK_PATH):
                with open(SIMILARITY_FEEDBACK_PATH, 'r') as f:
                    import json
                    self.feedback_data = json.load(f)
                logger.info(f"[Similar Ticket Finder] Loaded {len(self.feedback_data)} feedback entries")
        except Exception as e:
            logger.warning(f"[Similar Ticket Finder] Could not load feedback: {e}")
            self.feedback_data = {}
    
    def _save_feedback(self):
        """Save human feedback to file for future runs."""
        try:
            import json
            with open(SIMILARITY_FEEDBACK_PATH, 'w') as f:
                json.dump(self.feedback_data, f, indent=2)
            logger.info(f"[Similar Ticket Finder] Saved {len(self.feedback_data)} feedback entries")
        except Exception as e:
            logger.warning(f"[Similar Ticket Finder] Could not save feedback: {e}")
    
    def record_feedback(self, ticket_id, similar_ticket_id, is_similar, notes=""):
        """
        Record human feedback on whether two tickets are actually similar.
        
        Args:
            ticket_id: The query ticket ID
            similar_ticket_id: The suggested similar ticket ID
            is_similar: True if human confirms similarity, False if rejected
            notes: Optional notes explaining the decision
        """
        feedback_key = f"{ticket_id}|{similar_ticket_id}"
        self.feedback_data[feedback_key] = {
            'ticket_id': str(ticket_id),
            'similar_ticket_id': str(similar_ticket_id),
            'is_similar': is_similar,
            'notes': notes,
            'timestamp': datetime.now().isoformat()
        }
        self._save_feedback()
    
    def get_feedback_adjusted_similarity(self, ticket_id, similar_ticket_id, base_similarity):
        """
        Adjust similarity score based on human feedback.
        
        If human has confirmed/rejected this pair before, use that feedback.
        """
        feedback_key = f"{ticket_id}|{similar_ticket_id}"
        reverse_key = f"{similar_ticket_id}|{ticket_id}"
        
        # Check both directions
        feedback = self.feedback_data.get(feedback_key) or self.feedback_data.get(reverse_key)
        
        if feedback:
            if feedback['is_similar']:
                # Human confirmed - boost similarity
                return min(1.0, base_similarity + 0.2)
            else:
                # Human rejected - heavily penalize
                return max(0.0, base_similarity - 0.5)
        
        return base_similarity
    
    def _calculate_days_to_resolution(self, row):
        """
        Calculate days from issue open to resolution.
        
        Returns: float (days) or None if not available
        """
        try:
            open_date = None
            close_date = None
            
            # Get open date
            if 'Issue_Date' in row.index and pd.notna(row.get('Issue_Date')):
                open_date = pd.to_datetime(row['Issue_Date'], errors='coerce')
            elif COL_DATETIME in row.index and pd.notna(row.get(COL_DATETIME)):
                open_date = pd.to_datetime(row[COL_DATETIME], errors='coerce')
            
            # Get close date
            if COL_CLOSE_DATE in row.index and pd.notna(row.get(COL_CLOSE_DATE)):
                close_date = pd.to_datetime(row[COL_CLOSE_DATE], errors='coerce')
            elif 'Close_Date' in row.index and pd.notna(row.get('Close_Date')):
                close_date = pd.to_datetime(row['Close_Date'], errors='coerce')
            elif 'Resolution_Date' in row.index and pd.notna(row.get('Resolution_Date')):
                close_date = pd.to_datetime(row['Resolution_Date'], errors='coerce')
            
            if open_date and close_date and pd.notna(open_date) and pd.notna(close_date):
                delta = (close_date - open_date).days
                return max(0, delta)  # Ensure non-negative
            
        except Exception:
            pass
        
        return None
    
    def _cosine_similarity(self, vec1, vec2):
        """Calculate cosine similarity between two vectors."""
        if vec1 is None or vec2 is None:
            return 0.0
        dot = np.dot(vec1, vec2)
        norm1 = np.linalg.norm(vec1)
        norm2 = np.linalg.norm(vec2)
        if norm1 == 0 or norm2 == 0:
            return 0.0
        return dot / (norm1 * norm2)
    
    def _get_embedding(self, text, ticket_id=None):
        """Get embedding for text, with caching."""
        if not text or pd.isna(text):
            return None
        
        text = str(text).strip()
        if not text:
            return None
        
        # Check cache first
        cache_key = ticket_id if ticket_id else hash(text)
        if cache_key in self.embedding_cache:
            return self.embedding_cache[cache_key]
        
        # Get embedding from AI engine
        if self.ai:
            embedding = self.ai.get_embedding(text)
            if embedding is not None and len(embedding) > 0:
                self.embedding_cache[cache_key] = embedding
                return embedding
        
        return None
    
    def _get_resolution_text(self, row):
        """
        Extract resolution/action text from a ticket row.
        Combines multiple resolution-related fields.
        """
        resolution_parts = []
        
        # Primary: Lesson learned title
        if COL_LESSON_TITLE in row.index:
            lesson = row.get(COL_LESSON_TITLE)
            if lesson and not pd.isna(lesson) and str(lesson).strip():
                resolution_parts.append(f"Lesson: {lesson}")
        
        # Action taken/required
        if 'Action_Required' in row.index:
            action = row.get('Action_Required')
            if action and not pd.isna(action) and str(action).strip() and action != 'Monitor':
                resolution_parts.append(f"Action: {action}")
        
        # Root cause (gives insight into resolution approach)
        root_cause = None
        if 'Root_Cause_Category' in row.index:
            root_cause = row.get('Root_Cause_Category')
        elif COL_ROOT_CAUSE in row.index:
            root_cause = row.get(COL_ROOT_CAUSE)
        
        if root_cause and not pd.isna(root_cause) and str(root_cause).strip():
            resolution_parts.append(f"Root Cause: {root_cause}")
        
        return " | ".join(resolution_parts) if resolution_parts else "No resolution documented"
    
    def _analyze_resolution_consistency(self, similar_tickets):
        """
        Analyze if similar tickets had consistent resolutions.
        Includes resolution time analysis.
        
        Returns:
            dict with consistency metrics, time analysis, and insights
        """
        if len(similar_tickets) < 2:
            return {
                'consistency_score': 1.0,
                'status': 'N/A - Not enough similar tickets',
                'unique_approaches': 1,
                'dominant_approach': None,
                'insight': None,
                'resolution_time': None
            }
        
        # Collect root causes, actions, and resolution times
        root_causes = []
        actions = []
        lessons = []
        resolution_times = []
        
        for ticket in similar_tickets:
            rc = ticket.get('root_cause', '')
            if rc and rc != 'Unknown':
                root_causes.append(rc)
            
            action = ticket.get('action', '')
            if action and action != 'Monitor' and action != 'No resolution documented':
                actions.append(action)
            
            lesson = ticket.get('lesson', '')
            if lesson:
                lessons.append(lesson)
            
            # Collect resolution times
            days = ticket.get('days_to_resolution')
            if days is not None:
                resolution_times.append(days)
        
        # Calculate consistency
        unique_root_causes = len(set(root_causes)) if root_causes else 0
        unique_actions = len(set(actions)) if actions else 0
        
        total_unique = unique_root_causes + unique_actions
        total_items = len(root_causes) + len(actions)
        
        if total_items == 0:
            consistency_score = 0.0
            status = '⚠️ No resolution data available'
        elif total_unique <= 1:
            consistency_score = 1.0
            status = '✅ Highly consistent resolutions'
        elif total_unique <= 2:
            consistency_score = 0.7
            status = '🟡 Mostly consistent with minor variations'
        else:
            consistency_score = max(0.3, 1.0 - (total_unique / total_items))
            status = '🔴 Inconsistent resolutions - review needed'
        
        # Find dominant approach
        all_approaches = root_causes + actions
        dominant = None
        if all_approaches:
            from collections import Counter
            counts = Counter(all_approaches)
            dominant = counts.most_common(1)[0][0] if counts else None
        
        # Analyze resolution times
        resolution_time_analysis = None
        if resolution_times:
            avg_time = np.mean(resolution_times)
            min_time = min(resolution_times)
            max_time = max(resolution_times)
            std_time = np.std(resolution_times) if len(resolution_times) > 1 else 0
            
            resolution_time_analysis = {
                'avg_days': round(avg_time, 1),
                'min_days': round(min_time, 1),
                'max_days': round(max_time, 1),
                'std_days': round(std_time, 1),
                'sample_size': len(resolution_times),
                'time_variance': 'High' if std_time > avg_time * 0.5 else 'Low'
            }
        
        # Generate insight
        insight = None
        if consistency_score < 0.5 and len(similar_tickets) >= 3:
            insight = f"⚠️ ALERT: {len(similar_tickets)} similar tickets resolved {unique_root_causes + unique_actions} different ways. Consider standardizing approach."
        elif consistency_score >= 0.8 and lessons:
            insight = f"✅ Consistent approach: {dominant if dominant else lessons[0][:50]}"
        
        # Add time insight if there's high variance
        if resolution_time_analysis and resolution_time_analysis['time_variance'] == 'High':
            time_insight = f"⏱️ Resolution time varies significantly ({resolution_time_analysis['min_days']}-{resolution_time_analysis['max_days']} days)"
            insight = f"{insight}\n{time_insight}" if insight else time_insight
        
        return {
            'consistency_score': consistency_score,
            'status': status,
            'unique_approaches': total_unique,
            'dominant_approach': dominant,
            'insight': insight,
            'has_lessons': len(lessons) > 0,
            'resolution_time': resolution_time_analysis
        }
    
    def find_similar(self, query_row, historical_df, exclude_self=True):
        """
        Find tickets similar to the query ticket.
        
        Args:
            query_row: Series or dict with query ticket data
            historical_df: DataFrame of historical tickets to search
            exclude_self: Whether to exclude the query ticket from results
            
        Returns:
            List of dicts with similar ticket info and similarity scores
        """
        # Get query text and embedding
        query_id = query_row.get(COL_SUMMARY, query_row.get('ID', ''))
        query_text = str(query_row.get(COL_SUMMARY, query_row.get('Text', ''))).strip()
        
        if not query_text:
            return []
        
        query_embedding = self._get_embedding(query_text, f"query_{hash(query_text)}")
        if query_embedding is None:
            return []
        
        similar_tickets = []
        
        for idx, hist_row in historical_df.iterrows():
            # Skip self if requested
            if exclude_self:
                hist_id = hist_row.get(COL_SUMMARY, hist_row.get('ID', ''))
                if hist_id == query_id:
                    continue
            
            # Get historical ticket text and embedding
            hist_text = str(hist_row.get(COL_SUMMARY, hist_row.get('Text', ''))).strip()
            if not hist_text:
                continue
            
            hist_embedding = self._get_embedding(hist_text, f"hist_{idx}")
            if hist_embedding is None:
                continue
            
            # Calculate base similarity
            base_similarity = self._cosine_similarity(query_embedding, hist_embedding)
            
            # Adjust similarity based on human feedback
            hist_ticket_id = hist_row.get('ID', str(idx))
            adjusted_similarity = self.get_feedback_adjusted_similarity(
                str(query_id), str(hist_ticket_id), base_similarity
            )
            
            if adjusted_similarity >= self.similarity_threshold:
                # Extract resolution info
                resolution_text = self._get_resolution_text(hist_row)
                
                # Calculate days to resolution
                days_to_resolution = self._calculate_days_to_resolution(hist_row)
                
                # Check if this pair has feedback
                feedback_key = f"{query_id}|{hist_ticket_id}"
                reverse_key = f"{hist_ticket_id}|{query_id}"
                has_feedback = feedback_key in self.feedback_data or reverse_key in self.feedback_data
                feedback_status = ""
                if has_feedback:
                    fb = self.feedback_data.get(feedback_key) or self.feedback_data.get(reverse_key)
                    if fb:
                        feedback_status = "✅ Confirmed" if fb.get('is_similar') else "❌ Rejected"
                
                similar_tickets.append({
                    'index': idx,
                    'id': hist_ticket_id,
                    'text': hist_text[:200] + '...' if len(hist_text) > 200 else hist_text,
                    'similarity': adjusted_similarity,
                    'base_similarity': base_similarity,
                    'category': hist_row.get('AI_Category', 'Unknown'),
                    'severity': hist_row.get('Severity_Norm', hist_row.get(COL_SEVERITY, 'Unknown')),
                    'root_cause': hist_row.get('Root_Cause_Category', hist_row.get(COL_ROOT_CAUSE, 'Unknown')),
                    'resolution': resolution_text,
                    'lesson': hist_row.get(COL_LESSON_TITLE, ''),
                    'action': hist_row.get('Action_Required', ''),
                    'date': hist_row.get('Issue_Date', hist_row.get(COL_DATETIME, '')),
                    'engineer': hist_row.get('Engineer', hist_row.get(COL_ENGINEER, '')),
                    'lob': hist_row.get('LOB', hist_row.get(COL_LOB, '')),
                    'friction_score': hist_row.get('Strategic_Friction_Score', 0),
                    'recurred': hist_row.get('Recurrence_Actual', hist_row.get('Is_Repeat', 'Unknown')),
                    'days_to_resolution': days_to_resolution,
                    'feedback_status': feedback_status
                })
        
        # Sort by similarity (highest first) and take top K
        similar_tickets.sort(key=lambda x: x['similarity'], reverse=True)
        return similar_tickets[:self.top_k]
    
    def analyze_ticket(self, query_row, historical_df):
        """
        Full analysis of a ticket: find similar and compare resolutions.
        
        Returns:
            dict with similar tickets, resolution analysis, and recommendations
        """
        similar = self.find_similar(query_row, historical_df)
        
        if not similar:
            return {
                'similar_tickets': [],
                'match_count': 0,
                'resolution_analysis': None,
                'recommendation': 'No sufficiently similar historical tickets found. This may be a new issue type.',
                'confidence': 'Low',
                'expected_resolution_days': None
            }
        
        # Analyze resolution consistency
        resolution_analysis = self._analyze_resolution_consistency(similar)
        
        # Generate recommendation
        if resolution_analysis['consistency_score'] >= 0.8 and resolution_analysis['dominant_approach']:
            recommendation = f"✅ RECOMMENDED: Follow established approach - {resolution_analysis['dominant_approach']}"
            confidence = 'High'
        elif resolution_analysis['consistency_score'] >= 0.5:
            recommendation = f"🟡 SUGGESTED: Review past resolutions. Most common: {resolution_analysis['dominant_approach']}"
            confidence = 'Medium'
        else:
            recommendation = "🔴 CAUTION: Inconsistent past resolutions. Consult with team lead before proceeding."
            confidence = 'Low'
        
        # Add resolution time estimate
        expected_days = None
        if resolution_analysis.get('resolution_time'):
            time_info = resolution_analysis['resolution_time']
            expected_days = time_info['avg_days']
            recommendation += f"\n\n⏱️ EXPECTED RESOLUTION TIME: {time_info['avg_days']} days (range: {time_info['min_days']}-{time_info['max_days']} days based on {time_info['sample_size']} similar cases)"
        
        # Add lessons learned if available
        lessons = [t['lesson'] for t in similar if t.get('lesson')]
        if lessons:
            recommendation += f"\n\n📚 LESSONS LEARNED:\n" + "\n".join([f"  • {l[:100]}" for l in lessons[:3]])
        
        # Check if similar tickets recurred
        recurrence_info = [t for t in similar if t.get('recurred') in ['Yes', True, 1]]
        if recurrence_info:
            recommendation += f"\n\n⚠️ WARNING: {len(recurrence_info)} of {len(similar)} similar tickets recurred!"
        
        return {
            'similar_tickets': similar,
            'match_count': len(similar),
            'avg_similarity': np.mean([t['similarity'] for t in similar]),
            'resolution_analysis': resolution_analysis,
            'recommendation': recommendation,
            'confidence': confidence,
            'expected_resolution_days': expected_days
        }
    
    def process_all_tickets(self, df, progress_callback=None):
        """
        Find similar tickets for all tickets in the dataset.
        
        Adds columns:
        - Similar_Ticket_Count: Number of similar historical tickets
        - Best_Match_Similarity: Highest similarity score found
        - Resolution_Consistency: How consistent resolutions were
        - Similar_Ticket_IDs: IDs of top similar tickets
        - Expected_Resolution_Days: Predicted time to resolve based on similar tickets
        - Resolution_Recommendation: Suggested approach
        - Similarity_Feedback: Column for human feedback (blank for user to fill)
        
        Returns:
            DataFrame with similarity columns added
        """
        logger.info(f"[Similar Ticket Finder] Processing {len(df)} tickets...")
        
        df = df.copy()
        
        # Initialize new columns
        df['Similar_Ticket_Count'] = 0
        df['Best_Match_Similarity'] = 0.0
        df['Resolution_Consistency'] = 'N/A'
        df['Similar_Ticket_IDs'] = ''
        df['Expected_Resolution_Days'] = np.nan
        df['Avg_Similar_Resolution_Days'] = np.nan
        df['Resolution_Recommendation'] = ''
        df['Similarity_Feedback'] = ''  # For human-in-the-loop: "Correct", "Wrong", or blank
        
        processed = 0
        total = len(df)
        
        for idx, row in df.iterrows():
            analysis = self.analyze_ticket(row, df)
            
            df.at[idx, 'Similar_Ticket_Count'] = analysis['match_count']
            df.at[idx, 'Best_Match_Similarity'] = analysis.get('avg_similarity', 0)
            
            if analysis['resolution_analysis']:
                df.at[idx, 'Resolution_Consistency'] = analysis['resolution_analysis']['status']
                
                # Add resolution time from similar tickets
                if analysis['resolution_analysis'].get('resolution_time'):
                    time_info = analysis['resolution_analysis']['resolution_time']
                    df.at[idx, 'Avg_Similar_Resolution_Days'] = time_info['avg_days']
            
            # Store expected resolution days
            if analysis.get('expected_resolution_days') is not None:
                df.at[idx, 'Expected_Resolution_Days'] = analysis['expected_resolution_days']
            
            # Store similar ticket IDs with resolution times
            if analysis['similar_tickets']:
                similar_info = []
                for t in analysis['similar_tickets'][:3]:
                    ticket_str = str(t['id'])
                    if t.get('days_to_resolution') is not None:
                        ticket_str += f" ({t['days_to_resolution']}d)"
                    similar_info.append(ticket_str)
                df.at[idx, 'Similar_Ticket_IDs'] = ', '.join(similar_info)
            
            df.at[idx, 'Resolution_Recommendation'] = analysis['recommendation'].split('\n')[0]  # First line only
            
            processed += 1
            if progress_callback and processed % 10 == 0:
                progress_callback(processed, total)
        
        # Log summary
        has_matches = (df['Similar_Ticket_Count'] > 0).sum()
        inconsistent = df['Resolution_Consistency'].str.contains('Inconsistent', na=False).sum()
        has_time_estimate = df['Expected_Resolution_Days'].notna().sum()
        
        logger.info(f"[Similar Ticket Finder] Complete:")
        logger.info(f"  → {has_matches}/{total} tickets have similar historical matches")
        logger.info(f"  → {inconsistent} tickets have inconsistent resolution patterns")
        logger.info(f"  → {has_time_estimate} tickets have resolution time estimates")
        
        return df
    
    def generate_similarity_report(self, df, top_n=20):
        """
        Generate a report of tickets with inconsistent resolutions.
        
        Returns:
            DataFrame with tickets needing resolution standardization
        """
        # Find tickets with inconsistent resolutions
        inconsistent_mask = df['Resolution_Consistency'].str.contains('Inconsistent', na=False)
        high_similarity_mask = df['Best_Match_Similarity'] >= 0.8
        
        priority_tickets = df[inconsistent_mask & high_similarity_mask].copy()
        
        if len(priority_tickets) == 0:
            # Fall back to any with matches
            priority_tickets = df[df['Similar_Ticket_Count'] > 2].copy()
        
        # Sort by friction score (highest priority issues first)
        if 'Strategic_Friction_Score' in priority_tickets.columns:
            priority_tickets = priority_tickets.sort_values('Strategic_Friction_Score', ascending=False)
        
        report_cols = [
            'ID' if 'ID' in priority_tickets.columns else COL_SUMMARY,
            'AI_Category',
            'Similar_Ticket_Count',
            'Best_Match_Similarity',
            'Resolution_Consistency',
            'Similar_Ticket_IDs',
            'Resolution_Recommendation'
        ]
        
        available_cols = [c for c in report_cols if c in priority_tickets.columns]
        
        return priority_tickets[available_cols].head(top_n)
    
    def load_feedback_from_excel(self, excel_path):
        """
        Load human feedback from the Similar Tickets feedback Excel file.
        
        The Excel file should have columns:
        - Ticket_ID
        - Similar_Ticket_ID
        - Is_Similar: "Similar" or "Not Similar" (dropdown)
        - Expected_Resolution_Days: Human-entered expected resolution time
        - Notes
        
        Returns:
            Tuple of (feedback_count, resolution_expectations dict)
        """
        resolution_expectations = {}  # category -> expected days
        
        try:
            if not os.path.exists(excel_path):
                return 0, resolution_expectations
            
            feedback_df = pd.read_excel(excel_path, sheet_name='Similarity Feedback')
            
            loaded = 0
            for _, row in feedback_df.iterrows():
                ticket_id = str(row.get('Ticket_ID', ''))
                similar_id = str(row.get('Similar_Ticket_ID', ''))
                category = str(row.get('Category', 'Unknown'))
                
                # Handle both old ("Human_Feedback": Correct/Wrong) and new ("Is_Similar": Similar/Not Similar) formats
                feedback = str(row.get('Is_Similar', row.get('Human_Feedback', ''))).strip().lower()
                notes = str(row.get('Notes', ''))
                
                # Parse feedback - support both formats
                is_similar = None
                if feedback in ['similar', 'correct']:
                    is_similar = True
                elif feedback in ['not similar', 'wrong']:
                    is_similar = False
                
                if ticket_id and similar_id and is_similar is not None:
                    self.record_feedback(ticket_id, similar_id, is_similar, notes)
                    loaded += 1
                
                # Load human expected resolution time by category
                expected_days = row.get('Expected_Resolution_Days', None)
                if pd.notna(expected_days) and expected_days != '':
                    try:
                        exp_days = float(expected_days)
                        if category not in resolution_expectations:
                            resolution_expectations[category] = []
                        resolution_expectations[category].append(exp_days)
                    except (ValueError, TypeError):
                        pass
            
            # Average the expectations by category
            for cat in resolution_expectations:
                resolution_expectations[cat] = np.mean(resolution_expectations[cat])
            
            logger.info(f"[Similar Ticket Finder] Loaded {loaded} feedback entries, {len(resolution_expectations)} category expectations")
            return loaded, resolution_expectations
            
        except Exception as e:
            logger.warning(f"[Similar Ticket Finder] Could not load Excel feedback: {e}")
            return 0, resolution_expectations
    
    def export_for_feedback(self, df, output_path):
        """
        Export similar ticket matches for human review with dropdowns.
        
        Creates an Excel file where users can:
        1. Mark whether AI similarity matches are correct or not (dropdown)
        2. Input expected resolution time for categories
        """
        feedback_rows = []
        
        for idx, row in df.iterrows():
            ticket_id = row.get('ID', str(idx))
            similar_ids_str = row.get('Similar_Ticket_IDs', '')
            similarity = row.get('Best_Match_Similarity', 0)
            category = row.get('AI_Category', 'Unknown')
            
            # Get actual resolution time if available
            actual_days = self._calculate_days_to_resolution(row)
            
            if similar_ids_str:
                # Parse similar ticket IDs (may include resolution days)
                for similar_info in similar_ids_str.split(', '):
                    # Extract ID (before any parentheses)
                    similar_id = similar_info.split('(')[0].strip()
                    
                    # Check existing feedback
                    feedback_key = f"{ticket_id}|{similar_id}"
                    existing = self.feedback_data.get(feedback_key, {})
                    existing_feedback = ""
                    if existing:
                        existing_feedback = "Similar" if existing.get('is_similar') else "Not Similar"
                    
                    feedback_rows.append({
                        'Ticket_ID': ticket_id,
                        'Ticket_Summary': str(row.get(COL_SUMMARY, ''))[:100],
                        'Category': category,
                        'Similar_Ticket_ID': similar_id,
                        'AI_Similarity': f"{similarity:.0%}",
                        'Is_Similar': existing_feedback,  # DROPDOWN: Similar / Not Similar
                        'Actual_Resolution_Days': actual_days if actual_days else '',
                        'Expected_Resolution_Days': '',  # Human input for expected time
                        'Notes': existing.get('notes', '')
                    })
        
        feedback_df = pd.DataFrame(feedback_rows)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Instructions sheet
            instructions = pd.DataFrame({
                'Instructions': [
                    'SIMILARITY FEEDBACK & RESOLUTION TIME INPUT',
                    '',
                    '📋 SIMILARITY FEEDBACK:',
                    '1. Review each ticket pair in the "Similarity Feedback" sheet',
                    '2. In "Is_Similar" column, use DROPDOWN to select:',
                    '   - "Similar" if the tickets ARE actually similar',
                    '   - "Not Similar" if the tickets are NOT related',
                    '3. AI learns from your feedback on next run!',
                    '',
                    '⏱️ RESOLUTION TIME INPUT:',
                    '1. Review "Actual_Resolution_Days" (from data)',
                    '2. Enter YOUR expected resolution time in "Expected_Resolution_Days"',
                    '3. This helps calibrate AI predictions vs human expectations',
                    '',
                    'FEEDBACK EFFECTS:',
                    '  - "Similar" → boosts similarity score (+10%)',
                    '  - "Not Similar" → penalizes false matches (-30%)',
                    '',
                    f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
                ]
            })
            instructions.to_excel(writer, sheet_name='Instructions', index=False)
            feedback_df.to_excel(writer, sheet_name='Similarity Feedback', index=False)
            
            # Apply dropdown validation to the Is_Similar column
            workbook = writer.book
            ws = writer.sheets['Similarity Feedback']
            
            # Create dropdown for Is_Similar column (Column F, starting row 2)
            similarity_validation = DataValidation(
                type='list',
                formula1='"Similar,Not Similar"',
                allow_blank=True,
                showDropDown=False
            )
            similarity_validation.error = 'Please select Similar or Not Similar'
            similarity_validation.errorTitle = 'Invalid Selection'
            similarity_validation.prompt = 'Select Similar or Not Similar'
            similarity_validation.promptTitle = 'Similarity Feedback'
            
            # Apply to column F (Is_Similar) for all data rows
            if len(feedback_df) > 0:
                similarity_validation.add(f'F2:F{len(feedback_df) + 1}')
                ws.add_data_validation(similarity_validation)
            
            # Set column widths
            ws.column_dimensions['A'].width = 15  # Ticket_ID
            ws.column_dimensions['B'].width = 60  # Ticket_Summary
            ws.column_dimensions['C'].width = 25  # Category
            ws.column_dimensions['D'].width = 18  # Similar_Ticket_ID
            ws.column_dimensions['E'].width = 14  # AI_Similarity
            ws.column_dimensions['F'].width = 15  # Is_Similar (dropdown)
            ws.column_dimensions['G'].width = 20  # Actual_Resolution_Days
            ws.column_dimensions['H'].width = 22  # Expected_Resolution_Days
            ws.column_dimensions['I'].width = 30  # Notes
            
            # Style the header row
            for col_idx in range(1, 10):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
            
            # Highlight the dropdown column
            for row_idx in range(2, len(feedback_df) + 2):
                cell = ws.cell(row=row_idx, column=6)  # Is_Similar column
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                # Expected_Resolution_Days column - light blue for input
                exp_cell = ws.cell(row=row_idx, column=8)
                exp_cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
        
        logger.info(f"[Similar Ticket Finder] Feedback file exported to {output_path}")
        return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# 4C. RESOLUTION TIME PREDICTOR - ML-based prediction
# ═══════════════════════════════════════════════════════════════════════════════

class ResolutionTimePredictor:
    """
    ML-based Resolution Time Predictor.
    
    Predicts resolution time for tickets based on:
    - Historical resolution times by category
    - Issue severity and complexity indicators
    - Similar ticket resolution patterns
    - Human-provided expected times (for calibration)
    
    Provides three metrics:
    - Actual: Real resolution time from data
    - Predicted: ML-predicted resolution time
    - Expected: Human-provided expectation (from feedback)
    """
    
    def __init__(self):
        self.model = None
        self.feature_columns = []
        self.category_stats = {}  # category -> {mean, median, std, count}
        self.severity_stats = {}  # severity -> {mean, median, std}
        self.human_expectations = {}  # category -> expected days from human input
        self.is_trained = False
        
        logger.info("[Resolution Predictor] Initialized")
    
    def _extract_features(self, row):
        """Extract features for resolution time prediction."""
        features = {}
        
        # Category encoding (one-hot would be better but simple for now)
        category = str(row.get('AI_Category', 'Unknown'))
        features['category_hash'] = hash(category) % 1000
        
        # Use category average if available
        if category in self.category_stats:
            features['category_avg_days'] = self.category_stats[category]['mean']
            features['category_median_days'] = self.category_stats[category]['median']
        else:
            features['category_avg_days'] = 5.0  # default
            features['category_median_days'] = 3.0
        
        # Severity indicator
        severity = str(row.get('AI_Severity', row.get(COL_SEVERITY, 'Medium')))
        severity_map = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1}
        features['severity_level'] = severity_map.get(severity, 2)
        
        # Complexity indicators from text
        summary = str(row.get(COL_SUMMARY, ''))
        features['text_length'] = len(summary)
        features['word_count'] = len(summary.split())
        
        # Complexity keywords
        complexity_keywords = ['complex', 'multiple', 'integration', 'migration', 'upgrade', 'critical']
        features['complexity_score'] = sum(1 for kw in complexity_keywords if kw in summary.lower())
        
        # Similar ticket resolution if available
        features['similar_resolution_days'] = row.get('Expected_Resolution_Days', 0) or 0
        features['similar_count'] = row.get('Similar_Ticket_Count', 0) or 0
        
        # Confidence from AI classification
        features['ai_confidence'] = row.get('AI_Confidence', 0.5) or 0.5
        
        # Recurrence risk as feature
        recurrence = str(row.get('AI_Recurrence_Risk', 'Medium'))
        recurrence_map = {'High': 3, 'Medium': 2, 'Low': 1, 'Unknown': 2}
        features['recurrence_risk'] = recurrence_map.get(recurrence, 2)
        
        return features
    
    def train(self, df):
        """
        Train the resolution time predictor on historical data.
        
        Args:
            df: DataFrame with historical tickets including resolution times
        """
        logger.info("[Resolution Predictor] Training model...")
        
        # Calculate actual resolution times
        training_data = []
        
        for idx, row in df.iterrows():
            # Get actual resolution time
            actual_days = self._calculate_resolution_days(row)
            if actual_days is None or actual_days <= 0 or actual_days > 365:  # Filter outliers
                continue
            
            features = self._extract_features(row)
            features['actual_days'] = actual_days
            training_data.append(features)
        
        if len(training_data) < 10:
            logger.warning("[Resolution Predictor] Insufficient training data (< 10 samples)")
            self.is_trained = False
            return False
        
        train_df = pd.DataFrame(training_data)
        
        # Build category statistics
        for category in df['AI_Category'].dropna().unique():
            cat_data = [t['actual_days'] for t in training_data 
                       if t.get('category_hash') == hash(category) % 1000]
            if cat_data:
                self.category_stats[category] = {
                    'mean': np.mean(cat_data),
                    'median': np.median(cat_data),
                    'std': np.std(cat_data) if len(cat_data) > 1 else 0,
                    'count': len(cat_data)
                }
        
        # Feature matrix
        self.feature_columns = [c for c in train_df.columns if c != 'actual_days']
        X = train_df[self.feature_columns].values
        y = train_df['actual_days'].values
        
        # Train Random Forest Regressor
        try:
            from sklearn.ensemble import RandomForestRegressor
            
            self.model = RandomForestRegressor(
                n_estimators=50,
                max_depth=8,
                min_samples_leaf=3,
                random_state=42,
                n_jobs=-1
            )
            self.model.fit(X, y)
            self.is_trained = True
            
            # Calculate training metrics
            predictions = self.model.predict(X)
            mae = np.mean(np.abs(predictions - y))
            rmse = np.sqrt(np.mean((predictions - y) ** 2))
            
            logger.info(f"[Resolution Predictor] Model trained on {len(training_data)} samples")
            logger.info(f"[Resolution Predictor] Training MAE: {mae:.2f} days, RMSE: {rmse:.2f} days")
            
            return True
            
        except Exception as e:
            logger.error(f"[Resolution Predictor] Training failed: {e}")
            self.is_trained = False
            return False
    
    def _calculate_resolution_days(self, row):
        """Calculate actual resolution days from ticket data."""
        try:
            # Try to get issue date and resolution date
            issue_date = row.get(COL_DATETIME)
            resolution_date = row.get(COL_RESOLUTION_DATE)
            
            if pd.isna(issue_date) or pd.isna(resolution_date):
                return None
            
            # Convert to datetime if needed
            if isinstance(issue_date, str):
                issue_date = pd.to_datetime(issue_date)
            if isinstance(resolution_date, str):
                resolution_date = pd.to_datetime(resolution_date)
            
            days = (resolution_date - issue_date).days
            return days if days >= 0 else None
            
        except Exception:
            return None
    
    def predict(self, row):
        """
        Predict resolution time for a single ticket.
        
        Returns:
            dict with:
            - predicted_days: ML predicted resolution time
            - confidence: prediction confidence (0-1)
            - method: 'ml' or 'fallback'
        """
        category = str(row.get('AI_Category', 'Unknown'))
        
        # ML prediction if model is trained
        if self.is_trained and self.model is not None:
            try:
                features = self._extract_features(row)
                X = np.array([[features[c] for c in self.feature_columns]])
                predicted = self.model.predict(X)[0]
                
                # Get prediction interval (approximate using category std)
                cat_std = self.category_stats.get(category, {}).get('std', 2.0)
                confidence = max(0.3, min(0.95, 1 - (cat_std / (predicted + 1))))
                
                return {
                    'predicted_days': max(0.5, round(predicted, 1)),
                    'confidence': confidence,
                    'method': 'ml'
                }
            except Exception as e:
                logger.debug(f"[Resolution Predictor] ML prediction failed: {e}")
        
        # Fallback to category statistics
        if category in self.category_stats:
            return {
                'predicted_days': round(self.category_stats[category]['median'], 1),
                'confidence': 0.5,
                'method': 'category_stats'
            }
        
        # Default fallback
        return {
            'predicted_days': 5.0,
            'confidence': 0.2,
            'method': 'default'
        }
    
    def set_human_expectations(self, expectations_dict):
        """
        Set human-provided expected resolution times by category.
        
        Args:
            expectations_dict: {category: expected_days}
        """
        self.human_expectations = expectations_dict
        logger.info(f"[Resolution Predictor] Loaded {len(expectations_dict)} human expectations")
    
    def process_all_tickets(self, df):
        """
        Process all tickets and add resolution time columns.
        
        Adds:
        - Actual_Resolution_Days: Real resolution time from data
        - Predicted_Resolution_Days: ML predicted time
        - Expected_Resolution_Days: Human expectation (if available)
        - Resolution_Prediction_Confidence: Confidence in ML prediction
        """
        logger.info(f"[Resolution Predictor] Processing {len(df)} tickets...")
        
        # First, train on the data
        self.train(df)
        
        # Initialize columns
        df['Actual_Resolution_Days'] = None
        df['Predicted_Resolution_Days'] = None
        df['Human_Expected_Days'] = None
        df['Resolution_Prediction_Confidence'] = None
        df['Resolution_Prediction_Method'] = None
        
        for idx, row in df.iterrows():
            # Actual resolution time
            actual = self._calculate_resolution_days(row)
            df.at[idx, 'Actual_Resolution_Days'] = actual
            
            # ML Prediction
            prediction = self.predict(row)
            df.at[idx, 'Predicted_Resolution_Days'] = prediction['predicted_days']
            df.at[idx, 'Resolution_Prediction_Confidence'] = prediction['confidence']
            df.at[idx, 'Resolution_Prediction_Method'] = prediction['method']
            
            # Human expectation by category
            category = str(row.get('AI_Category', 'Unknown'))
            if category in self.human_expectations:
                df.at[idx, 'Human_Expected_Days'] = self.human_expectations[category]
        
        # Summary stats
        has_actual = df['Actual_Resolution_Days'].notna().sum()
        has_predicted = df['Predicted_Resolution_Days'].notna().sum()
        has_expected = df['Human_Expected_Days'].notna().sum()
        
        logger.info(f"[Resolution Predictor] Complete:")
        logger.info(f"  → {has_actual} tickets with actual resolution times")
        logger.info(f"  → {has_predicted} tickets with ML predictions")
        logger.info(f"  → {has_expected} tickets with human expectations")
        
        return df
    
    def get_accuracy_metrics(self, df):
        """
        Calculate accuracy metrics comparing actual vs predicted.
        
        Returns:
            dict with MAE, RMSE, and per-category metrics
        """
        valid = df.dropna(subset=['Actual_Resolution_Days', 'Predicted_Resolution_Days'])
        
        if len(valid) < 5:
            return None
        
        actual = valid['Actual_Resolution_Days'].values
        predicted = valid['Predicted_Resolution_Days'].values
        
        metrics = {
            'mae': np.mean(np.abs(predicted - actual)),
            'rmse': np.sqrt(np.mean((predicted - actual) ** 2)),
            'mape': np.mean(np.abs((actual - predicted) / (actual + 0.1))) * 100,
            'sample_count': len(valid),
            'correlation': np.corrcoef(actual, predicted)[0, 1] if len(valid) > 2 else 0
        }
        
        # Per-category metrics
        metrics['by_category'] = {}
        for category in valid['AI_Category'].dropna().unique():
            cat_data = valid[valid['AI_Category'] == category]
            if len(cat_data) >= 3:
                cat_actual = cat_data['Actual_Resolution_Days'].values
                cat_pred = cat_data['Predicted_Resolution_Days'].values
                metrics['by_category'][category] = {
                    'mae': np.mean(np.abs(cat_pred - cat_actual)),
                    'count': len(cat_data)
                }
        
        return metrics


# Global resolution time predictor instance
resolution_time_predictor = None


# Global similar ticket finder instance
similar_ticket_finder = None


def apply_similar_ticket_analysis(df, ai_engine=None):
    """
    Apply similar ticket analysis to the dataframe.
    
    Args:
        df: DataFrame with ticket data
        ai_engine: AI engine instance for embeddings
        
    Returns:
        DataFrame with similarity columns added
    """
    global similar_ticket_finder
    
    logger.info("[Similar Ticket Finder] Initializing similar ticket analysis...")
    
    # Initialize finder with AI engine
    similar_ticket_finder = SimilarTicketFinder(
        ai_engine=ai_engine,
        top_k=5,
        similarity_threshold=0.70  # 70% similarity minimum
    )
    
    # Try to load existing feedback from previous runs
    similarity_feedback_excel = "similarity_feedback.xlsx"
    if os.path.exists(similarity_feedback_excel):
        similar_ticket_finder.load_feedback_from_excel(similarity_feedback_excel)
    
    # Process all tickets
    df = similar_ticket_finder.process_all_tickets(df)
    
    # Export feedback file for human review
    try:
        similar_ticket_finder.export_for_feedback(df, similarity_feedback_excel)
    except Exception as e:
        logger.warning(f"Could not export similarity feedback file: {e}")
    
    return df


def apply_resolution_time_prediction(df, human_expectations=None):
    """
    Apply ML-based resolution time prediction to the dataframe.
    
    Args:
        df: DataFrame with ticket data
        human_expectations: dict of {category: expected_days} from human feedback
        
    Returns:
        DataFrame with resolution time columns added:
        - Actual_Resolution_Days: Real resolution time from data
        - Predicted_Resolution_Days: ML predicted time
        - Human_Expected_Days: Human expectation (if available)
        - Resolution_Prediction_Confidence: Confidence in prediction
    """
    global resolution_time_predictor
    
    logger.info("[Resolution Predictor] Initializing resolution time prediction...")
    
    # Initialize predictor
    resolution_time_predictor = ResolutionTimePredictor()
    
    # Set human expectations if provided
    if human_expectations:
        resolution_time_predictor.set_human_expectations(human_expectations)
    
    # Process all tickets (includes training and prediction)
    df = resolution_time_predictor.process_all_tickets(df)
    
    # Log accuracy metrics
    metrics = resolution_time_predictor.get_accuracy_metrics(df)
    if metrics:
        logger.info(f"[Resolution Predictor] Accuracy Metrics:")
        logger.info(f"  → MAE: {metrics['mae']:.2f} days")
        logger.info(f"  → RMSE: {metrics['rmse']:.2f} days")
        logger.info(f"  → Correlation: {metrics['correlation']:.2f}")
    
    return df


# ==========================================
# 5. VISUALIZATION
# ==========================================
def generate_plots(df, output_dir):
    logger.info("[Vis Engine] Generating 14 Executive Charts...")
    plot_dir = os.path.join(output_dir, "plots")
    os.makedirs(plot_dir, exist_ok=True)
    paths: List[Optional[str]] = [None] * 15  # 15 charts total
   
    # Global Style Settings for "McKinsey Look"
    sns.set_theme(style="whitegrid")
    plt.rcParams.update({'font.size': 10, 'figure.dpi': 100}) # Low DPI prevents massive images
   
    # --- CHART 1: FRICTION PARETO (Top 10 by AI Category) ---
    try:
        plt.figure(figsize=(8, 5))
        risk_df = df[df['Strategic_Friction_Score'] > 0]
        if not risk_df.empty and 'AI_Category' in df.columns:
            cat_scores = risk_df.groupby('AI_Category')['Strategic_Friction_Score'].sum().nlargest(10)
            sns.barplot(x=cat_scores.values, y=cat_scores.index, palette="Reds_r")
            plt.title('Top Strategic Friction Sources (AI Classified)', fontweight='bold')
            plt.xlabel('Weighted Risk Score')
            plt.tight_layout()
            p1 = os.path.join(plot_dir, "friction.png")
            plt.savefig(p1)
            plt.close()
            paths[0] = p1
    except Exception as e: 
        logger.warning(f"Chart 1 Skipped: {e}")
        plt.close()

    # --- CHART 2: RISK ORIGIN (External vs Internal) ---
    try:
        plt.figure(figsize=(6, 4))
        if COL_ORIGIN in df.columns:
            ext = df[df[COL_ORIGIN]=='External']['Strategic_Friction_Score'].sum()
            int_ = df[df[COL_ORIGIN]=='Internal']['Strategic_Friction_Score'].sum()
            if ext + int_ > 0:
                plt.pie([ext, int_], labels=['External', 'Internal'], colors=['#D9534F', MC_BLUE],
                        autopct='%1.1f%%', startangle=90)
                plt.title('Risk Origin (Weighted)', fontweight='bold')
                plt.tight_layout()
                p2 = os.path.join(plot_dir, "origin.png")
                plt.savefig(p2)
                plt.close()
                paths[1] = p2
    except Exception as e:
        logger.warning(f"Chart 2 Skipped: {e}")
        plt.close()

    # --- CHART 3: LEARNING INTEGRITY (Recidivism - Tiered) ---
    try:
        plt.figure(figsize=(7, 4))
        # Count by tier
        high_conf = df['Learning_Status'].astype(str).str.contains('REPEAT OFFENSE').sum()
        medium_conf = df['Learning_Status'].astype(str).str.contains('POSSIBLE REPEAT').sum()
        new_issues = len(df) - high_conf - medium_conf
        
        # Only show meaningful segments
        labels = []
        sizes = []
        colors = []
        
        if new_issues > 0:
            labels.append(f'New Issues ({new_issues})')
            sizes.append(new_issues)
            colors.append('#5CB85C')  # Green
        if high_conf > 0:
            labels.append(f'Confirmed Repeats ({high_conf})')
            sizes.append(high_conf)
            colors.append('#D9534F')  # Red
        if medium_conf > 0:
            labels.append(f'Possible Repeats ({medium_conf})')
            sizes.append(medium_conf)
            colors.append('#F0AD4E')  # Orange
            
        if sizes:
            plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=60)
            plt.title('Institutional Learning Integrity', fontweight='bold')
            plt.tight_layout()
            p3 = os.path.join(plot_dir, "learning.png")
            plt.savefig(p3)
            plt.close()
            paths[2] = p3
    except Exception as e:
        logger.warning(f"Chart 3 Skipped: {e}")
        plt.close()

    # --- CHART 4: 7-DAY RISK TREND ---
    try:
        plt.figure(figsize=(10, 4))
        if COL_DATETIME in df.columns:
            df_temp = df.copy()
            df_temp['Date'] = pd.to_datetime(df_temp[COL_DATETIME], errors='coerce')
            daily = df_temp.groupby('Date')['Strategic_Friction_Score'].sum().rolling(7, min_periods=1).mean()
            if not daily.empty:
                sns.lineplot(data=daily, color=MC_BLUE, linewidth=2.5)
                plt.fill_between(daily.index, daily.values, alpha=0.1, color=MC_BLUE)
                plt.title('7-Day Rolling Risk Trend', fontweight='bold')
                plt.ylabel('Avg Daily Friction')
                plt.tight_layout()
                p4 = os.path.join(plot_dir, "trend.png")
                plt.savefig(p4)
                plt.close()
                paths[3] = p4
    except Exception as e:
        logger.warning(f"Chart 4 Skipped: {e}")
        plt.close()

    # --- CHART 5: CATEGORY × SEVERITY HEATMAP ---
    try:
        plt.figure(figsize=(10, 6))
        if 'AI_Category' in df.columns and 'Severity_Norm' in df.columns:
            # Create pivot table of counts
            heatmap_data = pd.crosstab(
                df['AI_Category'], 
                df['Severity_Norm'],
                values=df['Strategic_Friction_Score'],
                aggfunc='sum'
            ).fillna(0)
            
            # Reorder severity columns
            severity_order = ['Critical', 'Major', 'Minor', 'Default']
            existing_cols = [c for c in severity_order if c in heatmap_data.columns]
            heatmap_data = heatmap_data[existing_cols]
            
            if not heatmap_data.empty:
                sns.heatmap(heatmap_data, annot=True, fmt='.0f', cmap='YlOrRd',
                           linewidths=0.5, cbar_kws={'label': 'Friction Score'})
                plt.title('Risk Heatmap: Category × Severity', fontweight='bold')
                plt.xlabel('Severity Level')
                plt.ylabel('AI Category')
                plt.tight_layout()
                p5 = os.path.join(plot_dir, "heatmap.png")
                plt.savefig(p5)
                plt.close()
                paths[4] = p5
    except Exception as e:
        logger.warning(f"Chart 5 Skipped: {e}")
        plt.close()

    # --- CHART 6: TOP ENGINEERS BY FRICTION (Human Error Accountability) ---
    try:
        plt.figure(figsize=(9, 5))
        if 'Engineer' in df.columns and 'Strategic_Friction_Score' in df.columns:
            # Filter to only escalations/concerns for human error tracking
            df_issues = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])] if 'Type_Norm' in df.columns else df
            
            if not df_issues.empty:
                engineer_friction = df_issues.groupby('Engineer').agg({
                    'Strategic_Friction_Score': 'sum',
                    'Engineer_Is_Repeat_Offender': 'first'
                }).reset_index()
                engineer_friction = engineer_friction.nlargest(10, 'Strategic_Friction_Score')
                
                # Color based on repeat offender status
                colors = ['#D9534F' if row['Engineer_Is_Repeat_Offender'] else MC_BLUE 
                         for _, row in engineer_friction.iterrows()]
                
                bars = plt.barh(engineer_friction['Engineer'], 
                               engineer_friction['Strategic_Friction_Score'],
                               color=colors)
                
                plt.xlabel('Total Friction Score')
                plt.title('Top 10 Engineers by Friction Contribution', fontweight='bold')
                
                # Add legend
                from matplotlib.patches import Patch
                legend_elements = [Patch(facecolor='#D9534F', label='Repeat Offender'),
                                  Patch(facecolor=MC_BLUE, label='Normal')]
                plt.legend(handles=legend_elements, loc='lower right')
                
                plt.tight_layout()
                p6 = os.path.join(plot_dir, "engineer_friction.png")
                plt.savefig(p6)
                plt.close()
                paths[5] = p6
    except Exception as e:
        logger.warning(f"Chart 6 Skipped: {e}")
        plt.close()

    # --- CHART 7: ENGINEER LEARNING BEHAVIOR (Learning Gap Analysis) ---
    try:
        plt.figure(figsize=(8, 5))
        if 'Engineer_Learning_Score' in df.columns and 'Engineer' in df.columns:
            df_issues = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])] if 'Type_Norm' in df.columns else df
            
            if not df_issues.empty:
                # Get unique engineers with their learning metrics
                engineer_learning = df_issues.groupby('Engineer').agg({
                    'Engineer_Learning_Score': 'first',
                    'Engineer_Issue_Count': 'first' if 'Engineer_Issue_Count' in df_issues.columns else 'size',
                    'Engineer_Lessons_Logged': 'first',
                    'Engineer_Is_Repeat_Offender': 'first'
                }).reset_index()
                
                # Filter to engineers with issues
                engineer_learning = engineer_learning[engineer_learning.get('Engineer_Issue_Count', 1) > 0]
                
                if len(engineer_learning) > 0:
                    # Create categories based on learning behavior
                    learning_categories = {
                        'Active Learners (Score ≥ 0.8)': len(engineer_learning[engineer_learning['Engineer_Learning_Score'] >= 0.8]),
                        'Moderate Learning (0.4-0.8)': len(engineer_learning[(engineer_learning['Engineer_Learning_Score'] >= 0.4) & 
                                                                              (engineer_learning['Engineer_Learning_Score'] < 0.8)]),
                        'Low Learning (0.1-0.4)': len(engineer_learning[(engineer_learning['Engineer_Learning_Score'] >= 0.1) & 
                                                                        (engineer_learning['Engineer_Learning_Score'] < 0.4)]),
                        'No Lessons Logged (0)': len(engineer_learning[engineer_learning['Engineer_Learning_Score'] == 0])
                    }
                    
                    # Filter out zero values
                    learning_categories = {k: v for k, v in learning_categories.items() if v > 0}
                    
                    if learning_categories:
                        colors_learning = ['#5CB85C', '#5BC0DE', '#F0AD4E', '#D9534F'][:len(learning_categories)]
                        plt.pie(list(learning_categories.values()), 
                               labels=list(learning_categories.keys()),
                               colors=colors_learning,
                               autopct='%1.1f%%', 
                               startangle=90)
                        plt.title('Engineer Learning Behavior Distribution', fontweight='bold')
                        plt.tight_layout()
                        p7 = os.path.join(plot_dir, "engineer_learning.png")
                        plt.savefig(p7)
                        plt.close()
                        paths[6] = p7
    except Exception as e:
        logger.warning(f"Chart 7 Skipped: {e}")
        plt.close()

    # =====================================================
    # LOB (LINE OF BUSINESS) STRATEGIC CHARTS
    # McKinsey-Style Organizational Performance Visualization
    # =====================================================
    
    # --- CHART 8: LOB FRICTION WATERFALL (Strategic Priority) ---
    try:
        plt.figure(figsize=(10, 6))
        if 'LOB' in df.columns and 'Strategic_Friction_Score' in df.columns:
            df_issues = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])] if 'Type_Norm' in df.columns else df
            
            if not df_issues.empty:
                lob_friction = df_issues.groupby('LOB').agg({
                    'Strategic_Friction_Score': 'sum',
                    'LOB_Risk_Tier': 'first'
                }).reset_index()
                lob_friction = lob_friction.nlargest(10, 'Strategic_Friction_Score')
                
                # Color by risk tier
                tier_colors = {
                    'Critical': '#D9534F',
                    'High': '#F0AD4E', 
                    'Medium': '#5BC0DE',
                    'Low': '#5CB85C'
                }
                colors = [tier_colors.get(row['LOB_Risk_Tier'], MC_BLUE) for _, row in lob_friction.iterrows()]
                
                # Horizontal bar chart (waterfall style)
                bars = plt.barh(lob_friction['LOB'], lob_friction['Strategic_Friction_Score'], color=colors)
                
                # Add value labels
                for bar, val in zip(bars, lob_friction['Strategic_Friction_Score']):
                    plt.text(val + 50, bar.get_y() + bar.get_height()/2, 
                            f'{val:,.0f}', va='center', fontsize=8)
                
                plt.xlabel('Total Strategic Friction')
                plt.title('LOB Risk Exposure (by Friction)', fontweight='bold')
                
                # Legend for risk tiers
                from matplotlib.patches import Patch
                legend_elements = [
                    Patch(facecolor='#D9534F', label='Critical'),
                    Patch(facecolor='#F0AD4E', label='High'),
                    Patch(facecolor='#5BC0DE', label='Medium'),
                    Patch(facecolor='#5CB85C', label='Low')
                ]
                plt.legend(handles=legend_elements, title='Risk Tier', loc='lower right', fontsize=8)
                
                plt.tight_layout()
                p8 = os.path.join(plot_dir, "lob_friction.png")
                plt.savefig(p8)
                plt.close()
                paths[7] = p8
    except Exception as e:
        logger.warning(f"Chart 8 Skipped: {e}")
        plt.close()

    # --- CHART 9: LOB EFFICIENCY MATRIX (Bubble Chart - Strategic View) ---
    try:
        plt.figure(figsize=(10, 7))
        if 'LOB' in df.columns and 'LOB_Efficiency_Score' in df.columns:
            df_issues = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])] if 'Type_Norm' in df.columns else df
            
            if not df_issues.empty:
                # Get unique LOB metrics
                lob_metrics = df_issues.groupby('LOB').agg({
                    'LOB_Efficiency_Score': 'first',
                    'LOB_Learning_Rate': 'first',
                    'LOB_Total_Friction': 'first',
                    'LOB_Risk_Tier': 'first',
                    'LOB_Issue_Count': 'first'
                }).reset_index()
                
                # Filter out unknown and small LOBs
                lob_metrics = lob_metrics[lob_metrics['LOB'] != 'Unknown']
                lob_metrics = lob_metrics[lob_metrics['LOB_Issue_Count'] >= 2]
                
                if len(lob_metrics) > 0:
                    # Bubble chart: X = Efficiency, Y = Learning Rate, Size = Friction, Color = Risk
                    tier_colors = {
                        'Critical': '#D9534F',
                        'High': '#F0AD4E', 
                        'Medium': '#5BC0DE',
                        'Low': '#5CB85C'
                    }
                    colors = [tier_colors.get(tier, MC_BLUE) for tier in lob_metrics['LOB_Risk_Tier']]
                    
                    # Normalize bubble sizes
                    max_friction = lob_metrics['LOB_Total_Friction'].max()
                    sizes = (lob_metrics['LOB_Total_Friction'] / max_friction * 1000) + 100
                    
                    scatter = plt.scatter(
                        lob_metrics['LOB_Efficiency_Score'],
                        lob_metrics['LOB_Learning_Rate'] * 100,  # Convert to percentage
                        s=sizes,
                        c=colors,
                        alpha=0.6,
                        edgecolors='black',
                        linewidth=1
                    )
                    
                    # Add LOB labels
                    for _, row in lob_metrics.iterrows():
                        plt.annotate(
                            row['LOB'][:15],  # Truncate long names
                            (row['LOB_Efficiency_Score'], row['LOB_Learning_Rate'] * 100),
                            xytext=(5, 5), textcoords='offset points',
                            fontsize=8, alpha=0.8
                        )
                    
                    plt.xlabel('Operational Efficiency Score (0-100)')
                    plt.ylabel('Learning Rate (%)')
                    plt.title('LOB Strategic Matrix: Efficiency vs Learning Culture', fontweight='bold')
                    
                    # Add quadrant lines
                    plt.axhline(y=50, color='gray', linestyle='--', alpha=0.5)
                    plt.axvline(x=50, color='gray', linestyle='--', alpha=0.5)
                    
                    # Quadrant labels
                    plt.text(75, 75, 'LEADERS', fontsize=10, fontweight='bold', 
                            color='green', alpha=0.7, ha='center')
                    plt.text(25, 75, 'LEARNERS', fontsize=10, fontweight='bold', 
                            color='blue', alpha=0.7, ha='center')
                    plt.text(75, 25, 'EFFICIENT', fontsize=10, fontweight='bold', 
                            color='orange', alpha=0.7, ha='center')
                    plt.text(25, 25, 'AT RISK', fontsize=10, fontweight='bold', 
                            color='red', alpha=0.7, ha='center')
                    
                    plt.xlim(0, 100)
                    plt.ylim(0, 100)
                    plt.tight_layout()
                    p9 = os.path.join(plot_dir, "lob_matrix.png")
                    plt.savefig(p9)
                    plt.close()
                    paths[8] = p9
    except Exception as e:
        logger.warning(f"Chart 9 Skipped: {e}")
        plt.close()

    # --- CHART 10: LOB ISSUE CATEGORY BREAKDOWN (Stacked Bar - Root Cause Analysis) ---
    try:
        plt.figure(figsize=(12, 6))
        if 'LOB' in df.columns and 'AI_Category' in df.columns:
            df_issues = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])] if 'Type_Norm' in df.columns else df
            
            if not df_issues.empty:
                # Create pivot table
                lob_category = pd.crosstab(df_issues['LOB'], df_issues['AI_Category'])
                
                # Filter to top 8 LOBs by total issues
                lob_totals = lob_category.sum(axis=1).nlargest(8)
                lob_category = lob_category.loc[lob_totals.index]
                
                # Filter to top 6 categories for readability
                cat_totals = lob_category.sum(axis=0).nlargest(6)
                lob_category = lob_category[cat_totals.index]
                
                if not lob_category.empty:
                    # McKinsey color palette
                    mckinsey_colors = ['#004C97', '#0078D4', '#00A6ED', '#6CC24A', 
                                       '#F0AD4E', '#D9534F'][:len(lob_category.columns)]
                    
                    lob_category.plot(kind='barh', stacked=True, color=mckinsey_colors, 
                                     figsize=(12, 6), width=0.7)
                    
                    plt.xlabel('Number of Issues')
                    plt.ylabel('Line of Business')
                    plt.title('LOB Issue Breakdown by Category (Root Cause Analysis)', fontweight='bold')
                    plt.legend(title='Issue Category', bbox_to_anchor=(1.02, 1), loc='upper left', fontsize=8)
                    
                    plt.tight_layout()
                    p10 = os.path.join(plot_dir, "lob_categories.png")
                    plt.savefig(p10)
                    plt.close()
                    paths[9] = p10
    except Exception as e:
        logger.warning(f"Chart 10 Skipped: {e}")
        plt.close()

    # =====================================================
    # ROOT CAUSE & PM PREDICTION CHARTS
    # =====================================================
    
    # --- CHART 11: ROOT CAUSE BREAKDOWN (Donut Chart) ---
    try:
        plt.figure(figsize=(9, 6))
        if 'Root_Cause_Category' in df.columns:
            df_issues = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])] if 'Type_Norm' in df.columns else df
            
            if not df_issues.empty:
                root_cause_counts = df_issues['Root_Cause_Category'].value_counts()
                
                # Filter out small categories for readability
                root_cause_counts = root_cause_counts[root_cause_counts >= 1]
                
                if not root_cause_counts.empty:
                    # Color mapping for root causes
                    rc_colors = {
                        'Human Error': '#D9534F',
                        'External Party': '#5BC0DE',
                        'Process Gap': '#F0AD4E',
                        'System/Technical': '#5CB85C',
                        'Training Gap': '#9B59B6',
                        'Communication': '#3498DB',
                        'Resource': '#E74C3C',
                        'Other': '#95A5A6',
                        'Unclassified': '#BDC3C7'
                    }
                    colors = [rc_colors.get(cat, '#95A5A6') for cat in root_cause_counts.index]
                    
                    # Create donut chart - autopct ensures 3 returns
                    pie_result = plt.pie(
                        root_cause_counts.values,
                        labels=root_cause_counts.index,
                        colors=colors,
                        autopct='%1.1f%%',
                        startangle=90,
                        pctdistance=0.75
                    )
                    wedges = pie_result[0]
                    
                    # Draw center circle for donut effect
                    from matplotlib.patches import Circle
                    centre_circle = Circle((0, 0), 0.50, fc='white')
                    plt.gca().add_patch(centre_circle)
                    
                    # Center text
                    plt.text(0, 0, f'{len(df_issues)}\nIssues', ha='center', va='center', 
                            fontsize=14, fontweight='bold')
                    
                    plt.title('Root Cause Analysis (PM Reported)', fontweight='bold')
                    plt.tight_layout()
                    p11 = os.path.join(plot_dir, "root_cause.png")
                    plt.savefig(p11)
                    plt.close()
                    paths[10] = p11
    except Exception as e:
        logger.warning(f"Chart 11 Skipped: {e}")
        plt.close()

    # --- CHART 12: PM PREDICTION ACCURACY (Confusion Matrix Style) ---
    try:
        plt.figure(figsize=(10, 6))
        if 'PM_Prediction_Accuracy' in df.columns and 'PM_Recurrence_Risk_Norm' in df.columns:
            df_with_predictions = df[df['PM_Recurrence_Risk_Norm'] != 'Unknown']
            
            if not df_with_predictions.empty:
                # Get accuracy distribution
                accuracy_counts = df_with_predictions['PM_Prediction_Accuracy'].value_counts()
                
                # Define order and colors
                accuracy_order = [
                    '✅ Correct (High→Recurred)',
                    '✅ Correct (Low→No Recurrence)',
                    '⚪ Moderate (Medium Risk)',
                    '🟡 Overestimate (High→No Recurrence)',
                    '🔴 MISSED (Low→Recurred!)',
                    'Inconclusive',
                    'No Prediction'
                ]
                
                # Reorder based on what exists
                ordered_cats = [cat for cat in accuracy_order if cat in accuracy_counts.index]
                accuracy_ordered = accuracy_counts.reindex(ordered_cats).dropna()
                
                if not accuracy_ordered.empty:
                    # Color mapping
                    acc_colors = {
                        '✅ Correct (High→Recurred)': '#5CB85C',
                        '✅ Correct (Low→No Recurrence)': '#5CB85C',
                        '⚪ Moderate (Medium Risk)': '#5BC0DE',
                        '🟡 Overestimate (High→No Recurrence)': '#F0AD4E',
                        '🔴 MISSED (Low→Recurred!)': '#D9534F',
                        'Inconclusive': '#95A5A6',
                        'No Prediction': '#BDC3C7'
                    }
                    colors = [acc_colors.get(cat, '#95A5A6') for cat in accuracy_ordered.index]
                    
                    bars = plt.barh(range(len(accuracy_ordered)), accuracy_ordered.values, color=colors)
                    plt.yticks(range(len(accuracy_ordered)), accuracy_ordered.index)
                    
                    # Add value labels
                    for bar, val in zip(bars, accuracy_ordered.values):
                        plt.text(val + 0.5, bar.get_y() + bar.get_height()/2, 
                                f'{int(val)}', va='center', fontsize=9)
                    
                    # Calculate and display accuracy rate
                    total = len(df_with_predictions)
                    correct = accuracy_ordered.get('✅ Correct (High→Recurred)', 0) + \
                             accuracy_ordered.get('✅ Correct (Low→No Recurrence)', 0)
                    missed = accuracy_ordered.get('🔴 MISSED (Low→Recurred!)', 0)
                    
                    accuracy_pct = (correct / total * 100) if total > 0 else 0
                    
                    plt.xlabel('Number of Issues')
                    plt.title(f'PM Recurrence Prediction Accuracy ({accuracy_pct:.1f}% Correct)', fontweight='bold')
                    
                    # Add warning annotation if there are missed predictions
                    if missed > 0:
                        plt.annotate(
                            f'⚠ {missed} issues PM underestimated!',
                            xy=(0.98, 0.02), xycoords='axes fraction',
                            fontsize=10, color='#D9534F', fontweight='bold',
                            ha='right'
                        )
                    
                    plt.tight_layout()
                    p12 = os.path.join(plot_dir, "pm_accuracy.png")
                    plt.savefig(p12)
                    plt.close()
                    paths[11] = p12
    except Exception as e:
        logger.warning(f"Chart 12 Skipped: {e}")
        plt.close()

    # --- CHART 13: AI RECURRENCE PREDICTION DISTRIBUTION ---
    try:
        plt.figure(figsize=(10, 6))
        if 'AI_Recurrence_Probability' in df.columns:
            # Only show for issues (exclude lessons learned)
            df_issues = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])]
            
            if not df_issues.empty and df_issues['AI_Recurrence_Probability'].sum() > 0:
                # Create histogram with risk zones
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
                
                # Left: Probability Distribution Histogram
                probs = df_issues['AI_Recurrence_Probability'] * 100
                
                # Create bins for histogram
                bins = [0, 20, 30, 50, 70, 100]
                colors_hist = ['#5CB85C', '#8BC34A', '#F0AD4E', '#FF9800', '#D9534F']
                
                n, bin_edges, patches = ax1.hist(probs, bins=bins, edgecolor='white', linewidth=1.2)
                
                # Color each bin
                for patch, color in zip(patches, colors_hist):
                    patch.set_facecolor(color)
                
                ax1.axvline(x=50, color='red', linestyle='--', linewidth=2, label='High Risk Threshold')
                ax1.set_xlabel('AI Predicted Recurrence Probability (%)')
                ax1.set_ylabel('Number of Tickets')
                ax1.set_title('AI Recurrence Risk Distribution', fontweight='bold')
                ax1.legend()
                
                # Add zone labels
                ax1.text(10, ax1.get_ylim()[1]*0.9, 'Low\nRisk', ha='center', fontsize=9, color='#5CB85C')
                ax1.text(40, ax1.get_ylim()[1]*0.9, 'Moderate', ha='center', fontsize=9, color='#F0AD4E')
                ax1.text(75, ax1.get_ylim()[1]*0.9, 'High\nRisk', ha='center', fontsize=9, color='#D9534F')
                
                # Right: Risk Category Breakdown
                if 'AI_Recurrence_Risk' in df_issues.columns:
                    risk_counts = df_issues['AI_Recurrence_Risk'].value_counts()
                    
                    risk_colors = {
                        '🔴 High Risk (>70%)': '#D9534F',
                        '🟠 Elevated (50-70%)': '#FF9800',
                        '🟡 Moderate (30-50%)': '#F0AD4E',
                        '🟢 Low (<30%)': '#5CB85C',
                        'Unknown': '#95A5A6'
                    }
                    
                    colors_bar = [risk_colors.get(cat, '#95A5A6') for cat in risk_counts.index]
                    bars = ax2.barh(risk_counts.index, risk_counts.values, color=colors_bar)
                    
                    # Add value labels
                    for bar, val in zip(bars, risk_counts.values):
                        pct = val / len(df_issues) * 100
                        ax2.text(val + 0.5, bar.get_y() + bar.get_height()/2, 
                                f'{int(val)} ({pct:.1f}%)', va='center', fontsize=9)
                    
                    ax2.set_xlabel('Number of Tickets')
                    ax2.set_title('AI Risk Classification', fontweight='bold')
                    
                    # Summary stats
                    high_risk = (df_issues['AI_Recurrence_Probability'] >= 0.5).sum()
                    avg_prob = df_issues['AI_Recurrence_Probability'].mean() * 100
                    
                    ax2.text(0.98, 0.02, 
                            f'Avg Probability: {avg_prob:.1f}%\n{high_risk} tickets ≥50% risk',
                            transform=ax2.transAxes, fontsize=9, 
                            verticalalignment='bottom', horizontalalignment='right',
                            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
                
                plt.tight_layout()
                p13 = os.path.join(plot_dir, "ai_recurrence_prediction.png")
                plt.savefig(p13)
                plt.close()
                paths[12] = p13
    except Exception as e:
        logger.warning(f"Chart 13 Skipped: {e}")
        plt.close()
    
    # --- CHART 14: RESOLUTION CONSISTENCY ANALYSIS ---
    try:
        plt.figure(figsize=(10, 5))
        
        if 'Resolution_Consistency' in df.columns:
            # Count resolution consistency statuses
            consistency_counts = df['Resolution_Consistency'].value_counts()
            
            # Clean up labels for display
            def clean_label(label):
                if pd.isna(label) or label == 'N/A':
                    return 'No Similar Tickets'
                # Remove emoji for cleaner chart
                return str(label).replace('✅ ', '').replace('🟡 ', '').replace('🔴 ', '').replace('⚠️ ', '')
            
            labels = [clean_label(l) for l in consistency_counts.index]
            values = consistency_counts.values
            
            # Color coding by consistency level
            colors = []
            for label in consistency_counts.index:
                label_str = str(label)
                if 'Highly consistent' in label_str or 'Consistent' in label_str:
                    colors.append('#28A745')  # Green
                elif 'Mostly consistent' in label_str or 'minor' in label_str:
                    colors.append('#FFC107')  # Yellow
                elif 'Inconsistent' in label_str:
                    colors.append('#DC3545')  # Red
                elif 'No resolution' in label_str or 'available' in label_str:
                    colors.append('#6C757D')  # Gray
                else:
                    colors.append('#17A2B8')  # Blue (default)
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
            
            # Left: Bar chart of consistency statuses
            bars = ax1.barh(labels, values, color=colors)
            ax1.set_xlabel('Number of Tickets')
            ax1.set_title('Resolution Consistency Distribution', fontweight='bold')
            ax1.invert_yaxis()  # Highest at top
            
            # Add count labels on bars
            for bar, val in zip(bars, values):
                ax1.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                        f'{int(val)}', va='center', fontsize=9)
            
            # Right: Similarity score distribution
            if 'Best_Match_Similarity' in df.columns:
                similarity_scores = df[df['Best_Match_Similarity'] > 0]['Best_Match_Similarity']
                if len(similarity_scores) > 0:
                    ax2.hist(similarity_scores, bins=20, color=MC_BLUE, alpha=0.7, edgecolor='white')
                    ax2.axvline(x=0.7, color='red', linestyle='--', label='Threshold (70%)')
                    ax2.axvline(x=similarity_scores.mean(), color='green', linestyle='-', 
                               label=f'Mean ({similarity_scores.mean():.0%})')
                    ax2.set_xlabel('Similarity Score')
                    ax2.set_ylabel('Frequency')
                    ax2.set_title('Similar Ticket Match Strength', fontweight='bold')
                    ax2.legend(loc='upper left')
                    
                    # Add summary stats
                    high_match = (similarity_scores >= 0.8).sum()
                    total = len(similarity_scores)
                    ax2.text(0.98, 0.98, 
                            f'High Matches (≥80%): {high_match}\nTotal with matches: {total}',
                            transform=ax2.transAxes, fontsize=9, 
                            verticalalignment='top', horizontalalignment='right',
                            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
            else:
                ax2.text(0.5, 0.5, 'No similarity data available',
                        ha='center', va='center', transform=ax2.transAxes)
                ax2.set_title('Similar Ticket Match Strength', fontweight='bold')
            
            plt.tight_layout()
            p14 = os.path.join(plot_dir, "resolution_consistency.png")
            plt.savefig(p14)
            plt.close()
            paths[13] = p14
        else:
            logger.info("Chart 14: No resolution consistency data available")
    except Exception as e:
        logger.warning(f"Chart 14 Skipped: {e}")
        plt.close()
    
    # --- CHART 15: RESOLUTION TIME COMPARISON (Expected vs Actual vs Predicted) ---
    try:
        # Check if we have the resolution time columns
        has_actual = 'Actual_Resolution_Days' in df.columns and df['Actual_Resolution_Days'].notna().sum() > 0
        has_predicted = 'Predicted_Resolution_Days' in df.columns and df['Predicted_Resolution_Days'].notna().sum() > 0
        has_expected = 'Human_Expected_Days' in df.columns and df['Human_Expected_Days'].notna().sum() > 0
        
        if has_actual or has_predicted:
            fig, axes = plt.subplots(1, 2, figsize=(14, 5))
            
            # Subplot 1: Bar chart by category - Expected vs Actual vs Predicted
            ax1 = axes[0]
            
            # Group by category and calculate means
            if 'AI_Category' in df.columns:
                category_data = []
                for cat in df['AI_Category'].dropna().unique()[:10]:  # Top 10 categories
                    cat_df = df[df['AI_Category'] == cat]
                    actual_mean = cat_df['Actual_Resolution_Days'].mean() if has_actual else 0
                    predicted_mean = cat_df['Predicted_Resolution_Days'].mean() if has_predicted else 0
                    expected_mean = cat_df['Human_Expected_Days'].mean() if has_expected else None
                    
                    category_data.append({
                        'Category': cat[:20],  # Truncate long names
                        'Actual': actual_mean if pd.notna(actual_mean) else 0,
                        'Predicted': predicted_mean if pd.notna(predicted_mean) else 0,
                        'Expected': expected_mean if pd.notna(expected_mean) else 0
                    })
                
                cat_summary = pd.DataFrame(category_data)
                if not cat_summary.empty:
                    x = np.arange(len(cat_summary))
                    width = 0.25
                    
                    bars1 = ax1.bar(x - width, cat_summary['Actual'], width, label='Actual', color='#004C97')
                    bars2 = ax1.bar(x, cat_summary['Predicted'], width, label='AI Predicted', color='#28A745')
                    if has_expected:
                        bars3 = ax1.bar(x + width, cat_summary['Expected'], width, label='Human Expected', color='#FFA500')
                    
                    ax1.set_xlabel('Category')
                    ax1.set_ylabel('Days to Resolution')
                    ax1.set_title('Resolution Time by Category', fontweight='bold')
                    ax1.set_xticks(x)
                    ax1.set_xticklabels(cat_summary['Category'], rotation=45, ha='right', fontsize=8)
                    ax1.legend()
                    ax1.grid(axis='y', alpha=0.3)
            
            # Subplot 2: Scatter plot - Actual vs Predicted
            ax2 = axes[1]
            if has_actual and has_predicted:
                valid_df = df.dropna(subset=['Actual_Resolution_Days', 'Predicted_Resolution_Days'])
                if len(valid_df) > 0:
                    actual = valid_df['Actual_Resolution_Days'].values
                    predicted = valid_df['Predicted_Resolution_Days'].values
                    
                    ax2.scatter(actual, predicted, alpha=0.5, color='#004C97', edgecolors='white')
                    
                    # Perfect prediction line
                    max_val = max(actual.max(), predicted.max())
                    ax2.plot([0, max_val], [0, max_val], 'r--', label='Perfect Prediction', linewidth=2)
                    
                    # Calculate R² and MAE
                    correlation = np.corrcoef(actual, predicted)[0, 1] if len(actual) > 2 else 0
                    mae = np.mean(np.abs(predicted - actual))
                    
                    ax2.set_xlabel('Actual Resolution Days')
                    ax2.set_ylabel('Predicted Resolution Days')
                    ax2.set_title(f'Prediction Accuracy (R²={correlation**2:.2f}, MAE={mae:.1f}d)', fontweight='bold')
                    ax2.legend()
                    ax2.grid(alpha=0.3)
                    
                    # Add color-coded regions
                    ax2.fill_between([0, max_val], [0, max_val], [0, max_val*1.2], alpha=0.1, color='orange', label='Over-predicted')
                    ax2.fill_between([0, max_val], [0, max_val*0.8], [0, max_val], alpha=0.1, color='green', label='Under-predicted')
            
            plt.tight_layout()
            p15 = os.path.join(plot_dir, "resolution_time_comparison.png")
            plt.savefig(p15)
            plt.close()
            paths[14] = p15
            logger.info("Chart 15: Resolution Time Comparison generated")
        else:
            logger.info("Chart 15: No resolution time data available")
    except Exception as e:
        logger.warning(f"Chart 15 Skipped: {e}")
        plt.close()
   
    return paths

# ==========================================
# 6. MAIN PIPELINE
# ==========================================
def validate_data_quality(df) -> bool:
    """Check data quality and warn about issues"""
    issues = []
    
    if len(df) == 0:
        logger.error("Dataset is empty!")
        return False
    
    if len(df) < 5:
        issues.append(f"Very small dataset ({len(df)} rows) - results may not be meaningful")
    
    # Check for mostly empty data
    empty_pct = df.isna().sum().sum() / (len(df) * len(df.columns)) * 100
    if empty_pct > 50:
        issues.append(f"High missing data rate: {empty_pct:.1f}%")
    
    # Check for required columns
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing_cols:
        issues.append(f"Missing recommended columns: {missing_cols}")
    
    # Log warnings
    if issues:
        logger.warning("Data Quality Issues Detected:")
        for issue in issues:
            logger.warning(f"  ⚠ {issue}")
    else:
        logger.info(f"✓ Data quality check passed ({len(df)} rows, {len(df.columns)} columns)")
    
    return True

def main_pipeline():
    root = tk.Tk()
    root.withdraw()
    
    try:
        # Pre-flight checks
        if not check_ollama_server():
            return
        
        file_path = filedialog.askopenfilename(title="Select Log File")
        if not file_path: 
            return

        ai = OllamaBrain()
        logger.info(f"AI Initialized. Embed: {ai.embed_model} | Gen: {ai.gen_model}")
        
        # Validate models before processing
        if not check_models(ai):
            return
        
        # Load feedback from previous runs (reinforcement learning)
        feedback_learner = get_feedback_learner()
        feedback_learner.load_feedback(ai)
        
        # Load price catalog for financial impact calculations
        price_catalog = get_price_catalog()
        price_catalog.load_catalog()

        # Read Data
        try:
            xls = pd.ExcelFile(file_path)
            sheet = next((s for s in xls.sheet_names if 'raw' in str(s).lower()), xls.sheet_names[0])
            logger.info(f"Loading Sheet: {sheet}")
            df = pd.read_excel(file_path, sheet_name=sheet)
            # Store original raw data before any processing
            df_raw = df.copy()
        except Exception as e:
            logger.warning(f"Excel read failed, trying CSV: {e}")
            df = pd.read_csv(file_path, engine='python')
            df_raw = df.copy()
        
        # Data quality validation
        if not validate_data_quality(df):
            messagebox.showerror("Data Error", "The selected file contains no usable data.")
            return

        # Prepare Text
        text_cols = [COL_SUMMARY, COL_CATEGORY]
        actual_cols = [c for c in df.columns if c.strip().lower() in [t.lower() for t in text_cols]]
       
        if actual_cols:
            df['Combined_Text'] = df[actual_cols].apply(lambda x: ' - '.join(x.dropna().astype(str)), axis=1)
        else:
            df['Combined_Text'] = df.iloc[:, 0].astype(str) # Fallback
       
        df['Combined_Text'] = df['Combined_Text'].apply(clean_text)

        # 1. AI Classification (Categorize rows using embeddings)
        df = classify_rows(df, ai)

        # 2. Scoring
        df = calculate_strategic_friction(df)

        # 3. Recidivism (Uses Embeddings)
        df = audit_learning(df, ai)

        # 4. LLM Synthesis (Uses Generator)
        logger.info(f"[GenAI] Drafting Executive Summary using {GEN_MODEL}...")
       
        # Prepare comprehensive context for the AI
        top_risks = df.nlargest(5, 'Strategic_Friction_Score')
        total_friction = df['Strategic_Friction_Score'].sum()
        total_tickets = len(df)
        
        # Safe division helper
        def safe_pct(num, denom): 
            return (num / denom * 100) if denom > 0 else 0
        
        # Origin analysis
        ext_sum = df[df[COL_ORIGIN]=='External']['Strategic_Friction_Score'].sum() if COL_ORIGIN in df.columns else 0
        int_sum = df[df[COL_ORIGIN]=='Internal']['Strategic_Friction_Score'].sum() if COL_ORIGIN in df.columns else 0
        ext_pct = safe_pct(ext_sum, total_friction)
        
        # Severity breakdown
        severity_counts = df['Severity_Norm'].value_counts().to_dict() if 'Severity_Norm' in df.columns else {}
        critical_count = severity_counts.get('Critical', 0)
        major_count = severity_counts.get('Major', 0)
        
        # Recidivism analysis
        repeat_high = df['Learning_Status'].str.contains('REPEAT OFFENSE', na=False).sum()
        repeat_possible = df['Learning_Status'].str.contains('POSSIBLE REPEAT', na=False).sum()
        total_repeats = repeat_high + repeat_possible
        repeat_pct = safe_pct(total_repeats, total_tickets)
        
        # AI Category breakdown (top 5)
        category_friction = df.groupby('AI_Category')['Strategic_Friction_Score'].agg(['sum', 'count']).sort_values('sum', ascending=False).head(5) if 'AI_Category' in df.columns else pd.DataFrame()
        
        # Type breakdown
        type_counts = df[COL_TYPE].value_counts().to_dict() if COL_TYPE in df.columns else {}
        escalation_count = sum(v for k, v in type_counts.items() if 'escalation' in str(k).lower())
        
        # Time analysis (if available)
        time_context = ""
        if COL_DATETIME in df.columns:
            df_temp = df.copy()
            df_temp['Date'] = pd.to_datetime(df_temp[COL_DATETIME], errors='coerce')
            if df_temp['Date'].notna().any():
                date_range = f"{df_temp['Date'].min().strftime('%Y-%m-%d')} to {df_temp['Date'].max().strftime('%Y-%m-%d')}"
                recent_7d = df_temp[df_temp['Date'] >= df_temp['Date'].max() - pd.Timedelta(days=7)]
                recent_friction = recent_7d['Strategic_Friction_Score'].sum()
                recent_pct = safe_pct(recent_friction, total_friction)
                time_context = f"""
TIME ANALYSIS:
- Date Range: {date_range}
- Last 7 Days Friction: {recent_friction:,.0f} ({recent_pct:.1f}% of total)
- Tickets in Last 7 Days: {len(recent_7d)}
"""
        
        # Build comprehensive context
        context_lines = [
            "=" * 60,
            "ESCALATION REPORT STATISTICAL SUMMARY",
            "=" * 60,
            "",
            "VOLUME & SEVERITY METRICS:",
            f"- Total Tickets Analyzed: {total_tickets}",
            f"- Total Weighted Friction Score: {total_friction:,.0f}",
            f"- Critical Severity Tickets: {critical_count} ({safe_pct(critical_count, total_tickets):.1f}%)",
            f"- Major Severity Tickets: {major_count} ({safe_pct(major_count, total_tickets):.1f}%)",
            f"- Active Escalations: {escalation_count}",
            "",
            "RISK ORIGIN ANALYSIS:",
            f"- External (Customer/Partner-Facing): {ext_pct:.1f}% of total friction",
            f"- Internal (Operations): {safe_pct(int_sum, total_friction):.1f}% of total friction",
            f"- External Risk Score: {ext_sum:,.0f}",
            "",
            "ORGANIZATIONAL LEARNING ASSESSMENT:",
            f"- Confirmed Repeat Offenses: {repeat_high} incidents",
            f"- Possible Repeat Issues: {repeat_possible} incidents", 
            f"- Total Potential Recidivism: {total_repeats} ({repeat_pct:.1f}% of all tickets)",
            f"- Learning Failure Indicator: {'HIGH - Systemic issue' if repeat_pct > 10 else 'MODERATE' if repeat_pct > 5 else 'LOW'}",
        ]
        
        # Add financial impact summary
        if 'Financial_Impact' in df.columns:
            total_financial = df['Financial_Impact'].sum()
            avg_financial = df['Financial_Impact'].mean()
            max_financial = df['Financial_Impact'].max()
            top_cost_category = df.groupby('AI_Category')['Financial_Impact'].sum().idxmax() if 'AI_Category' in df.columns else 'Unknown'
            top_cost_amount = df.groupby('AI_Category')['Financial_Impact'].sum().max() if 'AI_Category' in df.columns else 0
            
            context_lines.extend([
                "",
                "FINANCIAL IMPACT ANALYSIS:",
                f"- Total Estimated Financial Impact: ${total_financial:,.0f}",
                f"- Average Impact per Ticket: ${avg_financial:,.0f}",
                f"- Highest Single Ticket Impact: ${max_financial:,.0f}",
                f"- Costliest Category: {top_cost_category} (${top_cost_amount:,.0f})",
            ])
        
        # Add category breakdown
        if not category_friction.empty:
            context_lines.extend([
                "",
                "TOP FAILURE CATEGORIES (AI-Classified):"
            ])
            for cat, row in category_friction.iterrows():
                context_lines.append(f"- {cat}: {row['sum']:,.0f} friction, {int(row['count'])} tickets")
        
        # Add time context if available
        if time_context:
            context_lines.append(time_context)
        
        # Add top specific issues
        context_lines.extend([
            "",
            "TOP 5 HIGHEST-RISK SPECIFIC ISSUES:"
        ])
        
        for i, (_, row) in enumerate(top_risks.iterrows(), 1):
            summary = str(row.get(COL_SUMMARY, 'N/A'))[:150] if COL_SUMMARY in top_risks.columns else 'N/A'
            score = row.get('Strategic_Friction_Score', 0)
            severity = row.get('Severity_Norm', 'Unknown')
            origin = row.get('Origin_Norm', 'Unknown')
            category = row.get('AI_Category', 'Unclassified')
            learning = row.get('Learning_Status', 'New')
            
            context_lines.append(f"\n{i}. [{severity}/{origin}] {summary}")
            context_lines.append(f"   Category: {category} | Score: {score:.0f} | Status: {learning}")
        
        context_lines.extend([
            "",
            "=" * 60,
            "END OF DATA SUMMARY",
            "=" * 60
        ])
        
        context = "\n".join(context_lines)
       
        exec_summary_text = ai.generate_synthesis(context)
        logger.info(f"AI Insight: {exec_summary_text[:100]}...")

        # Output
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Strategic_Report.xlsx")
        if not save_path: 
            return
       
        plot_paths = generate_plots(df, os.path.dirname(save_path))
       
        wb = Workbook()
       
        # =====================================================
        # SHEET 1: EXECUTIVE SUMMARY (Clean, dedicated page)
        # =====================================================
        ws_summary = wb.create_sheet("Executive Summary", 0)
        ws_summary.sheet_view.showGridLines = False
        
        # Title with branding
        report_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws_summary['A1'] = REPORT_TITLE
        ws_summary['A1'].font = Font(bold=True, size=24, color="004C97")
        ws_summary.merge_cells('A1:H1')
        
        ws_summary['A2'] = f"Generated: {report_timestamp} | Version: {REPORT_VERSION} | AI Model: {GEN_MODEL}"
        ws_summary['A2'].font = Font(size=10, italic=True, color="666666")
        ws_summary.merge_cells('A2:H2')
        
        # Horizontal line effect
        ws_summary.row_dimensions[3].height = 5
        
        # Key Metrics Summary Box
        ws_summary['A4'] = "KEY METRICS AT A GLANCE"
        ws_summary['A4'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_summary['A4'].fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
        ws_summary.merge_cells('A4:H4')
        
        # Calculate total financial impact
        total_financial = df['Financial_Impact'].sum() if 'Financial_Impact' in df.columns else 0
        
        # Metrics in 2 rows
        metrics_row1 = [
            ("Total Tickets", f"{total_tickets}"),
            ("Friction Score", f"{total_friction:,.0f}"),
            ("Critical Issues", f"{critical_count}"),
            ("Financial Impact", f"${total_financial:,.0f}"),
        ]
        metrics_row2 = [
            ("External Risk", f"{ext_pct:.1f}%"),
            ("Repeat Offenses", f"{total_repeats}"),
            ("Unclassified", f"{len(df[df['AI_Category']=='Unclassified'])}"),
            ("Major Issues", f"{major_count}"),
        ]
        
        for i, (label, val) in enumerate(metrics_row1):
            col = 1 + (i * 2)  # A, C, E, G
            ws_summary.cell(row=5, column=col).value = label
            ws_summary.cell(row=5, column=col).font = Font(bold=True, size=9, color="666666")
            ws_summary.cell(row=6, column=col).value = val
            ws_summary.cell(row=6, column=col).font = Font(bold=True, size=14, color="004C97")
        
        for i, (label, val) in enumerate(metrics_row2):
            col = 1 + (i * 2)
            ws_summary.cell(row=7, column=col).value = label
            ws_summary.cell(row=7, column=col).font = Font(bold=True, size=9, color="666666")
            ws_summary.cell(row=8, column=col).value = val
            ws_summary.cell(row=8, column=col).font = Font(bold=True, size=14, color="004C97")
        
        # Spacer row
        ws_summary.row_dimensions[9].height = 10
        
        # AI Executive Synthesis Section
        ws_summary['A10'] = "AI EXECUTIVE SYNTHESIS"
        ws_summary['A10'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_summary['A10'].fill = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
        ws_summary.merge_cells('A10:H10')
        
        # Split synthesis into paragraphs for better formatting
        synthesis_paragraphs = exec_summary_text.split('\n\n')
        current_row = 11
        
        for para in synthesis_paragraphs:
            if para.strip():
                # Check if it's a section header (all caps or ends with :)
                is_header = para.strip().isupper() or para.strip().endswith(':')
                
                ws_summary[f'A{current_row}'] = para.strip()
                ws_summary[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws_summary.merge_cells(f'A{current_row}:H{current_row}')
                
                if is_header:
                    ws_summary[f'A{current_row}'].font = Font(bold=True, size=11, color="004C97")
                    ws_summary.row_dimensions[current_row].height = 20
                else:
                    ws_summary[f'A{current_row}'].font = Font(size=10)
                    # Estimate row height based on text length
                    text_len = len(para)
                    row_height = max(45, min(120, text_len // 3))
                    ws_summary.row_dimensions[current_row].height = row_height
                
                current_row += 1
        
        # Footer with next steps
        current_row += 1
        ws_summary[f'A{current_row}'] = "→ See 'Dashboard' sheet for visual analysis | 'Scored Data' for details | 'Financial Analysis' for cost breakdown"
        ws_summary[f'A{current_row}'].font = Font(size=9, italic=True, color="666666")
        ws_summary.merge_cells(f'A{current_row}:H{current_row}')
        
        # Set column widths
        ws_summary.column_dimensions['A'].width = 18
        ws_summary.column_dimensions['B'].width = 12
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 18
        ws_summary.column_dimensions['F'].width = 12
        ws_summary.column_dimensions['G'].width = 18
        ws_summary.column_dimensions['H'].width = 12
        
        # =====================================================
        # SHEET 2: DASHBOARD (Charts and Visual Metrics Only)
        # =====================================================
        ws = wb.create_sheet("Dashboard", 1)
        ws.sheet_view.showGridLines = False
       
        # Title
        ws['A1'] = "VISUAL ANALYTICS DASHBOARD"
        ws['A1'].font = Font(bold=True, size=18, color="004C97")
        ws['A2'] = f"Report: {report_timestamp} | {total_tickets} tickets analyzed"
        ws['A2'].font = Font(size=9, color="888888")
       
        # Compact metrics row (Row 4)
        ws['A4'] = "Quick Stats:"
        ws['A4'].font = Font(bold=True, size=10)
        ws['B4'] = f"Friction: {total_friction:,.0f}"
        ws['D4'] = f"External: {ext_pct:.1f}%"
        ws['F4'] = f"Repeats: {total_repeats}"
        ws['H4'] = f"Critical: {critical_count}"
        ws['J4'] = f"Financial: ${total_financial:,.0f}"
        
        for cell in ['B4', 'D4', 'F4', 'H4', 'J4']:
            ws[cell].font = Font(size=10, color="004C97")
       
        # --- IMAGE INSERTION (10 Charts) ---
        # Row 6: Charts 1 & 2 (Friction Pareto + Risk Origin)
        # Row 22: Charts 3 & 4 (Learning Integrity + 7-Day Trend)
        # Row 38: Charts 5 & 6 (Category×Severity Heatmap + Engineer Friction)
        # Row 54: Chart 7 (Engineer Learning)
        # Row 70: LOB Section - Charts 8 & 9 (LOB Friction + LOB Matrix)
        # Row 86: Chart 10 (LOB Category Breakdown)
        img_positions = [
            'A6', 'G6',      # Charts 1-2: Core Risk Analysis
            'A22', 'G22',    # Charts 3-4: Learning & Trends
            'A38', 'G38',    # Charts 5-6: Deep Dive & Engineer
            'A54',           # Chart 7: Engineer Learning
            'A70', 'G70',    # Charts 8-9: LOB Analysis
            'A86'            # Chart 10: LOB Categories
        ]
       
        for path, pos in zip(plot_paths, img_positions):
            if path and os.path.exists(path):
                img = OpenpyxlImage(path)
                img.width = 450
                img.height = 260
                ws.add_image(img, pos)
        
        # Add section headers
        ws['A5'] = "CORE RISK ANALYSIS"
        ws['A5'].font = Font(bold=True, size=12, color="004C97")
        
        ws['A21'] = "LEARNING & TREND ANALYSIS"
        ws['A21'].font = Font(bold=True, size=11, color="004C97")
        
        ws['A37'] = "DEEP DIVE ANALYTICS"
        ws['A37'].font = Font(bold=True, size=11, color="004C97")
        
        ws['A53'] = "ENGINEER ACCOUNTABILITY"
        ws['A53'].font = Font(bold=True, size=11, color="004C97")
        
        ws['A69'] = "LINE OF BUSINESS (LOB) STRATEGIC ANALYSIS"
        ws['A69'].font = Font(bold=True, size=12, color="1F4E79")
        ws.merge_cells('A69:G69')
        
        ws['A85'] = "LOB ROOT CAUSE BREAKDOWN"
        ws['A85'].font = Font(bold=True, size=11, color="1F4E79")
        
        # =====================================================
        # SHEET 3: RAW DATA (Original import, unmodified)
        # =====================================================
        ws_raw = wb.create_sheet("Raw Data")
        ws_raw['A1'] = "ORIGINAL IMPORTED DATA"
        ws_raw['A1'].font = Font(bold=True, size=12, color="004C97")
        ws_raw.merge_cells('A1:E1')
        ws_raw['A2'] = f"Source: {os.path.basename(file_path)}"
        ws_raw['A2'].font = Font(size=9, italic=True, color="666666")
        
        # Write raw data starting at row 4
        for r_idx, r in enumerate(dataframe_to_rows(df_raw, index=False, header=True)):
            for c_idx, value in enumerate(r, 1):
                ws_raw.cell(row=r_idx + 4, column=c_idx, value=value)
        
        # Format header row
        header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        for col in range(1, len(df_raw.columns) + 1):
            ws_raw.cell(row=4, column=col).fill = header_fill
            ws_raw.cell(row=4, column=col).font = Font(bold=True)
        
        logger.info(f"  → Raw Data sheet added ({len(df_raw)} rows)")
               
        # =====================================================
        # SHEET 4: SCORED DATA (Processed with AI classifications)
        # =====================================================
        ws_data = wb.create_sheet("Scored Data")
        
        # Prepare dataframe for Excel (convert numpy arrays to strings)
        df_export = df.copy()
        for col in df_export.columns:
            if df_export[col].apply(lambda x: isinstance(x, np.ndarray)).any():
                df_export[col] = df_export[col].apply(
                    lambda x: str(x.tolist()) if isinstance(x, np.ndarray) else x
                )
        
        for r in dataframe_to_rows(df_export, index=False, header=True): 
            ws_data.append(r)
        
        # FINANCIAL ANALYSIS SHEET
        if 'Financial_Impact' in df.columns:
            ws_financial = wb.create_sheet("Financial Analysis")
            
            # Title
            ws_financial['A1'] = "FINANCIAL IMPACT ANALYSIS"
            ws_financial['A1'].font = Font(bold=True, size=14)
            ws_financial.merge_cells('A1:E1')
            
            # Summary metrics
            ws_financial['A3'] = "SUMMARY METRICS"
            ws_financial['A3'].font = Font(bold=True)
            
            total_fin = df['Financial_Impact'].sum()
            avg_fin = df['Financial_Impact'].mean()
            max_fin = df['Financial_Impact'].max()
            
            summary_data = [
                ["Total Estimated Impact", f"${total_fin:,.2f}"],
                ["Average per Ticket", f"${avg_fin:,.2f}"],
                ["Maximum Single Ticket", f"${max_fin:,.2f}"],
                ["Tickets Analyzed", len(df)],
            ]
            for i, (label, value) in enumerate(summary_data, start=4):
                ws_financial[f'A{i}'] = label
                ws_financial[f'B{i}'] = value
            
            # Category breakdown
            ws_financial['A10'] = "IMPACT BY CATEGORY"
            ws_financial['A10'].font = Font(bold=True)
            
            category_fin = df.groupby('AI_Category')['Financial_Impact'].agg(['sum', 'mean', 'count']).sort_values('sum', ascending=False)
            ws_financial['A11'] = "Category"
            ws_financial['B11'] = "Total Impact"
            ws_financial['C11'] = "Avg Impact"
            ws_financial['D11'] = "Ticket Count"
            
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col in ['A', 'B', 'C', 'D']:
                ws_financial[f'{col}11'].fill = header_fill
                ws_financial[f'{col}11'].font = header_font
            
            for i, (cat, row) in enumerate(category_fin.iterrows(), start=12):
                ws_financial[f'A{i}'] = cat
                ws_financial[f'B{i}'] = f"${row['sum']:,.2f}"
                ws_financial[f'C{i}'] = f"${row['mean']:,.2f}"
                ws_financial[f'D{i}'] = int(row['count'])
            
            # Severity breakdown
            severity_row = 12 + len(category_fin) + 2
            ws_financial[f'A{severity_row}'] = "IMPACT BY SEVERITY"
            ws_financial[f'A{severity_row}'].font = Font(bold=True)
            
            severity_fin = df.groupby('Severity_Norm')['Financial_Impact'].agg(['sum', 'mean', 'count']).sort_values('sum', ascending=False)
            ws_financial[f'A{severity_row+1}'] = "Severity"
            ws_financial[f'B{severity_row+1}'] = "Total Impact"
            ws_financial[f'C{severity_row+1}'] = "Avg Impact"
            ws_financial[f'D{severity_row+1}'] = "Ticket Count"
            
            for col in ['A', 'B', 'C', 'D']:
                ws_financial[f'{col}{severity_row+1}'].fill = header_fill
                ws_financial[f'{col}{severity_row+1}'].font = header_font
            
            for i, (sev, row) in enumerate(severity_fin.iterrows(), start=severity_row+2):
                ws_financial[f'A{i}'] = sev
                ws_financial[f'B{i}'] = f"${row['sum']:,.2f}"
                ws_financial[f'C{i}'] = f"${row['mean']:,.2f}"
                ws_financial[f'D{i}'] = int(row['count'])
            
            # Column widths
            ws_financial.column_dimensions['A'].width = 30
            ws_financial.column_dimensions['B'].width = 18
            ws_financial.column_dimensions['C'].width = 18
            ws_financial.column_dimensions['D'].width = 14
            
            logger.info("  → Financial Analysis sheet added to report")
        
        # =====================================================
        # ENGINEER ACCOUNTABILITY SHEET (Human Error Analysis)
        # =====================================================
        if 'Engineer' in df.columns and COL_ENGINEER in df.columns:
            ws_engineer = wb.create_sheet("Engineer Accountability")
            
            # Define styles for this section
            eng_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            eng_header_font = Font(bold=True, color="FFFFFF")
            
            # Title
            ws_engineer['A1'] = "ENGINEER ACCOUNTABILITY ANALYSIS"
            ws_engineer['A1'].font = Font(bold=True, size=14, color="004C97")
            ws_engineer.merge_cells('A1:H1')
            
            ws_engineer['A2'] = "Human Error & Learning Pattern Tracking for Escalations & Concerns"
            ws_engineer['A2'].font = Font(size=10, italic=True, color="666666")
            
            # Filter to only escalations and concerns
            df_human_error = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])].copy()
            
            # Summary metrics
            ws_engineer['A4'] = "SUMMARY"
            ws_engineer['A4'].font = Font(bold=True)
            
            total_engineers = df_human_error['Engineer'].nunique()
            total_issues = len(df_human_error)
            
            # Use new engineer pattern columns if available
            if 'Engineer_Is_Repeat_Offender' in df_human_error.columns:
                repeat_offenders = df_human_error[df_human_error['Engineer_Is_Repeat_Offender'] == True]['Engineer'].nunique()
            else:
                repeat_offenders = len(df_human_error[df_human_error.get('Engineer_Issue_Count', pd.Series([0]*len(df_human_error))) >= 3]['Engineer'].unique())
            
            human_error_count = len(df_human_error[df_human_error['Is_Human_Error'] == 'Yes'])
            
            # Calculate learning metrics
            avg_learning_gap = df_human_error['Engineer_Learning_Gap'].mean() if 'Engineer_Learning_Gap' in df_human_error.columns else 0
            engineers_with_lessons = (df_human_error['Engineer_Lessons_Logged'] > 0).sum() if 'Engineer_Lessons_Logged' in df_human_error.columns else 0
            
            summary_data = [
                ["Total Engineers Involved", total_engineers],
                ["Total Escalations/Concerns", total_issues],
                ["Repeat Offender Engineers (3+ issues)", repeat_offenders],
                ["Confirmed Human Errors", human_error_count],
                ["Human Error Rate", f"{(human_error_count/total_issues*100):.1f}%" if total_issues > 0 else "N/A"],
                ["Engineers Logging Lessons", engineers_with_lessons],
                ["Avg Learning Gap (Issues - Lessons)", f"{avg_learning_gap:.1f}"],
            ]
            for i, (label, value) in enumerate(summary_data, start=5):
                ws_engineer[f'A{i}'] = label
                ws_engineer[f'B{i}'] = value
            
            # Engineer ranking table with learning metrics
            ws_engineer['A14'] = "ENGINEER RANKING (by Friction Contribution & Learning Behavior)"
            ws_engineer['A14'].font = Font(bold=True)
            
            # Build aggregation dict dynamically based on available columns
            agg_dict = {
                'Strategic_Friction_Score': ['sum', 'count', 'mean'],
                'Is_Human_Error': lambda x: (x == 'Yes').sum()
            }
            if 'Financial_Impact' in df_human_error.columns:
                agg_dict['Financial_Impact'] = 'sum'
            if 'Engineer_Lessons_Logged' in df_human_error.columns:
                agg_dict['Engineer_Lessons_Logged'] = 'first'
            if 'Engineer_Learning_Gap' in df_human_error.columns:
                agg_dict['Engineer_Learning_Gap'] = 'first'
            if 'Engineer_Learning_Score' in df_human_error.columns:
                agg_dict['Engineer_Learning_Score'] = 'first'
            
            engineer_stats = df_human_error.groupby('Engineer').agg(agg_dict).reset_index()
            engineer_stats.columns = ['_'.join(col).strip('_') if isinstance(col, tuple) else col for col in engineer_stats.columns]
            
            # Standardize column names
            col_renames = {
                'Strategic_Friction_Score_sum': 'Total_Friction',
                'Strategic_Friction_Score_count': 'Issue_Count',
                'Strategic_Friction_Score_mean': 'Avg_Friction',
                'Financial_Impact_sum': 'Financial_Impact',
                'Is_Human_Error_<lambda>': 'Human_Errors',
                'Engineer_Lessons_Logged_first': 'Lessons_Logged',
                'Engineer_Learning_Gap_first': 'Learning_Gap',
                'Engineer_Learning_Score_first': 'Learning_Score'
            }
            engineer_stats = engineer_stats.rename(columns=col_renames)
            engineer_stats = engineer_stats.sort_values('Total_Friction', ascending=False)
            
            # Header row with new columns
            headers = ['Engineer', 'Issues', 'Lessons', 'Gap', 'Learn Score', 'Total Friction', 'Financial Impact', 'Human Errors', 'Status']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_engineer.cell(row=15, column=col_idx, value=header)
                cell.fill = eng_header_fill
                cell.font = eng_header_font
            
            # Data rows
            for row_idx, (_, eng_row) in enumerate(engineer_stats.head(20).iterrows(), start=16):
                ws_engineer.cell(row=row_idx, column=1, value=eng_row['Engineer'])
                ws_engineer.cell(row=row_idx, column=2, value=int(eng_row.get('Issue_Count', 0)))
                ws_engineer.cell(row=row_idx, column=3, value=int(eng_row.get('Lessons_Logged', 0)))
                ws_engineer.cell(row=row_idx, column=4, value=int(eng_row.get('Learning_Gap', 0)))
                ws_engineer.cell(row=row_idx, column=5, value=f"{eng_row.get('Learning_Score', 0):.0%}")
                ws_engineer.cell(row=row_idx, column=6, value=f"{eng_row.get('Total_Friction', 0):,.0f}")
                ws_engineer.cell(row=row_idx, column=7, value=f"${eng_row.get('Financial_Impact', 0):,.0f}")
                ws_engineer.cell(row=row_idx, column=8, value=int(eng_row.get('Human_Errors', 0)))
                
                # Status based on issues and learning behavior
                issue_count = eng_row.get('Issue_Count', 0)
                lessons_logged = eng_row.get('Lessons_Logged', 0)
                learning_gap = eng_row.get('Learning_Gap', 0)
                
                status = ""
                if issue_count >= 5 and learning_gap >= 3:
                    status = "🔴 Repeat + No Learning"
                    ws_engineer.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif issue_count >= 3:
                    if lessons_logged == 0:
                        status = "🟠 Multiple + No Lessons"
                        ws_engineer.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFE4CC", end_color="FFE4CC", fill_type="solid")
                    else:
                        status = "🟡 Multiple Issues"
                        ws_engineer.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif learning_gap > 0:
                    status = "⚪ Needs to Log Lessons"
                elif lessons_logged > 0:
                    status = "✅ Learning"
                ws_engineer.cell(row=row_idx, column=9, value=status)
            
            # Issue breakdown by engineer (top 5 repeat offenders with details)
            repeat_engineers = engineer_stats[engineer_stats.get('Issue_Count', pd.Series([0])) >= 3].head(5)['Engineer'].tolist()
            
            if repeat_engineers:
                detail_start_row = 16 + min(20, len(engineer_stats)) + 3
                ws_engineer[f'A{detail_start_row}'] = "REPEAT OFFENDER DETAILS (with Learning History)"
                ws_engineer[f'A{detail_start_row}'].font = Font(bold=True)
                
                current_row = detail_start_row + 1
                for eng in repeat_engineers:
                    eng_issues = df_human_error[df_human_error['Engineer'] == eng]
                    lessons = eng_issues['Engineer_Lessons_Logged'].iloc[0] if 'Engineer_Lessons_Logged' in eng_issues.columns and len(eng_issues) > 0 else 0
                    gap = eng_issues['Engineer_Learning_Gap'].iloc[0] if 'Engineer_Learning_Gap' in eng_issues.columns and len(eng_issues) > 0 else 0
                    
                    ws_engineer[f'A{current_row}'] = f"▸ {eng} ({len(eng_issues)} issues, {lessons} lessons logged, gap: {gap})"
                    ws_engineer[f'A{current_row}'].font = Font(bold=True, color="D9534F")
                    current_row += 1
                    
                    for _, issue in eng_issues.head(5).iterrows():
                        summary = str(issue.get(COL_SUMMARY, 'N/A'))[:80]
                        category = issue.get('AI_Category', 'Unknown')
                        learning_status = issue.get('Learning_Status', 'Unknown')[:30]
                        ws_engineer[f'B{current_row}'] = f"[{category}] {summary}... | {learning_status}"
                        ws_engineer[f'B{current_row}'].font = Font(size=9, color="666666")
                        current_row += 1
                    current_row += 1
            
            # Column widths
            ws_engineer.column_dimensions['A'].width = 25
            ws_engineer.column_dimensions['B'].width = 10
            ws_engineer.column_dimensions['C'].width = 10
            ws_engineer.column_dimensions['D'].width = 8
            ws_engineer.column_dimensions['E'].width = 12
            ws_engineer.column_dimensions['F'].width = 14
            ws_engineer.column_dimensions['G'].width = 16
            ws_engineer.column_dimensions['H'].width = 14
            ws_engineer.column_dimensions['I'].width = 22
            
            logger.info(f"  → Engineer Accountability sheet added ({total_engineers} engineers, {repeat_offenders} repeat offenders)")
        
        # =====================================================
        # LOB STRATEGIC ANALYSIS SHEET (McKinsey-Style)
        # =====================================================
        if 'LOB' in df.columns and COL_LOB in df.columns:
            ws_lob = wb.create_sheet("LOB Analysis")
            
            # Define styles
            lob_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            lob_header_font = Font(bold=True, color="FFFFFF")
            critical_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            high_fill = PatternFill(start_color="FFE4CC", end_color="FFE4CC", fill_type="solid")
            
            # Title
            ws_lob['A1'] = "LINE OF BUSINESS STRATEGIC ANALYSIS"
            ws_lob['A1'].font = Font(bold=True, size=16, color="1F4E79")
            ws_lob.merge_cells('A1:J1')
            
            ws_lob['A2'] = "McKinsey-Style Organizational Performance & Risk Assessment"
            ws_lob['A2'].font = Font(size=11, italic=True, color="666666")
            
            # Get LOB data from escalations/concerns
            df_lob = df[df['Type_Norm'].isin(['Escalations', 'Concerns'])].copy() if 'Type_Norm' in df.columns else df.copy()
            
            # Summary Section
            ws_lob['A4'] = "EXECUTIVE SUMMARY"
            ws_lob['A4'].font = Font(bold=True, size=12)
            
            total_lobs = df_lob['LOB'].nunique()
            critical_lobs = len(df_lob[df_lob['LOB_Risk_Tier'] == 'Critical']['LOB'].unique())
            high_risk_lobs = len(df_lob[df_lob['LOB_Risk_Tier'] == 'High']['LOB'].unique())
            total_lob_friction = df_lob['Strategic_Friction_Score'].sum()
            avg_learning_rate = df_lob['LOB_Learning_Rate'].mean() * 100 if 'LOB_Learning_Rate' in df_lob.columns else 0
            
            summary_data = [
                ["Total Lines of Business", total_lobs],
                ["Critical Risk LOBs", critical_lobs],
                ["High Risk LOBs", high_risk_lobs],
                ["Total Organizational Friction", f"{total_lob_friction:,.0f}"],
                ["Avg Organizational Learning Rate", f"{avg_learning_rate:.1f}%"],
            ]
            for i, (label, value) in enumerate(summary_data, start=5):
                ws_lob[f'A{i}'] = label
                ws_lob[f'A{i}'].font = Font(bold=True)
                ws_lob[f'B{i}'] = value
            
            # LOB Rankings Table
            ws_lob['A12'] = "LOB PERFORMANCE RANKING (Strategic Priority Order)"
            ws_lob['A12'].font = Font(bold=True, size=11)
            
            # Get unique LOB metrics
            lob_stats = df_lob.groupby('LOB').agg({
                'Strategic_Friction_Score': ['sum', 'count', 'mean'],
                'LOB_Efficiency_Score': 'first',
                'LOB_Learning_Rate': 'first',
                'LOB_Risk_Tier': 'first',
                'LOB_Engineer_Count': 'first',
                'LOB_Repeat_Offender_Rate': 'first',
                'Financial_Impact': 'sum' if 'Financial_Impact' in df_lob.columns else lambda x: 0
            }).reset_index()
            
            # Flatten column names
            lob_stats.columns = ['LOB', 'Total_Friction', 'Issue_Count', 'Avg_Friction',
                                'Efficiency', 'Learning_Rate', 'Risk_Tier', 'Engineers',
                                'Repeat_Rate', 'Financial_Impact']
            lob_stats = lob_stats.sort_values('Total_Friction', ascending=False)
            
            # Header row
            headers = ['LOB', 'Risk', 'Issues', 'Friction', 'Efficiency', 
                      'Learn%', 'Engineers', 'Repeat%', 'Financial', 'Recommendation']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_lob.cell(row=13, column=col_idx, value=header)
                cell.fill = lob_header_fill
                cell.font = lob_header_font
            
            # Data rows
            for row_idx, (_, lob_row) in enumerate(lob_stats.head(15).iterrows(), start=14):
                ws_lob.cell(row=row_idx, column=1, value=lob_row['LOB'])
                ws_lob.cell(row=row_idx, column=2, value=lob_row['Risk_Tier'])
                ws_lob.cell(row=row_idx, column=3, value=int(lob_row['Issue_Count']))
                ws_lob.cell(row=row_idx, column=4, value=f"{lob_row['Total_Friction']:,.0f}")
                ws_lob.cell(row=row_idx, column=5, value=f"{lob_row['Efficiency']:.0f}/100")
                ws_lob.cell(row=row_idx, column=6, value=f"{lob_row['Learning_Rate']*100:.0f}%")
                ws_lob.cell(row=row_idx, column=7, value=int(lob_row['Engineers']))
                ws_lob.cell(row=row_idx, column=8, value=f"{lob_row['Repeat_Rate']:.0f}%")
                ws_lob.cell(row=row_idx, column=9, value=f"${lob_row['Financial_Impact']:,.0f}")
                
                # Strategic Recommendation based on metrics
                if lob_row['Risk_Tier'] == 'Critical':
                    recommendation = "🔴 IMMEDIATE ACTION REQUIRED"
                    ws_lob.cell(row=row_idx, column=1).fill = critical_fill
                elif lob_row['Risk_Tier'] == 'High':
                    if lob_row['Learning_Rate'] < 0.3:
                        recommendation = "🟠 Focus: Learning Culture"
                    elif lob_row['Repeat_Rate'] > 30:
                        recommendation = "🟠 Focus: Talent Quality"
                    else:
                        recommendation = "🟠 Process Improvement"
                    ws_lob.cell(row=row_idx, column=1).fill = high_fill
                elif lob_row['Efficiency'] < 50:
                    recommendation = "🟡 Efficiency Optimization"
                elif lob_row['Learning_Rate'] >= 0.7:
                    recommendation = "✅ Best Practice Leader"
                else:
                    recommendation = "→ Monitor & Maintain"
                
                ws_lob.cell(row=row_idx, column=10, value=recommendation)
            
            # Best Practice Section
            best_row = 14 + min(15, len(lob_stats)) + 2
            ws_lob[f'A{best_row}'] = "BEST PRACTICE IDENTIFICATION"
            ws_lob[f'A{best_row}'].font = Font(bold=True, size=11)
            
            # Find best performing LOBs (high efficiency, good learning rate)
            good_lobs = lob_stats[(lob_stats['Efficiency'] >= 60) & (lob_stats['Learning_Rate'] >= 0.5)]
            if not good_lobs.empty:
                best_row += 1
                ws_lob[f'A{best_row}'] = "Top Performing LOBs (Benchmark Candidates):"
                ws_lob[f'A{best_row}'].font = Font(italic=True)
                for _, lob in good_lobs.head(3).iterrows():
                    best_row += 1
                    ws_lob[f'B{best_row}'] = f"✅ {lob['LOB']}: Efficiency {lob['Efficiency']:.0f}/100, Learning {lob['Learning_Rate']*100:.0f}%"
                    ws_lob[f'B{best_row}'].font = Font(color="2E7D32")
            
            # Intervention Needed Section
            int_row = best_row + 2
            ws_lob[f'A{int_row}'] = "PRIORITY INTERVENTIONS REQUIRED"
            ws_lob[f'A{int_row}'].font = Font(bold=True, size=11, color="D9534F")
            
            critical_lob_list = lob_stats[lob_stats['Risk_Tier'] == 'Critical']
            if not critical_lob_list.empty:
                for _, lob in critical_lob_list.head(3).iterrows():
                    int_row += 1
                    ws_lob[f'B{int_row}'] = f"🔴 {lob['LOB']}: {int(lob['Issue_Count'])} issues, ${lob['Financial_Impact']:,.0f} impact, {lob['Repeat_Rate']:.0f}% repeat rate"
                    ws_lob[f'B{int_row}'].font = Font(color="D9534F")
            else:
                int_row += 1
                ws_lob[f'B{int_row}'] = "No critical LOBs identified at this time."
            
            # Column widths
            ws_lob.column_dimensions['A'].width = 22
            ws_lob.column_dimensions['B'].width = 12
            ws_lob.column_dimensions['C'].width = 10
            ws_lob.column_dimensions['D'].width = 12
            ws_lob.column_dimensions['E'].width = 12
            ws_lob.column_dimensions['F'].width = 10
            ws_lob.column_dimensions['G'].width = 12
            ws_lob.column_dimensions['H'].width = 10
            ws_lob.column_dimensions['I'].width = 14
            ws_lob.column_dimensions['J'].width = 28
            
            logger.info(f"  → LOB Analysis sheet added ({total_lobs} LOBs, {critical_lobs} critical)")
        
        # =====================================================
        # SIMILAR TICKETS ANALYSIS SHEET
        # =====================================================
        if 'Similar_Ticket_Count' in df.columns:
            ws_similar = wb.create_sheet("Similar Tickets")
            ws_similar.sheet_view.showGridLines = False
            
            # Title
            ws_similar['A1'] = "SIMILAR TICKET ANALYSIS & RESOLUTION COMPARISON"
            ws_similar['A1'].font = Font(bold=True, size=16, color="004C97")
            ws_similar.merge_cells('A1:J1')
            
            ws_similar['A2'] = "AI-powered analysis identifying similar historical tickets and comparing resolution approaches"
            ws_similar['A2'].font = Font(italic=True, size=10, color="666666")
            ws_similar.merge_cells('A2:J2')
            
            # Summary metrics box
            ws_similar['A4'] = "SUMMARY METRICS"
            ws_similar['A4'].font = Font(bold=True, size=11, color="FFFFFF")
            ws_similar['A4'].fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
            ws_similar.merge_cells('A4:E4')
            
            # Calculate metrics
            has_matches = (df['Similar_Ticket_Count'] > 0).sum()
            total_tickets = len(df)
            avg_similarity = df[df['Best_Match_Similarity'] > 0]['Best_Match_Similarity'].mean() if (df['Best_Match_Similarity'] > 0).any() else 0
            
            inconsistent_count = 0
            consistent_count = 0
            if 'Resolution_Consistency' in df.columns:
                inconsistent_count = df['Resolution_Consistency'].str.contains('Inconsistent', na=False).sum()
                consistent_count = df['Resolution_Consistency'].str.contains('consistent', case=False, na=False).sum()
            
            # Calculate resolution time stats
            avg_resolution_days = None
            if 'Avg_Similar_Resolution_Days' in df.columns:
                valid_times = df['Avg_Similar_Resolution_Days'].dropna()
                if len(valid_times) > 0:
                    avg_resolution_days = valid_times.mean()
            
            ws_similar['A5'] = f"Tickets with Similar Matches: {has_matches}/{total_tickets} ({has_matches/total_tickets*100:.1f}%)"
            ws_similar['A6'] = f"Average Match Similarity: {avg_similarity:.1%}"
            ws_similar['A7'] = f"Consistent Resolutions: {consistent_count}"
            ws_similar['A8'] = f"⚠️ Inconsistent Resolutions: {inconsistent_count}"
            ws_similar['A8'].font = Font(bold=True, color="D9534F") if inconsistent_count > 0 else Font()
            
            if avg_resolution_days is not None:
                ws_similar['A9'] = f"⏱️ Avg Resolution Time (similar tickets): {avg_resolution_days:.1f} days"
                ws_similar['A9'].font = Font(color="004C97")
            
            # Priority tickets with inconsistent resolutions (need standardization)
            ws_similar['A10'] = "🔴 PRIORITY: TICKETS WITH INCONSISTENT RESOLUTIONS"
            ws_similar['A10'].font = Font(bold=True, size=12, color="FFFFFF")
            ws_similar['A10'].fill = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
            ws_similar.merge_cells('A10:J10')
            
            ws_similar['A11'] = "These tickets have similar historical cases that were resolved differently - consider standardizing the approach"
            ws_similar['A11'].font = Font(italic=True, size=9, color="666666")
            ws_similar.merge_cells('A11:J11')
            
            # Get priority tickets (inconsistent resolutions with high similarity)
            inconsistent_mask = df['Resolution_Consistency'].str.contains('Inconsistent', na=False) if 'Resolution_Consistency' in df.columns else pd.Series([False] * len(df))
            priority_tickets = df[inconsistent_mask].copy()
            
            if len(priority_tickets) == 0:
                # Fall back to any tickets with multiple matches
                priority_tickets = df[df['Similar_Ticket_Count'] >= 2].copy()
            
            # Sort by friction score if available
            if 'Strategic_Friction_Score' in priority_tickets.columns:
                priority_tickets = priority_tickets.sort_values('Strategic_Friction_Score', ascending=False)
            
            # Headers for priority table with resolution time and feedback
            headers = ['ID/Summary', 'Category', 'Similar Count', 'Similarity', 'Consistency', 
                      'Exp. Days', 'Feedback', 'Similar Tickets', 'Recommendation']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_similar.cell(row=13, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
            
            # Populate priority tickets
            row_num = 14
            for idx, ticket in priority_tickets.head(20).iterrows():
                # Get ticket ID or summary snippet
                ticket_id = str(ticket.get('ID', ticket.get(COL_SUMMARY, '')))[:50]
                
                ws_similar.cell(row=row_num, column=1).value = ticket_id
                ws_similar.cell(row=row_num, column=2).value = ticket.get('AI_Category', 'Unknown')
                ws_similar.cell(row=row_num, column=3).value = ticket.get('Similar_Ticket_Count', 0)
                ws_similar.cell(row=row_num, column=4).value = f"{ticket.get('Best_Match_Similarity', 0):.0%}"
                ws_similar.cell(row=row_num, column=5).value = ticket.get('Resolution_Consistency', 'N/A')
                
                # Expected resolution days
                exp_days = ticket.get('Expected_Resolution_Days', None)
                if exp_days and exp_days > 0:
                    ws_similar.cell(row=row_num, column=6).value = f"{exp_days:.1f}"
                else:
                    ws_similar.cell(row=row_num, column=6).value = "N/A"
                
                # Feedback status
                has_feedback = ticket.get('Has_Feedback', False)
                if has_feedback:
                    ws_similar.cell(row=row_num, column=7).value = "✓ Verified"
                    ws_similar.cell(row=row_num, column=7).font = Font(color="28A745")
                else:
                    ws_similar.cell(row=row_num, column=7).value = "Pending"
                    ws_similar.cell(row=row_num, column=7).font = Font(color="FFA500")
                
                ws_similar.cell(row=row_num, column=8).value = ticket.get('Similar_Ticket_IDs', '')
                ws_similar.cell(row=row_num, column=9).value = str(ticket.get('Resolution_Recommendation', ''))[:100]
                
                # Color-code consistency column
                consistency_cell = ws_similar.cell(row=row_num, column=5)
                consistency_val = str(ticket.get('Resolution_Consistency', ''))
                if 'Inconsistent' in consistency_val:
                    consistency_cell.font = Font(color="D9534F", bold=True)
                elif 'consistent' in consistency_val.lower():
                    consistency_cell.font = Font(color="28A745")
                
                row_num += 1
            
            # Add insights section
            insight_row = row_num + 2
            ws_similar[f'A{insight_row}'] = "📊 RESOLUTION PATTERN INSIGHTS"
            ws_similar[f'A{insight_row}'].font = Font(bold=True, size=12, color="004C97")
            ws_similar.merge_cells(f'A{insight_row}:J{insight_row}')
            
            insight_row += 1
            if inconsistent_count > 0:
                ws_similar[f'A{insight_row}'] = f"⚠️ {inconsistent_count} tickets have inconsistent resolutions - these represent opportunities to standardize processes"
                ws_similar[f'A{insight_row}'].font = Font(color="D9534F")
                insight_row += 1
            
            if has_matches > total_tickets * 0.5:
                ws_similar[f'A{insight_row}'] = f"✅ {has_matches/total_tickets*100:.0f}% of tickets have similar historical cases - leverage past resolutions"
                ws_similar[f'A{insight_row}'].font = Font(color="28A745")
                insight_row += 1
            
            if avg_similarity > 0.8:
                ws_similar[f'A{insight_row}'] = f"🎯 High average similarity ({avg_similarity:.0%}) indicates recurring issue patterns - consider preventive measures"
                insight_row += 1
            
            # Resolution time insight
            if avg_resolution_days and avg_resolution_days > 0:
                ws_similar[f'A{insight_row}'] = f"⏱️ Average expected resolution time: {avg_resolution_days:.1f} days based on similar historical tickets"
                ws_similar[f'A{insight_row}'].font = Font(color="004C97")
                insight_row += 1
            
            # Feedback instructions
            insight_row += 1
            ws_similar[f'A{insight_row}'] = "💡 HUMAN FEEDBACK LOOP"
            ws_similar[f'A{insight_row}'].font = Font(bold=True, size=11, color="004C97")
            insight_row += 1
            ws_similar[f'A{insight_row}'] = "To improve similarity matching: 1) Open 'similarity_feedback.xlsx' 2) Mark pairs as 'Correct' or 'Wrong' in Human_Feedback column 3) Re-run analysis"
            ws_similar[f'A{insight_row}'].font = Font(italic=True)
            
            # Column widths for new columns
            ws_similar.column_dimensions['A'].width = 50
            ws_similar.column_dimensions['B'].width = 20
            ws_similar.column_dimensions['C'].width = 12
            ws_similar.column_dimensions['D'].width = 12
            ws_similar.column_dimensions['E'].width = 25
            ws_similar.column_dimensions['F'].width = 12  # Exp. Days
            ws_similar.column_dimensions['G'].width = 12  # Feedback
            ws_similar.column_dimensions['H'].width = 25  # Similar Tickets
            ws_similar.column_dimensions['I'].width = 60  # Recommendation
            
            logger.info(f"  → Similar Tickets sheet added ({has_matches} with matches, {inconsistent_count} inconsistent)")
        
        # ══════════════════════════════════════════════════════════════
        # RESOLUTION TIME ANALYSIS SHEET
        # ══════════════════════════════════════════════════════════════
        if 'Predicted_Resolution_Days' in df.columns or 'Actual_Resolution_Days' in df.columns:
            ws_resolution = wb.create_sheet("Resolution Time")
            ws_resolution['A1'] = "⏱️ RESOLUTION TIME ANALYSIS"
            ws_resolution['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_resolution['A1'].fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
            ws_resolution.merge_cells('A1:H1')
            
            ws_resolution['A2'] = "Comparison of Actual, AI-Predicted, and Human-Expected resolution times"
            ws_resolution['A2'].font = Font(italic=True, size=10, color="666666")
            ws_resolution.merge_cells('A2:H2')
            
            # Summary metrics
            ws_resolution['A4'] = "PREDICTION ACCURACY METRICS"
            ws_resolution['A4'].font = Font(bold=True, size=11, color="FFFFFF")
            ws_resolution['A4'].fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            ws_resolution.merge_cells('A4:D4')
            
            # Calculate metrics
            has_actual = 'Actual_Resolution_Days' in df.columns
            has_predicted = 'Predicted_Resolution_Days' in df.columns
            has_expected = 'Human_Expected_Days' in df.columns
            
            valid_predictions = df.dropna(subset=['Actual_Resolution_Days', 'Predicted_Resolution_Days']) if has_actual and has_predicted else pd.DataFrame()
            
            if len(valid_predictions) > 0:
                actual = valid_predictions['Actual_Resolution_Days'].values
                predicted = valid_predictions['Predicted_Resolution_Days'].values
                mae = np.mean(np.abs(predicted - actual))
                rmse = np.sqrt(np.mean((predicted - actual) ** 2))
                correlation = np.corrcoef(actual, predicted)[0, 1] if len(actual) > 2 else 0
                
                ws_resolution['A5'] = f"Mean Absolute Error (MAE): {mae:.2f} days"
                ws_resolution['A6'] = f"Root Mean Square Error (RMSE): {rmse:.2f} days"
                ws_resolution['A7'] = f"Correlation (R): {correlation:.2f}"
                ws_resolution['A8'] = f"Samples with both Actual & Predicted: {len(valid_predictions)}"
                
                # Color-code MAE
                if mae < 2:
                    ws_resolution['A5'].font = Font(color="28A745", bold=True)  # Good
                elif mae < 5:
                    ws_resolution['A5'].font = Font(color="FFA500", bold=True)  # OK
                else:
                    ws_resolution['A5'].font = Font(color="D9534F", bold=True)  # Needs improvement
            else:
                ws_resolution['A5'] = "Insufficient data for accuracy metrics (need actual resolution dates)"
                ws_resolution['A5'].font = Font(italic=True, color="666666")
            
            # Category breakdown
            row_start = 10
            ws_resolution[f'A{row_start}'] = "RESOLUTION TIME BY CATEGORY"
            ws_resolution[f'A{row_start}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws_resolution[f'A{row_start}'].fill = PatternFill(start_color="004C97", end_color="004C97", fill_type="solid")
            ws_resolution.merge_cells(f'A{row_start}:H{row_start}')
            
            # Headers
            headers = ['Category', 'Count', 'Avg Actual', 'Avg Predicted', 'Human Expected', 'Difference', 'Accuracy']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_resolution.cell(row=row_start + 1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Data rows by category
            if 'AI_Category' in df.columns:
                row_num = row_start + 2
                for cat in df['AI_Category'].dropna().unique()[:15]:  # Top 15 categories
                    cat_df = df[df['AI_Category'] == cat]
                    count = len(cat_df)
                    
                    avg_actual = cat_df['Actual_Resolution_Days'].mean() if has_actual else None
                    avg_predicted = cat_df['Predicted_Resolution_Days'].mean() if has_predicted else None
                    avg_expected = cat_df['Human_Expected_Days'].mean() if has_expected else None
                    
                    # Calculate difference (predicted - actual)
                    diff = None
                    accuracy = "N/A"
                    if avg_actual and avg_predicted and avg_actual > 0:
                        diff = avg_predicted - avg_actual
                        mape = abs(diff) / avg_actual * 100
                        if mape < 20:
                            accuracy = "✅ Excellent"
                        elif mape < 40:
                            accuracy = "🟡 Good"
                        else:
                            accuracy = "⚠️ Needs Calibration"
                    
                    ws_resolution.cell(row=row_num, column=1).value = cat[:30]
                    ws_resolution.cell(row=row_num, column=2).value = count
                    ws_resolution.cell(row=row_num, column=3).value = f"{avg_actual:.1f}" if avg_actual else "N/A"
                    ws_resolution.cell(row=row_num, column=4).value = f"{avg_predicted:.1f}" if avg_predicted else "N/A"
                    ws_resolution.cell(row=row_num, column=5).value = f"{avg_expected:.1f}" if avg_expected else "N/A"
                    ws_resolution.cell(row=row_num, column=6).value = f"{diff:+.1f}" if diff else "N/A"
                    ws_resolution.cell(row=row_num, column=7).value = accuracy
                    
                    # Color-code difference
                    if diff:
                        diff_cell = ws_resolution.cell(row=row_num, column=6)
                        if abs(diff) < 1:
                            diff_cell.font = Font(color="28A745")
                        elif diff > 0:
                            diff_cell.font = Font(color="FFA500")  # Over-predicted
                        else:
                            diff_cell.font = Font(color="D9534F")  # Under-predicted
                    
                    row_num += 1
            
            # Instructions for human input
            insight_row = row_num + 2
            ws_resolution[f'A{insight_row}'] = "💡 HOW TO IMPROVE PREDICTIONS"
            ws_resolution[f'A{insight_row}'].font = Font(bold=True, size=11, color="004C97")
            insight_row += 1
            ws_resolution[f'A{insight_row}'] = "1. Open 'similarity_feedback.xlsx'"
            insight_row += 1
            ws_resolution[f'A{insight_row}'] = "2. Enter YOUR expected resolution times in 'Expected_Resolution_Days' column"
            insight_row += 1
            ws_resolution[f'A{insight_row}'] = "3. Re-run analysis - AI will calibrate predictions to human expectations!"
            insight_row += 1
            ws_resolution[f'A{insight_row}'] = "Your input helps identify if AI is over/under-estimating by category"
            ws_resolution[f'A{insight_row}'].font = Font(italic=True, color="666666")
            
            # Column widths
            ws_resolution.column_dimensions['A'].width = 30
            ws_resolution.column_dimensions['B'].width = 10
            ws_resolution.column_dimensions['C'].width = 15
            ws_resolution.column_dimensions['D'].width = 15
            ws_resolution.column_dimensions['E'].width = 18
            ws_resolution.column_dimensions['F'].width = 12
            ws_resolution.column_dimensions['G'].width = 18
            
            logger.info("  → Resolution Time sheet added")
        
        # AI CONTEXT SHEET - For audit trail
        ws_context = wb.create_sheet("AI Context")
        ws_context['A1'] = "DATA CONTEXT PROVIDED TO AI"
        ws_context['A1'].font = Font(bold=True, size=14)
        
        # Split context into chunks to avoid Excel 32K character limit (Err:510)
        MAX_CELL_CHARS = 30000
        context_lines = context.split('\n')
        current_chunk = []
        current_length = 0
        row_num = 3
        
        for line in context_lines:
            line_len = len(line) + 1  # +1 for newline
            if current_length + line_len > MAX_CELL_CHARS:
                # Write current chunk and start new one
                ws_context[f'A{row_num}'] = '\n'.join(current_chunk)
                ws_context[f'A{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
                row_num += 1
                current_chunk = [line]
                current_length = line_len
            else:
                current_chunk.append(line)
                current_length += line_len
        
        # Write remaining chunk
        if current_chunk:
            ws_context[f'A{row_num}'] = '\n'.join(current_chunk)
            ws_context[f'A{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        ws_context.column_dimensions['A'].width = 100
        
        # Remove default "Sheet" if it still exists
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Save feedback file for reinforcement learning
        output_dir = os.path.dirname(save_path)
        feedback_path = feedback_learner.save_for_review(df, output_dir)
       
        wb.save(save_path)
        ai.unload()
        logger.info(f"Report saved to: {save_path}")
        
        # Show completion with feedback instructions
        feedback_msg = (
            f"Strategic Report Generated!\n\n"
            f"Saved to:\n{save_path}\n\n"
            f"📝 FEEDBACK FOR AI IMPROVEMENT:\n"
            f"A feedback Excel file has been created at:\n"
            f"{feedback_path}\n\n"
            f"To improve AI classification:\n"
            f"1. Open the Excel file\n"
            f"2. Go to 'Classifications' sheet\n"
            f"3. Yellow rows = low confidence (review these first)\n"
            f"4. Fill in 'Corrected_Category' column for wrong ones\n"
            f"5. Save and re-run - AI will learn!"
        )
        messagebox.showinfo("Complete", feedback_msg)
        
    finally:
        # Ensure Tkinter root is destroyed
        root.destroy()

if __name__ == "__main__":
    main_pipeline()