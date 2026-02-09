import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from scipy.interpolate import make_interp_spline
import openpyxl

# ═══════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="CSE Unit – Pulse Tracker",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ═══════════════════════════════════════════════════════════════════
# THEME / CSS
# ═══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap');
    
    /* Global dark theme */
    .stApp { 
        background: linear-gradient(135deg, #0a0f1a 0%, #0d1526 50%, #0a1628 100%);
        color: #E2E8F0; 
        font-family: 'DM Sans', sans-serif; 
    }
    
    /* Header bar with gradient background */
    .header-container {
        background: linear-gradient(90deg, #0c1929 0%, #162a4a 50%, #1a3a5c 100%);
        border-radius: 8px;
        padding: 0.6rem 1.5rem;
        margin-bottom: 0.8rem;
        display: flex;
        justify-content: space-between;
        align-items: center;
        border: 1px solid #1e3a5f;
        box-shadow: 0 4px 20px rgba(0,0,0,0.3);
    }
    .header-left {
        display: flex;
        align-items: center;
        gap: 1.5rem;
    }
    .logo-section {
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .logo-icon {
        width: 45px;
        height: 45px;
        background: linear-gradient(135deg, #3b82f6 0%, #8b5cf6 100%);
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.2rem;
    }
    .logo-text {
        font-size: 1.4rem;
        font-weight: 700;
        color: #60a5fa;
        letter-spacing: -0.5px;
    }
    .logo-subtext {
        font-size: 0.7rem;
        color: #94a3b8;
        letter-spacing: 2px;
        text-transform: uppercase;
    }
    .filter-row {
        display: flex;
        gap: 1rem;
        align-items: center;
    }
    .filter-box {
        background: #0f172a;
        border: 1px solid #334155;
        border-radius: 4px;
        padding: 0.3rem 0.8rem;
        min-width: 80px;
    }
    .filter-label {
        font-size: 0.65rem;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .filter-value {
        font-size: 0.85rem;
        color: #e2e8f0;
        font-weight: 500;
    }
    .header-right {
        display: flex;
        align-items: center;
        gap: 1rem;
    }
    .brand-text {
        text-align: right;
    }
    .brand-amdocs {
        font-size: 0.75rem;
        color: #94a3b8;
        letter-spacing: 1px;
    }
    .brand-tagline {
        font-size: 1.1rem;
        font-weight: 700;
        background: linear-gradient(90deg, #f472b6, #fb923c, #facc15);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        letter-spacing: 1px;
    }
    
    /* Section titles */
    .section-title {
        font-size: 0.85rem;
        font-weight: 600;
        color: #94a3b8;
        margin: 0.8rem 0 0.4rem 0;
        padding-bottom: 0.3rem;
        border-bottom: 2px solid #1e3a5f;
    }
    
    /* Matrix table styling */
    .matrix-container {
        background: #0d1526;
        border: 1px solid #1e3a5f;
        border-radius: 8px;
        overflow: hidden;
    }
    .matrix-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.75rem;
    }
    .matrix-table th {
        background: #162a4a;
        color: #94a3b8;
        font-weight: 600;
        padding: 0.5rem 0.4rem;
        text-align: center;
        border-bottom: 2px solid #2563eb;
        font-size: 0.7rem;
        text-transform: uppercase;
        letter-spacing: 0.3px;
    }
    .matrix-table td {
        padding: 0.35rem 0.4rem;
        border-bottom: 1px solid #1e293b;
        text-align: center;
        color: #cbd5e1;
    }
    .matrix-table tr:hover td {
        background: rgba(37, 99, 235, 0.1);
    }
    .matrix-table .region-row td {
        background: #111d32;
        font-weight: 600;
    }
    .matrix-table .area-row td {
        background: transparent;
        padding-left: 1.5rem;
    }
    .matrix-table .total-row td {
        background: #1a365d;
        font-weight: 700;
        border-top: 2px solid #2563eb;
    }
    .matrix-table .expand-icon {
        cursor: pointer;
        color: #60a5fa;
        margin-right: 0.3rem;
    }
    .matrix-table .region-name {
        text-align: left !important;
        font-weight: 600;
        color: #e2e8f0;
    }
    .matrix-table .area-name {
        text-align: left !important;
        color: #94a3b8;
    }
    
    /* Sparkline cell */
    .sparkline-cell {
        width: 80px;
        height: 25px;
    }
    
    /* Score cells with conditional formatting */
    .score-cell {
        font-weight: 600;
        padding: 0.2rem 0.4rem;
        border-radius: 3px;
        min-width: 35px;
        display: inline-block;
    }
    .pulse-green { background: #065f46; color: #6ee7b7; }
    .pulse-yellow { background: #78350f; color: #fcd34d; }
    .pulse-red { background: #7f1d1d; color: #fca5a5; }
    .pulse-blue { background: #1e3a5f; color: #93c5fd; }
    
    .score-high { background: rgba(34, 197, 94, 0.2); color: #86efac; }
    .score-mid { background: rgba(59, 130, 246, 0.15); color: #93c5fd; }
    .score-low { background: rgba(251, 191, 36, 0.2); color: #fcd34d; }
    .score-critical { background: rgba(239, 68, 68, 0.2); color: #fca5a5; }
    
    /* Weekly trend heatmap */
    .heatmap-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.75rem;
    }
    .heatmap-table th {
        background: #162a4a;
        color: #94a3b8;
        font-weight: 600;
        padding: 0.4rem 0.3rem;
        text-align: center;
        font-size: 0.7rem;
        border-bottom: 2px solid #2563eb;
    }
    .heatmap-table td {
        padding: 0.3rem;
        text-align: center;
        font-weight: 600;
        font-size: 0.7rem;
        border-bottom: 1px solid #1e293b;
    }
    .heatmap-table .region-col {
        text-align: left;
        padding-left: 0.5rem;
        color: #e2e8f0;
        min-width: 60px;
    }
    .heatmap-cell {
        border-radius: 2px;
        padding: 0.2rem 0.3rem;
    }
    .heat-green { background: #166534; color: #bbf7d0; }
    .heat-lime { background: #3f6212; color: #d9f99d; }
    .heat-yellow { background: #854d0e; color: #fef08a; }
    .heat-orange { background: #9a3412; color: #fed7aa; }
    .heat-red { background: #991b1b; color: #fecaca; }
    
    /* Legend tables */
    .legend-container {
        display: flex;
        gap: 1rem;
        flex-wrap: wrap;
        margin-top: 0.5rem;
    }
    .legend-table {
        background: #0d1526;
        border: 1px solid #1e3a5f;
        border-radius: 6px;
        overflow: hidden;
        font-size: 0.7rem;
        flex: 1;
        min-width: 180px;
    }
    .legend-table th {
        background: #162a4a;
        color: #94a3b8;
        padding: 0.4rem 0.5rem;
        text-align: left;
        font-weight: 600;
        border-bottom: 1px solid #2563eb;
    }
    .legend-table td {
        padding: 0.3rem 0.5rem;
        border-bottom: 1px solid #1e293b;
        color: #cbd5e1;
    }
    .legend-table .rating-badge {
        display: inline-block;
        width: 18px;
        height: 18px;
        border-radius: 3px;
        text-align: center;
        line-height: 18px;
        font-weight: 700;
        font-size: 0.65rem;
        margin-right: 0.4rem;
    }
    .rating-0 { background: #dc2626; color: white; }
    .rating-1 { background: #f97316; color: white; }
    .rating-2 { background: #3b82f6; color: white; }
    .rating-3 { background: #22c55e; color: white; }
    
    /* Notes box */
    .notes-box {
        background: #0d1526;
        border: 1px solid #1e3a5f;
        border-radius: 6px;
        padding: 0.6rem 0.8rem;
        font-size: 0.75rem;
        color: #94a3b8;
        line-height: 1.4;
    }
    .notes-title {
        font-weight: 600;
        color: #e2e8f0;
        margin-bottom: 0.3rem;
    }
    
    /* Drill-down panel */
    .drilldown-panel {
        background: linear-gradient(180deg, #0f1d32 0%, #0d1526 100%);
        border: 1px solid #2563eb;
        border-radius: 8px;
        padding: 1rem;
        margin-top: 0.5rem;
    }
    .drilldown-header {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        margin-bottom: 0.8rem;
        padding-bottom: 0.5rem;
        border-bottom: 1px solid #1e3a5f;
    }
    .drilldown-badge {
        background: linear-gradient(135deg, #2563eb 0%, #7c3aed 100%);
        color: white;
        padding: 0.25rem 0.75rem;
        border-radius: 4px;
        font-weight: 600;
        font-size: 0.8rem;
    }
    .drilldown-context {
        color: #64748b;
        font-size: 0.8rem;
    }
    
    /* Hide default streamlit elements */
    #MainMenu, footer, header { visibility: hidden; }
    .block-container { padding: 0.5rem 1rem 1rem 1rem; max-width: 100%; }
    
    /* Custom scrollbar */
    ::-webkit-scrollbar { width: 6px; height: 6px; }
    ::-webkit-scrollbar-track { background: #0d1526; }
    ::-webkit-scrollbar-thumb { background: #334155; border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: #475569; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════
@st.cache_data
def load_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Project Pulse']
    data = [row[:27] for row in ws.iter_rows(min_row=2, values_only=True)]
    col_names = [c.value for c in ws[1][:27]]
    df = pd.DataFrame(data, columns=col_names)
    df['Wk'] = pd.to_numeric(df['Wk'], errors='coerce')
    df = df.dropna(subset=['Wk'])
    df['Wk'] = df['Wk'].astype(int)
    df = df.rename(columns={'PM Performance': 'PM Perf'})
    score_cols = ['Design', 'IX', 'PAG', 'RF Opt', 'Field', 'CSAT', 'PM Perf', 'Potential']
    for c in score_cols:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    df['Total Score'] = pd.to_numeric(df['Total Score'], errors='coerce')
    return df


SCORE_COLS = ['Design', 'IX', 'PAG', 'RF Opt', 'Field', 'CSAT', 'PM Perf', 'Potential']

COLORS = {
    'Design':    '#1B5E7B',
    'IX':        '#7B2D8E',
    'PAG':       '#00BCD4',
    'RF Opt':    '#8D6E63',
    'Field':     '#FFC107',
    'CSAT':      '#26A69A',
    'PM Perf':   '#9E9E9E',
    'Potential': '#E040FB',
}

# Load data
DATA_PATH = "Pulse_Tracker.xlsx"
try:
    df = load_data(DATA_PATH)
except:
    DATA_PATH = "/mnt/user-data/uploads/Pulse_Tracker.xlsx"
    df = load_data(DATA_PATH)

all_weeks = sorted(df['Wk'].unique())
all_regions = sorted(df['Region'].dropna().unique())


# ═══════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════
if 'expanded_regions' not in st.session_state:
    st.session_state.expanded_regions = set(all_regions)
if 'selected_drill' not in st.session_state:
    st.session_state.selected_drill = None  # {'type': 'region'|'area'|'dimension', 'value': ..., 'week': ...}


# ═══════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════
# Filters in columns
col_head1, col_head2, col_head3, col_head4, col_head5 = st.columns([1.5, 0.8, 0.8, 0.8, 1.5])

with col_head1:
    st.markdown("""
    <div style="display:flex; align-items:center; gap:0.8rem;">
        <div style="font-size:2rem;">🏗️</div>
        <div>
            <div style="font-size:1.3rem; font-weight:700; color:#60a5fa;">CSE UNIT</div>
            <div style="font-size:0.65rem; color:#64748b; letter-spacing:1px;">PROJECT PULSE TRACKER</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_head2:
    month_filter = st.selectbox("Month", ["All"], key="month_filter", label_visibility="visible")

with col_head3:
    week_range = st.select_slider("Week", options=all_weeks, value=(all_weeks[0], all_weeks[-1]), key="week_range")

with col_head4:
    region_filter = st.multiselect("Region", ["All"] + list(all_regions), default=["All"], key="region_filter")

with col_head5:
    st.markdown("""
    <div style="text-align:right; padding-top:0.5rem;">
        <div style="font-size:0.7rem; color:#64748b;">amdocs</div>
        <div style="font-size:1rem; font-weight:700; background:linear-gradient(90deg,#f472b6,#fb923c,#facc15);
             -webkit-background-clip:text; -webkit-text-fill-color:transparent;">make it amazing</div>
    </div>
    """, unsafe_allow_html=True)

# Apply filters
if "All" in region_filter or len(region_filter) == 0:
    selected_regions = all_regions
else:
    selected_regions = [r for r in region_filter if r != "All"]

mask = (
    (df['Wk'] >= week_range[0]) &
    (df['Wk'] <= week_range[1]) &
    (df['Region'].isin(selected_regions))
)
fdf = df[mask].copy()


# ═══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════
def pulse_class(score):
    """Return CSS class for pulse score."""
    if pd.isna(score): return ""
    s = float(score)
    if s >= 17: return "pulse-green"
    elif s >= 14: return "pulse-yellow"
    else: return "pulse-red"

def heat_class(score):
    """Return CSS class for heatmap cell."""
    if pd.isna(score): return ""
    s = float(score)
    if s >= 18: return "heat-green"
    elif s >= 17: return "heat-lime"
    elif s >= 16: return "heat-yellow"
    elif s >= 14: return "heat-orange"
    else: return "heat-red"

def score_class(score, col=None):
    """Return CSS class for individual score cell."""
    if pd.isna(score): return ""
    s = float(score)
    if s >= 2.5: return "score-high"
    elif s >= 2.0: return "score-mid"
    elif s >= 1.5: return "score-low"
    else: return "score-critical"

def mini_sparkline_svg(values, width=70, height=20, color="#60a5fa"):
    """Generate inline SVG sparkline."""
    if len(values) < 2:
        return ""
    vals = [v for v in values if pd.notna(v)]
    if len(vals) < 2:
        return ""
    mn, mx = min(vals), max(vals)
    rng = mx - mn if mx != mn else 1
    
    points = []
    for i, v in enumerate(vals):
        x = (i / (len(vals) - 1)) * (width - 4) + 2
        y = height - 3 - ((v - mn) / rng) * (height - 6)
        points.append(f"{x},{y}")
    
    polyline = " ".join(points)
    svg = f'''<svg width="{width}" height="{height}" style="display:block;">
        <polyline points="{polyline}" fill="none" stroke="{color}" stroke-width="1.5" stroke-linecap="round"/>
        <circle cx="{points[-1].split(',')[0]}" cy="{points[-1].split(',')[1]}" r="2" fill="{color}"/>
    </svg>'''
    return svg


# ═══════════════════════════════════════════════════════════════════
# PULSE RANKING CHART
# ═══════════════════════════════════════════════════════════════════
st.markdown('<div class="section-title">Pulse Ranking</div>', unsafe_allow_html=True)

weekly = fdf.groupby('Wk')[SCORE_COLS].mean().reset_index().sort_values('Wk')
weeks = weekly['Wk'].values

if len(weeks) >= 4:
    # Build ranked stacks
    week_stacks = []
    for _, row in weekly.iterrows():
        vals = {c: round(row[c], 2) for c in SCORE_COLS}
        sorted_dims = sorted(vals.items(), key=lambda x: x[1])
        cum_y = 0
        stack = {}
        for dim, val in sorted_dims:
            stack[dim] = {'y0': cum_y, 'y1': cum_y + val, 'val': val}
            cum_y += val
        week_stacks.append(stack)

    x_smooth = np.linspace(weeks[0], weeks[-1], 300)
    fig_pulse = go.Figure()

    avg_vals = {c: weekly[c].mean() for c in SCORE_COLS}
    draw_order = sorted(avg_vals, key=lambda x: avg_vals[x])

    for dim in draw_order:
        y0_pts = [week_stacks[i][dim]['y0'] for i in range(len(weeks))]
        y1_pts = [week_stacks[i][dim]['y1'] for i in range(len(weeks))]
        val_pts = [week_stacks[i][dim]['val'] for i in range(len(weeks))]

        k = min(3, len(weeks) - 1)
        spl_y0 = make_interp_spline(weeks, y0_pts, k=k)
        spl_y1 = make_interp_spline(weeks, y1_pts, k=k)
        y0_sm = spl_y0(x_smooth)
        y1_sm = spl_y1(x_smooth)

        fig_pulse.add_trace(go.Scatter(
            x=np.concatenate([x_smooth, x_smooth[::-1]]),
            y=np.concatenate([y1_sm, y0_sm[::-1]]),
            fill='toself',
            fillcolor=COLORS[dim],
            line=dict(color='rgba(255,255,255,0.25)', width=0.5),
            name=dim,
            showlegend=True,
            hoverinfo='skip',
        ))

        # Clickable markers
        mid_pts = [(y0_pts[i] + y1_pts[i]) / 2 for i in range(len(weeks))]
        fig_pulse.add_trace(go.Scatter(
            x=list(weeks),
            y=mid_pts,
            mode='markers+text',
            marker=dict(size=22, color=COLORS[dim], opacity=0.01, symbol='square'),
            text=[f"{v:.2f}" for v in val_pts],
            textfont=dict(size=8, color='white', family='DM Sans'),
            textposition='middle center',
            name=dim,
            showlegend=False,
            customdata=[[dim, int(w), v] for w, v in zip(weeks, val_pts)],
            hovertemplate=f"<b>{dim}</b><br>Week %{{customdata[1]}}<br>Avg: %{{customdata[2]:.2f}}<extra></extra>",
        ))

    fig_pulse.update_layout(
        xaxis=dict(
            title='WeekNum',
            tickmode='array',
            tickvals=[int(w) for w in weeks],
            ticktext=[str(int(w)) for w in weeks],
            range=[weeks[0] - 0.5, weeks[-1] + 0.5],
            gridcolor='rgba(255,255,255,0.03)',
            color='#64748b',
            tickfont=dict(size=9),
        ),
        yaxis=dict(
            range=[0, 20],
            dtick=10,
            gridcolor='rgba(255,255,255,0.05)',
            griddash='dot',
            color='#64748b',
            tickfont=dict(size=9),
        ),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(family='DM Sans', color='#94A3B8'),
        legend=dict(
            orientation='h',
            yanchor='bottom',
            y=1.02,
            xanchor='left',
            x=0,
            font=dict(size=10, color='#CBD5E1'),
        ),
        height=280,
        margin=dict(l=35, r=15, t=35, b=45),
        hovermode='closest',
    )

    event = st.plotly_chart(fig_pulse, use_container_width=True, on_select="rerun", key="pulse_chart")

    # Handle chart click
    if event and event.selection and event.selection.points:
        pt = event.selection.points[0]
        if 'customdata' in pt and pt['customdata']:
            st.session_state.selected_drill = {
                'type': 'dimension',
                'value': pt['customdata'][0],
                'week': pt['customdata'][1]
            }
else:
    st.warning("Not enough data for the selected filters")


# ═══════════════════════════════════════════════════════════════════
# MAIN TABLES SECTION
# ═══════════════════════════════════════════════════════════════════
col_left, col_right = st.columns([2.2, 1])

# ─────────────────────────────────────────────────────────────────
# PROJECT PULSE - REGION | AREA TABLE
# ─────────────────────────────────────────────────────────────────
with col_left:
    st.markdown('<div class="section-title">Project Pulse – Region | Area</div>', unsafe_allow_html=True)
    
    # Build hierarchical data
    region_data = []
    for region in selected_regions:
        rdf = fdf[fdf['Region'] == region]
        if len(rdf) == 0:
            continue
        
        # Region-level aggregates
        region_trend = rdf.groupby('Wk')['Total Score'].mean().sort_index().values
        region_avg = rdf['Total Score'].mean()
        region_scores = {c: rdf[c].mean() for c in SCORE_COLS}
        
        region_data.append({
            'level': 'region',
            'name': region,
            'trend': region_trend,
            'pulse': region_avg,
            **region_scores
        })
        
        # Area-level (if region is expanded)
        if region in st.session_state.expanded_regions:
            areas = rdf['Area'].dropna().unique()
            for area in sorted(areas):
                adf = rdf[rdf['Area'] == area]
                if len(adf) == 0:
                    continue
                area_trend = adf.groupby('Wk')['Total Score'].mean().sort_index().values
                area_avg = adf['Total Score'].mean()
                area_scores = {c: adf[c].mean() for c in SCORE_COLS}
                
                region_data.append({
                    'level': 'area',
                    'name': area,
                    'parent': region,
                    'trend': area_trend,
                    'pulse': area_avg,
                    **area_scores
                })
    
    # Add total row
    total_trend = fdf.groupby('Wk')['Total Score'].mean().sort_index().values
    total_avg = fdf['Total Score'].mean()
    total_scores = {c: fdf[c].mean() for c in SCORE_COLS}
    region_data.append({
        'level': 'total',
        'name': 'Total',
        'trend': total_trend,
        'pulse': total_avg,
        **total_scores
    })
    
    # Build HTML table
    table_html = '''<div class="matrix-container"><table class="matrix-table">
    <thead><tr>
        <th style="text-align:left; min-width:100px;">Region</th>
        <th style="min-width:75px;">Pulse Trend</th>
        <th>Project Pulse</th>
        <th>Design</th>
        <th>IX</th>
        <th>PAG</th>
        <th>RF Opt</th>
        <th>Field</th>
        <th>CSAT</th>
        <th>PM Perf</th>
        <th>Potential</th>
    </tr></thead><tbody>'''
    
    for row in region_data:
        level = row['level']
        row_class = f"{level}-row"
        
        if level == 'region':
            is_expanded = row['name'] in st.session_state.expanded_regions
            icon = "⊟" if is_expanded else "⊞"
            name_cell = f'<td class="region-name"><span class="expand-icon">{icon}</span>{row["name"]}</td>'
        elif level == 'area':
            name_cell = f'<td class="area-name">{row["name"]}</td>'
        else:
            name_cell = f'<td class="region-name" style="color:#60a5fa;">{row["name"]}</td>'
        
        sparkline = mini_sparkline_svg(row['trend'], color="#f472b6" if row['pulse'] < 16 else "#60a5fa")
        
        pulse_cls = pulse_class(row['pulse'])
        pulse_val = f"{row['pulse']:.2f}" if pd.notna(row['pulse']) else "—"
        
        score_cells = ""
        for col in SCORE_COLS:
            val = row.get(col)
            cls = score_class(val)
            display = f"{val:.2f}" if pd.notna(val) else "—"
            score_cells += f'<td><span class="score-cell {cls}">{display}</span></td>'
        
        table_html += f'''<tr class="{row_class}">
            {name_cell}
            <td>{sparkline}</td>
            <td><span class="score-cell {pulse_cls}">{pulse_val}</span></td>
            {score_cells}
        </tr>'''
    
    table_html += '</tbody></table></div>'
    
    st.markdown(table_html, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
# WEEKLY TREND HEATMAP
# ─────────────────────────────────────────────────────────────────
with col_right:
    st.markdown('<div class="section-title">Project Pulse – Weekly Trend</div>', unsafe_allow_html=True)
    
    # Get weeks to display (last 7-8 weeks that fit)
    display_weeks = sorted(fdf['Wk'].unique())[-8:]
    
    # Build heatmap data
    heat_data = []
    for region in selected_regions:
        rdf = fdf[fdf['Region'] == region]
        row = {'Region': region}
        for wk in display_weeks:
            wdf = rdf[rdf['Wk'] == wk]
            row[wk] = wdf['Total Score'].mean() if len(wdf) > 0 else None
        heat_data.append(row)
    
    # Total row
    total_row = {'Region': 'Total'}
    for wk in display_weeks:
        wdf = fdf[fdf['Wk'] == wk]
        total_row[wk] = wdf['Total Score'].mean() if len(wdf) > 0 else None
    heat_data.append(total_row)
    
    # Build heatmap table
    heat_html = '<div class="matrix-container"><table class="heatmap-table"><thead><tr>'
    heat_html += '<th style="text-align:left;">Region</th>'
    for wk in display_weeks:
        heat_html += f'<th>{int(wk)}</th>'
    heat_html += '</tr></thead><tbody>'
    
    region_colors = {'Central': '🟦', 'NE': '🟩', 'South': '🟨', 'West': '🟪', 'Total': '⬜'}
    
    for row in heat_data:
        is_total = row['Region'] == 'Total'
        row_class = "total-row" if is_total else ""
        icon = region_colors.get(row['Region'], '')
        
        heat_html += f'<tr class="{row_class}">'
        heat_html += f'<td class="region-col">{icon} {row["Region"]}</td>'
        
        for wk in display_weeks:
            val = row.get(wk)
            if pd.notna(val):
                cls = heat_class(val)
                heat_html += f'<td><span class="heatmap-cell {cls}">{val:.2f}</span></td>'
            else:
                heat_html += '<td>—</td>'
        
        heat_html += '</tr>'
    
    heat_html += '</tbody></table></div>'
    st.markdown(heat_html, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# LEGEND SECTION
# ═══════════════════════════════════════════════════════════════════
st.markdown('<div class="section-title" style="margin-top:0.8rem;">Ratings Reference</div>', unsafe_allow_html=True)

leg_col1, leg_col2, leg_col3, leg_col4, leg_col5, leg_col6 = st.columns([1.2, 1, 1, 1.2, 1, 1.5])

with leg_col1:
    st.markdown('''
    <table class="legend-table">
        <tr><th colspan="2">Ratings – LOB</th></tr>
        <tr><td><span class="rating-badge rating-0">0</span>Escalation</td></tr>
        <tr><td><span class="rating-badge rating-1">1</span>Complaint/Concern</td></tr>
        <tr><td><span class="rating-badge rating-2">2</span>Constructive / BAU / NA</td></tr>
        <tr><td><span class="rating-badge rating-3">3</span>Appreciation</td></tr>
    </table>
    ''', unsafe_allow_html=True)

with leg_col2:
    st.markdown('''
    <table class="legend-table">
        <tr><th>CSAT</th></tr>
        <tr><td>Escalation</td></tr>
        <tr><td>Complaint</td></tr>
        <tr><td>Mixed</td></tr>
        <tr><td>Positive</td></tr>
    </table>
    ''', unsafe_allow_html=True)

with leg_col3:
    st.markdown('''
    <table class="legend-table">
        <tr><th>PM Perf</th></tr>
        <tr><td>Escalation</td></tr>
        <tr><td>Issues</td></tr>
        <tr><td>On-time / Good</td></tr>
        <tr><td>Exceptional</td></tr>
    </table>
    ''', unsafe_allow_html=True)

with leg_col4:
    st.markdown('''
    <table class="legend-table">
        <tr><th>Potential</th></tr>
        <tr><td>Declining / At Risk</td></tr>
        <tr><td>Stagnant</td></tr>
        <tr><td>Moderate Opportunity</td></tr>
        <tr><td>Strong Future ROI</td></tr>
    </table>
    ''', unsafe_allow_html=True)

with leg_col5:
    st.markdown('''
    <table class="legend-table">
        <tr><th colspan="2">Project Pulse</th></tr>
        <tr><td><span class="score-cell pulse-red" style="margin-right:0.4rem;">1-13</span></td></tr>
        <tr><td><span class="score-cell pulse-yellow" style="margin-right:0.4rem;">14-16</span></td></tr>
        <tr><td><span class="score-cell" style="background:#1e4d3a;color:#6ee7b7;margin-right:0.4rem;">16-20</span></td></tr>
        <tr><td><span class="score-cell pulse-green" style="margin-right:0.4rem;">20-24</span></td></tr>
    </table>
    ''', unsafe_allow_html=True)

with leg_col6:
    st.markdown('''
    <div class="notes-box">
        <div class="notes-title">Notes</div>
        Pulse Ranking – Ranking of each contributor in Project Pulse per week.
        Click any band in the chart or cell in the tables to drill down.
    </div>
    ''', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# DRILL-DOWN PANEL
# ═══════════════════════════════════════════════════════════════════
if st.session_state.selected_drill:
    drill = st.session_state.selected_drill
    
    st.markdown('<div class="drilldown-panel">', unsafe_allow_html=True)
    
    if drill['type'] == 'dimension':
        dim = drill['value']
        wk = drill.get('week')
        dim_color = COLORS.get(dim, '#2563eb')
        
        st.markdown(f'''
        <div class="drilldown-header">
            <span class="drilldown-badge" style="background:{dim_color};">{dim}</span>
            <span class="drilldown-context">Drill-down • Week {wk if wk else 'All'} • All regions</span>
            <span style="margin-left:auto; color:#64748b; cursor:pointer;" onclick="window.location.reload();">✕ Close</span>
        </div>
        ''', unsafe_allow_html=True)
        
        drill_df = fdf.copy()
        if wk:
            drill_df = drill_df[drill_df['Wk'] == wk]
        
        # Summary by region
        col_d1, col_d2 = st.columns([1, 1.5])
        
        with col_d1:
            st.markdown(f"**{dim} – By Region**")
            region_summary = drill_df.groupby('Region').agg(
                Projects=('Project', 'count'),
                Avg=(dim, 'mean'),
            ).reset_index().sort_values('Avg')
            
            fig_bar = go.Figure(go.Bar(
                x=region_summary['Region'],
                y=region_summary['Avg'],
                marker_color=dim_color,
                text=[f"{v:.2f}" for v in region_summary['Avg']],
                textposition='outside',
                textfont=dict(color='#cbd5e1', size=11),
            ))
            fig_bar.update_layout(
                height=200,
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94a3b8', size=10),
                xaxis=dict(color='#64748b', gridcolor='rgba(255,255,255,0.03)'),
                yaxis=dict(range=[0, 3.2], color='#64748b', gridcolor='rgba(255,255,255,0.05)', griddash='dot'),
                margin=dict(l=30, r=15, t=10, b=30),
            )
            st.plotly_chart(fig_bar, use_container_width=True)
        
        with col_d2:
            st.markdown(f"**{dim} – Score Distribution**")
            dist = drill_df[dim].value_counts().sort_index()
            dist_colors = {0: '#ef4444', 1: '#f97316', 2: '#3b82f6', 3: '#22c55e'}
            
            fig_dist = go.Figure(go.Bar(
                x=[f"Score {int(s)}" for s in dist.index],
                y=dist.values,
                marker_color=[dist_colors.get(int(s), '#64748b') for s in dist.index],
                text=dist.values,
                textposition='outside',
                textfont=dict(color='#cbd5e1', size=11),
            ))
            fig_dist.update_layout(
                height=200,
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94a3b8', size=10),
                xaxis=dict(color='#64748b'),
                yaxis=dict(color='#64748b', gridcolor='rgba(255,255,255,0.05)', griddash='dot'),
                margin=dict(l=30, r=15, t=10, b=30),
            )
            st.plotly_chart(fig_dist, use_container_width=True)
        
        # Project detail table
        st.markdown(f"**Project Details – {dim}**")
        detail = drill_df[['Wk', 'Region', 'Area', 'Project', 'PM Name', dim, 'Total Score']].copy()
        detail = detail.sort_values([dim, 'Total Score'], ascending=[True, True]).head(50)
        
        detail_html = '<div class="matrix-container" style="max-height:250px; overflow-y:auto;"><table class="matrix-table">'
        detail_html += f'<thead><tr><th>Wk</th><th>Region</th><th>Area</th><th style="text-align:left;">Project</th><th>PM</th><th>{dim}</th><th>Pulse</th></tr></thead><tbody>'
        
        for _, r in detail.iterrows():
            dim_val = r[dim]
            pulse_val = r['Total Score']
            dim_cls = f"rating-{int(dim_val)}" if pd.notna(dim_val) else ""
            pulse_cls = pulse_class(pulse_val)
            project_name = str(r['Project'])[:40] if pd.notna(r['Project']) else '—'
            pm_name = str(r['PM Name'])[:15] if pd.notna(r['PM Name']) else '—'
            
            detail_html += f'''<tr>
                <td>{int(r['Wk'])}</td>
                <td>{r['Region']}</td>
                <td>{r.get('Area', '—')}</td>
                <td style="text-align:left;" title="{r['Project']}">{project_name}</td>
                <td>{pm_name}</td>
                <td><span class="rating-badge {dim_cls}">{int(dim_val) if pd.notna(dim_val) else '—'}</span></td>
                <td><span class="score-cell {pulse_cls}">{int(pulse_val) if pd.notna(pulse_val) else '—'}</span></td>
            </tr>'''
        
        detail_html += '</tbody></table></div>'
        st.markdown(detail_html, unsafe_allow_html=True)
        
        # Weekly trend by region
        st.markdown(f"**{dim} – Weekly Trend by Region**")
        trend = fdf.groupby(['Wk', 'Region'])[dim].mean().reset_index()
        fig_trend = go.Figure()
        region_colors_line = {'Central': '#60A5FA', 'NE': '#34D399', 'South': '#FBBF24', 'West': '#F472B6'}
        for region in selected_regions:
            rd = trend[trend['Region'] == region].sort_values('Wk')
            if len(rd) > 0:
                fig_trend.add_trace(go.Scatter(
                    x=rd['Wk'], y=rd[dim],
                    mode='lines+markers',
                    name=region,
                    line=dict(color=region_colors_line.get(region, '#94a3b8'), width=2),
                    marker=dict(size=4),
                ))
        fig_trend.update_layout(
            height=200,
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#94a3b8', size=10),
            xaxis=dict(title='Week', color='#64748b', gridcolor='rgba(255,255,255,0.03)'),
            yaxis=dict(range=[1.5, 2.5], color='#64748b', gridcolor='rgba(255,255,255,0.05)', griddash='dot'),
            legend=dict(orientation='h', yanchor='bottom', y=1.02, font=dict(color='#cbd5e1', size=9)),
            margin=dict(l=30, r=15, t=30, b=40),
        )
        st.plotly_chart(fig_trend, use_container_width=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Clear drill-down button
    if st.button("✕ Close Drill-Down", key="close_drill"):
        st.session_state.selected_drill = None
        st.rerun()


# ═══════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════
st.markdown('''
<div style="text-align:center; padding:1rem; color:#334155; font-size:0.65rem; margin-top:1rem; border-top:1px solid #1e293b;">
    CSE Unit • Pulse Tracker • Powered by Streamlit + Plotly
</div>
''', unsafe_allow_html=True)
